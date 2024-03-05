#!/usr/bin/python3
#
#  Copyright (C) 2018-2023 Sustainable Energy Now Inc., Angus King
#
#  powermatch.py - This file is part of SIREN.
#
#  SIREN is free software: you can redistribute it and/or modify
#  it under the terms of the GNU Affero General Public License as
#  published by the Free Software Foundation, either version 3 of
#  the License, or (at your option) any later version.
#
#  SIREN is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU Affero General Public License for more details.
#
#  You should have received a copy of the GNU Affero General
#  Public License along with SIREN.  If not, see
#  <http://www.gnu.org/licenses/>.
#
# Note: Batch process is all rather messy.
from decimal import Decimal
import random
import time
import matplotlib.pyplot as plt
import matplotlib
import numpy as np
import openpyxl as oxl
import sys

# This import registers the 3D projection, but is otherwise unused.

tech_names = ['Load', 'Onshore Wind', 'Offshore Wind', 'Rooftop PV', 'Fixed PV', 'Single Axis PV',
              'Dual Axis PV', 'Biomass', 'Geothermal', 'Other1', 'CST', 'Shortfall']
target_keys = ['lcoe', 'load_pct', 'surplus_pct', 're_pct', 'cost', 'co2']
target_names = ['LCOE', 'Load%', 'Surplus%', 'RE%', 'Cost', 'CO2']
target_fmats = ['$%.2f', '%.1f%%', '%.1f%%', '%.1f%%', '$%.1fpwr_chr', '%.1fpwr_chr']
target_titles = ['LCOE ($)', 'Load met %', 'Surplus %', 'RE %', 'Total Cost ($)', 'tCO2e']
headers = ['Facility', 'Capacity\n(Gen, MW;\nStor, MWh)', 'To meet\nLoad (MWh)',
           'Subtotal\n(MWh)', 'CF', 'Cost ($/yr)', 'LCOG\nCost\n($/MWh)', 'LCOE\nCost\n($/MWh)',
           'Emissions\n(tCO2e)', 'Emissions\nCost', 'LCOE With\nCO2 Cost\n($/MWh)', 'Max.\nMWH',
           'Max.\nBalance', 'Capital\nCost', 'Lifetime\nCost', 'Lifetime\nEmissions',
           'Lifetime\nEmissions\nCost', 'Reference\nLCOE', 'Reference\nCF']
# set up columns for summary table. Hopefully to make it easier to add / alter columns
st_fac = 0
st_cap = 1
st_tml = 2
st_sub = 3
st_cfa = 4
st_cst = 5
st_lcg = 6
st_lco = 7
st_emi = 8
st_emc = 9
st_lcc = 10
st_max = 11
st_bal = 12
st_cac = 13
st_lic = 14
st_lie = 15
st_lec = 16
st_rlc = 17
st_rcf = 18

C = 0 # Constraints - xls or xlsx
G = 1 # Generators - xls or xlsx
O = 2 # Optimisation - xls or xlsx
D = 3 # Data - xlsx
R = 4 # Results - xlsx
B = 5 # Batch input - xlsx
col_letters = ' ABCDEFGHIJKLMNOPQRSTUVWXYZ'
def ss_col(col, base=1):
    if base == 1:
        col -= 1
    c1 = 0
    c2, c3 = divmod(col, 26)
    c3 += 1
    if c2 > 26:
        c1, c2 = divmod(c2, 26)
    return (col_letters[c1] + col_letters[c2] + col_letters[c3]).strip()

# Define a class 
class Constraint:
    def __init__(self, name, category, capacity_min, capacity_max, rampup_max, rampdown_max,
                 recharge_max, recharge_loss, discharge_max, discharge_loss, parasitic_loss,
                 min_run_time, warm_time) -> None:
        self.name = name.strip()
        self.category = category
        try:
            self.capacity_min = Decimal(capacity_min) # minimum run_rate for generator; don't drain below for storage
        except:
            self.capacity_min = Decimal(0)
        try:
            self.capacity_max = Decimal(capacity_max) # maximum run_rate for generator; don't drain more than this for storage
        except:
            self.capacity_max = Decimal(1)
        try:
            self.recharge_max = Decimal(recharge_max) # can't charge more than this per hour
        except:
            self.recharge_max = Decimal(1)
        try:
            self.recharge_loss = Decimal(recharge_loss)
        except:
            self.recharge_loss = Decimal(0)
        try:
            self.discharge_max = Decimal(discharge_max) # can't discharge more than this per hour
        except:
            self.discharge_max = Decimal(1)
        try:
            self.discharge_loss = Decimal(discharge_loss)
        except:
            self.discharge_loss = Decimal(0)
        try:
            self.parasitic_loss = Decimal(parasitic_loss) # daily parasitic loss / hourly ?
        except:
            self.parasitic_loss = Decimal(0)
        try:
            self.rampup_max = Decimal(rampup_max)
        except:
            self.rampup_max = Decimal(1)
        try:
            self.rampdown_max = Decimal(rampdown_max)
        except:
            self.rampdown_max = Decimal(1)
        try:
            self.min_run_time = int(min_run_time)
        except:
            self.min_run_time = Decimal(0)
        try:
            self.warm_time = Decimal(warm_time)
            if self.warm_time >= 1:
                self.warm_time = self.warm_time / 60
                if self.warm_time > 1:
                    self.warm_time = 1
            elif self.warm_time > 0:
                if self.warm_time <= 1 / 24.:
                    self.warm_time = self.warm_time * 24
        except:
            self.warm_time = 0

class Facility:
    def __init__(self, **kwargs):
        kwargs = {**kwargs}
      #  return
        self.name = ''
        self.constr = ''
        self.order = 0
        self.lifetime = 20
        for attr in ['category', 'capacity', 'lcoe', 'lcoe_cf', 'emissions', 'initial', 'capex',
                     'fixed_om', 'variable_om', 'fuel', 'disc_rate', 'lifetime']:
            setattr(self, attr, 0.)
        for key, value in kwargs.items():
            if value != '':
                if key == 'lifetime' and value == 0:
                    setattr(self, key, 20)
                else:
                    setattr(self, key, value)
                    
class PM_Facility:
    def __init__(self, name, generator, capacity, fac_type, col, multiplier):
        self.name = name
        if name.find('.') > 0:
            self.zone = name[:name.find('.')]
        else:
            self.zone = ''
        self.generator = generator
        self.capacity = capacity
        self.fac_type = fac_type
        self.col = col
        self.multiplier = multiplier
        
class Optimisation:
    def __init__(self, name, approach, values): #capacity=None, cap_min=None, cap_max=None, cap_step=None, caps=None):
        self.name = name.strip()
        self.approach = approach
        if approach == 'Discrete':
            caps = values.split()
            self.capacities = []
            cap_max = 0.
            for cap in caps:
                try:
                    self.capacities.append(Decimal(cap))
                    cap_max += Decimal(cap)
                except:
                    pass
            self.capacity_min = 0
            self.capacity_max = round(cap_max, 3)
            self.capacity_step = None
        elif approach == 'Range':
            caps = values.split()
            try:
                self.capacity_min = Decimal(caps[0])
            except:
                self.capacity_min = 0.
            try:
                self.capacity_max = Decimal(caps[1])
            except:
                self.capacity_max = 0.
            try:
                self.capacity_step = Decimal(caps[2])
            except:
                self.capacity_step = 0.
            self.capacities = None
        else:
            self.capacity_min = 0.
            self.capacity_max = 0.
            self.capacity_step = 0.
            self.capacities = None
        self.capacity = 0.

class powerMatch():
    def __init__ (self, settings=dict[str, dict[str, str]], constraints=dict, generators=dict,):
        super(powerMatch, self).__init__()
        self.file_labels = ['Constraints', 'Generators', 'Optimisation', 'Data', 'Results', 'Batch']
        self.more_details = False
        self.constraints = constraints
        self.optimisation = None
        self.adjustto = None # adjust capacity to this
        self.adjust_cap = 25
        self.adjust_gen = False
        self.change_res = True
        self.adjusted_lcoe = True
        self.optimise_debug = False
        self.carbon_price_max = 200.

        self.pmss_details = {} # contains name, generator, capacity, fac_type, col, multiplier
        self.pmss_data = []
        self.show_multipliers = False
        # it's easier for the user to understand while for the program logic surplus is easier
        self.scenarios = 'C:/Users/Paul/Local Sites/Powermatch/'
        iorder = []
        self.targets = {}
        self.optimise_progress = 0
        for t in range(len(target_keys)):
            if target_keys[t] in ['re_pct', 'surplus_pct']:
                self.targets[target_keys[t]] = [target_names[t], 0., -1, 0., 0, target_fmats[t],
                                                 target_titles[t]]
            else:
                self.targets[target_keys[t]] = [target_names[t], 0., 0., -1, 0, target_fmats[t],
                                                 target_titles[t]]
        try:
            items = settings['Powermatch']
            for key, value in items.items():
                if key == 'adjusted_capacities':
                    self.adjustto = {}
                    bits = value.split(',')
                    for bit in bits:
                        bi = bit.split('=')
                        self.adjustto[bi[0]] = Decimal(bi[1])
                elif key == 'carbon_price':
                    try:
                        carbon_price = Decimal(value)
                    except:
                        pass
                elif key == 'carbon_price_max':
                    try:
                        carbon_price_max = Decimal(value)
                    except:
                        pass
                elif key == 'change_results':
                    if value.lower() in ['false', 'off', 'no']:
                        self.change_res = False
                elif key == 'adjusted_lcoe' or key == 'corrected_lcoe':
                    if value.lower() in ['false', 'no', 'off']:
                        self.adjusted_lcoe = False
                elif key == 'discount_rate':
                    try:
                        self.discount_rate = Decimal(value)
                    except:
                        pass
                elif key == 'more_details':
                    if value.lower() in ['true', 'yes', 'on']:
                        self.more_details = True
                elif key == 'show_multipliers':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.show_multipliers = True
                elif key == 'shortfall_sign':
                    if value[0] == '+' or value[0].lower() == 'p':
                        self.surplus_sign = -1
        except:
            pass
        try:
            adjust_cap = settings['Power']['adjust_cap']
            try:
                self.adjust_cap = Decimal(adjust_cap)
            except:
                try:
                    self.adjust_cap = eval(adjust_cap)
                except:
                    pass
            if self.adjust_cap < 0:
                self.adjust_cap = pow(10, 12)
        except:
            pass
        if self.adjust_gen and self.adjustto is None:
            self.adjustto = {}
            self.adjustto['Load'] = 0
            for gen in tech_names:
                try:
                    if generators[gen].capacity > 0:
                        self.adjustto[gen] = generators[gen].capacity
                except:
                    pass
            for gen in iorder:
                try:
                    if generators[gen].capacity > 0:
                        self.adjustto[gen] = generators[gen].capacity
                except:
                    pass

        self.labels = [None] * len(self.file_labels)
        self.files = [None] * len(self.file_labels)
        self.sheets = self.file_labels[:]
        self.progressbar = None
        del self.sheets[-2:]

    def changes(self):
        self.updated = True

    def quitClicked(self):
        self.close()
        
    @staticmethod
    def doDispatch(settings, year, option, pmss_details, pmss_data, generators, re_order, dispatch_order,
                   pm_data_file, data_file, title=None):
        def calcLCOE(annual_output, capital_cost, annual_operating_cost, discount_rate, lifetime):
            # Compute levelised cost of electricity
            if discount_rate > 0:
                annual_cost_capital = capital_cost * discount_rate * pow(1 + discount_rate, lifetime) / \
                                      (pow(1 + discount_rate, lifetime) - 1)
            else:
                annual_cost_capital = capital_cost / lifetime
            total_annual_cost = annual_cost_capital + annual_operating_cost
            return total_annual_cost / annual_output

        def format_period(per):
            hr = per % 24
            day = int((per - hr) / 24)
            mth = 0
            while day > the_days[mth] - 1:
                day -= the_days[mth]
                mth += 1
            return '{}-{:02d}-{:02d} {:02d}:00'.format(year, mth+1, day+1, hr)

        def summary_totals(title=''):
            sp_d = [' '] * len(headers)
            sp_d[st_fac] = title + 'Total'
            sp_d[st_cap] = cap_sum
            sp_d[st_tml] = tml_sum
            sp_d[st_sub] = gen_sum
            sp_d[st_cst] = cost_sum
            sp_d[st_lcg] = gs
            sp_d[st_lco] = gsw
            sp_d[st_emi] = co2_sum
            sp_d[st_emc] = co2_cost_sum
            sp_d[st_lcc] = gswc
            sp_d[st_cac] = capex_sum
            sp_d[st_lic] = lifetime_sum
            sp_d[st_lie] = lifetime_co2_sum
            sp_d[st_lec] = lifetime_co2_cost
            sp_data.append(sp_d)
            if (carbon_price > 0 or option == 'B'):
                sp_d = [' '] * len(headers)
                cc = co2_sum * carbon_price
                cl = cc * max_lifetime
                if adjusted_lcoe and tml_sum > 0:
                    cs = (cost_sum + cc) / tml_sum
                else:
                    if gen_sum > 0:
                        cs = (cost_sum + cc) / gen_sum
                    else:
                        cs = ''
                sp_d[st_fac] = title + 'Total incl. Carbon Cost'
                sp_d[st_cst] = cost_sum + cc
                sp_d[st_lic] = lifetime_sum + cl
                sp_data.append(sp_d)
            if tml_sum > 0:
                sp_d = [' '] * len(headers)
             #   sp_d[st_fac] = 'RE Direct Contribution to ' + title + 'Load'
                sp_d[st_fac] = 'RE %age'
                re_pct = (tml_sum - sto_sum - ff_sum) / tml_sum
                sp_d[st_cap] = '{:.1f}%'.format(re_pct * Decimal(100))
                sp_d[st_tml] = tml_sum - ff_sum - sto_sum
                sp_data.append(sp_d)
                if sto_sum > 0:
                    sp_d = [' '] * len(headers)
                 #   sp_d[st_fac] = 'RE Contribution to ' + title + 'Load via Storage'
                    sp_d[st_fac] = 'Storage %age'
                    sp_d[st_cap] = '{:.1f}%'.format(sto_sum * Decimal(100) / tml_sum)
                    sp_d[st_tml] = sto_sum
                    sp_data.append(sp_d)
            sp_data.append([' '])
            sp_data.append([title + 'Load Analysis'])
            if sp_load != 0:
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = title + 'Load met'
                load_pct = (sp_load - sf_sums[0]) / sp_load
                sp_d[st_cap] = '{:.1f}%'.format(load_pct * 100)
                sp_d[st_tml] = sp_load - sf_sums[0]
                sp_data.append(sp_d)
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = 'Shortfall'
                sp_d[st_cap] = '{:.1f}%'.format(sf_sums[0] * 100 / sp_load)
                sp_d[st_tml] = sf_sums[0]
                sp_data.append(sp_d)
                if option == 'B':
                    sp_d = [' '] * len(headers)
                    sp_d[st_fac] = title + 'Total Load'
                    sp_d[st_tml] = sp_load
                    if title == '':
                        sp_d[st_max] = load_max
                    sp_data.append(sp_d)
                else:
                    load_mult = ''
                    try:
                        mult = round(pmss_details['Load'].multiplier, 3)
                        if mult != 1:
                            load_mult = ' x ' + str(mult)
                    except:
                        pass
                    sp_d = [' '] * len(headers)
                    sp_d[st_fac] = 'Total ' + title + 'Load - ' + str(year) + load_mult
                    sp_d[st_tml] = sp_load
                    if title == '' or option == 'S':
                        sp_d[st_max] = load_max
                        sp_d[st_bal] = ' (' + format_period(load_hr)[5:] + ')'
                    sp_data.append(sp_d)
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = 'RE %age of Total ' + title + 'Load'
                sp_d[st_cap] = '{:.1f}%'.format((sp_load - sf_sums[0] - ff_sum) * Decimal(100) / sp_load)
                sp_data.append(sp_d)
                sp_data.append(' ')
                if tot_sto_loss != 0:
                    sp_d = [' '] * len(headers)
                    sp_d[st_fac] = 'Storage losses'
                    sp_d[st_sub] = tot_sto_loss
                    sp_data.append(sp_d)
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = title + 'Surplus'
                surp_pct = -sf_sums[1] / sp_load
                sp_d[st_cap] = '{:.1f}%'.format(surp_pct * 100)
                sp_d[st_sub] = -sf_sums[1]
                sp_data.append(sp_d)
            else:
                load_pct = 0
                surp_pct = 0
                re_pct = 0
            max_short = [0, 0]
            for h in range(len(shortfall)):
                if shortfall[h] > max_short[1]:
                    max_short[0] = h
                    max_short[1] = shortfall[h]
            if max_short[1] > 0:
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = 'Largest Shortfall'
                sp_d[st_sub] = round(max_short[1], 2)
                sp_d[st_cfa] = ' (' + format_period(max_short[0])[5:] + ')'
                sp_data.append(sp_d)
            if option == 'O' or option == '1':
                return load_pct, surp_pct, re_pct

        def do_detail(fac, col, ss_row):
            if fac in generators.keys():
                gen = fac
            else:
                gen = pmss_details[fac].generator
            col += 1
            sp_cols.append(fac)
            sp_cap.append(pmss_details[fac].capacity * pmss_details[fac].multiplier)
            if do_zone and pmss_details[fac].zone != '':
                ns.cell(row=zone_row, column=col).value = pmss_details[fac].zone
                ns.cell(row=zone_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
            try:
                ns.cell(row=what_row, column=col).value = fac[fac.find('.') + 1:]
            except:
                ns.cell(row=what_row, column=col).value = fac # gen
            ns.cell(row=what_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
            ns.cell(row=cap_row, column=col).value = sp_cap[-1]
            ns.cell(row=cap_row, column=col).number_format = '#,##0.00'
            # capacity
            ns.cell(row=sum_row, column=col).value = '=SUM(' + ss_col(col) \
                    + str(hrows) + ':' + ss_col(col) + str(hrows + 8759) + ')'
            ns.cell(row=sum_row, column=col).number_format = '#,##0'
            # To meet load MWh
            ns.cell(row=tml_row, column=col).value = fac_tml[fac]
            ns.cell(row=tml_row, column=col).number_format = '#,##0'
            ns.cell(row=cf_row, column=col).value = '=IF(' + ss_col(col) + str(cap_row) + '>0,' + \
                    ss_col(col) + str(sum_row) + '/' + ss_col(col) + str(cap_row) + '/8760,"")'
            ns.cell(row=cf_row, column=col).number_format = '#,##0.0%'
            # subtotal MWh
            ns.cell(row=cf_row, column=col).value = '=IF(' + ss_col(col) + str(cap_row) + '>0,' + \
                    ss_col(col) + str(sum_row) +'/' + ss_col(col) + str(cap_row) + '/8760,"")'
            ns.cell(row=cf_row, column=col).number_format = '#,##0.0%'
            if gen not in generators.keys():
                return col
            if generators[gen].capex > 0 or generators[gen].fixed_om > 0 \
              or generators[gen].variable_om > 0 or generators[gen].fuel > 0:
                disc_rate = generators[gen].disc_rate
                if disc_rate == 0:
                    disc_rate = discount_rate
                if disc_rate == 0:
                    cst_calc = '/' + str(generators[gen].lifetime)
                else:
                    pwr_calc = 'POWER(1+' + str(disc_rate) + ',' + str(generators[gen].lifetime) + ')'
                    cst_calc = '*' + str(disc_rate) + '*' + pwr_calc + '/SUM(' + pwr_calc + ',-1)'
                ns.cell(row=cost_row, column=col).value = '=IF(' + ss_col(col) + str(cf_row) + \
                        '>0,' + ss_col(col) + str(cap_row) + '*' + str(generators[gen].capex) + \
                        cst_calc + '+' + ss_col(col) + str(cap_row) + '*' + \
                        str(generators[gen].fixed_om) + '+' + ss_col(col) + str(sum_row) + '*(' + \
                        str(generators[gen].variable_om) + '+' + str(generators[gen].fuel) + \
                        '),0)'
                ns.cell(row=cost_row, column=col).number_format = '$#,##0'
                ns.cell(row=lcoe_row, column=col).value = '=IF(AND(' + ss_col(col) + str(cf_row) + \
                        '>0,' + ss_col(col) + str(cap_row) + '>0),' + ss_col(col) + \
                        str(cost_row) + '/' + ss_col(col) + str(sum_row) + ',"")'
                ns.cell(row=lcoe_row, column=col).number_format = '$#,##0.00'
            elif generators[gen].lcoe > 0:
                if ss_row >= 0:
                    ns.cell(row=cost_row, column=col).value = '=IF(' + ss_col(col) + str(cf_row) + \
                            '>0,' + ss_col(col) + str(sum_row) + '*Summary!' + ss_col(st_rlc + 1) + str(ss_row) + \
                        '*Summary!' + ss_col(st_rcf + 1) + str(ss_row) + '/' + ss_col(col) + str(cf_row) + ',0)'
                    ns.cell(row=cost_row, column=col).number_format = '$#,##0'
                ns.cell(row=lcoe_row, column=col).value = '=IF(AND(' + ss_col(col) + str(cf_row) + '>0,' \
                        + ss_col(col) + str(cap_row) + '>0),' + ss_col(col) + str(cost_row) + '/8760/' \
                        + ss_col(col) + str(cf_row) +'/' + ss_col(col) + str(cap_row) + ',"")'
                ns.cell(row=lcoe_row, column=col).number_format = '$#,##0.00'
            elif generators[gen].lcoe_cf == 0: # no cost facility
                if ss_row >= 0:
                    ns.cell(row=cost_row, column=col).value = '=IF(' + ss_col(col) + str(cf_row) + \
                            '>0,' + ss_col(col) + str(sum_row) + '*Summary!' + ss_col(st_rlc + 1) + str(ss_row) + \
                        '*Summary!' + ss_col(st_rcf + 1) + str(ss_row) + '/' + ss_col(col) + str(cf_row) + ',0)'
                    ns.cell(row=cost_row, column=col).number_format = '$#,##0'
                ns.cell(row=lcoe_row, column=col).value = '=IF(AND(' + ss_col(col) + str(cf_row) + '>0,' \
                        + ss_col(col) + str(cap_row) + '>0),' + ss_col(col) + str(cost_row) + '/8760/' \
                        + ss_col(col) + str(cf_row) +'/' + ss_col(col) + str(cap_row) + ',"")'
                ns.cell(row=lcoe_row, column=col).number_format = '$#,##0.00'
            if generators[gen].emissions > 0:
                ns.cell(row=emi_row, column=col).value = '=' + ss_col(col) + str(sum_row) \
                        + '*' + str(generators[gen].emissions)
                ns.cell(row=emi_row, column=col).number_format = '#,##0'
            ns.cell(row=max_row, column=col).value = '=MAX(' + ss_col(col) + str(hrows) + \
                                           ':' + ss_col(col) + str(hrows + 8759) + ')'
            ns.cell(row=max_row, column=col).number_format = '#,##0.00'
            ns.cell(row=hrs_row, column=col).value = '=COUNTIF(' + ss_col(col) + str(hrows) + \
                                           ':' + ss_col(col) + str(hrows + 8759) + ',">0")'
            ns.cell(row=hrs_row, column=col).number_format = '#,##0'
            di = pmss_details[fac].col
            if pmss_details[fac].multiplier == 1:
                for row in range(hrows, 8760 + hrows):
                    ns.cell(row=row, column=col).value = pmss_data[di][row - hrows]
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            else:
                for row in range(hrows, 8760 + hrows):
                    ns.cell(row=row, column=col).value = pmss_data[di][row - hrows] * \
                                                         pmss_details[fac].multiplier
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            return col

        def do_detail_summary(fac, col, ss_row, dd_tml_sum, dd_re_sum):
            if do_zone and pmss_details[fac].zone != '':
                ss.cell(row=ss_row, column=st_fac+1).value = '=Detail!' + ss_col(col) + str(zone_row) + \
                                                      '&"."&Detail!' + ss_col(col) + str(what_row)
            else:
                ss.cell(row=ss_row, column=st_fac+1).value = '=Detail!' + ss_col(col) + str(what_row)
            if fac in generators.keys():
                gen = fac
            else:
                gen = pmss_details[fac].generator
            # capacity
            ss.cell(row=ss_row, column=st_cap+1).value = '=Detail!' + ss_col(col) + str(cap_row)
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.00'
            # To meet load MWh
            ss.cell(row=ss_row, column=st_tml+1).value = '=Detail!' + ss_col(col) + str(tml_row)
            ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
            dd_tml_sum += ss_col(st_tml+1) + str(ss_row) + '+'
            # subtotal MWh
            ss.cell(row=ss_row, column=st_sub+1).value = '=IF(Detail!' + ss_col(col) + str(sum_row) \
                                                  + '>0,Detail!' + ss_col(col) + str(sum_row) + ',"")'
            ss.cell(row=ss_row, column=st_sub+1).number_format = '#,##0'
            dd_re_sum += ss_col(st_sub+1) + str(ss_row) + '+'
            # CF
            ss.cell(row=ss_row, column=st_cfa+1).value = '=IF(Detail!' + ss_col(col) + str(cf_row) \
                                                  + '>0,Detail!' + ss_col(col) + str(cf_row) + ',"")'
            ss.cell(row=ss_row, column=st_cfa+1).number_format = '#,##0.0%'
            if gen not in generators.keys():
                return dd_tml_sum, dd_re_sum
            if generators[gen].capex > 0 or generators[gen].fixed_om > 0 \
              or generators[gen].variable_om > 0 or generators[gen].fuel > 0:
                disc_rate = generators[gen].disc_rate
                if disc_rate == 0:
                    disc_rate = discount_rate
                if disc_rate == 0:
                    cst_calc = '/' + str(generators[gen].lifetime)
                else:
                    pwr_calc = 'POWER(1+' + str(disc_rate) + ',' + str(generators[gen].lifetime) + ')'
                    cst_calc = '*' + str(disc_rate) + '*' + pwr_calc + '/SUM(' + pwr_calc + ',-1)'
                # cost / yr
                ss.cell(row=ss_row, column=st_cst+1).value = '=Detail!' + ss_col(col) + str(cost_row)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                # lcog
                ss.cell(row=ss_row, column=st_lcg+1).value = '=IF(Detail!' + ss_col(col) + str(lcoe_row) \
                                                      + '>0,Detail!' + ss_col(col) + str(lcoe_row) + ',"")'
                ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
                # capital cost
                ss.cell(row=ss_row, column=st_cac+1).value = '=IF(Detail!' + ss_col(col) + str(cap_row) \
                                                        + '>0,Detail!' + ss_col(col) + str(cap_row) + '*'  \
                                                        + str(generators[gen].capex) + ',"")'
                ss.cell(row=ss_row, column=st_cac+1).number_format = '$#,##0'
            elif generators[gen].lcoe > 0:
                # cost / yr
                ss.cell(row=ss_row, column=st_cst+1).value = '=Detail!' + ss_col(col) + str(cost_row)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                # lcog
                ss.cell(row=ss_row, column=st_lcg+1).value = '=Detail!' + ss_col(col) + str(lcoe_row)
                ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
                # ref lcoe
                ss.cell(row=ss_row, column=st_rlc+1).value = generators[gen].lcoe
                ss.cell(row=ss_row, column=st_rlc+1).number_format = '$#,##0.00'
                # ref cf
                ss.cell(row=ss_row, column=st_rcf+1).value = generators[gen].lcoe_cf
                ss.cell(row=ss_row, column=st_rcf+1).number_format = '#,##0.0%'
            elif generators[gen].lcoe_cf == 0: # no cost facility
                # cost / yr
                ss.cell(row=ss_row, column=st_cst+1).value = '=Detail!' + ss_col(col) + str(cost_row)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                # lcog
                ss.cell(row=ss_row, column=st_lcg+1).value = '=Detail!' + ss_col(col) + str(lcoe_row)
                ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
                # ref lcoe
                ss.cell(row=ss_row, column=st_rlc+1).value = generators[gen].lcoe
                ss.cell(row=ss_row, column=st_rlc+1).number_format = '$#,##0.00'
                # ref cf
                ss.cell(row=ss_row, column=st_rcf+1).value = generators[gen].lcoe_cf
                ss.cell(row=ss_row, column=st_rcf+1).number_format = '#,##0.0%'
            # lifetime cost
            ss.cell(row=ss_row, column=st_lic+1).value = '=IF(Detail!' + ss_col(col) + str(sum_row) \
                                                    + '>0,Detail!' + ss_col(col) + str(cost_row) + '*lifetime,"")'
            ss.cell(row=ss_row, column=st_lic+1).number_format = '$#,##0'
            # max mwh
            ss.cell(row=ss_row, column=st_max+1).value = '=IF(Detail!' + ss_col(col) + str(sum_row) \
                                                   + '>0,Detail!' + ss_col(col) + str(max_row) + ',"")'
            ss.cell(row=ss_row, column=st_max+1).number_format = '#,##0.00'
            if generators[gen].emissions > 0:
                ss.cell(row=ss_row, column=st_emi+1).value = '=Detail!' + ss_col(col) + str(emi_row)
                ss.cell(row=ss_row, column=st_emi+1).number_format = '#,##0'
                if carbon_price > 0:
                    ss.cell(row=ss_row, column=st_emc+1).value = '=IF(AND(' + ss_col(st_emi+1) + str(ss_row) + '<>"",' + \
                                                                 ss_col(st_emi+1) + str(ss_row) + '>0),' + \
                                                                 ss_col(st_emi+1) + str(ss_row) + '*carbon_price,"")'
                    ss.cell(row=ss_row, column=st_emc+1).number_format = '$#,##0'
            ss.cell(row=ss_row, column=st_lie+1).value = '=IF(AND(' + ss_col(st_emi+1) + str(ss_row) + '<>"",' + \
                                                         ss_col(st_emi+1) + str(ss_row) + '>0),' + \
                                                         ss_col(st_emi+1) + str(ss_row) + '*lifetime,"")'
            ss.cell(row=ss_row, column=st_lie+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_lec+1).value = '=IF(AND(' + ss_col(st_emi+1) + str(ss_row) + '<>"",' + \
                                                         ss_col(st_emi+1) + str(ss_row) + '>0),' + \
                                                         ss_col(st_emc+1) + str(ss_row) + '*lifetime,"")'
            ss.cell(row=ss_row, column=st_lec+1).number_format = '$#,##0'
            return dd_tml_sum, dd_re_sum

        def detail_summary_total(ss_row, title='', base_row='', back_row=''):
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = title + 'Total'
            for col in range(1, len(headers) + 1):
                ss.cell(row=3, column=col).font = bold
                ss.cell(row=ss_row, column=col).font = bold
            for col in [st_cap, st_tml, st_sub, st_cst, st_emi, st_emc, st_cac, st_lic, st_lie, st_lec]:
                if back_row != '':
                    strt = ss_col(col, base=0) + back_row + '+'
                else:
                    strt = ''
                ss.cell(row=ss_row, column=col+1).value = '=' + strt + 'SUM(' + ss_col(col, base=0) + \
                        base_row + ':' + ss_col(col, base=0) + str(ss_row - 1) + ')'
                if col in [st_cap]:
                    ss.cell(row=ss_row, column=col+1).number_format = '#,##0.00'
                elif col in [st_tml, st_sub, st_emi, st_lie]:
                    ss.cell(row=ss_row, column=col+1).number_format = '#,##0'
                else:
                    ss.cell(row=ss_row, column=col+1).number_format = '$#,##0'
            ss.cell(row=ss_row, column=st_lcg+1).value = '=' + ss_col(st_cst+1) + str(ss_row) + \
                                                         '/' + ss_col(st_sub+1) + str(ss_row)
            ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
            ss.cell(row=ss_row, column=st_lco+1).value = '=' + ss_col(st_cst+1) + str(ss_row) + \
                                                         '/' + ss_col(st_tml+1) + str(ss_row)
            ss.cell(row=ss_row, column=st_lco+1).number_format = '$#,##0.00'
            if carbon_price > 0:
                ss.cell(row=ss_row, column=st_lcc+1).value = '=(' + ss_col(st_cst+1) + str(ss_row) + \
                    '+' + ss_col(st_emc+1) + str(ss_row) + ')/' + ss_col(st_tml+1) + str(ss_row)
                ss.cell(row=ss_row, column=st_lcc+1).number_format = '$#,##0.00'
                ss.cell(row=ss_row, column=st_lcc+1).font = bold
            last_col = ss_col(ns.max_column)
            r = 1
            if carbon_price > 0:
                ss_row += 1
                ss.cell(row=ss_row, column=1).value = title + 'Total incl. Carbon Cost'
                ss.cell(row=ss_row, column=st_cst+1).value = '=' + ss_col(st_cst+1) + str(ss_row - 1) + \
                        '+' + ss_col(st_emc+1) + str(ss_row - 1)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                ss.cell(row=ss_row, column=st_lic+1).value = '=' + ss_col(st_lic+1) + str(ss_row - r) + \
                                                             '+' + ss_col(st_lec+1) + str(ss_row - 1)
                ss.cell(row=ss_row, column=st_lic+1).number_format = '$#,##0'
                r += 1
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = title + 'RE %age'
            ss.cell(row=ss_row, column=st_tml+1).value = ns_tml_sum[:-1] + ')'
            ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_cap+1).value = '=' + ss_col(st_tml+1) + str(ss_row) + '/' +\
                                                         ss_col(st_tml+1) + str(ss_row - r)
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.0%'
            ss_re_row = ss_row
            # if storage
            if ns_sto_sum != '':
                ss_row += 1
                ss.cell(row=ss_row, column=1).value = title + 'Storage %age'
                ss.cell(row=ss_row, column=st_tml+1).value = '=' + ns_sto_sum[1:]
                ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
                ss.cell(row=ss_row, column=st_cap+1).value = '=(' + ns_sto_sum[1:] + ')/' + ss_col(st_tml+1) + \
                                                             str(ss_row - r - 1)
                ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.0%'
                ss_sto_row = ss_row
            # now do the LCOE and LCOE with CO2 stuff
            if base_row == '4':
                base_col = 'C'
                if ss_sto_row >= 0:
                    for rw in range(ss_re_fst_row, ss_re_lst_row + 1):
                        ss.cell(row=rw, column=st_lco+1).value = '=IF(' + ss_col(st_lcg+1) + str(rw) + '>0,' + \
                                ss_col(st_cst+1) + str(rw) + '/(' + ss_col(st_tml+1) + str(rw) + '+(' + \
                                ss_col(st_tml+1) + '$' + str(ss_sto_row) + '*' + ss_col(st_tml+1) + str(rw) + \
                                ')/' + ss_col(st_tml+1) + '$' + str(ss_re_row) + '),"")'
                        ss.cell(row=rw, column=st_lco+1).number_format = '$#,##0.00'
                        if carbon_price > 0:
                            ss.cell(row=rw, column=st_lcc+1).value = '=IF(' + ss_col(st_emc+1) + str(rw) + '>0,(' + \
                                    ss_col(st_cst+1) + str(rw) + '+' + ss_col(st_emc+1) + str(rw) + ')/(' + \
                                    ss_col(st_tml+1) + str(rw) + '+(' + ss_col(st_tml+1) + '$' + str(ss_sto_row) + \
                                    '*' + ss_col(st_tml+1) + str(rw) + ')/' + ss_col(st_tml+1) + '$' + \
                                    str(ss_re_row) + '),"")'
                            ss.cell(row=rw, column=st_lcc+1).number_format = '$#,##0.00'
                else:
                    for rw in range(ss_re_fst_row, ss_re_lst_row):
                        ss.cell(row=rw, column=st_lco+1).value = '=IF(' + ss_col(st_lcg+1) + str(rw) + '>0,' + \
                                ss_col(st_cst+1) + str(rw) + '/' + ss_col(st_tml+1) + str(rw) + '),"")'
                        ss.cell(row=rw, column=st_lco+1).number_format = '$#,##0.00'
                        if carbon_price > 0:
                            ss.cell(row=rw, column=st_lcc+1).value = '=IF(' + ss_col(st_emc+1) + str(rw) + '>0,(' + \
                                    ss_col(st_cst+1) + str(rw) + ss_col(st_emc+1) + str(rw) + ')/' + \
                                    ss_col(st_tml+1) + str(rw) + '),"")'
                            ss.cell(row=rw, column=st_lcc+1).number_format = '$#,##0.00'
                for rw in range(ss_re_lst_row + 1, ss_lst_row + 1):
                    ss.cell(row=rw, column=st_lco+1).value = '=' + ss_col(st_lcg+1) + str(rw)
                    ss.cell(row=rw, column=st_lco+1).number_format = '$#,##0.00'
                    if carbon_price > 0:
                        ss.cell(row=rw, column=st_lcc+1).value = '=IF(' + ss_col(st_emc+1) + str(rw) + '>0,(' + \
                                ss_col(st_cst+1) + str(rw) + '+' + ss_col(st_emc+1) + str(rw) + ')/' + \
                                ss_col(st_tml+1) + str(rw) + ',"")'
                        ss.cell(row=rw, column=st_lcc+1).number_format = '$#,##0.00'
            else:
                base_col = ss_col(next_col)
                for rw in range(ul_fst_row, ul_lst_row + 1):
                    ss.cell(row=rw, column=st_lco+1).value = '=' + ss_col(st_cst+1) + str(rw) + \
                                                             '/' + ss_col(st_tml+1) + str(rw)
                    ss.cell(row=rw, column=st_lco+1).number_format = '$#,##0.00'
                    if carbon_price > 0:
                        ss.cell(row=rw, column=st_lcc+1).value = '=(' + ss_col(st_cst+1) + str(rw) + \
                            '+' + ss_col(st_emc+1) + str(rw) + ')/' + ss_col(st_tml+1) + str(rw)
                        ss.cell(row=rw, column=st_lcc+1).number_format = '$#,##0.00'
            ss_row += 2
            ss.cell(row=ss_row, column=1).value = title + 'Load Analysis'
            ss.cell(row=ss_row, column=1).font = bold
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = title + 'Load met'
      ##      lm_row = ss_row
      #      if surplus_sign < 0:
      #          addsub = ')+' + base_col
      #      else:
      #          addsub = ')-' + base_col
      #      ss.cell(row=ss_row, column=st_tml+1).value = '=SUMIF(Detail!' + last_col + str(hrows) + ':Detail!' \
      #          + last_col + str(hrows + 8759) + ',"' + sf_test[0] + '=0",Detail!C' + str(hrows) \
      #          + ':Detail!C' + str(hrows + 8759) + ')+SUMIF(Detail!' + last_col + str(hrows) + ':Detail!' \
      #          + last_col + str(hrows + 8759) + ',"' + sf_test[1] + '0",Detail!C' + str(hrows) + ':Detail!C' \
      #          + str(hrows + 8759) + addsub + str(ss_row + 1)
            ss.cell(row=ss_row, column=st_tml+1).value = '=Detail!' + base_col + str(sum_row) + '-' + base_col + str(ss_row + 1)
            ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_cap+1).value = '=' + ss_col(st_tml+1) + str(ss_row) + '/' + ss_col(st_tml+1) + \
                                                         str(ss_row + 2)
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.0%'
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = title + 'Shortfall'
            sf_text = 'SUMIF(Detail!' + last_col + str(hrows) + ':Detail!' + last_col \
                      + str(hrows + 8759) + ',"' + sf_test[0] + '0",Detail!' + last_col \
                      + str(hrows) + ':Detail!' + last_col + str(hrows + 8759) + ')'
            if surplus_sign > 0:
                ss.cell(row=ss_row, column=st_tml+1).value = '=-' + sf_text
            else:
                ss.cell(row=ss_row, column=st_tml+1).value = '=' + sf_text
            ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_cap+1).value = '=' + ss_col(st_tml+1) + str(ss_row) + '/' + ss_col(st_tml+1) + \
                                                         str(ss_row + 1)
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.0%'
            ss_row += 1
            ld_row = ss_row
            load_mult = ''
            try:
                mult = round(pmss_details['Load'].multiplier, 3)
                if mult != 1:
                    load_mult = ' x ' + str(mult)
            except:
                pass
            ss.cell(row=ss_row, column=1).value = 'Total ' + title + 'Load - ' + year + load_mult
            ss.cell(row=ss_row, column=1).font = bold
            ss.cell(row=ss_row, column=st_tml+1).value = '=SUM(' + ss_col(st_tml+1) + str(ss_row - 2) + ':' + \
                                                         ss_col(st_tml+1) + str(ss_row - 1) + ')'
            ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_tml+1).font = bold
            ss.cell(row=ss_row, column=st_max+1).value = '=Detail!' + base_col + str(max_row)
            ss.cell(row=ss_row, column=st_max+1).number_format = '#,##0.00'
            ss.cell(row=ss_row, column=st_max+1).font = bold
            ss.cell(row=ss_row, column=st_bal+1).value = '=" ("&OFFSET(Detail!B' + str(hrows - 1) + ',MATCH(Detail!' + \
                    base_col + str(max_row) + ',Detail!' + base_col + str(hrows) + ':Detail!' + base_col + \
                    str(hrows + 8759) + ',0),0)&")"'
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = 'RE %age of Total ' + title + 'Load'
            ss.cell(row=ss_row, column=1).font = bold
            if ns_sto_sum == '':
                ss.cell(row=ss_row, column=st_cap+1).value = ss_col(st_tml+1) + str(ss_re_row - 1) + \
                                                             '/' + ss_col(st_tml+1) + str(ss_row - 1)
            else:
                ss.cell(row=ss_row, column=st_cap+1).value = '=(' + ss_col(st_tml+1) + str(ss_re_row) + '+' + \
                                                             ss_col(st_tml+1) + str(ss_sto_row) + ')/' + \
                                                             ss_col(st_tml+1) + str(ss_row - 1)
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.0%'
            ss.cell(row=ss_row, column=st_cap+1).font = bold
            ss_row += 2
            if ns_loss_sum != '':
                ss.cell(row=ss_row, column=1).value = title + 'Storage Losses'
                ss.cell(row=ss_row, column=st_sub+1).value = '=' + ns_loss_sum[1:]
                ss.cell(row=ss_row, column=st_sub+1).number_format = '#,##0'
                ss_row += 1
            ss.cell(row=ss_row, column=1).value = title + 'Surplus'
            sf_text = 'SUMIF(Detail!' + last_col + str(hrows) + ':Detail!' + last_col \
                      + str(hrows + 8759) + ',"' + sf_test[1] + '0",Detail!' + last_col + str(hrows) \
                      + ':Detail!' + last_col + str(hrows + 8759) + ')'
            if surplus_sign < 0:
                ss.cell(row=ss_row, column=st_sub+1).value = '=-' + sf_text
            else:
                ss.cell(row=ss_row, column=st_sub+1).value = '=' + sf_text
            ss.cell(row=ss_row, column=st_sub+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_cap+1).value = '=' + ss_col(st_sub+1) + str(ss_row) + '/' + ss_col(st_tml+1) + str(ld_row)
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.0%'
            max_short = [0, 0]
            for h in range(len(shortfall)):
                if shortfall[h] > max_short[1]:
                    max_short[0] = h
                    max_short[1] = shortfall[h]
            if max_short[1] > 0:
                ss_row += 1
                ss.cell(row=ss_row, column=1).value = 'Largest ' + title + 'Shortfall:'
                ss.cell(row=ss_row, column=st_sub+1).value = '=Detail!' + last_col + str(hrows + max_short[0])
                ss.cell(row=ss_row, column=st_sub+1).number_format = '#,##0.00'
                ss.cell(row=ss_row, column=st_cfa+1).value = '=" ("&OFFSET(Detail!B' + str(hrows - 1) + \
                        ',MATCH(' + ss_col(st_sub+1) + str(ss_row) + ',Detail!' + last_col + str(hrows) + \
                        ':Detail!' + last_col + str(hrows + 8759) + ',0),0)&")"'
            return ss_row, ss_re_row

    # The "guts" of Powermatch processing. Have a single calculation algorithm
    # for Summary, Powermatch (detail), and Optimise. The detail makes it messy
    # Note: For Batch pmss_data is reused so don't update it in doDispatch
        the_days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        progress_text = "Operation in progress. Please wait."
        surplus_sign = 1
        underlying = ['Rooftop PV']
        operational = []
        show_correlation = False
        carbon_price = Decimal(settings['Powermatch']['carbon_price'])
        discount_rate = Decimal(settings['Powermatch']['discount_rate'])
        adjusted_lcoe = True
        optimise_debug = False
        scenarios = 'C:/Users/Paul/Local Sites/Powermatch/'
        results_prefix = ''
        save_tables = settings['Powermatch']['save_tables']
        # if (self.progressbar is None):
        #     self.progressbar = st.progress(0, text=progress_text)
        if surplus_sign < 0:
            sf_test = ['>', '<']
            sf_sign = ['+', '-']
        else:
            sf_test = ['<', '>']
            sf_sign = ['-', '+']
        sp_cols = []
        sp_cap = []
        shortfall = [Decimal(0)] * 8760
        re_tml_sum = 0. # keep tabs on how much RE is used
        start_time = time.time()
        do_zone = False # could pass as a parameter
        max_lifetime = 0
        # find max. lifetime years for all technologies selected
        for key in pmss_details.keys():
            if key == 'Load'or key == 'Total':
                continue
            if pmss_details[key].capacity * pmss_details[key].multiplier > 0:
             #   gen = key.split('.')[-1]
                gen = pmss_details[key].generator
                max_lifetime = max(max_lifetime, generators[gen].lifetime)
        for key in pmss_details.keys():
            if key.find('.') > 0:
                do_zone = True
                break
        underlying_facs = []
        undercol = [] * len(underlying)
        operational_facs = []
        fac_tml = {}
        for fac in re_order:
            if fac == 'Load':
                continue
            fac_tml[fac] = 0.
            if fac in operational:
              #  operational_facs.append(fac)
                continue
            if fac.find('.') > 0:
                if fac[fac.find('.') + 1:] in underlying:
                    underlying_facs.append(fac)
                    continue
            elif fac in underlying:
                underlying_facs.append(fac)
                continue
        load_col = pmss_details['Load'].col
        for h in range(len(pmss_data[0])):
            load_h = pmss_data[load_col][h] * pmss_details['Load'].multiplier
            shortfall[h] = load_h
            for fac in fac_tml.keys():
                if fac in underlying_facs:
                    continue
                shortfall[h] -= pmss_data[pmss_details[fac].col][h] * pmss_details[fac].multiplier
            if shortfall[h] >= 0:
                alloc = Decimal(1)
            else:
                alloc = load_h / (load_h - shortfall[h])
            for fac in fac_tml.keys():
                if fac in underlying_facs:
                    fac_tml[fac] += pmss_data[pmss_details[fac].col][h] * pmss_details[fac].multiplier
                else:
                    pmss_data[pmss_details[fac].col][h] * pmss_details[fac].multiplier * alloc
            line = ''
        fac_tml_sum = 0
        for fac in fac_tml.keys():
            fac_tml_sum += fac_tml[fac]
        if show_correlation:
            col = pmss_details['Load'].col
            if pmss_details['Load'].multiplier == 1:
                df1 = pmss_data[col]
            else:
                tgt = []
                for h in range(len(pmss_data[col])):
                    tgt.append(pmss_data[col][h] * Decimal(pmss_details['Load'].multiplier))
                df1 = tgt
            corr_src = []
            for h in range(len(shortfall)):
                if shortfall[h] < 0:
                    corr_src.append(pmss_data[col][h])
                else:
                    corr_src.append(pmss_data[col][h] - shortfall[h])
            try:
                corr = np.corrcoef(df1, corr_src)
                if np.isnan(corr.item((0, 1))):
                    corr = Decimal(0)
                else:
                    corr = corr.item((0, 1))
            except:
                corr = Decimal(0)
            corr_data = [['Correlation To Load']]
            corr_data.append(['RE Contribution', corr])
        else:
            corr_data = None
        if option == 'D':
            wb = oxl.Workbook()
            ns = wb.active
            ns.title = 'Detail'
            normal = oxl.styles.Font(name='Arial')
            bold = oxl.styles.Font(name='Arial', bold=True)
            ss = wb.create_sheet('Summary', 0)
            ns_re_sum = '=('
            ns_tml_sum = '=('
            ns_sto_sum = ''
            ns_loss_sum = ''
            ns_not_sum = ''
            cap_row = 1
            ns.cell(row=cap_row, column=2).value = 'Capacity (MW/MWh)' #headers[1].replace('\n', ' ')
            ss.row_dimensions[3].height = 40
            ss.cell(row=3, column=st_fac+1).value = headers[st_fac] # facility
            ss.cell(row=3, column=st_cap+1).value = headers[st_cap] # capacity
            ini_row = 2
            ns.cell(row=ini_row, column=2).value = 'Initial Capacity'
            tml_row = 3
            ns.cell(row=tml_row, column=2).value = headers[st_tml].replace('\n', ' ')
            ss.cell(row=3, column=st_tml+1).value = headers[st_tml] # to meet load
            sum_row = 4
            ns.cell(row=sum_row, column=2).value = headers[st_sub].replace('\n', ' ')
            ss.cell(row=3, column=st_sub+1).value = headers[st_sub] # subtotal MWh
            cf_row = 5
            ns.cell(row=cf_row, column=2).value = headers[st_cfa].replace('\n', ' ')
            ss.cell(row=3, column=st_cfa+1).value = headers[st_cfa] # CF
            cost_row = 6
            ns.cell(row=cost_row, column=2).value = headers[st_cst].replace('\n', ' ')
            ss.cell(row=3, column=st_cst+1).value = headers[st_cst] # Cost / yr
            lcoe_row = 7
            ns.cell(row=lcoe_row, column=2).value = headers[st_lcg].replace('\n', ' ')
            ss.cell(row=3, column=st_lcg+1).value = headers[st_lcg] # LCOG
            ss.cell(row=3, column=st_lco+1).value = headers[st_lco] # LCOE
            emi_row = 8
            ns.cell(row=emi_row, column=2).value = headers[st_emi].replace('\n', ' ')
            ss.cell(row=3, column=st_emi+1).value = headers[st_emi] # emissions
            ss.cell(row=3, column=st_emc+1).value = headers[st_emc] # emissions cost
            ss.cell(row=3, column=st_lcc+1).value = headers[st_lcc] # LCOE with CO2
            ss.cell(row=3, column=st_max+1).value = headers[st_max] # max. MWh
            ss.cell(row=3, column=st_bal+1).value = headers[st_bal] # max. balance
            ss.cell(row=3, column=st_cac+1).value = headers[st_cac] # capital cost
            ss.cell(row=3, column=st_lic+1).value = headers[st_lic] # lifetime cost
            ss.cell(row=3, column=st_lie+1).value = headers[st_lie] # lifetime emissions
            ss.cell(row=3, column=st_lec+1).value = headers[st_lec] # lifetime emissions cost
            ss.cell(row=3, column=st_rlc+1).value = headers[st_rlc] # reference lcoe
            ss.cell(row=3, column=st_rcf+1).value = headers[st_rcf] # reference cf
            ss_row = 3
            ss_re_fst_row = 4
            fall_row = 9
            ns.cell(row=fall_row, column=2).value = 'Shortfall periods'
            max_row = 10
            ns.cell(row=max_row, column=2).value = 'Maximum (MW/MWh)'
            hrs_row = 11
            ns.cell(row=hrs_row, column=2).value = 'Hours of usage'
            if do_zone:
                zone_row = 12
                what_row = 13
                hrows = 14
                ns.cell(row=zone_row, column=1).value = 'Zone'
            else:
                what_row = 12
                hrows = 13
            ns.cell(row=what_row, column=1).value = 'Hour'
            ns.cell(row=what_row, column=2).value = 'Period'
            ns.cell(row=what_row, column=3).value = 'Load'
            ns.cell(row=sum_row, column=3).value = '=SUM(' + ss_col(3) + str(hrows) + \
                                                   ':' + ss_col(3) + str(hrows + 8759) + ')'
            ns.cell(row=sum_row, column=3).number_format = '#,##0'
            ns.cell(row=max_row, column=3).value = '=MAX(' + ss_col(3) + str(hrows) + \
                                                   ':' + ss_col(3) + str(hrows + 8759) + ')'
            ns.cell(row=max_row, column=3).number_format = '#,##0.00'
            o = 4
            col = 3
            # hour, period
            for row in range(hrows, 8760 + hrows):
                ns.cell(row=row, column=1).value = row - hrows + 1
                ns.cell(row=row, column=2).value = format_period(row - hrows)
            # and load
            load_col = pmss_details['Load'].col
            if pmss_details['Load'].multiplier == 1:
                for row in range(hrows, 8760 + hrows):
                    ns.cell(row=row, column=3).value = pmss_data[load_col][row - hrows]
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            else:
                for row in range(hrows, 8760 + hrows):
                    ns.cell(row=row, column=3).value = pmss_data[load_col][row - hrows] * \
                            pmss_details['Load'].multiplier
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            # here we're processing renewables (so no storage)
            for fac in re_order:
                if fac == 'Load':
                    continue
                if fac in underlying_facs:
                    continue
                if pmss_details[fac].col <= 0:
                    continue
                ss_row += 1
                col = do_detail(fac, col, ss_row)
                ns_tml_sum, ns_re_sum = do_detail_summary(fac, col, ss_row, ns_tml_sum, ns_re_sum)
            ss_re_lst_row = ss_row
            col += 1
            shrt_col = col
            ns.cell(row=fall_row, column=shrt_col).value = '=COUNTIF(' + ss_col(shrt_col) \
                            + str(hrows) + ':' + ss_col(shrt_col) + str(hrows + 8759) + \
                            ',"' + sf_test[0] + '0")'
            ns.cell(row=fall_row, column=shrt_col).number_format = '#,##0'
            ns.cell(row=what_row, column=shrt_col).value = 'Shortfall (' + sf_sign[0] \
                    + ') /\nSurplus (' + sf_sign[1] + ')'
            ns.cell(row=max_row, column=shrt_col).value = '=MAX(' + ss_col(shrt_col) + str(hrows) + \
                                           ':' + ss_col(shrt_col) + str(hrows + 8759) + ')'
            ns.cell(row=max_row, column=shrt_col).number_format = '#,##0.00'
            for col in range(3, shrt_col + 1):
                ns.cell(row=what_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                        vertical='bottom', horizontal='center')
                ns.cell(row=row, column=shrt_col).value = shortfall[row - hrows] * -surplus_sign
                for col in range(3, shrt_col + 1):
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            for row in range(hrows, 8760 + hrows):
                ns.cell(row=row, column=shrt_col).value = shortfall[row - hrows] * -surplus_sign
                ns.cell(row=row, column=col).number_format = '#,##0.00'
            col = shrt_col + 1
            ns.cell(row=tml_row, column=col).value = '=SUM(' + ss_col(col) + str(hrows) + \
                                                   ':' + ss_col(col) + str(hrows + 8759) + ')'
            ns.cell(row=tml_row, column=col).number_format = '#,##0'
            ns.cell(row=max_row, column=col).value = '=MAX(' + ss_col(col) + str(hrows) + \
                                           ':' + ss_col(col) + str(hrows + 8759) + ')'
            ns.cell(row=max_row, column=col).number_format = '#,##0.00'
            ns.cell(row=hrs_row, column=col).value = '=COUNTIF(' + ss_col(col) + str(hrows) + \
                                           ':' + ss_col(col) + str(hrows + 8759) + ',">0")'
            ns.cell(row=hrs_row, column=col).number_format = '#,##0'
            ns.cell(row=what_row, column=col).value = 'RE Contrib.\nto Load'
            ns.cell(row=what_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
            for row in range(hrows, 8760 + hrows):
                if shortfall[row - hrows] < 0:
                    if pmss_details['Load'].multiplier == 1:
                        rec = pmss_data[load_col][row - hrows]
                    else:
                        rec = pmss_data[load_col][row - hrows] * Decimal(pmss_details['Load'].multiplier)
                else:
                    if pmss_details['Load'].multiplier == 1:
                        rec = pmss_data[load_col][row - hrows] - shortfall[row - hrows]
                    else:
                        rec = pmss_data[load_col][row - hrows] * Decimal(pmss_details['Load'].multiplier) - \
                              shortfall[row - hrows]
                ns.cell(row=row, column=col).value = rec
               # the following formula will do the same computation
               # ns.cell(row=row, column=col).value = '=IF(' + ss_col(shrt_col) + str(row) + '>0,' + \
               #                            ss_col(3) + str(row) + ',' + ss_col(3) + str(row) + \
               #                            '+' + ss_col(shrt_col) + str(row) + ')'
                ns.cell(row=row, column=col).number_format = '#,##0.00'
          #  shrt_col += 1
           # col = shrt_col + 1
            ul_re_sum = ns_re_sum
            ul_tml_sum = ns_tml_sum
            nsul_sums = ['C']
            nsul_sum_cols = [3]
            for fac in underlying_facs:
                if pmss_details[fac].capacity * pmss_details[fac].multiplier == 0:
                    continue
                col = do_detail(fac, col, -1)
                nsul_sums.append(ss_col(col))
                nsul_sum_cols.append(col)
            if col > shrt_col + 1: # underlying
                col += 1
                ns.cell(row=what_row, column=col).value = 'Underlying\nLoad'
                ns.cell(row=what_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                        vertical='bottom', horizontal='center')
                ns.cell(row=sum_row, column=col).value = '=SUM(' + ss_col(col) + str(hrows) + \
                                                         ':' + ss_col(col) + str(hrows + 8759) + ')'
                ns.cell(row=sum_row, column=col).number_format = '#,##0'
                ns.cell(row=max_row, column=col).value = '=MAX(' + ss_col(col) + str(hrows) + \
                                                         ':' + ss_col(col) + str(hrows + 8759) + ')'
                ns.cell(row=max_row, column=col).number_format = '#,##0.00'
                for row in range(hrows, 8760 + hrows):
                    txt = '='
                    for c in nsul_sums:
                        txt += c + str(row) + '+'
                    ns.cell(row=row, column=col).value = txt[:-1]
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            next_col = col
            col += 1
        else:
            sp_data = []
            sp_load = Decimal(0) # load from load curve
            hrows = 10
            load_max = Decimal(0)
            load_hr = Decimal(0)
            load_col = Decimal(0)
            tml = Decimal(0)
            for fac in re_order:
                if fac in underlying_facs:
                    continue
                if fac == 'Load':
                    load_col = pmss_details[fac].col
                    sp_load = sum(pmss_data[load_col]) * pmss_details[fac].multiplier
                    load_max = 0
                    for h in range(len(pmss_data[0])):
                        amt = pmss_data[load_col][h] * pmss_details[fac].multiplier
                        if amt > load_max:
                            load_max = amt
                            load_hr = h
                    continue
                if pmss_details[fac].capacity * pmss_details[fac].multiplier == 0:
                    continue
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = fac
                sp_d[st_cap] = pmss_details[fac].capacity * pmss_details[fac].multiplier
                try:
                    sp_d[st_tml] = fac_tml[fac]
                except:
                    pass
                sp_d[st_sub] = sum(pmss_data[pmss_details[fac].col]) * pmss_details[fac].multiplier
                sp_d[st_max] = max(pmss_data[pmss_details[fac].col]) * pmss_details[fac].multiplier
                sp_data.append(sp_d)
            for h in range(len(shortfall)):
                if shortfall[h] < 0:
                    tml += Decimal(pmss_data[load_col][h]) * Decimal(pmss_details['Load'].multiplier)
                else:
                    tml += Decimal(pmss_data[load_col][h]) * Decimal(pmss_details['Load'].multiplier) - Decimal(shortfall[h])
        if option not in ['O', '1', 'B']:
            progress_text = "Operation in progress. Please wait."
        elif (option == 'O'):
            progress_text = "Optimisation in progress. Please wait."
        else:
            progress_text = "Iterative Batch run in progress. Please wait."
        storage_names = []
        # find any minimum generation for generators
        short_taken = {}
        short_taken_tot = 0
        for gen in dispatch_order:
            if pmss_details[gen].fac_type == 'G': # generators
                if generators[gen].capacity_min != 0:
                    try:
                        short_taken[gen] = pmss_details[gen].capacity * pmss_details[gen].multiplier * \
                            generators[gen].capacity_min
                    except:
                        short_taken[gen] = pmss_details[gen].capacity * \
                            generators[gen].capacity_min
                    short_taken_tot += short_taken[gen]
                    for row in range(8760):
                        shortfall[row] = shortfall[row] - short_taken[gen]
        tot_sto_loss = Decimal(0)
        for gen in dispatch_order:
         #   min_after = [0, 0, -1, 0, 0, 0] # initial, low balance, period, final, low after, period
         #  Min_after is there to see if storage is as full at the end as at the beginning
            try:
                capacity = pmss_details[gen].capacity * pmss_details[gen].multiplier
            except:
                try:
                    capacity = pmss_details[gen].capacity
                except:
                    continue
            if gen not in generators.keys():
                continue
            if generators[gen].category == 'Storage': # storage
                storage_names.append(gen)
                storage = [Decimal(0), Decimal(0), Decimal(0), Decimal(0)] # capacity, initial, min level, max drain
                storage[0] = capacity
                if option == 'D':
                    ns.cell(row=cap_row, column=col + 2).value = capacity
                    ns.cell(row=cap_row, column=col + 2).number_format = '#,##0.00'
                try:
                    storage[1] = generators[gen].initial * pmss_details[gen].multiplier
                except:
                    storage[1] = generators[gen].initial
                if generators[gen].capacity_min > 0:
                    storage[2] = capacity * generators[gen].capacity_min
                if generators[gen].capacity_max > 0:
                    storage[3] = capacity * generators[gen].capacity_max
                else:
                    storage[3] = capacity
                recharge = [0., 0.] # cap, loss
                if generators[gen].recharge_max > 0:
                    recharge[0] = capacity * generators[gen].recharge_max
                else:
                    recharge[0] = capacity
                if generators[gen].recharge_loss > 0:
                    recharge[1] = generators[gen].recharge_loss
                discharge = [Decimal(0), Decimal(0)] # cap, loss
                if generators[gen].discharge_max > 0:
                    discharge[0] = capacity * generators[gen].discharge_max
                if generators[gen].discharge_loss > 0:
                    discharge[1] = generators[gen].discharge_loss
                if generators[gen].parasitic_loss > 0:
                    parasite = generators[gen].parasitic_loss / 24.
                else:
                    parasite = Decimal(0)
                in_run = [False, False]
                min_run_time = generators[gen].min_runtime
                in_run[0] = True # start off in_run
                if min_run_time > 0 and generators[gen].initial == 0:
                    in_run[0] = False
                warm_time = generators[gen].warm_time
                storage_carry = storage[1] # generators[gen].initial
                if option == 'D':
                    ns.cell(row=ini_row, column=col + 2).value = storage_carry
                    ns.cell(row=ini_row, column=col + 2).number_format = '#,##0.00'
                storage_bal = []
                storage_can = Decimal(0)
                use_max = [Decimal(0), None]
                sto_max = storage_carry
                for row in range(8760):
                    storage_loss = Decimal(0)
                    storage_losses = Decimal(0)
                    if storage_carry > 0:
                        loss = storage_carry * parasite
                        # for later: record parasitic loss
                        storage_carry = storage_carry - loss
                        storage_losses -= loss
                    if shortfall[row] < 0:  # excess generation
                        if min_run_time > 0:
                            in_run[0] = False
                        if warm_time > 0:
                            in_run[1] = False
                        can_use = - (storage[0] - storage_carry) * (1 / (1 - recharge[1]))
                        if can_use < 0: # can use some
                            if shortfall[row] > can_use:
                                can_use = shortfall[row]
                            if can_use < - recharge[0] * (1 / (1 - recharge[1])):
                                can_use = - recharge[0]
                        else:
                            can_use = 0.
                        # for later: record recharge loss
                        storage_losses += can_use * recharge[1]
                        storage_carry -= (can_use * (1 - recharge[1]))
                        shortfall[row] -= can_use
                        if corr_data is not None:
                            corr_src[row] += can_use
                    else: # shortfall
                        if min_run_time > 0 and shortfall[row] > 0:
                            if not in_run[0]:
                                if row + min_run_time <= 8759:
                                    for i in range(row + 1, row + min_run_time + 1):
                                        if shortfall[i] <= 0:
                                            break
                                    else:
                                        in_run[0] = True
                        if in_run[0]:
                            can_use = shortfall[row] * (1 / (1 - discharge[1]))
                            can_use = min(can_use, discharge[0])
                            if can_use > storage_carry - storage[2]:
                                can_use = storage_carry - storage[2]
                            if warm_time > 0 and not in_run[1]:
                                in_run[1] = True
                                can_use = can_use * (1 - warm_time)
                        else:
                            can_use = 0
                        if can_use > 0:
                            storage_loss = can_use * discharge[1]
                            storage_losses -= storage_loss
                            storage_carry -= can_use
                            can_use = can_use - storage_loss
                            shortfall[row] -= can_use
                            if corr_data is not None:
                                corr_src[row] += can_use
                            if storage_carry < 0:
                                storage_carry = Decimal(0)
                        else:
                            can_use = Decimal(0)
                    if can_use < 0:
                        if use_max[1] is None or can_use < use_max[1]:
                            use_max[1] = can_use
                    elif can_use > use_max[0]:
                        use_max[0] = can_use
                    storage_bal.append(storage_carry)
                    if storage_bal[-1] > sto_max:
                        sto_max = storage_bal[-1]
                    if option == 'D':
                        if can_use > 0:
                            ns.cell(row=row + hrows, column=col).value = 0
                            ns.cell(row=row + hrows, column=col + 2).value = can_use * surplus_sign
                        else:
                            ns.cell(row=row + hrows, column=col).value = can_use * -surplus_sign
                            ns.cell(row=row + hrows, column=col + 2).value = 0
                        ns.cell(row=row + hrows, column=col + 1).value = storage_losses
                        ns.cell(row=row + hrows, column=col + 3).value = storage_carry
                        ns.cell(row=row + hrows, column=col + 4).value = (shortfall[row] + short_taken_tot) * -surplus_sign
                        for ac in range(5):
                            ns.cell(row=row + hrows, column=col + ac).number_format = '#,##0.00'
                            ns.cell(row=max_row, column=col + ac).value = '=MAX(' + ss_col(col + ac) + \
                                    str(hrows) + ':' + ss_col(col + ac) + str(hrows + 8759) + ')'
                            ns.cell(row=max_row, column=col + ac).number_format = '#,##0.00'
                    else:
                        tot_sto_loss += storage_losses
                        if can_use > 0:
                            storage_can += can_use
                if option == 'D':
                    ns.cell(row=sum_row, column=col).value = '=SUMIF(' + ss_col(col) + \
                            str(hrows) + ':' + ss_col(col) + str(hrows + 8759) + ',">0")'
                    ns.cell(row=sum_row, column=col).number_format = '#,##0'
                    ns.cell(row=sum_row, column=col + 1).value = '=SUMIF(' + ss_col(col + 1) + \
                            str(hrows) + ':' + ss_col(col + 1) + str(hrows + 8759) + ',"<0")'
                    ns.cell(row=sum_row, column=col + 1).number_format = '#,##0'
                    ns.cell(row=sum_row, column=col + 2).value = '=SUMIF(' + ss_col(col + 2) + \
                            str(hrows) + ':' + ss_col(col + 2) + str(hrows + 8759) + ',">0")'
                    ns.cell(row=sum_row, column=col + 2).number_format = '#,##0'
                    ns.cell(row=cf_row, column=col + 2).value = '=IF(' + ss_col(col + 2) + str(cap_row) + '>0,' + \
                            ss_col(col + 2) + str(sum_row) + '/' + ss_col(col + 2) + '1/8760,"")'
                    ns.cell(row=cf_row, column=col + 2).number_format = '#,##0.0%'
                    ns.cell(row=max_row, column=col).value = '=MAX(' + ss_col(col) + \
                            str(hrows) + ':' + ss_col(col) + str(hrows + 8759) + ')'
                    ns.cell(row=max_row, column=col).number_format = '#,##0.00'
                    ns.cell(row=hrs_row, column=col + 2).value = '=COUNTIF(' + ss_col(col + 2) + \
                            str(hrows) + ':' + ss_col(col + 2) + str(hrows + 8759) + ',">0")'
                    ns.cell(row=hrs_row, column=col + 2).number_format = '#,##0'
                    ns.cell(row=hrs_row, column=col + 3).value = '=' + ss_col(col + 2) + \
                            str(hrs_row) + '/8760'
                    ns.cell(row=hrs_row, column=col + 3).number_format = '#,##0.0%'
                    col += 5
                else:
                    if storage[0] == 0:
                        continue
               #     tml_tot += storage_can
                    sp_d = [' '] * len(headers)
                    sp_d[st_fac] = gen
                    sp_d[st_cap] = storage[0]
                    sp_d[st_tml] = storage_can
                    sp_d[st_max] = use_max[0]
                    sp_d[st_bal] = sto_max
                    sp_data.append(sp_d)
            else: # generator
                try:
                    if generators[gen].capacity_max > 0:
                        cap_capacity = capacity * generators[gen].capacity_max
                    else:
                        cap_capacity = capacity
                except:
                    cap_capacity = capacity
                if gen in short_taken.keys():
                    for row in range(8760):
                        shortfall[row] = shortfall[row] + short_taken[gen]
                    short_taken_tot -= short_taken[gen]
                    min_gen = short_taken[gen]
                else:
                    min_gen = 0
                if option == 'D':
                    ns.cell(row=cap_row, column=col).value = capacity
                    ns.cell(row=cap_row, column=col).number_format = '#,##0.00'
                    for row in range(8760):
                        if shortfall[row] >= 0: # shortfall?
                            if shortfall[row] >= cap_capacity:
                                shortfall[row] = shortfall[row] - cap_capacity
                                ns.cell(row=row + hrows, column=col).value = cap_capacity
                            elif shortfall[row] < min_gen:
                                ns.cell(row=row + hrows, column=col).value = min_gen
                                shortfall[row] -= min_gen
                            else:
                                ns.cell(row=row + hrows, column=col).value = shortfall[row]
                                shortfall[row] = 0
                        else:
                            shortfall[row] -= min_gen
                            ns.cell(row=row + hrows, column=col).value = min_gen
                        ns.cell(row=row + hrows, column=col + 1).value = (shortfall[row] + short_taken_tot) * -surplus_sign
                        ns.cell(row=row + hrows, column=col).number_format = '#,##0.00'
                        ns.cell(row=row + hrows, column=col + 1).number_format = '#,##0.00'
                    ns.cell(row=sum_row, column=col).value = '=SUM(' + ss_col(col) + str(hrows) + \
                            ':' + ss_col(col) + str(hrows + 8759) + ')'
                    ns.cell(row=sum_row, column=col).number_format = '#,##0'
                    ns.cell(row=cf_row, column=col).value = '=IF(' + ss_col(col) + str(cap_row) + '>0,' + \
                            ss_col(col) + str(sum_row) + '/' + ss_col(col) + str(cap_row) + '/8760,"")'
                    ns.cell(row=cf_row, column=col).number_format = '#,##0.0%'
                    ns.cell(row=max_row, column=col).value = '=MAX(' + ss_col(col) + \
                                str(hrows) + ':' + ss_col(col) + str(hrows + 8759) + ')'
                    ns.cell(row=max_row, column=col).number_format = '#,##0.00'
                    ns.cell(row=hrs_row, column=col).value = '=COUNTIF(' + ss_col(col) + \
                            str(hrows) + ':' + ss_col(col) + str(hrows + 8759) + ',">0")'
                    ns.cell(row=hrs_row, column=col).number_format = '#,##0'
                    ns.cell(row=hrs_row, column=col + 1).value = '=' + ss_col(col) + \
                            str(hrs_row) + '/8760'
                    ns.cell(row=hrs_row, column=col + 1).number_format = '#,##0.0%'
                    col += 2
                else:
                    gen_can = Decimal(0)
                    gen_max = Decimal(0)
                    for row in range(8760):
                        if shortfall[row] >= 0: # shortfall?
                            if shortfall[row] >= cap_capacity:
                                shortfall[row] = shortfall[row] - cap_capacity
                                gen_can += cap_capacity
                                if cap_capacity > gen_max:
                                    gen_max = cap_capacity
                            elif shortfall[row] < min_gen:
                                gen_can += min_gen
                                if min_gen > gen_max:
                                    gen_max = min_gen
                                shortfall[row] -= min_gen
                            else:
                                gen_can += shortfall[row]
                                if shortfall[row] > gen_max:
                                    gen_max = shortfall[row]
                                shortfall[row] = 0
                        else:
                            if min_gen > gen_max:
                                gen_max = min_gen
                            gen_can += min_gen
                            shortfall[row] -= min_gen # ??
                    if capacity == 0:
                        continue
                    sp_d = [' '] * len(headers)
                    sp_d[st_fac] = gen
                    sp_d[st_cap] = capacity
                    sp_d[st_tml] = gen_can
                    sp_d[st_sub] = gen_can
                    sp_d[st_max] = gen_max
                    sp_data.append(sp_d)
#        if option == 'D': # Currently calculated elsewhere
#            if surplus_sign > 0:
#                maxmin = 'MIN'
#            else:
#                maxmin = 'MAX'
#            ns.cell(row=max_row, column=col-1).value = '=' + maxmin + '(' + \
#                    ss_col(col-1) + str(hrows) + ':' + ss_col(col - 1) + str(hrows + 8759) + ')'
#            ns.cell(row=max_row, column=col-1).number_format = '#,##0.00'
        # if option not in ['O', '1', 'B']:
        #     self.progressbar.progress(8, text=progress_text)
        if corr_data is not None:
            try:
                corr = np.corrcoef(df1, corr_src)
                if np.isnan(corr.item((0, 1))):
                    corr = Decimal(0)
                else:
                    corr = corr.item((0, 1))
            except:
                corr = Decimal(0)
            corr_data.append(['RE plus Storage', corr])
            col = pmss_details['Load'].col
            corr_src = []
            for h in range(len(shortfall)):
                if shortfall[h] < 0:
                    corr_src.append(pmss_data[col][h])
                else:
                    corr_src.append(pmss_data[col][h] - shortfall[h])
            try:
                corr = np.corrcoef(df1, corr_src)
                if np.isnan(corr.item((0, 1))):
                    corr = Decimal(0)
                else:
                    corr = corr.item((0, 1))
            except:
                corr = Decimal(0)
            corr_data.append(['To Meet Load', corr])
            for c in range(1, len(corr_data)):
                if abs(corr_data[c][1]) < 0.1:
                    corr_data[c].append('None')
                elif abs(corr_data[c][1]) < 0.3:
                    corr_data[c].append('Little if any')
                elif abs(corr_data[c][1]) < 0.5:
                    corr_data[c].append('Low')
                elif abs(corr_data[c][1]) < 0.7:
                    corr_data[c].append('Moderate')
                elif abs(corr_data[c][1]) < 0.9:
                    corr_data[c].append('High')
                else:
                    corr_data[c].append('Very high')
        if option != 'D':
            load_col = pmss_details['Load'].col
            cap_sum = Decimal(0)
            gen_sum = Decimal(0)
            re_sum = Decimal(0)
            tml_sum = Decimal(0)
            ff_sum = Decimal(0)
            sto_sum = Decimal(0)
            cost_sum = Decimal(0)
            co2_sum = Decimal(0.0)
            co2_cost_sum = Decimal(0)
            capex_sum = Decimal(0)
            lifetime_sum = Decimal(0)
            lifetime_co2_sum = Decimal(0)
            lifetime_co2_cost = Decimal(0)
            for sp in range(len(sp_data)):
                gen = sp_data[sp][st_fac]
                if gen in storage_names:
                    sto_sum += sp_data[sp][2]
                else:
                    try:
                        gen2 = gen[gen.find('.') + 1:]
                    except:
                        gen2 = gen
                    if gen in tech_names or gen2 in tech_names:
                        re_sum += sp_data[sp][st_sub]
            for sp in range(len(sp_data)):
                gen = sp_data[sp][st_fac]
                if gen in storage_names:
                    ndx = 2
                else:
                    if gen in generators.keys():
                        pass
                    else:
                        try:
                            gen = gen[gen.find('.') + 1:]
                        except:
                            pass
                    ndx = 3
                try:
                    if sp_data[sp][st_cap] > 0:
                        cap_sum += sp_data[sp][st_cap]
                        if generators[gen].lcoe > 0:
                            sp_data[sp][st_cfa] = sp_data[sp][ndx] / sp_data[sp][st_cap] / 8760 # need number for now
                        else:
                            sp_data[sp][st_cfa] = '{:.1f}%'.format(sp_data[sp][ndx] / sp_data[sp][st_cap] / 8760 * 100)
                    gen_sum += sp_data[sp][st_sub]
                except:
                    pass
                try:
                    tml_sum += sp_data[sp][st_tml]
                except:
                    pass
                if gen not in generators.keys():
                    continue
                ndx = 3
                if gen in storage_names:
                    ndx = 2
                else:
                    try:
                        gen2 = gen[gen.find('.') + 1:]
                    except:
                        gen2 = gen
                    if gen not in tech_names and gen2 not in tech_names:
                        ff_sum += sp_data[sp][ndx]
                if not generators[gen].fuel:
                    generators[gen].fuel = Decimal(0)
                if generators[gen].capex > 0 or generators[gen].fixed_om > 0 \
                  or generators[gen].variable_om > 0 or generators[gen].fuel > 0:
                    capex = sp_data[sp][st_cap] * generators[gen].capex
                    capex_sum += capex
                    opex = sp_data[sp][st_cap] * generators[gen].fixed_om \
                           + sp_data[sp][ndx] * generators[gen].variable_om \
                           + sp_data[sp][ndx] * generators[gen].fuel
                    disc_rate = generators[gen].disc_rate
                    if disc_rate == 0:
                        disc_rate = Decimal(settings['Powermatch']['discount_rate'])
                    lifetime = generators[gen].lifetime
                    if sp_data[sp][ndx] > 0:
                        sp_data[sp][st_lcg] = calcLCOE(sp_data[sp][ndx], capex, opex, disc_rate, lifetime)
                    if sp_data[sp][st_lcg] != ' ':
                        sp_data[sp][st_cst] = sp_data[sp][ndx] * sp_data[sp][st_lcg]
                    if (gen in tech_names and fac_tml_sum > 0):
                        sp_data[sp][st_lco] = sp_data[sp][st_cst] / (sp_data[sp][st_tml] + (sto_sum * sp_data[sp][st_tml] / fac_tml_sum))
                    else:
                        sp_data[sp][st_lco] = sp_data[sp][st_lcg]
                    if sp_data[sp][st_cst] != ' ':
                        cost_sum += sp_data[sp][st_cst]
                    sp_data[sp][st_cac] = capex
                elif generators[gen].lcoe > 0:
                    if generators[gen].lcoe_cf > 0:
                        lcoe_cf = generators[gen].lcoe_cf
                    else:
                        lcoe_cf = sp_data[sp][st_cfa]
                    sp_data[sp][st_cst] = generators[gen].lcoe * lcoe_cf * 8760 * sp_data[sp][st_cap]
                    if sp_data[sp][st_cfa] > 0:
                        sp_data[sp][st_lcg] = sp_data[sp][st_cst] / sp_data[sp][ndx]
                        sp_data[sp][st_lco] = sp_data[sp][st_lcg]
                    sp_data[sp][st_cfa] = '{:.1f}%'.format(sp_data[sp][st_cfa] * Decimal(100))
                    cost_sum += sp_data[sp][st_cst]
                    sp_data[sp][st_rlc] = generators[gen].lcoe
                    sp_data[sp][st_rcf] = '{:.1f}%'.format(lcoe_cf * Decimal(100))
                elif generators[gen].lcoe_cf == 0: # no cost facility
                    lcoe_cf = sp_data[sp][st_cfa]
                    sp_data[sp][st_cst] = 0
                    cost_sum += sp_data[sp][st_cst]
                if sp_data[sp][st_cst] != ' ':
                    sp_data[sp][st_lic] = sp_data[sp][st_cst] * max_lifetime
                if sp_data[sp][st_lic] != ' ':
                    lifetime_sum += sp_data[sp][st_lic]
                if generators[gen].emissions > 0:
                    sp_data[sp][st_emi] = sp_data[sp][ndx] * generators[gen].emissions
                    co2_sum += sp_data[sp][st_emi]
                    sp_data[sp][st_emc] = sp_data[sp][st_emi] * carbon_price
                    if sp_data[sp][st_cst] == 0:
                        sp_data[sp][st_lcc] = sp_data[sp][st_emc] / sp_data[sp][st_tml]
                    elif sp_data[sp][st_cst] != ' ':
                        sp_data[sp][st_lcc] = sp_data[sp][st_lco] * ((sp_data[sp][st_cst] + sp_data[sp][st_emc]) / sp_data[sp][st_cst])
                    co2_cost_sum += sp_data[sp][st_emc]
                    sp_data[sp][st_lie] = sp_data[sp][st_emi] * max_lifetime
                    lifetime_co2_sum += sp_data[sp][st_lie]
                    sp_data[sp][st_lec] = sp_data[sp][st_lie] * carbon_price
                    lifetime_co2_cost += sp_data[sp][st_lec]
                else:
                    sp_data[sp][st_lcc] = sp_data[sp][st_lco]
            sf_sums = [Decimal(0), Decimal(0), Decimal(0)]
            for sf in range(len(shortfall)):
                if shortfall[sf] > 0:
                    sf_sums[0] += shortfall[sf]
                    sf_sums[2] += pmss_data[load_col][sf] * pmss_details['Load'].multiplier
                else:
                    sf_sums[1] += shortfall[sf]
                    sf_sums[2] += pmss_data[load_col][sf] * pmss_details['Load'].multiplier
            if gen_sum > 0:
                gs = cost_sum / gen_sum
            else:
                gs = ''
            if tml_sum > 0:
                gsw = cost_sum / tml_sum # LCOE
                gswc = (cost_sum + co2_cost_sum) / tml_sum
            else:
                gsw = ''
                gswc = ''
            if option == 'O' or option == '1':
                load_pct, surp_pct, re_pct = summary_totals()
            else:
                summary_totals()
            do_underlying = False
            if len(underlying_facs) > 0:
                for fac in underlying_facs:
                    if pmss_details[fac].capacity * pmss_details[fac].multiplier > 0:
                        do_underlying = True
                        break
            if do_underlying:
                sp_data.append(' ')
                sp_data.append('Additional Underlying Load')
                for fac in underlying_facs:
                    if pmss_details[fac].capacity * pmss_details[fac].multiplier == 0:
                        continue
                    if fac in generators.keys():
                        gen = fac
                    else:
                        gen = pmss_details[fac].generator
                    col = pmss_details[fac].col
                    sp_d = [' '] * len(headers)
                    sp_d[st_fac] = fac
                    sp_d[st_cap] = pmss_details[fac].capacity * pmss_details[fac].multiplier
                    cap_sum += sp_d[st_cap]
                    sp_d[st_tml] = sum(pmss_data[pmss_details[fac].col]) * pmss_details[fac].multiplier
                    tml_sum += sp_d[st_tml]
                    sp_d[st_sub] = sp_d[st_tml]
                    gen_sum += sp_d[st_tml]
                    sp_load += sp_d[st_tml]
                    sp_d[st_cfa] = '{:.1f}%'.format(sp_d[st_sub] / sp_d[st_cap] / 8760 * Decimal(100))
                    sp_d[st_max] = max(pmss_data[pmss_details[fac].col]) * pmss_details[fac].multiplier
                    if generators[gen].capex > 0 or generators[gen].fixed_om > 0 \
                      or generators[gen].variable_om > 0 or generators[gen].fuel > 0:
                        capex = sp_d[st_cap] * generators[gen].capex
                        capex_sum += capex
                        opex = sp_d[st_cap] * generators[gen].fixed_om \
                               + sp_d[st_tml] * generators[gen].variable_om \
                               + sp_d[st_tml] * generators[gen].fuel
                        disc_rate = generators[gen].disc_rate
                        if disc_rate == 0:
                            disc_rate = discount_rate
                        lifetime = generators[gen].lifetime
                        sp_d[st_lcg] = calcLCOE(sp_d[st_tml], capex, opex, disc_rate, lifetime)
                        sp_d[st_cst] = sp_d[st_tml] * sp_d[st_lcg]
                        cost_sum += sp_d[st_cst]
                        sp_d[st_lco] = sp_d[st_lcg]
                        sp_d[st_cac] = capex
                    elif generators[gen].lcoe > 0:
                        if generators[gen].lcoe_cf > 0:
                            lcoe_cf = generators[gen].lcoe_cf
                        else:
                            lcoe_cf = sp_d[st_cfa]
                        sp_d[st_cst] = generators[gen].lcoe * lcoe_cf * 8760 * sp_d[st_cap]
                        cost_sum += sp_d[st_cst]
                        if sp_d[st_cfa] > 0:
                            sp_d[st_lcg] = sp_d[st_cst] / sp_d[st_tml]
                            sp_d[st_lco] = sp_d[st_lcg]
                        sp_d[st_cfa] = '{:.1f}%'.format(sp_d[st_cfa] * Decimal(100))
                        sp_d[st_rlc] = generators[gen].lcoe
                        sp_d[st_rcf] = '{:.1f}%'.format(lcoe_cf * Decimal(100))
                    elif generators[gen].lcoe_cf == 0: # no cost facility
                        sp_d[st_cst] = 0
                        sp_d[st_lcg] = 0
                        sp_d[st_lco] = 0
                        sp_d[st_rlc] = generators[gen].lcoe
                    sp_d[st_lic] = sp_d[st_cst] * max_lifetime
                    lifetime_sum += sp_d[st_lic]
                    if generators[gen].emissions > 0:
                        sp_d[st_emi] = sp_d[st_tml] * generators[gen].emissions
                        co2_sum += sp_d[st_emi]
                        sp_d[st_emc] = sp_d[st_emi] * carbon_price
                        if sp_d[st_cst] > 0:
                            sp_d[st_lcc] = sp_d[st_lco] * ((sp_d[st_cst] + sp_d[st_emc]) / sp_d[st_cst])
                        else:
                            sp_d[st_lcc] = sp_d[st_emc] / sp_d[st_tml]
                        co2_cost_sum += sp_d[st_emc]
                        sp_d[st_lie] = sp_d[st_emi] * max_lifetime
                        lifetime_co2_sum += sp_d[st_lie]
                        sp_d[st_lec] = sp_d[st_lie] * carbon_price
                        lifetime_co2_cost += sp_d[st_lec]
                    else:
                        sp_d[st_lcc] = sp_d[st_lco]
                    sp_data.append(sp_d)
                if gen_sum > 0:
                    gs = cost_sum / gen_sum
                else:
                    gs = ''
                if tml_sum > 0:
                    gsw = cost_sum / tml_sum # LCOE
                    gswc = (cost_sum + co2_cost_sum) / tml_sum
                else:
                    gsw = ''
                    gswc = ''
                # find maximum underlying load
                if option == 'S':
                    load_max = 0
                    load_hr = 0
                    for h in range(len(pmss_data[0])):
                        amt = pmss_data[load_col][h] * pmss_details['Load'].multiplier
                        for fac in underlying_facs:
                            amt += pmss_data[pmss_details[fac].col][h] * pmss_details[fac].multiplier
                        if amt > load_max:
                            load_max = amt
                            load_hr = h
                summary_totals('Underlying ')
            if corr_data is not None:
                sp_data.append(' ')
                sp_data = sp_data + corr_data
            sp_data.append(' ')
            sp_data.append(['Static Variables'])
            if carbon_price > 0:
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = 'Carbon Price ($/tCO2e)'
                sp_d[st_cap] = carbon_price
                sp_data.append(sp_d)
            sp_d = [' '] * len(headers)
            sp_d[st_fac] = 'Lifetime (years)'
            sp_d[st_cap] = max_lifetime
            sp_data.append(sp_d)
            sp_d = [' '] * len(headers)
            sp_d[st_fac] = 'Discount Rate'
            sp_d[st_cap] = '{:.2%}'.format(discount_rate)
            sp_data.append(sp_d)
            if option == 'B':
                if optimise_debug:
                    sp_pts = [0] * len(headers)
                    for p in [st_cap, st_lcg, st_lco, st_lcc, st_max, st_bal, st_rlc]:
                        sp_pts[p] = 2
                    if corr_data is not None:
                        sp_pts[st_cap] = 3 # compromise between capacity (2) and correlation (4)
                    dialog = displaytable.Table(sp_data, title=title, fields=headers,
                             save_folder=scenarios, sortby='', decpts=sp_pts)
                    dialog.exec_()
                # return sp_data
            if option == 'O' or option == '1':
                op_load_tot = pmss_details['Load'].capacity * pmss_details['Load'].multiplier
                if gswc != '':
                    lcoe = gswc
                elif adjusted_lcoe:
                    lcoe = gsw # target is lcoe
                else:
                    lcoe = gs
                if gen_sum == 0:
                    re_pct = 0
                    load_pct = 0
                    re_pct = 0
                multi_value = {'lcoe': lcoe, #lcoe. lower better
                    'load_pct': load_pct, #load met. 100% better
                    'surplus_pct': surp_pct, #surplus. lower better
                    're_pct': re_pct, # RE pct. higher better
                    'cost': cost_sum, # cost. lower better
                    'co2': co2_sum} # CO2. lower better
                if option == 'O':
                    if multi_value['lcoe'] == '':
                        multi_value['lcoe'] = 0
                    return multi_value, sp_data, None
                else:
                    extra = [gsw, op_load_tot, sto_sum, re_sum, re_pct, sf_sums]
                    return multi_value, sp_data, extra
        #    list(map(list, list(zip(*sp_data))))

        #    list(map(list, list(zip(*sp_data))))
            sp_pts = [0] * len(headers)
            for p in [st_cap, st_lcg, st_lco, st_lcc, st_max, st_bal, st_rlc]:
                sp_pts[p] = 2
            if corr_data is not None:
                sp_pts[st_cap] = 3 # compromise between capacity (2) and correlation (4)
            if title is not None:
                atitle = title
            elif results_prefix != '':
                atitle = results_prefix + '_' + ' '
            else:
                pass

            #dialog = displaytable.Table(sp_data, title='title', fields=headers,
            #         save_folder=scenarios, sortby='', decpts=sp_pts)

            return sp_data, headers, sp_pts# finish if not detailed spreadsheet
        col = next_col + 1
        is_storage = False
        ss_sto_rows = []
        ss_st_row = -1
        for gen in dispatch_order:
            ss_row += 1
            try:
                if generators[gen].category == 'Storage':
                    ss_sto_rows.append(ss_row)
                    nc = 2
                    ns.cell(row=what_row, column=col).value = 'Charge\n' + gen
                    ns.cell(row=what_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                            vertical='bottom', horizontal='center')
                    ns.cell(row=what_row, column=col + 1).value = gen + '\nLosses'
                    ns.cell(row=what_row, column=col + 1).alignment = oxl.styles.Alignment(wrap_text=True,
                            vertical='bottom', horizontal='center')
                    is_storage = True
                    ns_sto_sum += '+' + ss_col(st_tml+1) + str(ss_row)
                    ns_loss_sum += '+Detail!' + ss_col(col + 1) + str(sum_row)
                else:
                    nc = 0
                    is_storage = False
                    ns_not_sum += '-' + ss_col(st_tml+1) + str(ss_row)
            except KeyError as err:
                msg = 'Key Error: No Constraint for ' + gen
                if title is not None:
                    msg += ' (model ' + title + ')'
                nc = 0
                is_storage = False
                ns_not_sum += '-' + ss_col(st_tml+1) + str(ss_row)
            ns.cell(row=what_row, column=col + nc).value = gen
            ss.cell(row=ss_row, column=st_fac+1).value = '=Detail!' + ss_col(col + nc) + str(what_row)
            # facility
            ss.cell(row=ss_row, column=st_cap+1).value = '=Detail!' + ss_col(col + nc) + str(cap_row)
            # capacity
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.00'
            # tml
            ss.cell(row=ss_row, column=st_tml+1).value = '=Detail!' + ss_col(col + nc) + str(sum_row)
            ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
            # subtotal
            try:
                if generators[gen].category != 'Storage':
                    ss.cell(row=ss_row, column=st_sub+1).value = '=Detail!' + ss_col(col + nc) + str(sum_row)
                    ss.cell(row=ss_row, column=st_sub+1).number_format = '#,##0'
            except KeyError as err:
                ss.cell(row=ss_row, column=st_sub+1).value = '=Detail!' + ss_col(col + nc) + str(sum_row)
                ss.cell(row=ss_row, column=st_sub+1).number_format = '#,##0'
            # cf
            ss.cell(row=ss_row, column=st_cfa+1).value = '=Detail!' + ss_col(col + nc) + str(cf_row)
            ss.cell(row=ss_row, column=st_cfa+1).number_format = '#,##0.0%'
            if generators[gen].capex > 0 or generators[gen].fixed_om > 0 \
              or generators[gen].variable_om > 0 or generators[gen].fuel > 0:
                disc_rate = generators[gen].disc_rate
                if disc_rate == 0:
                    disc_rate = discount_rate
                if disc_rate == 0:
                    cst_calc = '/' + str(generators[gen].lifetime)
                else:
                    pwr_calc = 'POWER(1+' + str(disc_rate) + ',' + str(generators[gen].lifetime) + ')'
                    cst_calc = '*' + str(disc_rate) + '*' + pwr_calc + '/SUM(' + pwr_calc + ',-1)'
                ns.cell(row=cost_row, column=col + nc).value = '=IF(' + ss_col(col + nc) + str(cf_row) + \
                        '>0,' + ss_col(col + nc) + str(cap_row) + '*' + str(generators[gen].capex) + \
                        cst_calc + '+' + ss_col(col + nc) + str(cap_row) + '*' + \
                        str(generators[gen].fixed_om) + '+' + ss_col(col + nc) + str(sum_row) + '*(' + \
                        str(generators[gen].variable_om) + '+' + str(generators[gen].fuel) + \
                        '),0)'
                ns.cell(row=cost_row, column=col + nc).number_format = '$#,##0'
                # cost / yr
                ss.cell(row=ss_row, column=st_cst+1).value = '=Detail!' + ss_col(col + nc) + str(cost_row)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                ns.cell(row=lcoe_row, column=col + nc).value = '=IF(AND(' + ss_col(col + nc) + str(cf_row) + \
                        '>0,' + ss_col(col + nc) + str(cap_row) + '>0),' + ss_col(col + nc) + \
                        str(cost_row) + '/' + ss_col(col + nc) + str(sum_row) + ',"")'
                ns.cell(row=lcoe_row, column=col + nc).number_format = '$#,##0.00'
                # lcog
                ss.cell(row=ss_row, column=st_lcg+1).value = '=Detail!' + ss_col(col + nc) + str(lcoe_row)
                ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
                # lcoe
                ss.cell(row=ss_row, column=st_lco+1).value = '=Detail!' + ss_col(col + nc) + str(lcoe_row)
                ss.cell(row=ss_row, column=st_lco+1).number_format = '$#,##0.00'
                # capital cost
                ss.cell(row=ss_row, column=st_cac+1).value = '=IF(Detail!' + ss_col(col + nc) + str(cap_row) \
                                                            + '>0,Detail!' + ss_col(col + nc) + str(cap_row) + '*'  \
                                                            + str(generators[gen].capex) + ',"")'
                ss.cell(row=ss_row, column=st_cac+1).number_format = '$#,##0'
            elif generators[gen].lcoe > 0:
                ns.cell(row=cost_row, column=col + nc).value = '=IF(' + ss_col(col + nc) + str(cf_row) + \
                        '>0,' + ss_col(col + nc) + str(sum_row) + '*Summary!' + ss_col(st_rlc + 1) + str(ss_row) + \
                        '*Summary!' + ss_col(st_rcf + 1) + str(ss_row) + '/' + ss_col(col + nc) + str(cf_row) + ',0)'
                ns.cell(row=cost_row, column=col + nc).number_format = '$#,##0'
                # cost / yr
                ss.cell(row=ss_row, column=st_cst+1).value = '=Detail!' + ss_col(col + nc) + str(cost_row)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                ns.cell(row=lcoe_row, column=col + nc).value = '=IF(AND(' + ss_col(col + nc) + str(cf_row) + '>0,' \
                            + ss_col(col + nc) + str(cap_row) + '>0),' + ss_col(col + nc) + str(cost_row) + '/8760/' \
                            + ss_col(col + nc) + str(cf_row) + '/' + ss_col(col + nc) + str(cap_row)+  ',"")'
                ns.cell(row=lcoe_row, column=col + nc).number_format = '$#,##0.00'
                # lcog
                ss.cell(row=ss_row, column=st_lcg+1).value = '=Detail!' + ss_col(col + nc) + str(lcoe_row)
                ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
                # lcoe
                ss.cell(row=ss_row, column=st_lco+1).value = '=Detail!' + ss_col(col + nc) + str(lcoe_row)
                ss.cell(row=ss_row, column=st_lco+1).number_format = '$#,##0.00'
                # ref lcoe
                ss.cell(row=ss_row, column=st_rlc+1).value = generators[gen].lcoe
                ss.cell(row=ss_row, column=st_rlc+1).number_format = '$#,##0.00'
                # ref cf
                if generators[gen].lcoe_cf == 0:
                    ss.cell(row=ss_row, column=st_rcf+1).value = '=' + ss_col(st_cfa+1) + str(ss_row)
                else:
                    ss.cell(row=ss_row, column=st_rcf+1).value = generators[gen].lcoe_cf
                ss.cell(row=ss_row, column=st_rcf+1).number_format = '#,##0.0%'
            elif generators[gen].lcoe_cf == 0: # no cost facility
                ns.cell(row=cost_row, column=col + nc).value = '=IF(' + ss_col(col + nc) + str(cf_row) + \
                        '>0,' + ss_col(col + nc) + str(sum_row) + '*Summary!' + ss_col(st_rlc + 1) + str(ss_row) + \
                        '*Summary!' + ss_col(st_rcf + 1) + str(ss_row) + '/' + ss_col(col + nc) + str(cf_row) + ',0)'
                ns.cell(row=cost_row, column=col + nc).number_format = '$#,##0'
                # cost / yr
                ss.cell(row=ss_row, column=st_cst+1).value = '=Detail!' + ss_col(col + nc) + str(cost_row)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                ns.cell(row=lcoe_row, column=col + nc).value = '=IF(AND(' + ss_col(col + nc) + str(cf_row) + '>0,' \
                            + ss_col(col + nc) + str(cap_row) + '>0),' + ss_col(col + nc) + str(cost_row) + '/8760/' \
                            + ss_col(col + nc) + str(cf_row) + '/' + ss_col(col + nc) + str(cap_row)+  ',"")'
                ns.cell(row=lcoe_row, column=col + nc).number_format = '$#,##0.00'
                # lcog
                ss.cell(row=ss_row, column=st_lcg+1).value = '=Detail!' + ss_col(col + nc) + str(lcoe_row)
                ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
                # lcoe
                ss.cell(row=ss_row, column=st_lco+1).value = '=Detail!' + ss_col(col + nc) + str(lcoe_row)
                ss.cell(row=ss_row, column=st_lco+1).number_format = '$#,##0.00'
                # ref lcoe
                ss.cell(row=ss_row, column=st_rlc+1).value = generators[gen].lcoe
                ss.cell(row=ss_row, column=st_rlc+1).number_format = '$#,##0.00'
                # ref cf
                if generators[gen].lcoe_cf == 0:
                    ss.cell(row=ss_row, column=st_rcf+1).value = '=' + ss_col(st_cfa+1) + str(ss_row)
                else:
                    ss.cell(row=ss_row, column=st_rcf+1).value = generators[gen].lcoe_cf
                ss.cell(row=ss_row, column=st_rcf+1).number_format = '#,##0.0%'
            if generators[gen].emissions > 0:
                ns.cell(row=emi_row, column=col + nc).value = '=' + ss_col(col + nc) + str(sum_row) \
                        + '*' + str(generators[gen].emissions)
                ns.cell(row=emi_row, column=col + nc).number_format = '#,##0'
                # emissions
                ss.cell(row=ss_row, column=st_emi+1).value = '=Detail!' + ss_col(col + nc) + str(emi_row)
                ss.cell(row=ss_row, column=st_emi+1).number_format = '#,##0'
                if carbon_price > 0:
                    ss.cell(row=ss_row, column=st_emc+1).value = '=IF(' + ss_col(st_emi+1) + str(ss_row) + '>0,' + \
                                                                 ss_col(st_emi+1) + str(ss_row) + '*carbon_price,"")'
                    ss.cell(row=ss_row, column=st_emc+1).number_format = '$#,##0'
            # max mwh
            ss.cell(row=ss_row, column=st_max+1).value = '=Detail!' + ss_col(col + nc) + str(max_row)
            ss.cell(row=ss_row, column=st_max+1).number_format = '#,##0.00'
            # max balance
            if nc > 0: # storage
                ss.cell(row=ss_row, column=st_bal+1).value = '=Detail!' + ss_col(col + nc + 1) + str(max_row)
                ss.cell(row=ss_row, column=st_bal+1).number_format = '#,##0.00'
            ns.cell(row=what_row, column=col + nc).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
            ns.cell(row=what_row, column=col + nc + 1).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
            if is_storage:
                # lifetime cost
                ss.cell(row=ss_row, column=st_lic+1).value = '=IF(Detail!' + ss_col(col + 2) + str(sum_row) \
                                                        + '>0,Detail!' + ss_col(col + 2) + str(cost_row) + '*lifetime,"")'
                ss.cell(row=ss_row, column=st_lic+1).number_format = '$#,##0'
                # ns.cell(row=what_row, column=col + 1).value = gen
                ns.cell(row=what_row, column=col + 3).value = gen + '\nBalance'
                ns.cell(row=what_row, column=col + 3).alignment = oxl.styles.Alignment(wrap_text=True,
                        vertical='bottom', horizontal='center')
                ns.cell(row=what_row, column=col + 4).value = 'After\n' + gen
                ns.cell(row=what_row, column=col + 4).alignment = oxl.styles.Alignment(wrap_text=True,
                        vertical='bottom', horizontal='center')
                ns.cell(row=fall_row, column=col + 4).value = '=COUNTIF(' + ss_col(col + 4) \
                        + str(hrows) + ':' + ss_col(col + 4) + str(hrows + 8759) + \
                        ',"' + sf_test[0] + '0")'
                ns.cell(row=fall_row, column=col + 4).number_format = '#,##0'
                col += 5
            else:
                # lifetime cost
                ss.cell(row=ss_row, column=st_lic+1).value = '=IF(Detail!' + ss_col(col) + str(sum_row) \
                                                        + '>0,Detail!' + ss_col(col) + str(cost_row) + '*lifetime,"")'
                ss.cell(row=ss_row, column=st_lic+1).number_format = '$#,##0'
                ns.cell(row=what_row, column=col + 1).value = 'After\n' + gen
                ns.cell(row=fall_row, column=col + 1).value = '=COUNTIF(' + ss_col(col + 1) \
                        + str(hrows) + ':' + ss_col(col + 1) + str(hrows + 8759) + \
                        ',"' + sf_test[0] + '0")'
                ns.cell(row=fall_row, column=col + 1).number_format = '#,##0'
                col += 2
            ss.cell(row=ss_row, column=st_lie+1).value = '=IF(' + ss_col(st_emi+1) + str(ss_row) + '>0,' + \
                                                         ss_col(st_emi+1) + str(ss_row) + '*lifetime,"")'
            ss.cell(row=ss_row, column=st_lie+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_lec+1).value = '=IF(' + ss_col(st_emc+1) + str(ss_row) + '>0,' + \
                                                         ss_col(st_emc+1) + str(ss_row) + '*lifetime,"")'
            ss.cell(row=ss_row, column=st_lec+1).number_format = '$#,##0'
        if is_storage:
            ns.cell(row=emi_row, column=col - 2).value = '=MIN(' + ss_col(col - 2) + str(hrows) + \
                    ':' + ss_col(col - 2) + str(hrows + 8759) + ')'
            ns.cell(row=emi_row, column=col - 2).number_format = '#,##0.00'
        for column_cells in ns.columns:
            length = 0
            value = ''
            row = 0
            sum_value = 0
            do_sum = False
            do_cost = False
            for cell in column_cells:
                if cell.row >= hrows:
                    if do_sum:
                        try:
                            sum_value += abs(cell.value)
                        except:
                            pass
                    else:
                        try:
                            value = str(round(cell.value, 2))
                            if len(value) > length:
                                length = len(value) + 2
                        except:
                            pass
                elif cell.row > 0:
                    if str(cell.value)[0] != '=':
                        values = str(cell.value).split('\n')
                        for value in values:
                            if cell.row == cost_row:
                                valf = value.split('.')
                                alen = int(len(valf[0]) * 1.6)
                            else:
                                alen = len(value) + 2
                            if alen > length:
                                length = alen
                    else:
                        if cell.row == cost_row:
                            do_cost = True
                        if cell.value[1:4] == 'SUM':
                            do_sum = True
            if sum_value > 0:
                alen = len(str(int(sum_value))) * 1.5
                if do_cost:
                    alen = int(alen * 1.5)
                if alen > length:
                    length = alen
            if isinstance(cell.column, int):
                cel = ss_col(cell.column)
            else:
                cel = cell.column
            ns.column_dimensions[cel].width = max(length, 10)
        ns.column_dimensions['A'].width = 6
        ns.column_dimensions['B'].width = 21
        st_row = hrows + 8760
        st_col = col
        for row in range(1, st_row):
            for col in range(1, st_col):
                try:
                    ns.cell(row=row, column=col).font = normal
                except:
                    pass
        # self.progressbar.progress(12, text=progress_text)
        ns.row_dimensions[what_row].height = 30
        ns.freeze_panes = 'C' + str(hrows)
        ns.activeCell = 'C' + str(hrows)
        if results_prefix != '':
            ss.cell(row=1, column=1).value = 'Powermatch - ' + results_prefix + ' Summary'
        else:
            ss.cell(row=1, column=1).value = 'Powermatch - Summary'
        ss.cell(row=1, column=1).font = bold
        ss_lst_row = ss_row + 1
        ss_row, ss_re_row = detail_summary_total(ss_row, base_row='4')
        if len(nsul_sum_cols) > 0:
            ss_row += 2
            ss.cell(row=ss_row, column=1).value = 'Additional Underlying Load'
            ss.cell(row=ss_row, column=1).font = bold
            base_row = str(ss_row + 1)
            for col in nsul_sum_cols[1:]:
                ss_row += 1
                ul_tml_sum, ul_re_sum = do_detail_summary(fac, col, ss_row, ul_tml_sum, ul_re_sum)
            ul_fst_row = int(base_row)
            ul_lst_row = ss_row
            ns_re_sum = ul_re_sum
            ns_tml_sum = ul_tml_sum
            ss_row, ss_re_row = detail_summary_total(ss_row, title='Underlying ', base_row=base_row,
                                          back_row=str(ss_lst_row))
        wider = [ss_col(st_cac + 1), ss_col(st_lic + 1)]
        for column_cells in ss.columns:
            length = 0
            value = ''
            for cell in column_cells:
                if str(cell.value)[0] != '=':
                    values = str(cell.value).split('\n')
                    for value in values:
                        if len(value) + 1 > length:
                            length = len(value) + 1
            if isinstance(cell.column, int):
                cel = ss_col(cell.column)
            else:
                cel = cell.column
            if cel in wider:
                ss.column_dimensions[cel].width = max(length, 10) * 1.5
            else:
                ss.column_dimensions[cel].width = max(length, 10) * 1.2

        if corr_data is not None:
            ss_row += 2
            for corr in corr_data:
                ss.cell(row=ss_row, column=1).value = corr[0]
                if len(corr) > 1:
                    ss.cell(row=ss_row, column=2).value = corr[1]
                    ss.cell(row=ss_row, column=2).number_format = '#0.0000'
                    ss.cell(row=ss_row, column=3).value = corr[2]
                ss_row += 1
        ss_row += 2
        ss.cell(row=ss_row, column=1).value = 'Static Variables'
        ss.cell(row=ss_row, column=1).font = bold
        if carbon_price > 0:
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = 'Carbon Price ($/tCO2e)'
            ss.cell(row=ss_row, column=st_cap+1).value = carbon_price
            ss.cell(row=ss_row, column=st_cap+1).number_format = '$#,##0.00'
            attr_text = 'Summary!$' + ss_col(st_cap+1) + '$' + str(ss_row)
            carbon_cell = oxl.workbook.defined_name.DefinedName('carbon_price', attr_text=attr_text)
            wb.defined_names.append(carbon_cell)
        ss_row += 1
        attr_text = 'Summary!$' + ss_col(st_cap+1) + '$' + str(ss_row)
        lifetime_cell = oxl.workbook.defined_name.DefinedName('lifetime', attr_text=attr_text)
        wb.defined_names.append(lifetime_cell)
        ss.cell(row=ss_row, column=1).value = 'Lifetime (years)'
        ss.cell(row=ss_row, column=st_cap+1).value = max_lifetime
        ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0'
        ss_row += 1
        ss.cell(row=ss_row, column=1).value = 'Discount Rate'
        ss.cell(row=ss_row, column=st_cap+1).value = discount_rate
        ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.00%'
        ss_row += 2
        # self.progressbar.progress(14, text=progress_text)
        for row in range(1, ss_row + 1):
            for col in range(1, len(headers) + 1):
                try:
                    if ss.cell(row=row, column=col).font.name != 'Arial':
                        ss.cell(row=row, column=col).font = normal
                except:
                    pass
        ss.freeze_panes = 'B4'
        ss.activeCell = 'B4'
        if save_tables:
            gens = []
            cons = []
            for fac in re_order:
                if fac == 'Load':
                    continue
                if pmss_details[fac].multiplier <= 0:
                    continue
                if fac.find('.') > 0:
                    gens.append(fac[fac.find('.') + 1:])
                else:
                    gens.append(fac)
                cons.append(generators[pmss_details[fac].generator].constr)
            for gen in dispatch_order:
                gens.append(gen)
                cons.append(generators[gen].constr)
            gs = wb.create_sheet(self.sheets[G])
            fields = []
            col = 1
            row = 1
            if hasattr(generators[list(generators.keys())[0]], 'name'):
                fields.append('name')
                gs.cell(row=row, column=col).value = 'Name'
                col += 1
            for prop in dir(generators[list(generators.keys())[0]]):
                if prop[:2] != '__' and prop[-2:] != '__':
                    if prop != 'name':
                        fields.append(prop)
                        txt = prop.replace('_', ' ').title()
                        txt = txt.replace('Cf', 'CF')
                        txt = txt.replace('Lcoe', 'LCOE')
                        txt = txt.replace('Om', 'OM')
                        gs.cell(row=row, column=col).value = txt
                        if prop == 'capex':
                            txt = txt + txt
                        gs.column_dimensions[ss_col(col)].width = max(len(txt) * 1.4, 10)
                        col += 1
            nme_width = 4
            con_width = 4
            for key, value in generators.items():
                if key in gens:
                    row += 1
                    col = 1
                    for field in fields:
                        gs.cell(row=row, column=col).value = getattr(value, field)
                        if field in ['name', 'constraint']:
                            txt = getattr(value, field)
                            if field == 'name':
                                if len(txt) > nme_width:
                                    nme_width = len(txt)
                                    gs.column_dimensions[ss_col(col)].width = nme_width * 1.4
                            else:
                                if len(txt) > con_width:
                                    con_width = len(txt)
                                    gs.column_dimensions[ss_col(col)].width = con_width * 1.4
                        elif field in ['capex', 'fixed_om']:
                            gs.cell(row=row, column=col).number_format = '$#,##0'
                        elif field in ['lcoe', 'variable_om', 'fuel']:
                            gs.cell(row=row, column=col).number_format = '$#,##0.00'
                        elif field in ['disc_rate']:
                            gs.cell(row=row, column=col).number_format = '#,##0.00%'
                        elif field in ['capacity', 'lcoe_cf', 'initial']:
                            gs.cell(row=row, column=col).number_format = '#,##0.00'
                        elif field in ['emissions']:
                            gs.cell(row=row, column=col).number_format = '#,##0.000'
                        elif field in ['lifetime', 'order']:
                            gs.cell(row=row, column=col).number_format = '#,##0'
                        col += 1
            for row in range(1, row + 1):
                for col in range(1, len(fields) + 1):
                    gs.cell(row=row, column=col).font = normal
            gs.freeze_panes = 'B2'
            gs.activeCell = 'B2'
            fields = []
            col = 1
            row = 1
            cs = wb.create_sheet(self.sheets[C])
            if hasattr(self.constraints[list(self.constraints.keys())[0]], 'name'):
                fields.append('name')
                cs.cell(row=row, column=col).value = 'Name'
                col += 1
            for prop in dir(self.constraints[list(self.constraints.keys())[0]]):
                if prop[:2] != '__' and prop[-2:] != '__':
                    if prop != 'name':
                        fields.append(prop)
                        if prop == 'warm_time':
                            cs.cell(row=row, column=col).value = 'Warmup Time'
                        else:
                            cs.cell(row=row, column=col).value = prop.replace('_', ' ').title()
                        cs.column_dimensions[ss_col(col)].width = max(len(prop) * 1.4, 10)
                        col += 1
            nme_width = 4
            cat_width = 4
            for key, value in self.constraints.items():
                if key in cons:
                    row += 1
                    col = 1
                    for field in fields:
                        cs.cell(row=row, column=col).value = getattr(value, field)
                        if field in ['name', 'category']:
                            txt = getattr(value, field)
                            if field == 'name':
                                if len(txt) > nme_width:
                                    nme_width = len(txt)
                                    cs.column_dimensions[ss_col(col)].width = nme_width * 1.4
                            else:
                                if len(txt) > cat_width:
                                    cat_width = len(txt)
                                    cs.column_dimensions[ss_col(col)].width = cat_width * 1.4
                        elif field == 'warm_time':
                            cs.cell(row=row, column=col).number_format = '#0.00'
                        elif field != 'category':
                            cs.cell(row=row, column=col).number_format = '#,##0%'
                        col += 1
            for row in range(1, row + 1):
                for col in range(1, len(fields) + 1):
                    try:
                        cs.cell(row=row, column=col).font = normal
                    except:
                        pass
            cs.freeze_panes = 'B2'
            cs.activeCell = 'B2'
        wb.save(data_file)
        self.progressbar.progress(20, text=progress_text)
        j = data_file.rfind('/')
        data_file = data_file[j + 1:]
        msg = '%s created (%.2f seconds)' % (data_file, time.time() - start_time)
        msg = '%s created.' % data_file
        self.progressbar.empty()

    def setStatus(self, text):
        print(text)

    def exit(self):
        self.updated = False
        self.close()
        

    def optClicked(self, in_year, in_option, in_pmss_details, in_pmss_data, in_re_order,
                   in_dispatch_order, OptParms, optimisation, pm_data_file, data_file):

        def create_starting_population(individuals, chromosome_length):
            # Set up an initial array of all zeros
            population = np.zeros((individuals, chromosome_length))
            # Loop through each row (individual)
            for i in range(individuals):
                # Choose a random number of ones to create but at least one 1
                ones = random.randint(1, chromosome_length)
                # Change the required number of zeros to ones
                population[i, 0:ones] = 1
                # Sfuffle row
                np.random.shuffle(population[i])
            return population

        def select_individual_by_tournament(population, *argv):
            # Get population size
            population_size = len(population)

            # Pick individuals for tournament
            fighter = [0, 0]
            fighter_fitness = [0, 0]
            fighter[0] = random.randint(0, population_size - 1)
            fighter[1] = random.randint(0, population_size - 1)

            # Get fitness score for each
            if len(argv) == 1:
                fighter_fitness[0] = argv[0][fighter[0]]
                fighter_fitness[1] = argv[0][fighter[1]]
            else:
                for arg in argv:
                    min1 = min(arg)
                    max1 = max(arg)
                    for f in range(len(fighter)):
                        try:
                            fighter_fitness[f] += (arg[f] - min1) / (max1 - min1)
                        except:
                            pass
            # Identify individual with lowest score
            # Fighter 1 will win if score are equal
            if fighter_fitness[0] <= fighter_fitness[1]:
                winner = fighter[0]
            else:
                winner = fighter[1]

            # Return the chromsome of the winner
            return population[winner, :]

        def breed_by_crossover(parent_1, parent_2, points=2):
            # Get length of chromosome
            chromosome_length = len(parent_1)

            # Pick crossover point, avoiding ends of chromsome
            if points == 1:
                crossover_point = random.randint(1,chromosome_length-1)
            # Create children. np.hstack joins two arrays
                child_1 = np.hstack((parent_1[0:crossover_point],
                                     parent_2[crossover_point:]))
                child_2 = np.hstack((parent_2[0:crossover_point],
                                     parent_1[crossover_point:]))
            else: # only do 2 at this    stage
                crossover_point_1 = random.randint(1, chromosome_length - 2)
                crossover_point_2 = random.randint(crossover_point_1 + 1, chromosome_length - 1)
                child_1 = np.hstack((parent_1[0:crossover_point_1],
                                     parent_2[crossover_point_1:crossover_point_2],
                                     parent_1[crossover_point_2:]))
                child_2 = np.hstack((parent_2[0:crossover_point_1],
                                     parent_1[crossover_point_1:crossover_point_2],
                                     parent_2[crossover_point_2:]))
            # Return children
            return child_1, child_2

        def randomly_mutate_population(population, mutation_probability):
            # Apply random mutation
            random_mutation_array = np.random.random(size=(population.shape))
            random_mutation_boolean = random_mutation_array <= mutation_probability
        #    random_mutation_boolean[0][:] = False # keep the best multi and lcoe
       #     random_mutation_boolean[1][:] = False
            population[random_mutation_boolean] = np.logical_not(population[random_mutation_boolean])
            # Return mutation populationself.debug:
            return population

        def calculate_fitness(population):
            lcoe_fitness_scores = [] # scores = LCOE values
            multi_fitness_scores = [] # scores = multi-variable weight
            multi_values = [] # values for each of the six variables
            if len(population) == 1:
                option = '1'
            else:
                option = 'O'
            if self.debug:
                self.popn += 1
                self.chrom = 0
            for chromosome in population:
                # now get random amount of generation per technology (both RE and non-RE)
                for fac, value in opt_order.items():
                    capacity = value[2]
                    for c in range(value[0], value[1]):
                        if chromosome[c]:
                            capacity = capacity + capacities[c]
                    try:
                        pmss_details[fac].multiplier = capacity / pmss_details[fac].capacity
                    except:
                        print('PME2:', gen, capacity, pmss_details[fac].capacity)
                multi_value, op_data, extra = self.doDispatch(year, option, pmss_details, pmss_data, re_order,
                                              dispatch_order, pm_data_file, data_file)
                if multi_value['load_pct'] < self.targets['load_pct'][3]:
                    if multi_value['load_pct'] == 0:
                        print('PME3:', multi_value['lcoe'], self.targets['load_pct'][3], multi_value['load_pct'])
                        lcoe_fitness_scores.append(1)
                    else:
                        try:
                            lcoe_fitness_scores.append(pow(multi_value['lcoe'],
                                self.targets['load_pct'][3] / multi_value['load_pct']))
                        except OverflowError as err:
                            self.setStatus(f"Overflow error: {err}; POW({multi_value['lcoe']:,}, " \
                                         + f"{self.targets['load_pct'][3] / multi_value['load_pct']:,}) " \
                                         + f"({self.targets['load_pct'][3]:,} / {multi_value['load_pct']:,} )")
                        except:
                            pass
                else:
                    lcoe_fitness_scores.append(multi_value['lcoe'])
                multi_values.append(multi_value)
                multi_fitness_scores.append(calc_weight(multi_value))
                if self.debug:
                    self.chrom += 1
                    line = str(self.popn) + ',' + str(self.chrom) + ','
                    for fac, value in opt_order.items():
                        try:
                            line += str(pmss_details[fac].capacity * pmss_details[fac].multiplier) + ','
                        except:
                            line += ','
                    for key in self.targets.keys():
                        try:
                            line += '{:.3f},'.format(multi_value[key])
                        except:
                            line += multi_value[key] + ','
                    line += '{:.5f},'.format(multi_fitness_scores[-1])
                    self.db_file.write(line + '\n')
            # alternative approach to calculating fitness
            multi_fitness_scores1 = []
            maxs = {}
            mins = {}
            tgts = {}
            for key in multi_value.keys():
                if key[-4:] == '_pct':
                    tgts[key] = abs(self.targets[key][2] - self.targets[key][3])
                else:
                    maxs[key] = 0
                    mins[key] = -1
                    for popn in multi_values:
                        try:
                            tgt = abs(self.targets[key][2] - popn[key])
                        except:
                            continue
                        if tgt > maxs[key]:
                            maxs[key] = tgt
                        if mins[key] < 0 or tgt < mins[key]:
                            mins[key] = tgt
            for popn in multi_values:
                weight = 0
                for key, value in multi_value.items():
                    if self.targets[key][1] <= 0:
                        continue
                    try:
                        tgt = abs(self.targets[key][2] - popn[key])
                    except:
                        continue
                    if key[-4:] == '_pct':
                        if tgts[key] != 0:
                            if tgt > tgts[key]:
                                weight += 1 * self.targets[key][1]
                            else:
                                try:
                                    weight += 1 - ((tgt / tgts[key]) * self.targets[key][1])
                                except:
                                    pass
                    else:
                        try:
                            weight += 1 - (((maxs[key] - tgt) / (maxs[key] - mins[key])) \
                                      * self.targets[key][1])
                        except:
                            pass
                multi_fitness_scores1.append(weight)

            if len(population) == 1: # return the table for best chromosome
                return op_data, multi_values
            else:
                return lcoe_fitness_scores, multi_fitness_scores, multi_values

        def optQuitClicked(event):
            self.optExit = True

        def calc_weight(multi_value, calc=0):
            weight = [0., 0.]
            if calc == 0:
                for key, value in self.targets.items():
                    if multi_value[key] == '':
                        continue
                    if value[1] <= 0:
                        continue
                    if value[2] == value[3]: # wants specific target
                        if multi_value[key] == value[2]:
                            w = 0
                        else:
                            w = 1
                    elif value[2] > value[3]: # wants higher target
                        if multi_value[key] > value[2]: # high no weight
                            w = 0.
                        elif multi_value[key] < value[3]: # low maximum weight
                            w = 2.
                        else:
                            w = 1 - (multi_value[key] - value[3]) / (value[2] - value[3])
                    else: # lower target
                        if multi_value[key] == -1 or multi_value[key] > value[3]: # high maximum weight
                            w = 2.
                        elif multi_value[key] < value[2]: # low no weight
                            w = 0.
                        else:
                            w = multi_value[key] / (value[3] - value[2])
                    weight[0] += w * value[1]
            elif calc == 1:
                for key, value in self.targets.items():
                    if multi_value[key] == '':
                        continue
                    if value[1] <= 0:
                        continue
                    if multi_value[key] < 0:
                        w = 1
                    elif value[2] == value[3]: # wants specific target
                        if multi_value[key] == value[2]:
                            w = 0
                        else:
                            w = 1
                    else: # target range
                        w = min(abs(value[2] - multi_value[key]) / abs(value[2] - value[3]), 1)
                    weight[1] += w * value[1]
            return weight[calc]

        def plot_multi(multi_scores, multi_best, multi_order, title):
            data = [[], [], []]
            max_amt = [0., 0.]
            for multi in multi_best:
                max_amt[0] = max(max_amt[0], multi['cost'])
                max_amt[1] = max(max_amt[1], multi['co2'])
            pwr_chr = ['', '']
            divisor = [1., 1.]
            pwr_chrs = ' KMBTPEZY'
            for m in range(2):
                for pwr in range(len(pwr_chrs) - 1, -1, -1):
                    if max_amt[m] > pow(10, pwr * 3):
                        pwr_chr[m] = pwr_chrs[pwr]
                        divisor[m] = 1. * pow(10, pwr * 3)
                        break
            self.targets['cost'][5] = self.targets['cost'][5].replace('pwr_chr', pwr_chr[0])
            self.targets['co2'][5] = self.targets['co2'][5].replace('pwr_chr', pwr_chr[1])
            for multi in multi_best:
                for axis in range(3): # only three axes in plot
                    if multi_order[axis] == 'cost':
                        data[axis].append(multi[multi_order[axis]] / divisor[0]) # cost
                    elif multi_order[axis] == 'co2':
                        data[axis].append(multi[multi_order[axis]] / divisor[1]) # co2
                    elif multi_order[axis][-4:] == '_pct': # percentage
                        data[axis].append(multi[multi_order[axis]] * Decimal(100))
                    else:
                        data[axis].append(multi[multi_order[axis]])
            # create colour map
            colours = multi_scores[:]
            cmax = max(colours)
            cmin = min(colours)
            if cmin == cmax:
                return
            for c in range(len(colours)):
                colours[c] = (colours[c] - cmin) / (cmax - cmin)
            scolours = sorted(colours)
            cvals  = [-1., 0, 1]
            colors = ['green' ,'orange', 'red']
            norm = plt.Normalize(min(cvals), max(cvals))
            tuples = list(zip(map(norm,cvals), colors))
            cmap = matplotlib.colors.LinearSegmentedColormap.from_list('', tuples)
            fig = plt.figure(title)
            mx = plt.axes(projection='3d')
            plt.title('\n' + title.title() + '\n')
            try:
                for i in range(len(data[0])):
                    mx.scatter3D(data[2][i], data[1][i], data[0][i], picker=True, color=cmap(colours[i]), cmap=cmap)
                    if title[:5] == 'start':
                        mx.text(data[2][i], data[1][i], data[0][i], '%s' % str(i+1))
                    else:
                        j = scolours.index(colours[i])
                        if j < 10:
                            mx.text(data[2][i], data[1][i], data[0][i], '%s' % str(j+1))
            except:
                return
            if self.optimise_multisurf:
                cvals_r  = [-1., 0, 1]
                colors_r = ['red' ,'orange', 'green']
                norm_r = plt.Normalize(min(cvals_r), max(cvals_r))
                tuples_r = list(zip(map(norm_r, cvals_r), colors_r))
                cmap_r = matplotlib.colors.LinearSegmentedColormap.from_list('', tuples_r)
                # https://www.fabrizioguerrieri.com/blog/surface-graphs-with-irregular-dataset/
                triang = mtri.Triangulation(data[2], data[1])
                mx.plot_trisurf(triang, data[0], cmap=cmap_r)
            mx.xaxis.set_major_formatter(FormatStrFormatter(self.targets[multi_order[2]][5]))
            mx.yaxis.set_major_formatter(FormatStrFormatter(self.targets[multi_order[1]][5]))
            mx.zaxis.set_major_formatter(FormatStrFormatter(self.targets[multi_order[0]][5]))
            mx.set_xlabel(self.targets[multi_order[2]][6])
            mx.set_ylabel(self.targets[multi_order[1]][6])
            mx.set_zlabel(self.targets[multi_order[0]][6])
            plt.show()

        def show_multitable(best_score_progress, multi_best, multi_order, title):
            def pwr_chr(amt):
                pwr_chrs = ' KMBTPEZY'
                pchr = ''
                divisor = 1.
                for pwr in range(len(pwr_chrs) - 1, -1, -1):
                    if amt > pow(10, pwr * 3):
                        pchr = pwr_chrs[pwr]
                        divisor = 1. * pow(10, pwr * 3)
                        break
                return amt / divisor, pchr

            def opt_fmat(amt, fmat):
                tail = ' '
                p = fmat.find('pwr_chr')
                if p > 0:
                    amt, tail = pwr_chr(amt)
                    fmat = fmat.replace('pwr_chr', '')
              #  d = fmat.find('d')
             #   if d > 0:
              #      amt = amt * 100.
                fmat = fmat.replace('.1f', '.2f')
                i = fmat.find('%')
                fmt = fmat[:i] + '{:> 8,' + fmat[i:].replace('%', '').replace('d', '.2%') + '}' + tail
                return fmt.format(amt)

            best_table = []
            best_fmate = []
            for b in range(len(multi_best)):
                bl = list(multi_best[b].values())
                bl.insert(0, best_score_progress[b])
                bl.insert(0, b + 1)
                best_table.append(bl)
                best_fmate.append([b + 1])
                best_fmate[-1].insert(1, best_score_progress[b])
                for f in range(2, len(best_table[-1])):
                    if target_keys[f - 2] in ['load_pct', 'surplus_pct', 're_pct']:
                        best_fmate[-1].append(opt_fmat(bl[f], '%d%%'))
                    else:
                        best_fmate[-1].append(opt_fmat(bl[f], target_fmats[f - 2]))
            fields = target_names[:]
            fields.insert(0, 'weight')
            fields.insert(0, 'iteration')
            dialog = displaytable.Table(best_fmate, fields=fields, txt_align='R', decpts=[0, 4],
                     title=title, sortby='weight')
            dialog.exec_()
            b = int(dialog.getItem(0)) - 1
            del dialog
            pick = [b]
            for fld in multi_order[:3]:
                i = target_keys.index(fld)
                pick.append(best_table[b][i + 2])
            return [pick]

#       optClicked mainline starts here
        year = in_year
        option = in_option
        pmss_details = dict(in_pmss_details)
        # pmss_data = in_pmss_data[:]
        pmss_data = dict(in_pmss_data)
        re_order = in_re_order[:]
        dispatch_order = in_dispatch_order[:]
        self.debug = False
        self.optimisation = optimisation
        self.optExit = False
        self.setStatus('Optimise processing started')
        err_msg = ''
        self.optLoad = OptParms['optLoad']
        pmss_details['Load'].multiplier = self.optLoad
        optPopn = OptParms['optPopn']
        self.optimise_population = int(optPopn)
        optGenn = OptParms['optGenn']
        self.optimise_generations = optGenn
        optMutn = OptParms['optMutn']
        self.optimise_mutation = Decimal(optMutn)
        optStop = OptParms['optStop']
        self.optimise_stop = int(optStop)
        optimise_choice = OptParms['optimise_choice']
        self.optimise_choice = optimise_choice
        self.optimise_progress = 0
             
        LCOE_Weight = OptParms['LCOE_Weight']
        Load_Weight = OptParms['Load_Weight']
        Surplus_Weight = OptParms['Surplus_Weight']
        RE_Weight = OptParms['RE_Weight']
        Cost_Weight = OptParms['Cost_Weight']
        CO2_Weight = OptParms['CO2_Weight']

        LCOE_Better = OptParms['LCOE_Better']
        Load_Better = OptParms['Load_Better']
        Surplus_Better = OptParms['Surplus_Better']
        RE_Better = OptParms['RE_Better']
        Cost_Better = OptParms['Cost_Better']
        CO2_Better = OptParms['CO2_Better']

        LCOE_Worse = OptParms['LCOE_Worse']
        Load_Worse = OptParms['Load_Worse']
        Surplus_Worse = OptParms['Surplus_Worse']
        RE_Worse = OptParms['RE_Worse']
        Cost_Worse = OptParms['Cost_Worse']
        CO2_Worse = OptParms['CO2_Worse']

        # check we have optimisation entries for generators and storage
        # update any changes to targets
        updates = {}
        lines = []
        lines.append('optimise_choice=' + self.optimise_choice)
        lines.append('optimise_generations=' + str(self.optimise_generations))
        lines.append('optimise_mutation=' + str(self.optimise_mutation))
        lines.append('optimise_population=' + str(self.optimise_population))
        lines.append('optimise_stop=' + str(self.optimise_stop))
        if self.optimise_choice == 'LCOE':
            do_lcoe = True
            do_multi = False
        elif self.optimise_choice == 'Multi':
            do_lcoe = False
            do_multi = True
        else:
            do_lcoe = True
            do_multi = True
        multi_order = []
        updates['Powermatch'] = lines
        multi_order.sort(reverse=True)
    #    multi_order = multi_order[:3] # get top three weighted variables - but I want them all
        multi_order = [o[4:] for o in multi_order]
        self.adjust_gen = True
        orig_load = []
        load_col = -1
        orig_tech = {}
        orig_capacity = {}
        opt_order = {} # rely on it being processed in added order
        # each entry = [first entry in chrom, last entry, minimum capacity]
        # first get original renewables generation from data sheet
        for fac in re_order:
            if fac == 'Load':
                continue
            opt_order[fac] = [0, 0, 0]
        # now add scheduled generation
        for gen in dispatch_order:
            opt_order[gen] = [0, 0, 0]
        capacities = []
        for gen in opt_order.keys():
            opt_order[gen][0] = len(capacities) # first entry
            try:
                if self.optimisation[gen].approach == 'Discrete':
                    capacities.extend(self.optimisation[gen].capacities)
                    opt_order[gen][1] = len(capacities) # last entry
                elif self.optimisation[gen].approach == 'Range':
                    if self.optimisation[gen].capacity_max == self.optimisation[gen].capacity_min:
                        capacities.extend([0])
                        opt_order[gen][1] = len(capacities)
                        opt_order[gen][2] = self.optimisation[gen].capacity_min
                        continue
                    ctr = int((self.optimisation[gen].capacity_max - self.optimisation[gen].capacity_min) / \
                              self.optimisation[gen].capacity_step)
                    if ctr < 1:
                        self.setStatus("Error with Optimisation table entry for '" + gen + "'")
                        return
                    capacities.extend([self.optimisation[gen].capacity_step] * ctr)
                    tot = self.optimisation[gen].capacity_step * ctr + self.optimisation[gen].capacity_min
                    if tot < self.optimisation[gen].capacity_max:
                        capacities.append(self.optimisation[gen].capacity_max - tot)
                    opt_order[gen][1] = len(capacities)
                    opt_order[gen][2] = self.optimisation[gen].capacity_min
                else:
                    opt_order[gen][1] = len(capacities)
            except KeyError as err:
                self.setStatus('Key Error: No Optimisation entry for ' + str(err))
                opt_order[gen] = [len(capacities), len(capacities) + 5, 0]
                capacities.extend([pmss_details[gen].capacity / 5.] * 5)
            except ZeroDivisionError as err:
                self.setStatus('Zero capacity: ' + gen + ' ignored')
            except:
                err = str(sys.exc_info()[0]) + ',' + str(sys.exc_info()[1]) + ',' + gen + ',' \
                      + str(opt_order[gen])
                self.setStatus('Error: ' + str(err))
                return
        # chromosome = [1] * int(len(capacities) / 2) + [0] * (len(capacities) - int(len(capacities) / 2))
        # we have the original data - from here down we can do our multiple optimisations
        # Set general parameters
        self.setStatus('Optimisation choice is ' + self.optimise_choice)
        chromosome_length = len(capacities)
        self.setStatus(f'Chromosome length: {chromosome_length}; {pow(2, chromosome_length):,} permutations')
        population_size = self.optimise_population
        maximum_generation = int(self.optimise_generations)
        lcoe_scores = []
        multi_scores = []
        multi_values = []
     #   if do_lcoe:
      #      lcoe_target = 0. # aim for this LCOE
        if do_multi:
            multi_best = [] # list of six variables for best weight
            multi_best_popn = [] # list of chromosomes for best weight
        progress_text = "Process Optimisation iterations"
        #self.progressbar = st.progress(0, text=progress_text)
        start_time = time.time()
        # Create starting population
        progress_text = "'Processing iteration 1'"
        self.progressbar.progress(1, text=progress_text)
        population = create_starting_population(population_size, chromosome_length)
        # calculate best score(s) in starting population
        # if do_lcoe best_score = lowest non-zero lcoe
        # if do_multi best_multi = lowest weight and if not do_lcoe best_score also = best_weight
        if self.debug:
            filename = scenarios + 'opt_debug_' + '.csv'
            self.db_file = open(filename, 'w')
            line0 = 'Popn,Chrom,'
            line1 = 'Weights,,'
            line2 = 'Targets,,'
            line3 = 'Range,' + str(population_size) + ','
            for gen, value in opt_order.items():
                 line0 += gen + ','
                 line1 += ','
                 line2 += ','
                 line3 += ','
            for key in self.targets.keys():
                 line0 += key + ','
                 line1 += str(self.targets[key][1]) + ','
                 line2 += str(self.targets[key][2]) + ','
                 if key[-4:] == '_pct':
                     line3 += str(abs(self.targets[key][2] - self.targets[key][3])) + ','
                 else:
                     line3 += ','
            line0 += 'Score'
            self.db_file.write(line0 + '\n' + line1 + '\n' + line2 + '\n' + line3 + '\n')
            self.popn = 0
            self.chrom = 0
        lcoe_scores, multi_scores, multi_values = calculate_fitness(population)
        if do_lcoe:
            try:
                best_score = np.min(lcoe_scores)
            except:
                print('PME4:', lcoe_scores)
            best_ndx = lcoe_scores.index(best_score)
            lowest_chrom = population[best_ndx]
            self.setStatus('Starting LCOE: $%.2f' % best_score)
        if do_multi:
            if self.more_details: # display starting population ?
                pick = plot_multi(multi_scores, multi_values, multi_order, 'starting population')
            # want maximum from first round to set base upper limit
            for key in self.targets.keys():
                if self.targets[key][2] < 0: # want a maximum from first round
                    setit = 0
                    for multi in multi_values:
                        setit = max(multi[key], setit)
                    self.targets[key][2] = setit
                if self.targets[key][3] < 0: # want a maximum from first round
                    setit = 0
                    for multi in multi_values:
                        setit = max(multi[key], setit)
                    self.targets[key][3] = setit
            # now we can find the best weighted result - lowest is best
            best_multi = np.min(multi_scores)
            best_mndx = multi_scores.index(best_multi)
            multi_lowest_chrom = population[best_mndx]
            multi_best_popn.append(multi_lowest_chrom)
            multi_best.append(multi_values[best_mndx])
            self.setStatus('Starting Weight: %.4f' % best_multi)
            multi_best_weight = best_multi
            best_multi_progress = [best_multi]
            if not do_lcoe:
                best_score = best_multi
            last_multi_score = best_multi
            lowest_multi_score = best_multi
            mud = '='
        # Add starting best score to progress tracker
        best_score_progress = [best_score]
        best_ctr = 1
        last_score = best_score
        lowest_score = best_score
        lud = '='
        # Now we'll go through the generations of genetic algorithm
        for generation in range(1, maximum_generation):
            lcoe_status = ''
            multi_status = ''
            if do_lcoe:
                lcoe_status = ' %s $%.2f ;' % (lud, best_score)
            if do_multi:
                multi_status = ' %s %.4f ;' % (mud, best_multi)
            tim = (time.time() - start_time)
            if tim < 60:
                tim = ' (%s%s %.1f secs)' % (lcoe_status, multi_status, tim)
            else:
                tim = ' (%s%s %.2f mins)' % (lcoe_status, multi_status, tim / 60.)
            progress_text = 'Processing iteration ' + str(generation + 1) + tim
            self.progressbar.progress(generation + 1, text=progress_text)
        # Create an empty list for new population
            new_population = []
        # Using elitism approach include best individual
            if do_lcoe:
                new_population.append(lowest_chrom)
            if do_multi:
                new_population.append(multi_lowest_chrom)
            # Create new population generating two children at a time
            if do_lcoe:
                if do_multi:
                    for i in range(int(population_size/2)):
                        parent_1 = select_individual_by_tournament(population, lcoe_scores,
                                                                   multi_scores)
                        parent_2 = select_individual_by_tournament(population, lcoe_scores,
                                                                   multi_scores)
                        child_1, child_2 = breed_by_crossover(parent_1, parent_2)
                        new_population.append(child_1)
                        new_population.append(child_2)
                else:
                    for i in range(int(population_size/2)):
                        parent_1 = select_individual_by_tournament(population, lcoe_scores)
                        parent_2 = select_individual_by_tournament(population, lcoe_scores)
                        child_1, child_2 = breed_by_crossover(parent_1, parent_2)
                        new_population.append(child_1)
                        new_population.append(child_2)
            else:
                for i in range(int(population_size/2)):
                    parent_1 = select_individual_by_tournament(population, multi_scores)
                    parent_2 = select_individual_by_tournament(population, multi_scores)
                    child_1, child_2 = breed_by_crossover(parent_1, parent_2)
                    new_population.append(child_1)
                    new_population.append(child_2)
            # get back to original size (after elitism adds)
            if do_lcoe:
                new_population.pop()
            if do_multi:
                new_population.pop()
            # Replace the old population with the new one
            population = np.array(new_population)
            if self.optimise_mutation > 0:
                population = randomly_mutate_population(population, self.optimise_mutation)
            # Score best solution, and add to tracker
            lcoe_scores, multi_scores, multi_values = calculate_fitness(population)
            if do_lcoe:
                best_lcoe = np.min(lcoe_scores)
                best_ndx = lcoe_scores.index(best_lcoe)
                best_score = best_lcoe
            # now we can find the best weighted result - lowest is best
            if do_multi:
                best_multi = np.min(multi_scores)
                best_mndx = multi_scores.index(best_multi)
                multi_lowest_chrom = population[best_mndx]
                multi_best_popn.append(multi_lowest_chrom)
                multi_best.append(multi_values[best_mndx])
           #     if multi_best_weight > best_multi:
                multi_best_weight = best_multi
                if not do_lcoe:
                    best_score = best_multi
                best_multi_progress.append(best_multi)
            best_score_progress.append(best_score)
            if best_score < lowest_score:
                lowest_score = best_score
                if do_lcoe:
                    lowest_chrom = population[best_ndx]
                else: #(do_multi only)
                    multi_lowest_chrom = population[best_mndx]
            if self.optimise_stop > 0:
                if best_score == last_score:
                    best_ctr += 1
                    if best_ctr >= self.optimise_stop:
                        break
                else:
                    last_score = best_score
                    best_ctr = 1
            last_score = best_score
            if do_lcoe:
                if best_score == best_score_progress[-2]:
                    lud = '='
                elif best_score < best_score_progress[-2]:
                    lud = '<html>&darr;</html>'
                else:
                    lud = '<html>&uarr;</html>'
            if do_multi:
                if best_multi == last_multi_score:
                    mud = '='
                elif best_multi < last_multi_score:
                    mud = '<html>&darr;</html>'
                else:
                    mud = '<html>&uarr;</html>'
                last_multi_score = best_multi
        if self.debug:
            try:
                self.db_file.close()
                optimiseDebug(self.db_file.name)
                os.remove(self.db_file.name)
            except:
                pass
            self.debug = False
        self.progressbar.empty()
        tim = (time.time() - start_time)
        if tim < 60:
            tim = '%.1f secs)' % tim
        else:
            tim = '%.2f mins)' % (tim / 60.)
        msg = 'Optimise completed (%0d iterations; %s' % (generation + 1, tim)
        if best_score > lowest_score:
            msg += ' Try more iterations.'
        # we'll keep two or three to save re-calculating_fitness
        op_data = [[], [], [], [], []]
        score_data = [None, None, None, None, None]
        if do_lcoe:
            op_data[0], score_data[0] = calculate_fitness([lowest_chrom])
        if do_multi:
            op_data[1], score_data[1] = calculate_fitness([multi_lowest_chrom])
        self.setStatus(msg)
        self.progressbar.empty()
        # GA has completed required generation
        if do_lcoe:
            self.setStatus('Final LCOE: $%.2f' % best_score)
            fig = 'optimise_lcoe'
            titl = 'Optimise LCOE using Genetic Algorithm'
            ylbl = 'Best LCOE ($/MWh)'
        else:
            fig = 'optimise_multi'
            titl = 'Optimise Multi using Genetic Algorithm'
            ylbl = 'Best Weight'
        if do_multi:
            self.setStatus('Final Weight: %.4f' % multi_best_weight)
        # Plot progress
        x = list(range(1, len(best_score_progress)+ 1))
        matplotlib.rcParams['savefig.directory'] = scenarios
        plt.figure(fig)
        lx = plt.subplot(111)
        plt.title(titl)
        lx.plot(x, best_score_progress)
        lx.set_xlabel('Optimise Cycle (' + str(len(best_score_progress)) + ' iterations)')
        lx.set_ylabel(ylbl)
        plt.show()
        pick = None
        pickf = None
        if do_multi:
            if self.optimise_multiplot:
                pick = plot_multi(best_multi_progress, multi_best, multi_order, 'best of each iteration')
                if self.more_details:
                    pickf = plot_multi(multi_scores, multi_values, multi_order, 'final iteration')
            if self.optimise_multitable:
                pick2 = show_multitable(best_multi_progress, multi_best, multi_order, 'best of each iteration')
                try:
                    pick = pick + pick2
                except:
                    pick = pick2
                if self.more_details:
                    pick2 = show_multitable(multi_scores, multi_values, multi_order, 'final iteration')
                    try:
                        pickf = pickf + pick2
                    except:
                        pickf = pick2
        op_pts = [0] * len(headers)
        for p in [st_lcg, st_lco, st_lcc, st_max, st_bal, st_rlc]:
            op_pts[p] = 2
        op_pts[st_cap] = 3
        if self.more_details:
            if do_lcoe:
                list(map(list, list(zip(*op_data[0]))))
                dialog = displaytable.Table(op_data[0], title=self.sender().text(), fields=headers,
                         save_folder=scenarios, sortby='', decpts=op_pts)
                dialog.exec_()
                del dialog
            if do_multi:
                list(map(list, list(zip(*op_data[1]))))
                dialog = displaytable.Table(op_data[1], title='Multi_' + self.sender().text(), fields=headers,
                         save_folder=scenarios, sortby='', decpts=op_pts)
                dialog.exec_()
                del dialog
        # now I'll display the resulting capacities for LCOE, lowest weight, picked
        # now get random amount of generation per technology (both RE and non-RE)
        its = {}
        for fac, value in opt_order.items():
            its[fac] = []
        chrom_hdrs = []
        chroms = []
        ndxes = []
        if do_lcoe:
            chrom_hdrs = ['Lowest LCOE']
            chroms = [lowest_chrom]
            ndxes = [0]
        if do_multi:
            chrom_hdrs.append('Lowest Weight')
            chroms.append(multi_lowest_chrom)
            ndxes.append(1)
        if pickf is not None:
            for p in range(len(pickf)):
                if pick is None:
                    pick = [pickf[f]]
                else:
                    pick.append(pickf[p][:])
                pick[-1][0] = len(multi_best_popn)
                multi_best_popn.append(population[pickf[p][0]])
        if pick is not None:
            # at present I'll calculate the best weight for the chosen picks. Could actually present all for user choice
            if len(pick) <= 3:
                multi_lowest_chrom = multi_best_popn[pick[0][0]]
                op_data[2], score_data[2] = calculate_fitness([multi_lowest_chrom])
                if self.more_details:
                    list(map(list, list(zip(*op_data[2]))))
                    dialog = displaytable.Table(op_data[2], title='Pick_' + self.sender().text(), fields=headers,
                             save_folder=scenarios, sortby='', decpts=op_pts)
                    dialog.exec_()
                    del dialog
                chrom_hdrs.append('Your pick')
                chroms.append(multi_lowest_chrom)
                ndxes.append(2)
                if len(pick) >= 2:
                    multi_lowest_chrom = multi_best_popn[pick[1][0]]
                    op_data[3], score_data[3] = calculate_fitness([multi_lowest_chrom])
                    if self.more_details:
                        list(map(list, list(zip(*op_data[3]))))
                        dialog = displaytable.Table(op_data[3], title='Pick_' + self.sender().text(), fields=headers,
                                 save_folder=scenarios, sortby='', decpts=op_pts)
                        dialog.exec_()
                        del dialog
                    chrom_hdrs.append('Your 2nd pick')
                    chroms.append(multi_lowest_chrom)
                    ndxes.append(3)
                if len(pick) == 3:
                    multi_lowest_chrom = multi_best_popn[pick[2][0]]
                    op_data[4], score_data[4] = calculate_fitness([multi_lowest_chrom])
                    if self.more_details:
                        list(map(list, list(zip(*op_data[4]))))
                        dialog = displaytable.Table(op_data[4], title='Pick_' + self.sender().text(), fields=headers,
                                 save_folder=scenarios, sortby='', decpts=op_pts)
                        dialog.exec_()
                        del dialog
                    chrom_hdrs.append('Your 3rd pick')
                    chroms.append(multi_lowest_chrom)
                    ndxes.append(4)
            else:
                picks = []
                for pck in pick:
                    picks.append(multi_best_popn[pck[0]])
                a, b, c = calculate_fitness(picks)
                best_multi = np.min(b)
                best_mndx = b.index(best_multi)
                multi_lowest_chrom = picks[best_mndx]
                op_data[2], score_data[2] = calculate_fitness([multi_lowest_chrom])
                if self.more_details:
                    list(map(list, list(zip(*op_data[2]))))
                    dialog = displaytable.Table(op_data[2], title='Pick_' + self.sender().text(), fields=headers,
                             save_folder=scenarios, sortby='', decpts=op_pts)
                    dialog.exec_()
                    del dialog
                chrom_hdrs.append('Your pick')
                chroms.append(multi_lowest_chrom)
                ndxes.append(2)
        for chromosome in chroms:
            for fac, value in opt_order.items():
                capacity = opt_order[fac][2]
                for c in range(value[0], value[1]):
                    if chromosome[c]:
                        capacity = capacity + capacities[c]
                its[fac].append(capacity / pmss_details[fac].capacity)
        max_amt = [0., 0.]
        if do_lcoe:
            max_amt[0] = score_data[0][0]['cost']
            max_amt[1] = score_data[0][0]['co2']
        if do_multi:
            for multi in multi_best:
                max_amt[0] = max(max_amt[0], multi['cost'])
                max_amt[1] = max(max_amt[1], multi['co2'])
        pwr_chr = ['', '']
        divisor = [1., 1.]
        pwr_chrs = ' KMBTPEZY'
        for m in range(2):
            for pwr in range(len(pwr_chrs) - 1, -1, -1):
                if max_amt[m] > pow(10, pwr * 3):
                    pwr_chr[m] = pwr_chrs[pwr]
                    divisor[m] = 1. * pow(10, pwr * 3)
                    break
        self.targets['cost'][5] = self.targets['cost'][5].replace('pwr_chr', pwr_chr[0])
        self.targets['co2'][5] = self.targets['co2'][5].replace('pwr_chr', pwr_chr[1])

     #  this is a big of a kluge but I couldn't get it to behave
        self.opt_choice = ''
        try:
            h = chrom_hdrs.index(self.opt_choice)
        except:
            return
        op_data[h], score_data[h] = calculate_fitness([chroms[h]]) # make it current
        msg = chrom_hdrs[h] + ': '
        for key in multi_order[:3]:
            msg += self.targets[key][0] + ': '
            if key == 'cost':
                amt = score_data[h][0][key] / divisor[0] # cost
            elif key == 'co2':
                amt = score_data[h][0][key] / divisor[1] # co2
            elif key[-4:] == '_pct': # percentage
                amt = score_data[h][0][key] * Decimal(100)
            else:
                amt = score_data[h][0][key]
            txt = self.targets[key][5]
            txt = txt % amt
            msg += txt + '; '
        self.setStatus(msg)
        list(map(list, list(zip(*op_data[h]))))
        op_data[h].append(' ')
        op_data[h].append('Optimisation Parameters')
        op_op_prm = len(op_data[h])
        op_data[h].append(['Population size', str(self.optimise_population)])
        op_data[h].append(['No. of iterations', str(self.optimise_generations)])
        op_data[h].append(['Mutation probability', '%0.4f' % self.optimise_mutation])
        op_data[h].append(['Exit if stable', str(self.optimise_stop)])
        op_data[h].append(['Optimisation choice', self.optimise_choice])
        op_data[h].append(['Variable', 'Weight', 'Better', 'Worse'])
        for i in range(len(target_keys)):
            op_data[h].append([])
            for j in range(4):
                if j == 0:
                    op_data[h][-1].append(self.targets[target_keys[i]][j])
                else:
                    op_data[h][-1].append('{:.2f}'.format(self.targets[target_keys[i]][j]))
        op_max_row = len(op_data[h])
        for key in self.optimisation.keys():
            op_data[h].append(['Max. ' + key, self.optimisation[key].capacity_max])
        dialog = displaytable.Table(op_data[h], title='Chosen_' + self.sender().text(), fields=headers,
                 save_folder=scenarios, sortby='', decpts=op_pts)
        if self.adjust.isChecked():
            self.adjustto = {}
            for fac, value in sorted(pmss_details.items()):
                self.adjustto[fac] = value.capacity * value.multiplier
        return
