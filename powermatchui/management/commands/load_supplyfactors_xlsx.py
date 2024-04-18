from django.utils import timezone
import openpyxl
from django.core.management.base import BaseCommand
from siren_web.models import supplyfactors, Scenarios, Technologies, Zones
from siren_web.database_operations import copy_technologies_from_year0

class Command(BaseCommand):
    help = 'Loads supplyfactors model from an XLSX file'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to the XLSX file')

    def create_scenario(self, scenario_string):
        scenario_parts = scenario_string.split(';')
        if len(scenario_parts) > 1:
            scenario_title = scenario_parts[1]
            scenario, created = Scenarios.objects.get_or_create(
                title=scenario_title,
                defaults={
                    'dateexported': timezone.now(),
                    'description': 'New scenario created from {file_path}'
                }
            )
            return scenario
        else:
            return None
        
    def handle(self, *args, **options):
        file_path = options['file_path']
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        zones0_inst = Zones.objects.get(
            pk=0
        )
        zones1_inst = Zones.objects.get(
            pk=1
        )
        for row in worksheet.iter_rows(min_row=1, values_only=True):
            if (row[0] == 'Scenario Title:'):
                scenario_obj = self.create_scenario(row[2])
            if (row[0] == 'Data Year:'):
                demand_year = row[2]
                break
        technologies = {}
        for row in worksheet.iter_rows(min_row=10, values_only=True):
            if (row[0] == 'Technology'):
                for column in range(2, 9):
                    tech_name = row[column]
                    if column > 2 and column < 7:
                        tech_name = 'Existing ' + row[column]
                    elif column > 6:
                        tech_name = 'Proposed ' + row[column]
                    if row[column]:
                        technologies[column] = copy_technologies_from_year0(tech_name, demand_year, scenario_obj.title)
            break

        for row in worksheet.iter_rows(min_row=16, values_only=True):
            hour_value = row[0] if row[0] is not None else None
            for column in range(2, 9):
                quantum_value = row[column] if row[column] is not None else None
                if (quantum_value is not None):
                    if column < 6:
                        zones_inst = zones0_inst
                    else:
                        zones_inst = zones1_inst
                    supplyfactors.objects.create(
                        idscenarios=scenario_obj,
                        idtechnologies=technologies[column],
                        idzones=zones_inst,
                        year=demand_year,
                        hour=hour_value,
                        supply=1,
                        quantum=quantum_value,
                        col=column - 2
                    )

        self.stdout.write(self.style.SUCCESS('Data loaded successfully'))