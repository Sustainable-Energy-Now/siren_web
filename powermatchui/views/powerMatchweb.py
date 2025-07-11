import openpyxl as oxl
import os
from siren_web.siren.powermatch.logic.logic import Constraint, Facility, PM_Facility, Optimisation
from siren_web.siren.powermatch.logic.processor import PowerMatchProcessor
from siren_web.siren.utilities.senutils import WorkBook
import time

target_keys = ['lcoe', 'load_pct', 'surplus_pct', 're_pct', 'cost', 'co2']
target_names = ['LCOE', 'Load%', 'Surplus%', 'RE%', 'Cost', 'CO2']
target_fmats = ['$%.2f', '%.1f%%', '%.1f%%', '%.1f%%', '$%.1fpwr_chr', '%.1fpwr_chr']
target_titles = ['LCOE ($)', 'Load met %', 'Surplus %', 'RE %', 'Total Cost ($)', 'tCO2e']
headers = ['Facility', 'Capacity\n(Gen, MW;\nStor, MWh)', 'To meet\nLoad (MWh)',
           'Subtotal\n(MWh)', 'CF', 'Cost ($/yr)', 'LCOG\nCost\n($/MWh)', 'LCOE\nCost\n($/MWh)',
           'Emissions\n(tCO2e)', 'Emissions\nCost', 'LCOE With\nCO2 Cost\n($/MWh)', 'Max.\nMWH',
           'Max.\nBalance', 'Capital\nCost', 'Lifetime\nCost', 'Lifetime\nEmissions',
           'Lifetime\nEmissions\nCost', 'Area (km^2)', 'Reference\nLCOE', 'Reference\nCF']
tech_names = ['Load', 'Onshore Wind', 'Offshore Wind', 'Rooftop PV', 'Fixed PV', 'Single Axis PV',
              'Dual Axis PV', 'Biomass', 'Geothermal', 'Other1', 'CST', 'Shortfall']
# set up columns for summary table. Hopefully to make it easier to add / alter columns
st_fac = 0 # Facility
st_cap = 1 # Capacity\n(Gen, MW;\nStor, MWh)
st_tml = 2 # To meet\nLoad (MWh)
st_sub = 3 # Subtotal\n(MWh)
st_cfa = 4 # CF
st_cst = 5 # Cost ($/yr)
st_lcg = 6 # LCOG\nCost\n($/MWh)
st_lco = 7 # LCOE\nCost\n($/MWh)
st_emi = 8 # Emissions\n(tCO2e)
st_emc = 9 # Emissions\nCost
st_lcc = 10 # LCOE With\nCO2 Cost\n($/MWh)
st_max = 11 # Max.\nMWH
st_bal = 12 # Max.\nBalance'
st_cac = 13 # Capital\nCost'
st_lic = 14 # Lifetime\nCost'
st_lie = 15 # Lifetime\nEmissions
st_lec = 16 # Lifetime\nEmissions\nCost
st_are = 17 # Area (km^2)
st_rlc = 18 # Reference\nLCOE
st_rcf = 19 # Reference\nCF

# same order as self.file_labels
C = 0 # Constraints - xls or xlsx
G = 1 # Generators - xls or xlsx
O = 2 # Optimisation - xls or xlsx
D = 3 # Data - xlsx
R = 4 # Results - xlsx
B = 5 # Batch input - xlsx
T = 6 # Transition input - xlsx
S = 'S' # Summary
O1 = 'O1'
SWIS_dir = './siren_web/siren_files/SWIS/'

class powerMatchWEB():
    
    def __init__(self, config, cleaned_data, help='help.html'):
        super().__init__()
        self.help = help
        self.sender = ''
        self.config = config
        self.cleaned_data = cleaned_data
        self.batch_template = config['files']['pmb_template']
        self.scenarios = config['files']['scenarios']
        
        log_status = True
        try:
            rw = config.get('Windows', 'log_status')
            if rw.lower() in ['false', 'no', 'off']:
                log_status = False
        except:
            pass
        self.log_status = log_status
        self.file_labels = ['Constraints', 'Generators', 'Optimisation', 'Data', 'Results', 'Batch']
        ifiles = [''] * len(self.file_labels)
        self.batch_new_file = False
        self.batch_prefix = False
        self.batch_file = self.cleaned_data['batch_file']
        self.more_details = False
        self.constraints = None
        self.constraints_file = self.cleaned_data['constraints_file']
        self.generators = None
        self.generators_file = self.cleaned_data['generators_file']
        self.generators_sheet = self.cleaned_data['generators_sheet']
        self.optimisation = None
        self.optimisation_file = self.cleaned_data['optimisation_file']
        self.adjustto = None # adjust capacity to this
        self.adjust_cap = 25
        self.adjust_gen = False
        self.change_res = True
        self.adjusted_lcoe = True
        self.carbon_price = 0.
        self.carbon_price_max = 200.
        self.discount_rate = 0.
        self.data_file = self.cleaned_data['data_file']
        self.load_year = self.cleaned_data['load_year']
        self.load_file = './siren_web/siren_files/SWIS/siren_data/swis_load_hourly_$year$_for_sam.csv'
        self.optimise_choice = 'LCOE'
        self.optimise_generations = 20
        self.optimise_mutation = 0.005
        self.optimise_population = 50
        self.optimise_stop = 0
        self.optimise_debug = False
        self.optimise_default = None
        self.optimise_multiplot = True
        self.optimise_multisurf = False
        self.optimise_multitable = False
        self.optimise_to_batch = True
        self.remove_cost = True
        self.results = ''
        self.results_prefix = ''
        self.results_file = self.cleaned_data['results_file']
        self.dispatchable = ['Biomass', 'Geothermal', 'Pumped Hydro', 'Solar Thermal', 'CST'] # RE dispatchable
        self.save_tables = False
        self.show_multipliers = False
        self.show_correlation = False
        self.summary_sources = True
        self.surplus_sign = 1 # Note: Preferences file has it called shortfall_sign
        # it's easier for the user to understand while for the program logic surplus is easier
        self.underlying = ['Rooftop PV'] # technologies contributing to underlying (but not operational) load
        self.operational = []
        iorder = []
        iorder = self.config['powermatch']['generators_left_column'].split(',')
        self.order = iorder
        self.targets = {}

    def initialize_processor(self, config):
        self.processor = PowerMatchProcessor(
            config, self.scenarios, self.generators, self.constraints, None,
            event_callback=None,
            status_callback=None  # Pass setStatus as a callback
            )
        
    def getGenerators(self, ws):
        if ws is None:
            self.generators = {}
            args = {'name': '<name>', 'constraint': '<constraint>'}
            self.generators['<name>'] = Facility(**args)
            return
        if ws.cell_value(0, 0) != 'Name':
            success_message = 'Not a ' + self.file_labels[G] + ' worksheet.'
            return
        args = ['name', 'order', 'constraint', 'capacity', 'lcoe', 'lcoe_cf', 'emissions', 'initial',
                'capex', 'fixed_om', 'variable_om', 'fuel', 'disc_rate', 'lifetime', 'area']
        possibles = {'name': 0}
        for col in range(ws.ncols):
            try:
                arg = ws.cell_value(0, col).lower()
            except:
                continue
            if arg in args:
                possibles[arg] = col
            elif ws.cell_value(0, col)[:9] == 'Capital':
                possibles['capex'] = col
            elif ws.cell_value(0, col)[:8] == 'Discount':
                possibles['disc_rate'] = col
            elif ws.cell_value(0, col)[:8] == 'Dispatch':
                possibles['order'] = col
            elif ws.cell_value(0, col)[:9] == 'Emissions':
                possibles['emissions'] = col
            elif ws.cell_value(0, col) == 'FOM':
                possibles['fixed_om'] = col
            elif ws.cell_value(0, col) == 'LCOE CF':
                possibles['lcoe_cf'] = col
            elif ws.cell_value(0, col)[:4] == 'LCOE':
                possibles['lcoe'] = col
            elif ws.cell_value(0, col) == 'VOM':
                possibles['variable_om'] = col
        self.generators = {}
        for row in range(1, ws.nrows):
            if ws.cell_value(row, 0) is None:
                continue
            in_args = {}
            for key, value in possibles.items():
                in_args[key] = ws.cell_value(row, value)
            self.generators[str(ws.cell_value(row, 0))] = Facility(**in_args)
        return

    def getOptimisation(self, ws):
        if ws is None:
            self.optimisation = {}
            self.optimisation['<name>'] = Optimisation('<name>', 'None', None)
            return
        if ws.cell_value(0, 0) != 'Name':
            success_message = 'Not an ' + self.file_labels[O] + ' worksheet.'
            return
        cols = ['Name', 'Approach', 'Values', 'Capacity Max', 'Capacity Min',
                'Capacity Step', 'Capacities']
        coln = [-1] * len(cols)
        for col in range(ws.ncols):
            try:
                i = cols.index(ws.cell_value(0, col))
                coln[i] = col
            except:
                pass
        if coln[0] < 0:
            success_message = 'Not an ' + self.file_labels[O] + ' worksheet.'
            return
        self.optimisation = {}
        for row in range(1, ws.nrows):
            tech = ws.cell_value(row, 0)
            if tech is None:
                continue
            if coln[2] > 0: # values format
                self.optimisation[tech] = Optimisation(tech,
                                     ws.cell_value(row, coln[1]),
                                     ws.cell_value(row, coln[2]))
            else:
                if ws.cell_value(row, coln[1]) == 'Discrete': # fudge values format
                    self.optimisation[tech] = Optimisation(tech,
                                         ws.cell_value(row, coln[1]),
                                         ws.cell_value(row, coln[-1]))
                else:
                    self.optimisation[tech] = Optimisation(tech, '', '')
                    for col in range(1, len(coln)):
                        if coln[col] > 0:
                            attr = cols[col].lower().replace(' ', '_')
                            setattr(self.optimisation[tech], attr,
                                    ws.cell_value(row, coln[col]))
            try:
                self.optimisation[tech].capacity = self.generators[tech].capacity
            except:
                pass
        return

    def getBatch(self, ws, option):
        global columns, rows, values
        def recurse(lvl):
            if lvl >= len(rows) - 1:
                return
            for i in range(len(values[lvl])):
                columns[lvl] = columns[lvl] + [values[lvl][i]] * cols[lvl+1]
                recurse(lvl + 1)

        def step_split(steps):
            bits = steps.split(',')
            if len(bits) == 1:
                bits = steps.split(';')
            try:
                strt = int(bits[0])
            except:
                return 0, 0, 0, -1
            try:
                stop = int(bits[1])
                step = int(bits[2])
                try:
                    frst = int(bits[3])
                except:
                    frst = -1
            except:
                return strt, strt, strt, frst
            return strt, stop, step, frst

        if ws is None:
            self.setStatus(self.file_labels[B] + ' worksheet missing.')
            return False
        istrt = 0
        year_row = -1
        for row in range(3):
            if ws.cell_value(row, 0) in ['Model', 'Model Label', 'Technology']:
                istrt = row + 1
                break
        else:
            success_message = 'Not a ' + self.file_labels[B] + ' worksheet.'
            return False
        self.batch_models = [{}] # cater for a range of capacities
        self.batch_report = [['Capacity (MW/MWh)', 1]]
        self.batch_tech = []
        istop = ws.nrows
        inrows = False
        for row in range(istrt, ws.nrows):
            tech = ws.cell_value(row, 0)
            if tech is not None and tech != '':
                if year_row < 0 and tech[:4].lower() == 'year':
                    year_row = row
                    continue
                inrows = True
                if tech[:8].lower() != 'capacity':
                    if tech.find('.') > 0:
                        tech = tech[tech.find('.') + 1:]
                    if tech != 'Total' and tech not in self.generators.keys():
                        success_message = 'Unknown technology - ' + tech + ' - in batch file.'
                        return False
                    self.batch_tech.append(ws.cell_value(row, 0))
                else:
                    self.batch_report[0][1] = row + 1
            elif inrows:
                istop = row
                break
            if tech[:5] == 'Total':
                istop = row + 1
                break
        if len(self.batch_tech) == 0:
            success_message = 'No input technologies found in ' + self.file_labels[B] + ' worksheet (try opening and re-saving the workbook).'
            return False
        carbon_row = -1
        discount_row = -1
        for row in range(istop, ws.nrows):
            if ws.cell_value(row, 0) is not None and ws.cell_value(row, 0) != '':
                if ws.cell_value(row, 0).lower() in ['chart', 'graph', 'plot']:
                    self.batch_report.append(['Chart', row + 1])
                    break
                if ws.cell_value(row, 0).lower() in ['carbon price', 'carbon price ($/tco2e)']:
                    carbon_row = row
                if ws.cell_value(row, 0).lower() == 'discount rate' or ws.cell_value(row, 0).lower() == 'wacc':
                    discount_row = row
                self.batch_report.append([techClean(ws.cell_value(row, 0), full=True), row + 1])
        range_rows = {}
        for col in range(1, ws.ncols):
            model = ws.cell_value(istrt - 1, col)
            if model is None:
                break
            self.batch_models[0][col] = {'name': model}
            if option == T and year_row < 0:
                self.batch_models[0][col]['year'] = str(model)
            for row in range(istrt, istop):
                if row == year_row:
                    if ws.cell_value(row, col) is not None and ws.cell_value(row, col) != '':
                        self.batch_models[0][col]['year'] = str(ws.cell_value(row, col))
                    continue
                tech = ws.cell_value(row, 0)
                try:
                    if ws.cell_value(row, col) > 0:
                        self.batch_models[0][col][tech] = ws.cell_value(row, col)
                except:
                    if ws.cell_value(row, col) is None:
                        pass
                    elif ws.cell_value(row, col).find(',') >= 0 or ws.cell_value(row, col).find(';') >= 0:
                        try:
                            range_rows[col].append(row)
                        except:
                            range_rows[col] = [row]
                        try:
                            strt, stop, step, frst = step_split(ws.cell_value(row, col))
                            self.batch_models[0][col][tech] = strt
                            if frst >= 0 and len(range_rows[col]) > 1:
                                del range_rows[col][-1]
                                range_rows[col].insert(0, row)
                        except:
                            pass
                    pass
            if carbon_row >= 0:
                if isinstance(ws.cell_value(carbon_row, col), float):
                    self.batch_models[0][col]['Carbon Price'] = ws.cell_value(carbon_row, col)
                elif isinstance(ws.cell_value(carbon_row, col), int):
                    self.batch_models[0][col]['Carbon Price'] = float(ws.cell_value(carbon_row, col))
            if discount_row >= 0:
                if isinstance(ws.cell_value(discount_row, col), float):
                    self.batch_models[0][col]['Discount Rate'] = ws.cell_value(discount_row, col)
                elif isinstance(ws.cell_value(discount_row, col), int):
                    self.batch_models[0][col]['Discount Rate'] = float(ws.cell_value(discount_row, col))
        if len(self.batch_models[0]) == 0:
            success_message = 'No models found in ' + self.file_labels[B] + ' worksheet (try opening and re-saving the workbook).'
            return False
        if len(range_rows) == 0:
            return True
        # cater for ranges - so multiple batch_models lists
        for rcol, ranges in range_rows.items():
            rows = {}
            for rw in ranges:
                rows[rw] = ws.cell_value(rw, rcol)
            if len(ranges) > 1: # create sheet for each range else one sheet
                values = []
                cols = [1]
                for i in range(len(ranges) -1, 0, -1):
                    strt, stop, step, frst = step_split(rows[ranges[i]])
                    values.insert(0, [])
                    for stp in range(strt, stop + step, step):
                        values[0].append(stp)
                    cols.insert(0, cols[0] * len(values[0]))
                columns = [[]] * len(rows)
                recurse(0)
                my_tech = ws.cell_value(ranges[0], 0)
                tech_2 = ws.cell_value(ranges[1], 0)
              # produce new batch_models entry for first range tech
                techs = {}
                for c in range(1, len(ranges)):
                    techs[ws.cell_value(ranges[c], 0)] = c - 1
                bits = my_tech.split('.')
                strt, stop, step, frst = step_split(rows[ranges[0]])
                for sht in range(strt, stop + step, step):
                    self.batch_models.append({})
                    for c2 in range(len(columns[0])):
                        self.batch_models[-1][c2] = {}
                        for key, value in self.batch_models[0][rcol].items():
                            self.batch_models[-1][c2][key] = value
                        self.batch_models[-1][c2][my_tech] = sht
                        for key, value in techs.items():
                            self.batch_models[-1][c2][key] = columns[value][c2]
                        self.batch_models[-1][c2]['name'] = f'{bits[-1]}_{sht}_{tech_2}'
            else:
                my_tech = ws.cell_value(ranges[0], 0)
                self.batch_models.append({})
                strt, stop, step, frst = step_split(rows[ranges[0]])
                c2 = -1
                for ctr in range(strt, stop + step, step):
                    c2 += 1
                    self.batch_models[-1][c2] = {}
                    if c2 == 0:
                        self.batch_models[-1][c2]['hdr'] = ws.cell_value(ranges[0], 0) # fudge to get header name
                    for key, value in self.batch_models[0][rcol].items():
                        self.batch_models[-1][c2][key] = value
                    self.batch_models[-1][c2][my_tech] = ctr
                #    for key, value in techs.items():
                 #       self.batch_models[-1][c2][key] = columns[value][c2]
                    self.batch_models[-1][c2]['name'] = f'Model {c2 + 1}'
        return True

    def data_sources(self, sheet, sheet_row, pm_data_file, option):
        normal = oxl.styles.Font(name='Arial')
        bold = oxl.styles.Font(name='Arial', bold=True)
        sheet.cell(row=sheet_row, column=1).value = 'Data sources'
        sheet.cell(row=sheet_row, column=1).font = bold
        sheet_row += 1
        sheet.cell(row=sheet_row, column=1).value = 'Scenarios folder'
        sheet.cell(row=sheet_row, column=1).font = normal
        sheet.cell(row=sheet_row, column=2).value = self.scenarios
        sheet.cell(row=sheet_row, column=2).font = normal
        sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
        sheet_row += 1
        sheet.cell(row=sheet_row, column=1).value = 'Powermatch data file'
        sheet.cell(row=sheet_row, column=1).font = normal
        if pm_data_file[: len(self.scenarios)] == self.scenarios:
            pm_data_file = pm_data_file[len(self.scenarios):]
        sheet.cell(row=sheet_row, column=2).value = pm_data_file
        sheet.cell(row=sheet_row, column=2).font = normal
        sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
        sheet_row += 1
        sheet.cell(row=sheet_row, column=1).value = 'Constraints worksheet'
        sheet.cell(row=sheet_row, column=1).font = normal
        sheet.cell(row=sheet_row, column=2).value = str(self.constraints_file)
        sheet.cell(row=sheet_row, column=2).font = normal
        sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
        sheet_row += 1
        sheet.cell(row=sheet_row, column=1).value = 'Generators worksheet'
        sheet.cell(row=sheet_row, column=1).font = normal
        if option == T:
            sheet.cell(row=sheet_row, column=2).value = self.generators
        else:
            sheet.cell(row=sheet_row, column=2).value = self.generators
        sheet.cell(row=sheet_row, column=2).font = normal
        sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
        return sheet_row

    def pmClicked(self, action):
        def get_load_data(load_file):
            try:
                load_file_act = load_file.replace('$year$', self.load_year)
                tf = open(load_file_act, 'r')
                lines = tf.readlines()
                tf.close()
            except Exception as e:
                # Log the error but don't prevent form from loading
                print(f"Error reading load file: {str(e)}")
                return None
            load_data = []
            bit = lines[0].rstrip().split(',')
            if len(bit) > 0: # multiple columns
                for b in range(len(bit)):
                    if bit[b][:4].lower() == 'load':
                        if bit[b].lower().find('kwh') > 0: # kWh not MWh
                            for i in range(1, len(lines)):
                                bit = lines[i].rstrip().split(',')
                                load_data.append(float(bit[b]) * 0.001)
                        else:
                            for i in range(1, len(lines)):
                                bit = lines[i].rstrip().split(',')
                                load_data.append(float(bit[b]))
            else:
                for i in range(1, len(lines)):
                    load_data.append(float(lines[i].rstrip()))
            return load_data

        def get_batch_prefix(report_group):
            if report_group == 'Lifetime Emissions':
                return 'LES_'
            if report_group in ['Correlation To Load', 'Static Variables']:
                return ''
            bits = report_group.split(' ')
            for i in range(len(bits) -1, -1, -1):
                if bits[i][0] == '(' and bits[i][-1] == ')':
                    del bits[i]
            if len(bits) == 1:
                abr = bits[0][0] + bits[0][-1]
            else:
                abr = ''
                for bit in bits:
                    abr += bit[0]
            return abr.upper() + '_'

        col_letters = ' ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        self.sender= action
        if action == 'Detail': # detailed spreadsheet?
            option = D
        elif action == 'Optimise': # do optimisation?
            option = O
            self.optExit = False #??
        elif action == 'Batch': # do batch processsing
            option = B
        elif action == 'Transition': # do transition processsing
            option = T
        else:
            option = S
        sender_name = action
        constraints = self.cleaned_data['constraints_file']
        constraints_sheet = self.cleaned_data['constraints_sheet']

        generators = self.cleaned_data['generators_file']
        generators_sheet = self.cleaned_data['generators_sheet']
        if not os.path.exists(generators):
            return None
        try:
            ts = WorkBook()
            ts.open_workbook(generators)
            ws = ts.sheet_by_name(generators_sheet)
            self.getGenerators(ws)
            ts.close()
            del ts
        except FileNotFoundError:
            err_msg = 'Generators file not found - ' + generators
            self.getGenerators(None)
        except Exception as e:
            err_msg = 'Error accessing Generators' + str(e)
            self.getGenerators(None)
        self.initialize_processor(self.config)
        
        if option == B or option == T: # has to be xlsx workbook
            batch_file = self.cleaned_data['batch_file']
            try:
                ts = WorkBook()
                bwbopen_start = time.time()
                ts.open_workbook(batch_file)
                ws = ts.sheet_by_index(0)
                tim = time.time() - bwbopen_start
                if tim < 60:
                    tim = '%.1f secs' % tim
                else:
                    hhmm = tim / 60.
                    tim = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
                ok = self.getBatch(ws, option)
                ts.close()
                del ts
                if not ok:
                    return
            except FileNotFoundError:
                err_msg = 'Batch file not found - ' + batch_file
            except Exception as e:
                err_msg = 'Error accessing Batch file ' + str(e)

        err_msg = ''
        if self.generators is None:
            try:
                ts = WorkBook()
                ts.open_workbook(self.generators_file)
                self.getGenerators(ws)
                ts.close()
                del ts
            except FileNotFoundError:
                if err_msg != '':
                    err_msg += ' nor Generators - ' + self.generators_file
                else:
                    err_msg = 'Generators file not found - ' + self.generators_file
                self.getGenerators(None)
            except:
                if err_msg != '':
                    err_msg += ' and Generators'
                else:
                    err_msg = 'Error accessing Generators'
                self.getGenerators(None)
                
        if option == B or option == T: # has to be xlsx workbook
            try:
                ts = WorkBook()
                bwbopen_start = time.time()
                ts.open_workbook(self.batch_file)
                ws = ts.sheet_by_index(0)
                tim = time.time() - bwbopen_start
                if tim < 60:
                    tim = '%.1f secs' % tim
                else:
                    hhmm = tim / 60.
                    tim = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
                self.setStatus(f'{self.file_labels[B]} workbook opened ({tim})')
                ok = self.getBatch(ws, option)
                ts.close()
                del ts
                if not ok:
                    return
            except FileNotFoundError:
                err_msg = 'Batch file not found - ' + self.batch_file
            except Exception as e:
                err_msg = 'Error accessing Batch file ' + str(e)
                
        if option == O and self.optimisation is None:
            optimisation_file = self.cleaned_data['optimisation_file']
            optimisation_sheet = self.cleaned_data['optimisation_sheet']
            try:
                ts = WorkBook()
                ts.open_workbook(self.optimisation_file)
                self.getOptimisation(ws)
                ts.close()
                del ts
                if self.optimisation_file is None:
                    if err_msg != '':
                        err_msg += ' not an Optimisation worksheet'
                    else:
                        err_msg = 'Not an optimisation worksheet'
            except FileNotFoundError:
                if err_msg != '':
                    err_msg += ' nor Optimisation - ' + self.optimisation_file
                else:
                    err_msg = 'Optimisation file not found - ' + self.optimisation_file
            except:
                if err_msg != '':
                    err_msg += ' and Optimisation'
                else:
                    err_msg = 'Error accessing Optimisation'
            if self.optimisation_file is None:
                self.getOptimisation(None)
        if err_msg != '':
            self.setStatus(err_msg)
            return
        pm_data_file = self.cleaned_data['data_file']
        if pm_data_file[-5:] != '.xlsx': #xlsx format only
            success_message = "Not a Powermatch data spreadsheet (1)"
            return
        if not os.path.exists(pm_data_file):
            success_message = 'Data file not found - ' + pm_data_file
            return
        try:
            ts = oxl.load_workbook(pm_data_file)
        except Exception as e:
            success_message = f"Error processing batch file {str(e)}"
            return
        ws = ts.worksheets[0]
        top_row = ws.max_row - 8760
        
        if top_row < 1 or (ws.cell(row=top_row, column=1).value != 'Hour' \
                           or ws.cell(row=top_row, column=2).value != 'Period'):
            success_message = f'Not a Powermatch data spreadsheet (2; {top_row})'
            return
        typ_row = top_row - 1
        gen_row = typ_row
        while typ_row > 0:
            if ws.cell(row=typ_row, column=1).value[:9] == 'Generated':
                gen_row = typ_row
            if ws.cell(row=typ_row, column=3).value in tech_names:
                break
            typ_row -= 1
        else:
            success_message = 'no suitable data'
            return
        
        do_zone = False
        zone_row = typ_row - 1
        try:
            if ws.cell(row=zone_row, column=1).value.lower() == 'zone':
                do_zone = True
                zone_techs = []
        except:
            pass
        
        icap_row = typ_row + 1
        while icap_row < top_row:
            if ws.cell(row=icap_row, column=1).value[:8] == 'Capacity':
                break
            icap_row += 1
        else:
            success_message = 'no capacity data'
            return
        
        year = ws.cell(row=top_row + 1, column=2).value[:4]
        pmss_details = {} # contains name, generator, capacity, tech_type, col, multiplier
        pmss_data = []
        re_order = [] # order for re technology
        dispatch_order = [] # order for dispatchable technology
        load_columns = {}
        load_col = -1
        strt_col = 3
        try:
            year = self.load_year
            strt_col = 4
            load_col = len(pmss_data)
            typ = 'L'
            capacity = 0
            fctr = 1
            pmss_details['Load'] = PM_Facility('Load', 'Load', 0, 'L', len(pmss_data), 1)
            load_columns[year] = len(pmss_data)
            pmss_data.append([])
        except:
            pass
        
        pmss_data[-1] = get_load_data(self.load_file)
        re_order.append('Load')
        zone = ''
        for col in range(strt_col, ws.max_column + 1):
            try:
                valu = ws.cell(row=typ_row, column=col).value.replace('-','')
                i = tech_names.index(valu)
            except:
                continue
            key = tech_names[i]
            if key == 'Load':
                load_col = len(pmss_data)
                typ = 'L'
                capacity = 0
                fctr = 1
            else:
                if do_zone:
                    cell = ws.cell(row=zone_row, column=col)
                    if type(cell).__name__ == 'MergedCell':
                        pass
                    else:
                        zone = ws.cell(row=zone_row, column=col).value
                    if zone is None or zone == '':
                        zone_tech = valu
                    else:
                        zone_tech = zone + '.' + valu
                    key = zone_tech
                    zone_techs.append(key)
                else: # temp
                    if len(self.re_capacity) > 0 and tech_names[i] not in self.re_capacity.keys():
                        continue
                try:
                    capacity = float(ws.cell(row=icap_row, column=col).value)
                except:
                    continue
                if capacity <= 0:
                    continue
                typ = 'R'
                if do_zone:
                    fctr = 1
                elif tech_names[i] in self.re_capacity and capacity > 0:
                    fctr = self.re_capacity[tech_names[i]] / capacity
                else:
                    fctr = 1
            pmss_details[key] = PM_Facility(key, tech_names[i], capacity, typ, len(pmss_data), fctr)
            if key == 'Load':
                load_columns[year] = len(pmss_data)
            pmss_data.append([])
            re_order.append(key)
            for row in range(top_row + 1, ws.max_row + 1):
                pmss_data[-1].append(ws.cell(row=row, column=col).value)
        pmss_details['Load'].capacity = sum(pmss_data[load_col])
        do_adjust = False
        if option == O:
            for itm in range(self.order.count()):
                gen = self.order.item(itm).text()
                try:
                    if self.generators[gen].capacity <= 0:
                        continue
                except KeyError as err:
                    success_message = 'Key Error: No Generator entry for ' + str(err)
                    continue
                try:
                    if self.generators[gen].constraint in self.constraints and \
                      self.constraints[self.generators[gen].constraint].category == 'Generator':
                        typ = 'G'
                    else:
                        typ = 'S'
                except:
                    continue
                dispatch_order.append(gen)
                pmss_details[gen] = PM_Facility(gen, gen, self.generators[gen].capacity, typ, -1, 1)
            if self.adjust.isChecked():
                 pmss_details['Load'].multiplier = self.adjustto['Load'] / pmss_details['Load'].capacity
            self.optClicked(year, option, pmss_details, pmss_data, re_order, dispatch_order,
                            None, None)
            return
        ts.close()
        
        if self.results_file == '':
            i = pm_data_file.rfind('/')
            if i >= 0:
                data_file = pm_data_file[i + 1:]
            else:
                data_file = pm_data_file
            data_file = data_file.replace('data', 'results')
            data_file = data_file.replace('Data', 'Results')
            if data_file == pm_data_file[i + 1:]:
                j = data_file.find(' ')
                if j > 0:
                    jnr = ' '
                else:
                    jnr = '_'
                j = data_file.rfind('.')
                data_file = data_file[:j] + jnr + 'Results' + data_file[j:]
            self.results_file = data_file
        else:
            data_file = self.results_file
        if self.results_prefix != '':
            j = data_file.rfind('/')
            data_file = data_file[: j + 1] + self.results_prefix + '_' + data_file[j + 1:]
        for itm in self.order:
            gen = itm
            try:
                if self.generators[gen].capacity <= 0:
                    continue
            except KeyError as err:
                success_message = 'Key Error: No Generator entry for ' + str(err)
                continue
            except:
                continue
            if do_adjust:
                try:
                    if self.adjustto[gen] <= 0:
                        continue
                except:
                    pass
            try:
                if self.generators[gen].constraint in self.constraints and \
                  self.constraints[self.generators[gen].constraint].category == 'Generator':
                    typ = 'G'
                else:
                    typ = 'S'
            except:
                continue
            dispatch_order.append(gen)
            pmss_details[gen] = PM_Facility(gen, gen, self.generators[gen].capacity, typ, -1, 1)
        if option == B or option == T:
            if option == T:
                gen_sheet = self.generators_sheet
                trn_year = ''
                newfile = self.generators
                gen_book = WorkBook()
                gen_book.open_workbook(newfile)
                pmss_details['Load'].multiplier = 1
            elif self.adjust.isChecked():
                generated = sum(pmss_data[load_col])
                datain = [['Load', 'L', generated]]
                adjustto = {'Load': generated}
                adjust = Adjustments(self, datain, adjustto, self.adjust_cap, None,
                                     show_multipliers=self.show_multipliers)
                adjust.exec_()
                adjustto = adjust.getValues()
                pmss_details['Load'].multiplier = adjustto['Load'] / pmss_details['Load'].capacity
       #     start_time = time.time() # just for fun
            batch_details = {'Capacity (MW/MWh)': [st_cap, '#,##0.00'],
                             'To Meet Load (MWh)': [st_tml, '#,##0'],
                             'Generation (MWh)': [st_sub, '#,##0'],
                             'Capacity Factor': [st_cfa, '#,##0.0%'],
                             'Cost ($/Yr)': [st_cst, '#,##0'],
                             'LCOG ($/MWh)': [st_lcg, '#,##0.00'],
                             'LCOE ($/MWh)': [st_lco, '#,##0.00'],
                             'Emissions (tCO2e)': [st_emi, '#,##0'],
                             'Emissions Cost': [st_emc, '#,##0'],
                             'LCOE With CO2 ($/MWh)': [st_lcc, '#,##0.00'],
                             'Max MWh': [st_max, '#,##0'],
                             'Capital Cost': [st_cac, '#,##0'],
                             'Lifetime Cost': [st_lic, '#,##0'],
                             'Lifetime Emissions': [st_lie, '#,##0'],
                             'Lifetime Emissions Cost': [st_lec, '#,##0'],
                             'Area': [st_are, '#,###0.00']}
            batch_extra = {'RE': ['#,##0.00', ['RE %age', st_cap], ['Storage %age', st_cap], ['RE %age of Total Load', st_cap]],
                           'Load Analysis': ['#,##0', ['Load met', st_tml], ['Load met %age', st_cap], ['Shortfall', st_tml], ['Total Load', st_tml],
                           ['Largest Shortfall', st_cap], ['Storage losses', st_sub], ['Surplus', st_sub], ['Surplus %age', st_cap]],
                           'Carbon': ['#,##0.00', ['Carbon Price', st_cap], ['Carbon Cost', st_emc], ['LCOE incl. Carbon Cost', st_lcc],
                           ['Lifetime Emissions Cost', st_lec]],
                           'Correlation To Load': ['0.0000', ['RE Contribution', st_cap], ['RE plus Storage', st_cap],
                           ['To Meet Load', st_cap]],
                           'Static Variables': ['#,##0.00', ['Carbon Price', st_cap], ['Lifetime', st_cap],
                           ['Discount Rate', st_cap]],
                           'Optimisation Parameters': ['#,##0.00', ['Population size', 1], ['No. of iterations', 1],
                           ['Mutation probability', 1], ['Exit if stable', 1], ['Optimisation choice', 1],
                           ['Variable', 1], ['LCOE', 1], ['Load%', 1], ['Surplus%', 1], ['RE%', 1],
                           ['Cost', 1], ['CO2', 1]]}
                           # LCOE (incl. CO2)
         #   batch_extra['Optimisation Parameters'] = []
            batch_extra['LCOE ($/MWh)'] = ['#,##0.00']
            for tech in self.batch_tech:
                if tech == 'Total':
                    batch_extra['LCOE ($/MWh)'].append([tech + ' LCOE ($/MWh)'])
                else:
                    batch_extra['LCOE ($/MWh)'].append([tech])
            batch_extra['LCOE ($/MWh)'].append(['LCOE', st_lco])
            batch_extra['LCOE With CO2 ($/MWh)'] = ['#,##0.00']
            for tech in self.batch_tech:
                batch_extra['LCOE With CO2 ($/MWh)'].append([tech])
            batch_extra['LCOE With CO2 ($/MWh)'].append(['LCOE incl. Carbon Cost', st_lcc])
         #   batch_extra['To Meet Load (MWh)'] = ['#,##0.00', ['Total', st_tml]]
            wbopen_start = time.time()
            wb = oxl.load_workbook(self.batch_file)
            tim = time.time() - wbopen_start
            if tim < 60:
                tim = '%.1f secs' % tim
            else:
                hhmm = tim / 60.
                tim = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
            self.setStatus(f'{self.file_labels[B]} workbook re-opened for update ({tim})')
            batch_input_sheet = wb.worksheets[0]
            rpt_time = QtCore.QDateTime.toString(QtCore.QDateTime.currentDateTime(), 'yyyy-MM-dd_hhmm')
            if self.batch_new_file:
                wb.close()
                i = self.batch_file('.')
                suffix = '_report_' + rpt_time
                batch_report_file = self.batch_file[:i] + suffix + self.batch_file[i:]
                if batch_report_file == '':
                    self.setStatus(self.sender + ' aborted')
                    return
                if batch_report_file[-5:] != '.xlsx':
                    batch_report_file += '.xlsx'
                if os.path.exists(batch_report_file) and not self.replace_last.isChecked():
                    wb = oxl.load_workbook(batch_report_file)
                    bs = wb.create_sheet('Results_' + rpt_time)
                else:
                    wb = oxl.Workbook()
                    bs = wb.active
                    bs.title = 'Results_' + rpt_time
            else:
                batch_report_file = self.batch_file
                if self.replace_last.isChecked():
                    del_sht = ''
                    for sht in wb.sheetnames:
                        if sht[:8] == 'Results_' and sht > del_sht:
                            del_sht = sht
                    if del_sht != '':
                        del wb[del_sht]
                        del_sht = del_sht.replace('Results', 'Charts')
                        if del_sht in wb.sheetnames:
                            del wb[del_sht]
                bs = wb.create_sheet('Results_' + rpt_time)
            start_time = time.time() # just for fun
            normal = oxl.styles.Font(name='Arial')
            bold = oxl.styles.Font(name='Arial', bold=True)
            grey = oxl.styles.colors.Color(rgb='00f2f2f2')
            grey_fill = oxl.styles.fills.PatternFill(patternType='solid', fgColor=grey)
            total_models = 0
            for sht in range(len(self.batch_models)):
                total_models = total_models + len(self.batch_models[sht])
            try:
                incr = 20 / total_models
            except:
                incr = .05
            prgv = incr
            prgv_int = 0
            model_row = False
            model_row_no = 0
            sht_nam_len = max(len(str(len(self.batch_models))), 2)
            for sht in range(len(self.batch_models)):
                sheet_start = time.time()
                if sht == 0: # normal case
                   # copy header rows to new worksheet
                   merged_cells = []
                   merge_cells = None
                   model_row = False
                   model_cols = len(self.batch_models[sht])
                   for row in range(1, self.batch_report[0][1] + 2):
                       if batch_input_sheet.cell(row=row, column=1).value in ['Model', 'Model Label', 'Technology']:
                           model_row = True
                           model_row_no = row
                       else:
                           model_row = False
                       for col in range(1, model_cols + 2):
                           cell = batch_input_sheet.cell(row=row, column=col)
                           if type(cell).__name__ == 'MergedCell':
                               if merge_cells is None:
                                   merge_cells = [row, col - 1, col]
                               else:
                                   merge_cells[2] = col
                               continue
                           if model_row and col > 1:
                               new_cell = bs.cell(row=row, column=col, value=self.batch_models[sht][col - 1]['name'])
                           else:
                               new_cell = bs.cell(row=row, column=col, value=cell.value)
                           if cell.has_style:
                               new_cell.font = copy(cell.font)
                               new_cell.border = copy(cell.border)
                               new_cell.fill = copy(cell.fill)
                               new_cell.number_format = copy(cell.number_format)
                               new_cell.protection = copy(cell.protection)
                               new_cell.alignment = copy(cell.alignment)
                           if merge_cells is not None:
                               bs.merge_cells(start_row=row, start_column=merge_cells[1], end_row=row, end_column=merge_cells[2])
                               merged_cells.append(merge_cells)
                               merge_cells = None
                       if merge_cells is not None:
                           bs.merge_cells(start_row=row, start_column=merge_cells[1], end_row=row, end_column=merge_cells[2])
                           merged_cells.append(merge_cells)
                           merge_cells = None
                   try:
                       normal = oxl.styles.Font(name=cell.font.name, sz=cell.font.sz)
                       bold = oxl.styles.Font(name=cell.font.name, sz=cell.font.sz, bold=True)
                   except:
                       pass
                else:
                    sheet_name = f'{sht:0{sht_nam_len}}'
                    if sheet_name in wb.sheetnames:
                        del wb[sheet_name]
                        if 'Charts_' + sheet_name in wb.sheetnames:
                            del wb['Charts_' + sheet_name]
                    bs = wb.create_sheet(sheet_name)
                    if model_row_no > 1:
                        title = self.batch_models[sht][0]['name']
                        tech_2 = title.split('_')
                        if len(tech_2) > 1:
                            tech_2 = tech_2[-1]
                            bits_2 = tech_2.split('.')[-1]
                            title = title.replace(tech_2, bits_2)
                            cap_2 = self.batch_models[sht][0][tech_2]
                            fst_col = 2
                            bs.cell(row=1, column=2).value = f'{title}_{cap_2}'
                            bs.cell(row=1, column=2).font = normal
                            bs.cell(row=1, column=2).alignment = oxl.styles.Alignment(wrap_text=True, vertical='bottom', horizontal='center')
                            g = 1
                            for i in range(1, len(self.batch_models[sht])):
                                if self.batch_models[sht][i][tech_2] != cap_2:
                                    bs.merge_cells(start_row=1, start_column=fst_col, end_row=1, end_column=i + 1)
                                    fst_col = i + 2
                                    cap_2 = self.batch_models[sht][i][tech_2]
                                    bs.cell(row=1, column=fst_col).value = f'{title}_{cap_2}'
                                    if g == 0:
                                        g = 1
                                    else:
                                        bs.cell(row=1, column=fst_col).fill = grey_fill
                                        g = 0
                                    bs.cell(row=1, column=fst_col).font = normal
                                    bs.cell(row=1, column=fst_col).alignment = oxl.styles.Alignment(wrap_text=True, vertical='bottom', horizontal='center')
                            bs.merge_cells(start_row=1, start_column=fst_col, end_row=1, end_column=i + 2)
                        else:
                            try:
                                title = self.batch_models[sht][0]['hdr'].split('.')[-1]
                                del self.batch_models[sht][0]['hdr']
                            except:
                                pass
                            bs.cell(row=1, column=2).value = f'{title}'
                            bs.cell(row=1, column=2).font = normal
                            bs.cell(row=1, column=2).alignment = oxl.styles.Alignment(wrap_text=True, vertical='bottom', horizontal='center')
                            bs.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(self.batch_models[sht]) + 1)
                column = 1
                gndx = self.batch_report[0][1] # Capacity group starting row
                do_opt_parms = [False, 0, 0, 0]
                total_load_row = 0
                if self.discount_rate > 0:
                    batch_disc_row = 0
                else:
                    batch_disc_row = -1
                if self.carbon_price > 0:
                    batch_carbon_row = 0
                else:
                    batch_carbon_row = -1
                batch_lifetime = False
                batch_data_sources_row = 0
                re_tml_row = 0
                max_load_row = -1
                report_keys = []
                for g in range(len(self.batch_report)):
                    report_keys.append(self.batch_report[g][0])
                if 'Lifetime Cost' in report_keys:
                    batch_lifetime = True
                for g in range(len(self.batch_report)):
                    if self.batch_report[g][0] == 'Chart':
                        continue
                    elif self.batch_report[g][0] == 'Carbon Price':
                        batch_carbon_row = self.batch_report[g][1]
                        continue
                    elif self.batch_report[g][0] == 'Discount Rate' or self.batch_report[g][0].lower() == 'wacc':
                        batch_disc_row = self.batch_report[g][1]
                        continue
                    elif self.batch_report[g][0].lower() == 'data sources':
                        batch_data_sources_row = gndx
                        gndx += 6
                        continue
                    if self.batch_report[g][0] not in batch_details.keys() and self.batch_report[g][0] not in batch_extra.keys():
                        continue
                    self.batch_report[g][1] = gndx
                    if self.batch_prefix:
                        batch_pfx = get_batch_prefix(self.batch_report[g][0])
                    else:
                        batch_pfx = ''
                    bs.cell(row=gndx, column=1).value = self.batch_report[g][0]
                    bs.cell(row=gndx, column=1).font = bold
                    if self.batch_report[g][0] in batch_extra.keys():
                        key = self.batch_report[g][0]
                        if self.batch_report[g][0] == 'Optimisation Parameters':
                            for row in range(1, batch_input_sheet.max_row + 1):
                                if batch_input_sheet.cell(row=row, column=1).value == 'Optimisation Parameters':
                                    do_opt_parms[0] = True
                                    do_opt_parms[1] = gndx
                                    do_opt_parms[2] = row
                                    break
                            for row in range(row, batch_input_sheet.max_row + 1):
                                gndx += 1
                                if batch_input_sheet.cell(row=row, column=1).value == '':
                                    break
                            do_opt_parms[3] = row
                            continue
                        for sp in range(1, len(batch_extra[key])):
                            if batch_extra[key][sp][0] == 'Total Load':
                                total_load_row = gndx + sp
                            elif batch_extra[key][sp][0] == 'Carbon Price':
                                bs.cell(row=gndx + sp, column=1).value = batch_pfx + batch_extra[key][sp][0] + ' ($/tCO2e)'
                            elif batch_extra[key][sp][0] == 'Lifetime':
                                bs.cell(row=gndx + sp, column=1).value = batch_pfx + batch_extra[key][sp][0] + ' (years)'
                            elif batch_extra[key][sp][0] == 'Total incl. Carbon Cost':
                                bs.cell(row=gndx + sp, column=1).value = batch_pfx + 'LCOE incl. Carbon Cost'
                            else:
                                bs.cell(row=gndx + sp, column=1).value = batch_pfx + batch_extra[key][sp][0]
                            if batch_extra[key][sp][0] in ['RE %age of Total Load', 'Total incl. Carbon Cost'] or \
                              batch_extra[key][sp][0].find('LCOE') >= 0 and batch_extra[key][sp][0].find('Total LCOE') < 0:
                                bs.cell(row=gndx + sp, column=1).font = bold
                            else:
                                bs.cell(row=gndx + sp, column=1).font = normal
                        gndx += len(batch_extra[key]) + 1
                        if key == 'Carbon':
                            if not batch_lifetime:
                                gndx -= 1
                                tot_carb_row = gndx - 3
                            else:
                                tot_carb_row = gndx - 4
                        elif key == 'LCOE ($/MWh)':
                            tot_lco_row = gndx - 2
                        elif key == 'LCOE With CO2 ($/MWh)':
                            tot_lcc_row = gndx - 2
                    else:
                        if self.batch_report[g][0] not in batch_details.keys():
                            continue
                        if self.batch_prefix:
                            batch_pfx = get_batch_prefix(self.batch_report[g][0])
                        else:
                            batch_pfx = ''
                        for sp in range(len(self.batch_tech)):
                        #    if self.batch_report[g][0] == 'To Meet Load (MWh)' and sp == 0:
                         #       bs.cell(row=gndx + sp + 1, column=1).value = 'RE Contribution To Load'
                            if self.batch_report[g][0] != 'Capacity Factor' or self.batch_tech[sp] != 'Total':
                                bs.cell(row=gndx + sp + 1, column=1).value = batch_pfx + self.batch_tech[sp]
                            if self.batch_report[g][0] == 'Max MWh' and self.batch_tech[sp] == 'Total':
                                max_load_row = gndx + sp + 1
                                bs.cell(row=max_load_row, column=1).value = batch_pfx + 'Max Load'
                            elif self.batch_tech[sp] == 'Total' and self.batch_report[g][0] != 'Capacity Factor':
                                bs.cell(row=gndx + sp + 1, column=1).value = batch_pfx + self.batch_tech[sp] + ' ' + self.batch_report[g][0]
                            bs.cell(row=gndx + sp + 1, column=1).font = normal
                        if self.batch_report[g][0] == 'Cost ($/Yr)' and batch_disc_row >= 0:
                            batch_disc_row = gndx + sp + 2
                            bs.cell(row=batch_disc_row, column=1).value = batch_pfx + 'Discount Rate'
                            bs.cell(row=batch_disc_row, column=1).font = normal
                        if self.batch_report[g][0] == 'Capacity Factor' and self.batch_tech[-1] == 'Total':
                            gndx += len(self.batch_tech) + 1
                        else:
                            gndx += len(self.batch_tech) + 2
                        if self.batch_report[g][0] == 'Cost ($/Yr)' and batch_disc_row >= 0:
                            gndx += 1
                        if self.batch_report[g][0] == 'To Meet Load (MWh)':
                            re_tml_row = gndx - 1
                            bs.cell(row=re_tml_row, column=1).value = batch_pfx + 'RE Contribution To Load'
                            bs.cell(row=re_tml_row, column=1).font = normal
                            bs.cell(row=re_tml_row + 1, column=1).value = batch_pfx + 'Storage Contribution To Load'
                            bs.cell(row=re_tml_row + 1, column=1).font = normal
                            gndx += 2
                merge_col = 1
                last_name = ''
                # find first varying capacity to create model name
                model_key = ''
                model_nme = ''
                if sht > 0:
                    for key in self.batch_models[sht][0].keys():
                        if key == 'name':
                            continue
                        try:
                            if self.batch_models[sht][0][key] != self.batch_models[sht][1][key]:
                                model_key = key
                                bits = key.split('.')[-1].split(' ')
                                for bit in bits:
                                    model_nme += bit.strip('()')[0]
                                model_nme += '-'
                                break
                        except:
                            pass
                if option == T:
                    capex_table = {}
                    for fac in pmss_details.keys():
                        capex_table[fac] = {'cum': 0}
                for model, capacities in self.batch_models[sht].items():
                    if option == T:
                        if capacities['year'] != trn_year:
                            # get generators and load for new year
                            trn_year = capacities['year']
                            year = str(trn_year)
                            ws = gen_book.sheet_by_name(gen_sheet.replace('$YEAR$', year))
                            self.getGenerators(ws)
                    for fac in pmss_details.keys():
                        if fac == 'Load':
                            pmss_details['Load'].capacity = sum(pmss_data[load_columns[year]])
                            pmss_details['Load'].col = load_columns[year]
                            continue
                        pmss_details[fac].multiplier = 0
                    column += 1
                    dispatch_order = []
                    for key, capacity in capacities.items(): # cater for zones
                        if key in ['Carbon Price', 'Discount Rate', 'Total']:
                            continue
                        if key == 'name' and model_row_no > 0:
                            if model_key != '':
                                bs.cell(row=model_row_no, column=column).value = f'{model_nme}{capacities[model_key]}'
                            elif option == T:
                                bs.cell(row=model_row_no, column=column).value = f'{capacity}'
                            else:
                                bs.cell(row=model_row_no, column=column).value = f'Model {model + 1}'
                            bs.cell(row=model_row_no, column=column).font = normal
                            bs.cell(row=model_row_no, column=column).alignment = oxl.styles.Alignment(wrap_text=True,
                                    vertical='bottom', horizontal='center')
                            continue
                        if key == 'year':
                            if capacity in load_columns.keys():
                                pmss_details['Load'].col = load_columns[capacity]
                            else:
                                load_columns[capacity] = len(pmss_data)
                                pmss_data.append([])
                                load_file = self.load_files.replace('$YEAR$', capacity)
                                pmss_data[-1] = get_load_data(load_file)
                                pmss_details['Load'].col = load_columns[capacity]
                            pmss_details['Load'].capacity = sum(pmss_data[pmss_details['Load'].col])
                            continue
                        if key not in re_order:
                            dispatch_order.append(key)
                        if key not in pmss_details.keys():
                            gen = key[key.find('.') + 1:]
                            if gen in re_order:
                                typ = 'R'
                            elif self.generators[gen].constraint in self.constraints and \
                              self.constraints[self.generators[gen].constraint].category == 'Generator':
                                typ = 'G'
                            else:
                                typ = 'S'
                            pmss_details[key] = PM_Facility(key, gen, capacity, typ, -1, 1)
                    for fac in pmss_details.keys():
                        if fac == 'Load':
                            continue
                        gen = pmss_details[fac].generator
                        try:
                            pmss_details[fac].multiplier = capacities[fac] * 1.0 / pmss_details[fac].capacity
                        except:
                            pass
                        if option == T:
                            if fac not in capex_table.keys():
                                capex_table[fac] = {'cum': 0}
                            if year not in capex_table[fac].keys():
                                try:
                                    capex_table[fac][year] = [self.generators[fac].capex, 0]
                                except:
                                    capex_table[fac][year] = [self.generators[fac[fac.find('.') + 1:]].capex, 0]
                            capx = pmss_details[fac].multiplier * pmss_details[fac].capacity
                            capex_table[fac][year][1] = capx - capex_table[fac]['cum']
                            capex_table[fac]['cum'] = capx
                    if option == T:
                        for fac in capex_table.keys():
                            if capex_table[fac]['cum'] == 0:
                                continue
                            capx = 0
                            for key, detail in capex_table[fac].items():
                                if key == 'cum':
                                    continue
                                capx = capx + detail[0] * detail[1]
                            capx = capx / capex_table[fac]['cum']
                            try:
                                self.generators[fac].capex = round(capx)
                            except:
                                self.generators[fac[fac.find('.') + 1:]].capex = round(capx)
                    save_carbon_price = None
                    if 'Carbon Price' in capacities.keys():
                        save_carbon_price = self.carbon_price
                        self.carbon_price = capacities['Carbon Price']
                    if 'Discount Rate' in capacities.keys():
                        save_discount_rate = self.discount_rate
                        self.discount_rate = capacities['Discount Rate']
                    sp_data, corr_data = self.processor.doDispatch(year, option, sender_name, pmss_details, pmss_data, re_order, 
                        dispatch_order, pm_data_file, data_file)
                    if 'Carbon Price' in capacities.keys():
                        self.carbon_price = save_carbon_price
                    # first the Facility/technology table at the top of sp_data
                    for sp in range(len(self.batch_tech) + 1):
                        if sp_data[sp][st_fac] in self.batch_tech:
                            tndx = self.batch_tech.index(sp_data[sp][st_fac]) + 1
                            for group in self.batch_report:
                                if group[0] in batch_details.keys():
                                    gndx = group[1]
                                    col = batch_details[group[0]][0]
                                    if group[0] == 'Capacity Factor' and sp_data[sp][0] == 'Total':
                                        continue
                                    if group[0] == 'Capacity Factor' and isinstance(sp_data[sp][col], str):
                                        bs.cell(row=gndx + tndx, column=column).value = float(sp_data[sp][col].strip('%')) / 100.
                                    else:
                                        bs.cell(row=gndx + tndx, column=column).value = sp_data[sp][col]
                                    bs.cell(row=gndx + tndx, column=column).number_format = batch_details[group[0]][1]
                                    bs.cell(row=gndx + tndx, column=column).font = normal
                        if sp_data[sp][st_fac] == 'Total':
                            break
                    if batch_disc_row > 1:
                         bs.cell(row=batch_disc_row, column=column).value = self.discount_rate
                         bs.cell(row=batch_disc_row, column=column).number_format = '#0.00%'
                         bs.cell(row=batch_disc_row, column=column).font = normal
                    # save details from Total row
                    for group in self.batch_report:
                        if group[0] == 'LCOE ($/MWh)':
                            try:
                                col = batch_details['LCOE ($/MWh)'][0]
                                bs.cell(row=tot_lco_row, column=column).value = sp_data[sp][col]
                                bs.cell(row=tot_lco_row, column=column).number_format = batch_details['LCOE ($/MWh)'][1]
                                bs.cell(row=tot_lco_row, column=column).font = bold
                            except:
                                pass
                        elif group[0] == 'LCOE With CO2 ($/MWh)':
                            try:
                                col = batch_details['LCOE With CO2 ($/MWh)'][0]
                                bs.cell(row=tot_lcc_row, column=column).value = sp_data[sp][col]
                                bs.cell(row=tot_lcc_row, column=column).number_format = batch_details['LCOE With CO2 ($/MWh)'][1]
                                bs.cell(row=tot_lcc_row, column=column).font = bold
                            except:
                                pass
                        elif group[0] == 'Carbon':
                            try:
                                bs.cell(row=tot_carb_row, column=column).value = sp_data[sp][st_emc]
                                bs.cell(row=tot_carb_row, column=column).number_format = '#,##0'
                                bs.cell(row=tot_carb_row, column=column).font = normal
                                bs.cell(row=tot_carb_row + 1, column=column).value = sp_data[sp][st_lcc]
                                bs.cell(row=tot_carb_row + 1, column=column).number_format = '#,##0.00'
                                bs.cell(row=tot_carb_row + 1, column=column).font = bold
                                bs.cell(row=tot_carb_row + 2, column=column).value = sp_data[sp][st_lec]
                                bs.cell(row=tot_carb_row + 2, column=column).number_format = '#,##0'
                                bs.cell(row=tot_carb_row + 2, column=column).font = normal
                            except:
                                pass
                    if 'Discount Rate' in capacities.keys():
                        self.discount_rate = save_discount_rate
                    # now the other stuff in sp_data
                    for sp in range(sp + 1, len(sp_data)):
                        if sp_data[sp][st_fac] == '':
                            continue
                        i = sp_data[sp][st_fac].find(' (')
                        if i >= 0:
                            tgt = sp_data[sp][st_fac][: i]
                        else:
                            tgt = sp_data[sp][st_fac]
                        if tgt == 'RE %age':
                            for group in self.batch_report:
                                if group[0] == 'To Meet Load (MWh)':
                                    try:
                                        col = batch_details['To Meet Load (MWh)'][0]
                                        bs.cell(row=re_tml_row, column=column).value = sp_data[sp][col]
                                        bs.cell(row=re_tml_row, column=column).number_format = batch_details['To Meet Load (MWh)'][1]
                                        bs.cell(row=re_tml_row, column=column).font = normal
                                    except:
                                        pass
                                    break
                        elif tgt == 'Storage %age':
                            for group in self.batch_report:
                                if group[0] == 'To Meet Load (MWh)':
                                    try:
                                        col = batch_details['To Meet Load (MWh)'][0]
                                        bs.cell(row=re_tml_row + 1, column=column).value = sp_data[sp][col]
                                        bs.cell(row=re_tml_row + 1, column=column).number_format = batch_details['To Meet Load (MWh)'][1]
                                        bs.cell(row=re_tml_row + 1, column=column).font = normal
                                    except:
                                        pass
                                    break
                        elif tgt == 'LCOE':
                            for group in self.batch_report:
                                if group[0] == 'LCOE ($/MWh)':
                                    try:
                                        col = batch_details['LCOE ($/MWh)'][0]
                                        bs.cell(row=re_tml_row + 1, column=column).value = sp_data[sp][col]
                                        bs.cell(row=re_tml_row + 1, column=column).number_format = batch_details['LCOE ($/MWh)'][1]
                                        bs.cell(row=re_tml_row + 1, column=column).font = normal
                                    except:
                                        pass
                                    break
                        elif tgt == 'Carbon Price':
                            for group in batch_extra['Carbon'][1:]:
                                if group[0] == 'Carbon Price':
                                    try:
                                        col = group[1]
                                        bs.cell(row=tot_carb_row - 1, column=column).value = sp_data[sp][col]
                                        bs.cell(row=tot_carb_row - 1, column=column).number_format = batch_extra['Carbon'][0]
                                        bs.cell(row=tot_carb_row - 1, column=column).font = normal
                                    except:
                                        pass
                                    break
                        elif tgt[:10] == 'Total Load':
                            for group in self.batch_report:
                                if group[0] == 'Max MWh':
                                    try:
                                        col = batch_details['Max MWh'][0]
                                        bs.cell(row=max_load_row, column=column).value = sp_data[sp][col]
                                        bs.cell(row=max_load_row, column=column).number_format = batch_extra['Max MWh'][0]
                                        bs.cell(row=max_load_row, column=column).font = normal
                                    except:
                                        pass
                                    break
                        for key, details in batch_extra.items():
                            try:
                                x = [x for x in details if tgt in x][0]
                                for group in self.batch_report:
                                    if group[0] == key:
                                        gndx = group[1]
                                        break
                                else:
                                    continue
                                tndx = details.index(x)
                                col = x[1]
                                bs.cell(row=gndx + tndx, column=column).value = sp_data[sp][col]
                                if key == 'RE' or (key == 'Static Variables' and x[0] == 'Discount Rate'):
                                    pct = float(sp_data[sp][col].strip('%')) / 100.
                                    bs.cell(row=gndx + tndx, column=column).value = pct
                                    bs.cell(row=gndx + tndx, column=column).number_format = '0.0%'
                                else:
                                    bs.cell(row=gndx + tndx, column=column).value = sp_data[sp][col]
                                    bs.cell(row=gndx + tndx, column=column).number_format = details[0]
                                bs.cell(row=gndx + tndx, column=column).font = normal
                                if sp_data[sp][st_fac] == 'RE %age of Total Load' or \
                                  sp_data[sp][st_fac].find('LCOE') >= 0 or \
                                  sp_data[sp][st_fac].find('incl.') >= 0:
                                    bs.cell(row=gndx + tndx, column=column).font = bold
                                else:
                                    bs.cell(row=gndx + tndx, column=column).font = normal
                                if key == 'Load Analysis':
                                    if x[0] in ['Load met', 'Surplus']:
                                        tndx += 1
                                        col = batch_extra['Load Analysis'][tndx][1]
                                        pct = float(sp_data[sp][col].strip('%')) / 100.
                                        bs.cell(row=gndx + tndx, column=column).value = pct
                                        bs.cell(row=gndx + tndx, column=column).number_format = '0.0%'
                                        bs.cell(row=gndx + tndx, column=column).font = normal
                            except:
                                pass
                tim = (time.time() - sheet_start)
                if tim < 60:
                    tim = '%.1f secs' % tim
                else:
                    hhmm = tim / 60.
                    tim = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
                timt = (time.time() - start_time)
                if timt < 60:
                    timt = '%.1f secs' % timt
                else:
                    hhmm = timt / 60.
                    timt = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
                self.setStatus(f'Processed sheet {sht + 1} of {len(self.batch_models)}; ({len(self.batch_models[sht])} models; {tim}. Total {timt})')
                if total_load_row > 0:
                    if self.batch_prefix:
                        batch_pfx = get_batch_prefix('Load Analysis')
                    if option == T:
                        bs.cell(row=total_load_row, column=1).value = batch_pfx + 'Total Load'
                    else:
                        load_mult = ''
                        try:
                            mult = round(pmss_details['Load'].multiplier, 3)
                            if mult != 1:
                                load_mult = ' x ' + str(mult)
                        except:
                            pass
                        bs.cell(row=total_load_row, column=1).value = batch_pfx + 'Total Load - ' + year + load_mult
                if do_opt_parms[0]:
                    t_row = do_opt_parms[1]
                    for row in range(do_opt_parms[2], do_opt_parms[3] + 1):
                        for col in range(1, batch_input_sheet.max_column + 1):
                            cell = batch_input_sheet.cell(row=row, column=col)
                            new_cell = bs.cell(row=t_row, column=col, value=cell.value)
                            if cell.has_style:
                                new_cell.font = copy(cell.font)
                                new_cell.border = copy(cell.border)
                                new_cell.fill = copy(cell.fill)
                                new_cell.number_format = copy(cell.number_format)
                                new_cell.protection = copy(cell.protection)
                                new_cell.alignment = copy(cell.alignment)
                        t_row += 1
                del_rows = []
                for group in self.batch_report:
                    if group[0] in ['Generation (MWh)']:
                        # remove storage or RE
                        gndx = group[1]
                        if group[0] == 'Generation (MWh)':
                            tst = 'S'
                        else:
                            tst = 'R' # probably redundant
                        for row in range(gndx, gndx + len(self.batch_tech)):
                            try:
                                if pmss_details[bs.cell(row=row, column=1).value].tech_type == tst:
                                    del_rows.append(row)
                            except:
                                pass
                for row in sorted(del_rows, reverse=True):
                    bs.delete_rows(row, 1)
                for column_cells in bs.columns:
                    length = 0
                    for cell in column_cells:
                        if cell.row < self.batch_report[0][1] - 1:
                            continue
                        try:
                            value = str(round(cell.value, 2))
                        except:
                            value = cell.value
                        if value is None:
                            continue
                        if len(value) > length:
                            length = len(value)
                    if isinstance(cell.column, int):
                        cel = ssCol(cell.column)
                    else:
                        cel = cell.column
                    bs.column_dimensions[cel].width = max(length * 1.05, 10)
                if batch_data_sources_row > 0:
                    i = self.data_sources(bs, batch_data_sources_row - len(del_rows), pm_data_file, option)
                bs.freeze_panes = 'B' + str(self.batch_report[0][1])
                bs.activeCell = 'B' + str(self.batch_report[0][1])
                for sheet in wb:
                    wb[sheet.title].views.sheetView[0].tabSelected = False
                wb.active = bs
                # check if any charts/graphs
                if self.batch_report[-1][0] == 'Chart':
                    bold = oxl.styles.Font(name='Arial', bold=True)
                    min_col = 2
                    max_col = len(self.batch_models[sht]) + 1
                    chs = None
                    in_chart = False
                    cht_cells = ['N', 'B']
                    cht_row = -27
                    tndx_rows = max(9, len(self.batch_tech) + 4)
                    cats = None
                    chart_group = ''
                    for row in range(self.batch_report[-1][1], batch_input_sheet.max_row + 1):
                        if batch_input_sheet.cell(row=row, column=1).value is None:
                            continue
                        if batch_input_sheet.cell(row=row, column=1).value.lower() in ['chart', 'graph', 'plot']:
                            if in_chart:
                                charts[-1].width = 20
                                charts[-1].height = 12
                                for s in range(len(charts[-1].series)):
                                    ser = charts[-1].series[s]
                                    ser.marker.symbol = 'circle' #'dot', 'plus', 'triangle', 'x', 'picture', 'star', 'diamond', 'square', 'circle', 'dash', 'auto'
                              #      ser.graphicalProperties.line.solidFill = "00AAAA"
                                if charts2[-1] is not None:
                                    for s in range(len(charts2[-1].series)):
                                        ser = charts2[-1].series[s]
                                        ser.marker.symbol = 'triangle'
                               #         ser.graphicalProperties.line.solidFill = "00AAAA"
                                    charts2[-1].y_axis.crosses = 'max'
                                    charts[-1] += charts2[-1]
                                if cats is not None:
                                    charts[-1].set_categories(cats)
                                if len(charts) % 2:
                                    cht_row += 30
                                if chart_group != '':
                                    cht_col = col_letters.index(cht_cells[len(charts) % 2])
                                    chs.cell(row=cht_row - 1, column=cht_col).value = chart_group
                                    chs.cell(row=cht_row - 1, column=cht_col).font = bold
                                chs.add_chart(charts[-1], cht_cells[len(charts) % 2] + str(cht_row))
                            in_chart = True
                            if chs is None:
                                if bs.title.find('Results') >= 0:
                                    txt = bs.title.replace('Results', 'Charts')
                                else:
                                    txt = 'Charts_' + bs.title
                                chs = wb.create_sheet(txt)
                                charts = []
                                charts2 = []
                            charts.append(LineChart())
                            charts2.append(None)
                            if batch_input_sheet.cell(row=row, column=2).value is None or len(merged_cells) == 0:
                                min_col = 2
                                max_col = len(self.batch_models[sht]) + 1
                                chart_group = ''
                            else:
                                merge_group = get_value(batch_input_sheet, row, 2)
                                for i in range(len(merged_cells) -1, -1, -1):
                                    merge_value = get_value(batch_input_sheet, merged_cells[i][0], merged_cells[i][1])
                                    if merge_value == merge_group:
                                        min_col = merged_cells[i][1]
                                        max_col = merged_cells[i][2]
                                        chart_group = merge_group
                                        break
                        elif not in_chart:
                            continue
                        elif batch_input_sheet.cell(row=row, column=1).value.lower() == 'title':
                            charts[-1].title = batch_input_sheet.cell(row=row, column=2).value
                        elif batch_input_sheet.cell(row=row, column=1).value.lower() == 'x-title':
                            charts[-1].x_axis.title = get_value(batch_input_sheet, row, 2)
                        elif batch_input_sheet.cell(row=row, column=1).value.lower() == 'y-title':
                            charts[-1].y_axis.title = batch_input_sheet.cell(row=row, column=2).value
                        elif batch_input_sheet.cell(row=row, column=1).value.lower() == 'y-title2':
                            if charts2[-1] is None:
                                charts2[-1] = LineChart()
                                charts2[-1].x_axis.title = None
                            charts2[-1].y_axis.axId = 200
                            charts2[-1].y_axis.title = batch_input_sheet.cell(row=row, column=2).value
                        elif batch_input_sheet.cell(row=row, column=1).value.lower() in ['categories', 'y-labels', 'data', 'data2']:
                            dgrp = get_value(batch_input_sheet, row, 2)
                            if batch_input_sheet.cell(row=row, column=1).value.lower() == 'categories' \
                              and dgrp.lower() in ['model', 'model label', 'technology']: # models as categories
                                rw = self.batch_report[0][1] - 1
                                cats = Reference(bs, min_col=min_col, min_row=rw, max_col=max_col, max_row=rw)
                                continue
                            if dgrp.lower() in ['capacity (mw)', 'capacity (mw/mwh)']:
                                gndx = self.batch_report[0][1]
                            else:
                                for group in self.batch_report:
                                    if group[0].lower() == dgrp.lower():
                                        gndx = group[1]
                                        break
                                else:
                                     continue
                                # backup a bit in case rows deleted
                                for r in range(len(del_rows)):
                                    try:
                                        if bs.cell(row=gndx, column=1).value.lower() == group[0].lower():
                                            break
                                    except:
                                        pass
                                    gndx -= 1
                            ditm = get_value(batch_input_sheet, row, 3)
                            for tndx in range(tndx_rows):
                                if bs.cell(row=gndx + tndx, column=1).value is None:
                                    break
                                if bs.cell(row=gndx + tndx, column=1).value.lower() == ditm.lower():
                                    if batch_input_sheet.cell(row=row, column=1).value.lower() == 'data':
                                        values = Reference(bs, min_col=min_col, min_row=gndx + tndx, max_col=max_col, max_row=gndx + tndx)
                                        series = Series(values)
                                        series.title = oxl.chart.series.SeriesLabel(oxl.chart.data_source.StrRef("'" + bs.title + "'!A" + str(gndx + tndx)))
                                        charts[-1].append(series)
                                    elif batch_input_sheet.cell(row=row, column=1).value.lower() == 'data2':
                                        if charts2[-1] is None:
                                            charts2[-1] = LineChart()
                                        values = Reference(bs, min_col=min_col, min_row=gndx + tndx, max_col=max_col, max_row=gndx + tndx)
                                        series = Series(values)
                                        series.title = oxl.chart.series.SeriesLabel(oxl.chart.data_source.StrRef("'" + bs.title + "'!A" + str(gndx + tndx)))
                                        charts2[-1].append(series)
                                    else:
                                        cats = Reference(bs, min_col=min_col, min_row=gndx + tndx, max_col=max_col, max_row=gndx + tndx)
                                    break
                    if in_chart:
                        charts[-1].width = 20
                        charts[-1].height = 12
                        for s in range(len(charts[-1].series)):
                            ser = charts[-1].series[s]
                            ser.marker.symbol = 'circle' #'dot', 'plus', 'triangle', 'x', 'picture', 'star', 'diamond', 'square', 'circle', 'dash', 'auto'
                        if charts2[-1] is not None:
                            for s in range(len(charts2[-1].series)):
                                ser = charts2[-1].series[s]
                                ser.marker.symbol = 'triangle'
                            charts2[-1].y_axis.crosses = 'max'
                            charts[-1] += charts2[-1]
                        if cats is not None:
                            charts[-1].set_categories(cats)
                        if len(charts) % 2:
                            cht_row += 30
                        if chart_group != '':
                            cht_col = col_letters.index(cht_cells[len(charts) % 2])
                            chs.cell(row=cht_row - 1, column=cht_col).value = chart_group
                            chs.cell(row=cht_row - 1, column=cht_col).font = bold
                        chs.add_chart(charts[-1], cht_cells[len(charts) % 2] + str(cht_row))
            if len(self.batch_models) > 1 and len(self.batch_models[0]) == 1:
                try:
                    del wb['Results_' + rpt_time]
                    del wb['Charts_' + rpt_time]
                except:
                    pass
            tim = (time.time() - start_time)
            if tim < 60:
                tim = '%.1f secs' % tim
            else:
                hhmm = tim / 60.
                tim = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
            self.setStatus(f'Saving {self.sender} report ({total_models:,} models; {tim})')
           #     success_message = 'Saving %s report' % (self.sender))
            wb.save(batch_report_file)
            tim = (time.time() - start_time)
            if tim < 60:
                tim = '%.1f secs' % tim
            else:
                hhmm = tim / 60.
                tim = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
            self.setStatus(f'{self.sender} completed ({len(self.batch_models)}, {total_models:,} models; {tim}). You may need to open and save the workbook to reprocess it.')
            return
        if do_adjust:
            if self.adjustto is not None:
                for fac, value in self.adjustto.items():
                    try:
                        pmss_details[fac].multiplier = value / pmss_details[fac].capacity
                    except:
                        pass
        if option == D:
            self.processor.doDispatch(year, option, sender_name, pmss_details, pmss_data, re_order, dispatch_order,
                pm_data_file, data_file)
        else:
            sp_data, corr_data, headers, sp_pts = self.processor.doDispatch(year, option, sender_name, pmss_details, pmss_data, re_order, dispatch_order,
                pm_data_file, data_file)
        title = None
        if option == B or option == T:
            return sp_data, None, None
        span = None
        if self.summary_sources and option != D: # want data sources
            sp_data.append(' ')
            sp_data.append('Data sources')
            span = 'Data sources'
            sp_data.append(['Scenarios folder', self.scenarios])
            if pm_data_file[: len(self.scenarios)] == self.scenarios:
                pm_data_file = pm_data_file[len(self.scenarios):]
            sp_data.append(['Powermatch data file', pm_data_file])
            sp_data.append(['Load file', self.load_file])
            sp_data.append(['Constraints worksheet', str(self.constraints_file)])
            sp_data.append(['Generators worksheet', str(self.generators_file)])
            sp_pts = [0] * len(headers)
            for p in [st_cap, st_lcg, st_lco, st_lcc, st_max, st_bal, st_rlc, st_are]:
                sp_pts[p] = 2
            if corr_data is not None:
                sp_pts[st_cap] = 3 # compromise between capacity (2) and correlation (4)
            self.setStatus(self.sender + ' completed')
            if title is not None:
                atitle = title
            elif self.results_prefix != '':
                atitle = self.results_prefix + '_' + self.sender
            else:
                atitle = self.sender
            return sp_data, headers, sp_pts # finish if not detailed spreadsheet
        return None, None, None
# Detailed Processing
    def setStatus(self, text):
        self.success_message = text
        return

    def optClicked(self, in_year, in_option, in_pmss_details, in_pmss_data, in_re_order,
                   in_dispatch_order, pm_data_file, data_file):
        sp_data = self.processor.run_optimise(self, in_year, in_option, in_pmss_details, in_pmss_data, in_re_order,
                   in_dispatch_order, pm_data_file, data_file)
        span = None
        dialog = displaytable.Table(sp_data, title=atitle, fields=headers,
            save_folder=self.scenarios, sortby='', decpts=sp_pts,
            span=span)
        dialog.exec_()
