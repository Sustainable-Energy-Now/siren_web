"""
Management command to export Facilities and all related models to an Excel workbook.

Sheets produced:
  Facilities             - core facility data
  FacilitySolar          - solar installations per facility
  FacilityStorage        - storage installations per facility
  FacilityWindTurbines   - wind turbine installations per facility
  WindTurbines           - wind turbine model catalogue
  FacilityGridConnections- grid connection details per facility
  ScenariosFacilities    - which scenarios each facility belongs to
  Zones                  - zone reference data
  Technologies           - technology reference data

Usage:
  python manage.py export_facilities_excel
  python manage.py export_facilities_excel --output /path/to/file.xlsx
"""

import os
from django.core.management.base import BaseCommand
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class Command(BaseCommand):
    help = 'Export Facilities and related models to an Excel workbook'

    def add_arguments(self, parser):
        parser.add_argument(
            '--output', '-o',
            type=str,
            default='facilities_export.xlsx',
            help='Output Excel file path (default: facilities_export.xlsx)'
        )

    def handle(self, *args, **options):
        from siren_web.models import (
            facilities, FacilitySolar, FacilityStorage,
            FacilityWindTurbines, WindTurbines, FacilityGridConnections,
            ScenariosFacilities, Zones, Technologies,
        )

        output_path = options['output']
        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        def write_sheet(ws, headers, rows):
            ws.append(headers)
            header_row = ws[1]
            for cell in header_row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            for row in rows:
                ws.append(row)
            # Auto-size columns (capped at 60)
            for col_idx, _ in enumerate(headers, 1):
                col_letter = get_column_letter(col_idx)
                max_len = len(str(headers[col_idx - 1]))
                for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value is not None:
                            max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

        # ------------------------------------------------------------------
        # 1. Facilities
        # ------------------------------------------------------------------
        ws = wb.create_sheet('Facilities')
        headers = [
            'ID', 'Facility Name', 'Facility Code', 'Participant Code',
            'Registered From', 'Active', 'Existing', 'Status',
            'Commissioning Date', 'Decommissioning Date', 'Commissioning Probability',
            'Zone', 'Capacity (MW)', 'Capacity Factor', 'Latitude', 'Longitude',
            'Emission Intensity', 'Power File', 'Direction',
            'Legacy Technology',
        ]
        rows = []
        for f in facilities.objects.select_related('idzones', 'idtechnologies').order_by('facility_name'):
            rows.append([
                f.idfacilities,
                f.facility_name,
                f.facility_code,
                f.participant_code,
                f.registered_from,
                f.active,
                f.existing,
                f.status,
                f.commissioning_date,
                f.decommissioning_date,
                f.commissioning_probability,
                f.idzones.name if f.idzones else None,
                f.capacity,
                f.capacityfactor,
                f.latitude,
                f.longitude,
                f.emission_intensity,
                f.power_file,
                f.direction,
                f.idtechnologies.technology_name if f.idtechnologies else None,
            ])
        write_sheet(ws, headers, rows)
        self.stdout.write(f'  Facilities: {len(rows)} rows')

        # ------------------------------------------------------------------
        # 2. FacilitySolar
        # ------------------------------------------------------------------
        ws = wb.create_sheet('FacilitySolar')
        headers = [
            'ID', 'Facility', 'Facility Code', 'Technology',
            'Installation Name', 'Nameplate Capacity (MW DC)', 'AC Capacity (MW)',
            'Panel Count', 'Panel Wattage (W)', 'Tilt Angle (°)', 'Azimuth Angle (°)',
            'Array Area (m²)', 'Inverter Count', 'Inverter Capacity Each (kW)',
            'Installation Date', 'Commissioning Date', 'Active', 'Notes',
        ]
        rows = []
        for s in FacilitySolar.objects.select_related('idfacilities', 'idtechnologies').order_by('idfacilities__facility_name', 'installation_name'):
            rows.append([
                s.idfacilitysolar,
                s.idfacilities.facility_name,
                s.idfacilities.facility_code,
                s.idtechnologies.technology_name,
                s.installation_name,
                s.nameplate_capacity,
                s.ac_capacity,
                s.panel_count,
                s.panel_wattage,
                s.tilt_angle,
                s.azimuth_angle,
                s.array_area,
                s.inverter_count,
                s.inverter_capacity_each,
                s.installation_date,
                s.commissioning_date,
                s.is_active,
                s.notes,
            ])
        write_sheet(ws, headers, rows)
        self.stdout.write(f'  FacilitySolar: {len(rows)} rows')

        # ------------------------------------------------------------------
        # 3. FacilityStorage
        # ------------------------------------------------------------------
        ws = wb.create_sheet('FacilityStorage')
        headers = [
            'ID', 'Facility', 'Facility Code', 'Technology',
            'Installation Name', 'Power Capacity (MW)', 'Energy Capacity (MWh)',
            'Duration (h)', 'Initial SOC',
            'Installation Date', 'Commissioning Date', 'Active', 'Notes',
        ]
        rows = []
        for s in FacilityStorage.objects.select_related('idfacilities', 'idtechnologies').order_by('idfacilities__facility_name', 'installation_name'):
            rows.append([
                s.idfacilitystorage,
                s.idfacilities.facility_name,
                s.idfacilities.facility_code,
                s.idtechnologies.technology_name,
                s.installation_name,
                s.power_capacity,
                s.energy_capacity,
                s.duration,
                s.initial_state_of_charge,
                s.installation_date,
                s.commissioning_date,
                s.is_active,
                s.notes,
            ])
        write_sheet(ws, headers, rows)
        self.stdout.write(f'  FacilityStorage: {len(rows)} rows')

        # ------------------------------------------------------------------
        # 4. FacilityWindTurbines
        # ------------------------------------------------------------------
        ws = wb.create_sheet('FacilityWindTurbines')
        headers = [
            'ID', 'Facility', 'Facility Code', 'Technology',
            'Turbine Model', 'Manufacturer',
            'Installation Name', 'No. Turbines', 'Nameplate Capacity (MW)',
            'Tilt (°)', 'Direction', 'Hub Height Override (m)',
            'Installation Date', 'Commissioning Date', 'Active', 'Notes',
        ]
        rows = []
        for w in FacilityWindTurbines.objects.select_related(
            'idfacilities', 'idwindturbines', 'idtechnologies'
        ).order_by('idfacilities__facility_name', 'installation_name'):
            rows.append([
                w.idfacilitywindturbines,
                w.idfacilities.facility_name,
                w.idfacilities.facility_code,
                w.idtechnologies.technology_name if w.idtechnologies else None,
                w.idwindturbines.turbine_model,
                w.idwindturbines.manufacturer,
                w.installation_name,
                w.no_turbines,
                w.nameplate_capacity,
                w.tilt,
                w.direction,
                w.hub_height_override,
                w.installation_date,
                w.commissioning_date,
                w.is_active,
                w.notes,
            ])
        write_sheet(ws, headers, rows)
        self.stdout.write(f'  FacilityWindTurbines: {len(rows)} rows')

        # ------------------------------------------------------------------
        # 5. WindTurbines (catalogue)
        # ------------------------------------------------------------------
        ws = wb.create_sheet('WindTurbines')
        headers = [
            'ID', 'Turbine Model', 'Manufacturer', 'Application',
            'Hub Height (m)', 'Rated Power (kW)', 'Rotor Diameter (m)',
            'Cut-In Speed (m/s)', 'Cut-Out Speed (m/s)',
        ]
        rows = []
        for t in WindTurbines.objects.order_by('manufacturer', 'turbine_model'):
            rows.append([
                t.idwindturbines,
                t.turbine_model,
                t.manufacturer,
                t.application,
                t.hub_height,
                t.rated_power,
                t.rotor_diameter,
                t.cut_in_speed,
                t.cut_out_speed,
            ])
        write_sheet(ws, headers, rows)
        self.stdout.write(f'  WindTurbines: {len(rows)} rows')

        # ------------------------------------------------------------------
        # 6. FacilityGridConnections
        # ------------------------------------------------------------------
        ws = wb.create_sheet('FacilityGridConnections')
        headers = [
            'ID', 'Facility', 'Facility Code', 'Grid Line',
            'Connection Type', 'Is Primary', 'Active',
            'Connection Voltage (kV)', 'Connection Capacity (MW)',
            'Transformer Capacity (MVA)', 'Connection Distance (km)',
            'Connection Lat', 'Connection Lon', 'Connection Date',
        ]
        rows = []
        for g in FacilityGridConnections.objects.select_related('idfacilities', 'idgridlines').order_by('idfacilities__facility_name'):
            rows.append([
                g.idfacilitygridconnections,
                g.idfacilities.facility_name,
                g.idfacilities.facility_code,
                g.idgridlines.line_name if hasattr(g.idgridlines, 'line_name') else str(g.idgridlines),
                g.connection_type,
                g.is_primary,
                g.active,
                g.connection_voltage_kv,
                g.connection_capacity_mw,
                g.transformer_capacity_mva,
                g.connection_distance_km,
                g.connection_point_latitude,
                g.connection_point_longitude,
                g.connection_date,
            ])
        write_sheet(ws, headers, rows)
        self.stdout.write(f'  FacilityGridConnections: {len(rows)} rows')

        # ------------------------------------------------------------------
        # 7. ScenariosFacilities
        # ------------------------------------------------------------------
        ws = wb.create_sheet('ScenariosFacilities')
        headers = ['ID', 'Scenario ID', 'Scenario Name', 'Facility ID', 'Facility', 'Facility Code']
        rows = []
        for sf in ScenariosFacilities.objects.select_related('idscenarios', 'idfacilities').order_by('idscenarios', 'idfacilities__facility_name'):
            rows.append([
                sf.idscenariosfacilities,
                sf.idscenarios_id,
                sf.idscenarios.title if hasattr(sf.idscenarios, 'title') else str(sf.idscenarios),
                sf.idfacilities_id,
                sf.idfacilities.facility_name,
                sf.idfacilities.facility_code,
            ])
        write_sheet(ws, headers, rows)
        self.stdout.write(f'  ScenariosFacilities: {len(rows)} rows')

        # ------------------------------------------------------------------
        # 8. Zones
        # ------------------------------------------------------------------
        ws = wb.create_sheet('Zones')
        headers = ['ID', 'Name', 'Description']
        rows = []
        for z in Zones.objects.order_by('name'):
            rows.append([z.idzones, z.name, z.description])
        write_sheet(ws, headers, rows)
        self.stdout.write(f'  Zones: {len(rows)} rows')

        # ------------------------------------------------------------------
        # 9. Technologies (relevant to facilities)
        # ------------------------------------------------------------------
        ws = wb.create_sheet('Technologies')
        headers = [
            'ID', 'Technology Name', 'Signature', 'Category',
            'Renewable', 'Dispatchable', 'Fuel Type',
            'Lifetime (yr)', 'Discount Rate', 'Emissions',
            'Area', 'Water Usage', 'Caption', 'Description',
        ]
        rows = []
        for t in Technologies.objects.order_by('category', 'technology_name'):
            rows.append([
                t.idtechnologies,
                t.technology_name,
                t.technology_signature,
                t.category,
                t.renewable,
                t.dispatchable,
                t.fuel_type,
                t.lifetime,
                t.discount_rate,
                t.emissions,
                t.area,
                t.water_usage,
                t.caption,
                t.description,
            ])
        write_sheet(ws, headers, rows)
        self.stdout.write(f'  Technologies: {len(rows)} rows')

        # ------------------------------------------------------------------
        # Save
        # ------------------------------------------------------------------
        wb.save(output_path)
        abs_path = os.path.abspath(output_path)
        self.stdout.write(self.style.SUCCESS(f'\nSaved: {abs_path}'))
