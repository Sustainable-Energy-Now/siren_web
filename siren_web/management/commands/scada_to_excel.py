# management/commands/scada_to_excel.py
import json
import os
from datetime import datetime
from django.core.management.base import BaseCommand, CommandError
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class Command(BaseCommand):
    help = 'Convert AEMO facility SCADA JSON data to Excel format'

    def add_arguments(self, parser):
        parser.add_argument(
            'json_file',
            type=str,
            help='Path to the AEMO SCADA JSON file'
        )
        parser.add_argument(
            '--output',
            '-o',
            type=str,
            help='Output Excel file path (default: scada_data_YYYYMMDD.xlsx)'
        )
        parser.add_argument(
            '--pivot',
            action='store_true',
            help='Create pivot table with facilities as columns'
        )
        parser.add_argument(
            '--separate-sheets',
            action='store_true',
            help='Create separate sheet for each facility'
        )

    def handle(self, *args, **options):
        json_file = options['json_file']
        
        if not os.path.exists(json_file):
            raise CommandError(f'JSON file not found: {json_file}')
        
        # Determine output file
        if options['output']:
            output_file = options['output']
        else:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = f'scada_data_{timestamp}.xlsx'
        
        try:
            self.stdout.write(f'Reading SCADA data from: {json_file}')
            with open(json_file, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            
            # Extract the data
            if 'data' in json_data:
                data = json_data['data']
            else:
                data = json_data
            
            if 'facilityScadaDispatchIntervals' not in data:
                raise CommandError('JSON must contain data.facilityScadaDispatchIntervals')
            
            intervals = data['facilityScadaDispatchIntervals']
            self.stdout.write(f'Found {len(intervals)} SCADA intervals')
            
            if options['separate_sheets']:
                self.create_separate_sheets(intervals, output_file)
            elif options['pivot']:
                self.create_pivot_format(intervals, output_file)
            else:
                self.create_standard_format(intervals, output_file)
            
            self.stdout.write(
                self.style.SUCCESS(f'Successfully created: {output_file}')
            )
            
        except json.JSONDecodeError as e:
            raise CommandError(f'Invalid JSON: {e}')
        except Exception as e:
            raise CommandError(f'Conversion error: {e}')
    
    def create_standard_format(self, intervals, output_file):
        """Create standard long-format Excel with all data in one sheet"""
        wb = Workbook()
        ws = wb.active
        ws.title = "SCADA Data"
        
        # Headers
        headers = ['Dispatch Interval', 'Facility Code', 'Quantity (MW)']
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row, interval in enumerate(intervals, start=2):
            ws.cell(row=row, column=1, value=interval['dispatchInterval'])
            ws.cell(row=row, column=2, value=interval['code'])
            ws.cell(row=row, column=3, value=interval['quantity'])
        
        # Format columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        wb.save(output_file)
        self.stdout.write(f'Created standard format with {len(intervals)} records')
    
    def create_pivot_format(self, intervals, output_file):
        """Create pivot table with timestamps as rows, facilities as columns"""
        wb = Workbook()
        ws = wb.active
        ws.title = "SCADA Pivot"
        
        # Get unique timestamps and facility codes
        timestamps = sorted(set(item['dispatchInterval'] for item in intervals))
        facilities = sorted(set(item['code'] for item in intervals))
        
        self.stdout.write(f'Pivot: {len(timestamps)} timestamps × {len(facilities)} facilities')
        
        # Headers
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell = ws.cell(row=1, column=1, value='Dispatch Interval')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
        for col, facility in enumerate(facilities, start=2):
            cell = ws.cell(row=1, column=col, value=facility)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        # Create lookup dict for faster access
        data_lookup = {}
        for item in intervals:
            key = (item['dispatchInterval'], item['code'])
            data_lookup[key] = item['quantity']
        
        # Fill data
        for row_idx, timestamp in enumerate(timestamps, start=2):
            ws.cell(row=row_idx, column=1, value=timestamp)
            
            for col_idx, facility in enumerate(facilities, start=2):
                value = data_lookup.get((timestamp, facility), None)
                if value is not None:
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Format
        ws.column_dimensions['A'].width = 25
        for col in range(2, len(facilities) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        ws.freeze_panes = 'B2'
        
        wb.save(output_file)
        self.stdout.write(f'Created pivot format: {len(timestamps)} rows × {len(facilities)} columns')
    
    def create_separate_sheets(self, intervals, output_file):
        """Create separate sheet for each facility"""
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Group by facility
        facilities_data = {}
        for item in intervals:
            code = item['code']
            if code not in facilities_data:
                facilities_data[code] = []
            facilities_data[code].append(item)
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        for facility_code, facility_intervals in sorted(facilities_data.items()):
            # Create sheet (truncate name if needed for Excel limit)
            sheet_name = facility_code[:31]
            ws = wb.create_sheet(title=sheet_name)
            
            # Headers
            headers = ['Dispatch Interval', 'Quantity (MW)']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Data - sort by timestamp
            sorted_intervals = sorted(facility_intervals, key=lambda x: x['dispatchInterval'])
            for row, item in enumerate(sorted_intervals, start=2):
                ws.cell(row=row, column=1, value=item['dispatchInterval'])
                ws.cell(row=row, column=2, value=item['quantity'])
            
            # Format
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 15
            ws.freeze_panes = 'A2'
        
        wb.save(output_file)
        self.stdout.write(f'Created {len(facilities_data)} sheets for facilities')