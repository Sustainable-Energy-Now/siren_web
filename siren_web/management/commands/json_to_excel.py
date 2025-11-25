# management/commands/json_to_excel.py
import json
import os
from django.core.management.base import BaseCommand, CommandError
from openpyxl import Workbook
from openpyxl.styles import Font


class Command(BaseCommand):
    help = 'Convert a JSON file to Excel format'

    def add_arguments(self, parser):
        parser.add_argument(
            'json_file',
            type=str,
            help='Path to the JSON file to convert'
        )
        parser.add_argument(
            '--output',
            '-o',
            type=str,
            help='Output Excel file path (default: same name as JSON with .xlsx extension)'
        )
        parser.add_argument(
            '--sheet-name',
            type=str,
            default='Data',
            help='Name for the Excel sheet (default: Data)'
        )

    def handle(self, *args, **options):
        json_file = options['json_file']
        
        # Validate input file
        if not os.path.exists(json_file):
            raise CommandError(f'JSON file not found: {json_file}')
        
        # Determine output file
        if options['output']:
            output_file = options['output']
        else:
            output_file = os.path.splitext(json_file)[0] + '.xlsx'
        
        try:
            # Read JSON file
            self.stdout.write(f'Reading JSON file: {json_file}')
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Convert to Excel
            self.stdout.write('Converting to Excel...')
            self.json_to_excel(data, output_file, options['sheet_name'])
            
            self.stdout.write(
                self.style.SUCCESS(f'Successfully created Excel file: {output_file}')
            )
            
        except json.JSONDecodeError as e:
            raise CommandError(f'Invalid JSON file: {e}')
        except Exception as e:
            raise CommandError(f'Error during conversion: {e}')
    
    def json_to_excel(self, data, output_file, sheet_name):
        """Convert JSON data to Excel file"""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Handle different JSON structures
        if isinstance(data, list):
            if not data:
                raise CommandError('JSON file contains an empty list')
            
            # Get headers from first item
            if isinstance(data[0], dict):
                headers = list(data[0].keys())
            else:
                # Simple list of values
                headers = ['Value']
                data = [{'Value': item} for item in data]
            
            # Write headers
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            
            # Write data rows
            for row_idx, item in enumerate(data, start=2):
                if isinstance(item, dict):
                    for col_idx, header in enumerate(headers, start=1):
                        value = item.get(header, '')
                        # Handle nested structures
                        if isinstance(value, (dict, list)):
                            value = json.dumps(value)
                        ws.cell(row=row_idx, column=col_idx, value=value)
                else:
                    ws.cell(row=row_idx, column=1, value=item)
        
        elif isinstance(data, dict):
            # Handle dict as key-value pairs
            ws.cell(row=1, column=1, value='Key').font = Font(bold=True)
            ws.cell(row=1, column=2, value='Value').font = Font(bold=True)
            
            for row_idx, (key, value) in enumerate(data.items(), start=2):
                ws.cell(row=row_idx, column=1, value=key)
                if isinstance(value, (dict, list)):
                    value = json.dumps(value)
                ws.cell(row=row_idx, column=2, value=value)
        
        else:
            raise CommandError('JSON must contain a list or dictionary')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(output_file)