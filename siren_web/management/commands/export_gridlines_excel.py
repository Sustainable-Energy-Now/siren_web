"""
Management command to export GridLines and Terminals models to an Excel workbook.

Sheets produced:
  GridLines   - grid line data including physical characteristics and capacity
  Terminals   - terminal/substation data

Usage:
  python manage.py export_gridlines_excel
  python manage.py export_gridlines_excel --output /path/to/file.xlsx
"""

import os
from django.core.management.base import BaseCommand
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class Command(BaseCommand):
    help = 'Export GridLines and Terminals models to an Excel workbook'

    def add_arguments(self, parser):
        parser.add_argument(
            '--output', '-o',
            type=str,
            default='gridlines_export.xlsx',
            help='Output Excel file path (default: gridlines_export.xlsx)'
        )

    def handle(self, *args, **options):
        from siren_web.models import GridLines, Terminals

        output_path = options['output']
        wb = Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        def write_sheet(ws, headers, rows):
            ws.append(headers)
            header_row = ws[1]
            for cell in header_row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            for row in rows:
                ws.append(row)
            for col_idx, _ in enumerate(headers, 1):
                col_letter = get_column_letter(col_idx)
                max_len = len(str(headers[col_idx - 1]))
                for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value is not None:
                            max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

        # ------------------------------------------------------------------
        # 1. GridLines
        # ------------------------------------------------------------------
        ws = wb.create_sheet('GridLines')
        headers = [
            'ID', 'Line Name', 'Line Code', 'Line Type', 'Voltage Level (kV)',
            'Length (km)', 'Resistance per km (Ω/km)', 'Reactance per km (Ω/km)',
            'Conductance per km (S/km)', 'Susceptance per km (S/km)',
            'Thermal Capacity (MW)', 'Emergency Capacity (MW)',
            'Thermal Capacity (MVA)', 'Emergency Capacity (MVA)',
            'From Latitude', 'From Longitude', 'To Latitude', 'To Longitude',
            'From Terminal', 'To Terminal',
            'Active', 'Status', 'Commissioning Date', 'Decommissioning Date',
            'Commissioning Probability', 'Owner', 'KML Geometry',
        ]
        rows = []
        for g in GridLines.objects.select_related('from_terminal', 'to_terminal').order_by('line_name'):
            rows.append([
                g.idgridlines,
                g.line_name,
                g.line_code,
                g.line_type,
                g.voltage_level,
                g.length_km,
                g.resistance_per_km,
                g.reactance_per_km,
                g.conductance_per_km,
                g.susceptance_per_km,
                g.thermal_capacity_mw,
                g.emergency_capacity_mw,
                g.thermal_capacity_mva,
                g.emergency_capacity_mva,
                g.from_latitude,
                g.from_longitude,
                g.to_latitude,
                g.to_longitude,
                g.from_terminal.terminal_name if g.from_terminal else None,
                g.to_terminal.terminal_name if g.to_terminal else None,
                g.active,
                g.status,
                g.commissioning_date,
                g.decommissioning_date,
                g.commissioning_probability,
                g.owner,
                g.kml_geometry,
            ])
        write_sheet(ws, headers, rows)
        self.stdout.write(f'  GridLines: {len(rows)} rows')

        # ------------------------------------------------------------------
        # 2. Terminals
        # ------------------------------------------------------------------
        ws = wb.create_sheet('Terminals')
        headers = [
            'ID', 'Terminal Name', 'Terminal Code', 'Terminal Type',
            'Latitude', 'Longitude', 'Elevation (m)',
            'Primary Voltage (kV)', 'Secondary Voltage (kV)', 'Voltage Class',
            'Transformer Capacity (MVA)', 'Short Circuit Capacity (MVA)', 'Bay Count',
            'Active', 'Status', 'Commissioning Date', 'Decommissioning Date',
            'Commissioning Probability', 'Owner', 'SCADA ID', 'Description',
        ]
        rows = []
        for t in Terminals.objects.order_by('terminal_name'):
            rows.append([
                t.idterminals,
                t.terminal_name,
                t.terminal_code,
                t.terminal_type,
                t.latitude,
                t.longitude,
                t.elevation,
                t.primary_voltage_kv,
                t.secondary_voltage_kv,
                t.voltage_class,
                t.transformer_capacity_mva,
                t.short_circuit_capacity_mva,
                t.bay_count,
                t.active,
                t.status,
                t.commissioning_date,
                t.decommissioning_date,
                t.commissioning_probability,
                t.owner,
                t.scada_id,
                t.description,
            ])
        write_sheet(ws, headers, rows)
        self.stdout.write(f'  Terminals: {len(rows)} rows')

        # ------------------------------------------------------------------
        # Save
        # ------------------------------------------------------------------
        wb.save(output_path)
        abs_path = os.path.abspath(output_path)
        self.stdout.write(self.style.SUCCESS(f'\nSaved: {abs_path}'))
