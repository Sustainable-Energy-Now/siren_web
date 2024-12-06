#!/usr/bin/python3
#
#  Copyright (C) 2015-2024 Sustainable Energy Now Inc., Angus King
#
#  makegrid.py - This file is part of SIREN.
#
#  SIREN is free software: you can redistribute it and/or modify
#  it under the terms of the GNU Affero General Public License as
#  published by the Free Software Foundation, either version 3 of
#  the License, or (at your option) any later version.
#
#  SIREN is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU Affero General Public License for more details.
#
#  You should have received a copy of the GNU Affero General
#  Public License along with SIREN.  If not, see
#  <http://www.gnu.org/licenses/>.
#

import openpyxl as oxl
import os
import sys
import time
import xlwt

import displayobject
from credits import fileVersion
from getmodels import getModelFile
from senutils import getParents, getUser, ssCol


class makeFile():

    def close(self):
        return

    def getLog(self):
        return self.log, self.property

    def show_progress(self, msg):
        if self.his_log is None:
            return
        tim = time.time() - self.started
        if tim < 60:
            tim = '%.1f secs' % tim
        else:
            hhmm = tim / 60.
            tim = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
        self.his_log.setText(f'{msg} ({tim})')
        QtWidgets.QApplication.processEvents()

    def __init__(self, src_year, src_dir, wnd_dir, tgt_fil, detail='Daily By Month', rain='', nonzero='False', log=None):
        seasons = []
        periods = []
        try:
            items = config.items('Power')
            for item, values in items:
                if item[:6] == 'season':
                    if item == 'season':
                        continue
                    i = int(item[6:]) - 1
                    if i >= len(seasons):
                        seasons.append([])
                    seasons[i] = values.split(',')
                    for j in range(1, len(seasons[i])):
                        seasons[i][j] = int(seasons[i][j]) - 1
                elif item[:6] == 'period':
                    if item == 'period':
                        continue
                    i = int(item[6:]) - 1
                    if i >= len(periods):
                        periods.append([])
                    periods[i] = values.split(',')
                    for j in range(1, len(periods[i])):
                        periods[i][j] = int(periods[i][j]) - 1
        except:
            pass
        if len(seasons) == 0:
            seasons = [['Summer', 11, 0, 1], ['Autumn', 2, 3, 4], ['Winter', 5, 6, 7], ['Spring', 8, 9, 10]]
        if len(periods) == 0:
            periods = [['Winter', 4, 5, 6, 7, 8, 9], ['Summer', 10, 11, 0, 1, 2, 3]]
        for i in range(len(periods)):
            for j in range(len(seasons)):
                if periods[i][0] == seasons[j][0]:
                    periods[i][0] += '2'
                    break
        self.his_log = log
        self.started = time.time()
        self.show_progress('Collecting solar data')
        self.log = ''
        self.property = ''
        self.src_year = src_year
        self.src_dir = src_dir
        self.wind_dir = wnd_dir
        self.rain_dir = rain
        if rain != '':
            self.do_rain = True
        else:
            self.do_rain = False
        self.nonzero = False
        if nonzero != '':
            if nonzero[0].upper() in ('T', 'Y'):
                self.nonzero = True
        self.tgt_fil = tgt_fil
        bits = detail.split()
        self.hourly = False
        self.daily = False
        if len(bits) < 3:
            if detail[0].lower() == 'h':
                self.hourly = True
                if detail[1].lower() == 'd':
                    self.daily = True
        else:
            if bits[0][0].lower() == 'h':
                self.hourly = True
                if bits[2][0].lower() == 'd':
                    self.daily = True
        the_cols = ['DHI (Diffuse)', 'DNI (Normal)', 'GHI (Direct)', 'Temperature', 'Wind Speed',
                    'Wind @ 50m', 'Rainfall']
        col = []
        for i in range(len(the_cols)):
            col.append(-1)
        col_min = [[], []]
        col_max = [[], []]
        for i in range(len(the_cols)):
            col_min[0].append(None)
            col_max[0].append(0.)
            col_min[1].append(None)
            col_max[1].append(0.)
        the_hrs = [1, 1, 1, 24, 24, 24, 1]
        the_days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        the_hour = [0]
        drop_rainfall = True
        all_values = {}
        if self.hourly:
            hrly_values = {}
        if self.daily:
            daily_values = {}
        for i in range(len(the_days)):
            the_hour.append(the_hour[-1] + the_days[i] * 24)
        fils = os.listdir(self.src_dir)
        for fil in fils:
            if fil[-4:] == '.csv' or fil[-4:] == '.smw':
                if fil[-4:] == '.csv':
                    tf = open(self.src_dir + '/' + fil, 'r')
                    line = tf.readline()
                    if line.find('Latitude') < 0 or line.find('Longitude') < 0 \
                      or line.find('Time Zone') < 0:
                        tf.close()
                        continue
                    tf.seek(0)
                tf = open(self.src_dir + '/' + fil, 'r')
                lines = tf.readlines()
                tf.close()
                valu = []
                cell = []
                for j in range(len(the_cols)):
                    cell.append(0.)
                for i in range(12):   # monthly averages
                    valu.append(cell[:])
                if self.hourly:
                    valh = []
                    for i in range(24):   # hourly averages
                        valh.append([])
                        for j in range(12):   # by month
                            valh[-1].append([])
                            for k in range(len(the_cols)):
                                valh[-1][-1].append(0.)
                if self.daily:
                    vald = []
                    for i in range(24):   # hourly averages
                        vald.append([])
                        for d in range(365):   # by month
                            vald[-1].append([])
                            for k in range(len(the_cols)):
                                vald[-1][-1].append(0.)
                fst_row = len(lines) - 8760
                if fst_row < 0: # probably not for us
                    continue
                calc_mth = False
                if fil[-4:] == '.smw':
                    calc_mth = True
                    col = [9, 8, 7, 0, 4, -1, -1]
                    wnd_col = 4
                    bits = lines[0].split(',')
                    src_lat = float(bits[4])
                    src_lon = float(bits[5])
                    src_yr = int(bits[8])
                    fst_row = 1
                elif fil[-10:] == '(TMY2).csv' or fil[-10:] == '(TMY3).csv' \
                  or fil[-10:] == '(INTL).csv' or fil[-4:] == '.csv':
                    ghi_col = -1
                    if fst_row < 3:
                        bits = lines[0].split(',')
                        src_lat = float(bits[4])
                        src_lon = float(bits[5])
                        if isinstance(bits[8], int):
                            src_yr = bits[8]
                        else:
                            src_yr = -1
                    else:
                        cols = lines[fst_row - 3].strip().split(',')
                        bits = lines[fst_row - 2].strip().split(',')
                        src_yr = -1
                        for i in range(len(cols)):
                            if cols[i].lower() in ['latitude', 'lat']:
                                src_lat = float(bits[i])
                            elif cols[i].lower() in ['longitude', 'lon', 'long', 'lng']:
                                src_lon = float(bits[i])
                    cols = lines[fst_row - 1].strip().split(',')
                    for i in range(len(cols)):
                        if cols[i].lower() in ['df', 'dhi', 'diffuse', 'diffuse horizontal',
                                               'diffuse horizontal irradiance']:
                            col[the_cols.index('DHI (Diffuse)')] = i
                        elif cols[i].lower() in ['dn', 'dni', 'beam', 'direct normal',
                                                 'direct normal irradiance']:
                            col[the_cols.index('DNI (Normal)')] = i
                        elif cols[i].lower() in ['gh', 'ghi', 'global', 'global horizontal',
                                                 'global horizontal irradiance']:
                            col[the_cols.index('GHI (Direct)')] = i
                        elif cols[i].lower() in ['temp', 'tdry']:
                            col[the_cols.index('Temperature')] = i
                        elif cols[i].lower() in ['wspd', 'wind speed']:
                            col[the_cols.index('Wind Speed')] = i
                        elif cols[i].lower() in ['year']:
                            yr_col = i
                        elif cols[i].lower() in ['month']:
                            mth_col = i
                        elif cols[i].lower() in ['rainfall', 'rainfall (mm)']:
                            drop_rainfall = False
                            col[the_cols.index('Rainfall')] = i
                else:
                    continue
                for i in range(fst_row, len(lines)):
                    bits = lines[i].split(',')
                    try:
                        if src_yr < 0:
                            src_yr = bits[yr_col]
                    except:
                        pass
                    if calc_mth:
                        mth = 11
                        hr = i - fst_row
                        for j in range(12):
                            if hr < the_hour[j]:
                                mth = j - 1
                                break
                    else:
                        mth = int(bits[mth_col]) - 1
                    for j in range(len(col)):
                        if col[j] >= 0:
                            if col_min[1][j] == None or float(bits[col[j]]) < col_min[1][j]:
                                if (self.nonzero and float(bits[col[j]]) > 0) \
                                  or j == the_cols.index('Temperature') :
                                    col_min[1][j] = float(bits[col[j]])
                            if float(bits[col[j]]) > col_max[1][j]:
                                col_max[1][j] = float(bits[col[j]])
                            valu[mth][j] += float(bits[col[j]])
                            if self.hourly:
                                hr = (i - fst_row) % 24
                                valh[hr][mth][j] += float(bits[col[j]])
                            if self.daily:
                                hr = (i - fst_row) % 24
                                dy = (i - fst_row) % 365
                                vald[hr][dy][j] += float(bits[col[j]])
                key = '%s_%s_%s' % ('{:0.4f}'.format(src_lat), '{:0.4f}'.format(src_lon), src_yr)
                all_values[key] = valu
                if self.hourly:
                    hrly_values[key] = valh
                if self.daily:
                    daily_values[key] = vald
# now get 50m wind speed
        self.show_progress('Collected solar data')
        fils = os.listdir(self.wind_dir)
        val_col = the_cols.index('Wind @ 50m')
        wind_values = {}
        if self.hourly:
            hrly_wind_values = {}
        if self.daily:
            daily_wind_values = {}
        for fil in fils:
            if fil[-4:] != '.srw':
                continue
            valu = []
            for i in range(12):
                valu.append(0.)
            if self.hourly:
                valh = []
                for i in range(24):
                    valh.append(valu[:])
            if self.daily:
                vald = []
                for i in range(24):
                    vald.append([])
                    for j in range(365):
                        vald[-1].append(0.)
            tf = open(self.wind_dir + '/' + fil, 'r')
            lines = tf.readlines()
            tf.close()
            fst_row = len(lines) - 8760
            bits = lines[0].split(',')
            src_lat = float(bits[5])
            src_lon = float(bits[6])
            src_yr = bits[4]
            units = lines[3].strip().split(',')
            heights = lines[4].strip().split(',')
            wnd_col = -1
            for j in range(len(units)):
                if units[j] == 'm/s':
                   if heights[j] == '50':
                       wnd_col = j
                       break
                   elif heights[j] == '100':
                       the_cols[val_col] = 'Wind @ 100m'
                       wnd_col = j
                       break
            for i in range(fst_row, len(lines)):
                bits = lines[i].split(',')
                mth = 11
                hr = i - fst_row
                for j in range(12):
                    if hr < the_hour[j]:
                        mth = j - 1
                        break
                if col_min[1][val_col] == None or float(bits[wnd_col]) < col_min[1][val_col]:
                    if self.nonzero and float(bits[wnd_col]) > 0:
                        col_min[1][val_col] = float(bits[wnd_col])
                if float(bits[wnd_col]) > col_max[1][val_col]:
                    col_max[1][val_col] = float(bits[wnd_col])
                valu[mth] += float(bits[wnd_col])
                if self.hourly:
                    hr = (i - fst_row) % 24
                    valh[hr][mth] += float(bits[wnd_col])
                if self.daily:
                    hr = (i - fst_row) % 24
                    dy = (i - fst_row) % 365
                    vald[hr][dy] += float(bits[wnd_col])
            key = '%s_%s_%s' % ('{:0.4f}'.format(src_lat), '{:0.4f}'.format(src_lon), src_yr)
            wind_values[key] = valu
            if self.hourly:
                hrly_wind_values[key] = valh
            if self.daily:
                daily_wind_values[key] = vald
# we've got some 50m wind values?
        wind_error = False
        if len(wind_values) > 0:
            for key in all_values:
                for mth in range(12):
                    try:
                        all_values[key][mth][val_col] = wind_values[key][mth]
                    except:
                        if not wind_error:
                            wind_error = True
                            self.log += 'Wind value missing: ' + key + '-' + str(mth)
                        all_values[key][mth][val_col] = 0.
        if self.hourly:
            if len(hrly_wind_values) > 0:
                for key in hrly_values:
                    for hr in range(24):
                        for mth in range(12):
                            try:
                                hrly_values[key][hr][mth][val_col] = hrly_wind_values[key][hr][mth]
                            except:
                                if not wind_error:
                                    wind_error = True
                                    self.log += 'Wind value missing: ' + key + '-' + str(mth) + ' ' + str(hr)
                                hrly_values[key][hr][mth][val_col] = 0.
        if self.daily:
            if len(daily_wind_values) > 0:
                for key in daily_values:
                    for hr in range(24):
                        for dy in range(365):
                            try:
                                daily_values[key][hr][dy][val_col] = daily_wind_values[key][hr][dy]
                            except:
                                if not wind_error:
                                    wind_error = True
                                    self.log += 'Wind value missing: ' + key + '-' + str(mth) + '-' + str(dy) + ' ' + str(hr)
                                daily_values[key][hr][dy][val_col] = 0.
        self.show_progress('Collected wind data')
# and possibly rain if we haven't already got it
        if self.do_rain and col[the_cols.index('Rainfall')] < 0:
            fils = os.listdir(self.rain_dir)
            val_col = the_cols.index('Rainfall')
            rain_values = {}
            if self.hourly:
                hrly_rain_values = {}
            if self.daily:
                daily_rain_values = {}
            for fil in fils:
                if fil[-4:] != '.csv':
                    continue
                valu = []
                for i in range(12):
                    valu.append(0.)
                if self.hourly:
                    valh = []
                    for i in range(24):
                        valh.append(valu[:])
                if self.daily:
                    vald = []
                    for i in range(24):
                        vald.append([])
                        for j in range(365):
                            vald[-1].append(0.)
                tf = open(self.rain_dir + '/' + fil, 'r')
                lines = tf.readlines()
                tf.close()
                fst_row = len(lines) - 8760
                if fst_row < 3:
                    bits = lines[0].split(',')
                    src_lat = float(bits[4])
                    src_lon = float(bits[5])
                    if isinstance(bits[8], int):
                        src_yr = bits[8]
                    else:
                        src_yr = -1
                else:
                    cols = lines[fst_row - 3].strip().split(',')
                    bits = lines[fst_row - 2].strip().split(',')
                    src_yr = -1
                    for i in range(len(cols)):
                        if cols[i].lower() in ['latitude', 'lat']:
                            src_lat = float(bits[i])
                        elif cols[i].lower() in ['longitude', 'lon', 'long', 'lng']:
                            src_lon = float(bits[i])
                cols = lines[fst_row - 1].strip().split(',')
                for i in range(len(cols)):
                    if cols[i].lower() in ['year']:
                        yr_col = i
                    elif cols[i].lower() in ['rainfall', 'rainfall (mm)']:
                        drop_rainfall = False
                        rain_col = i
                for i in range(fst_row, len(lines)):
                    bits = lines[i].split(',')
                    if src_yr < 0:
                        src_yr = bits[yr_col]
                    mth = 11
                    hr = i - fst_row
                    for j in range(12):
                        if hr < the_hour[j]:
                            mth = j - 1
                            break
                    if col_min[1][val_col] == None or float(bits[rain_col]) < col_min[1][val_col]:
                        if self.nonzero and float(bits[rain_col]) > 0:
                            col_min[1][val_col] = float(bits[rain_col])
                    if float(bits[rain_col]) > col_max[1][val_col]:
                        col_max[1][val_col] = float(bits[rain_col])
                    valu[mth] += float(bits[rain_col])
                    if self.hourly:
                        hr = (i - fst_row) % 24
                        valh[hr][mth] += float(bits[rain_col])
                    if self.daily:
                        hr = (i - fst_row) % 24
                        dy = (i - fst_row) % 365
                        vald[hr][dy] += float(bits[rain_col])
                key = '%s_%s_%s' % ('{:0.4f}'.format(src_lat), '{:0.4f}'.format(src_lon), src_yr)
                rain_values[key] = valu
                if self.hourly:
                    hrly_rain_values[key] = valh
                if self.daily:
                    daily_rain_values[key] = vald
            if len(rain_values) > 0:
                for key in all_values:
                    for mth in range(12):
                        try:
                            all_values[key][mth][val_col] = rain_values[key][mth]
                        except:
                            all_values[key][mth][val_col] = 0.
            if self.hourly:
                if len(hrly_rain_values) > 0:
                    for key in hrly_values:
                        for hr in range(24):
                            for mth in range(12):
                                try:
                                    hrly_values[key][hr][mth][val_col] = hrly_rain_values[key][hr][mth]
                                except:
                                    hrly_values[key][hr][mth][val_col] = 0.
            if self.daily:
                if len(daily_rain_values) > 0:
                    for key in daily_values:
                        for hr in range(24):
                            for dy in range(365):
                                try:
                                    daily_values[key][hr][dy][val_col] = daily_rain_values[key][hr][dy]
                                except:
                                    daily_values[key][hr][dy][val_col] = 0.
            self.show_progress('Collected rainfall data')
        self.property = 'resource_grid=' + tgt_fil
        self.property = tgt_fil
        if tgt_fil[-5:] == '.xlsx':
            yr_ndx = tgt_fil.rfind(self.src_year) + 4
            if yr_ndx < 4:
                yr_ndx = tgt_fil.rfind('.') - 1
            wb = oxl.Workbook()
            normal = oxl.styles.Font(name='Arial', size='10')
            bold = oxl.styles.Font(name='Arial', bold=True, size='10')
            ws = wb.active
            ws.title = self.src_year
            ws.cell(row=1, column=1).value = 'Latitude'
            ws.cell(row=1, column=1).font = normal
            ws.cell(row=1, column=2).value = 'Longitude'
            ws.cell(row=1, column=2).font = normal
            ws.cell(row=1, column=3).value = 'Period'
            ws.cell(row=1, column=3).font = normal
            per_len = 6
            for i in range(len(the_cols)):
                if drop_rainfall:
                    if i == the_cols.index('Rainfall'):
                        continue
                ws.cell(row=1, column=i + 4).value = the_cols[i]
                ws.cell(row=1, column=i + 4).font = normal
            row = 3   # allow two rows for min & max
            for key in all_values:
                where = key.split('_')
                value = all_values[key]
                for i in range(len(value)):
                    ws.cell(row=row + 1, column=1).value = where[0]
                    ws.cell(row=row + 1, column=1).font = normal
                    ws.cell(row=row + 1, column=2).value = where[1]
                    ws.cell(row=row + 1, column=2).font = normal
                    ws.cell(row=row + 1, column=3).value = where[2] + '-' + '{0:02d}'.format(i + 1)
                    ws.cell(row=row + 1, column=3).font = normal
                    for j in range(len(value[i])):
                        if drop_rainfall:
                            if j == the_cols.index('Rainfall'):
                                continue
                        valu = round(value[i][j] / (the_days[i] * the_hrs[j]), 1)
                        ws.cell(row=row + 1, column=j + 4).value = valu
                        ws.cell(row=row + 1, column=j + 4).font = normal
                        if col_min[0][j] == None or valu < col_min[0][j]:
                            if (self.nonzero and valu > 0) or j == the_cols.index('Temperature'):
                                col_min[0][j] = valu
                        if valu > col_max[0][j]:
                            col_max[0][j] = valu
                    row += 1
                for j in range(len(seasons)):
                    ws.cell(row=row + 1, column=1).value = where[0]
                    ws.cell(row=row + 1, column=1).font = normal
                    ws.cell(row=row + 1, column=2).value = where[1]
                    ws.cell(row=row + 1, column=2).font = normal
                    ssn = where[2] + '-' + seasons[j][0]
                    per_len = max(per_len, len(ssn))
                    ws.cell(row=row + 1, column=3).value = ssn
                    ws.cell(row=row + 1, column=3).font = normal
                    valu = []
                    days = 0
                    for k in range(len(the_cols)):
                        valu.append(0)
                    for k in range(1, len(seasons[j])):
                        for l in range(len(the_cols)):
                            valu[l] += value[seasons[j][k]][l]
                        days += the_days[seasons[j][k]]
                    for k in range(len(valu)):
                        if drop_rainfall:
                            if k == the_cols.index('Rainfall'):
                                continue
                        val = round(valu[k] / (days * the_hrs[k]), 1)
                        ws.cell(row=row + 1, column=k + 4).value = val
                        ws.cell(row=row + 1, column=k + 4).font = normal
                    row += 1
                for j in range(len(periods)):
                    ws.cell(row=row + 1, column=1).value = where[0]
                    ws.cell(row=row + 1, column=1).font = normal
                    ws.cell(row=row + 1, column=2).value = where[1]
                    ws.cell(row=row + 1, column=2).font = normal
                    ssn = where[2] + '-' + periods[j][0]
                    per_len = max(per_len, len(ssn))
                    ws.cell(row=row + 1, column=3).value = ssn
                    ws.cell(row=row + 1, column=3).font = normal
                    valu = []
                    days = 0
                    for k in range(len(the_cols)):
                        valu.append(0)
                    for k in range(1, len(periods[j])):
                        for l in range(len(the_cols)):
                            valu[l] += value[periods[j][k]][l]
                        days += the_days[periods[j][k]]
                    for k in range(len(valu)):
                        if drop_rainfall:
                            if k == the_cols.index('Rainfall'):
                                continue
                        val = round(valu[k] / (days * the_hrs[k]), 1)
                        ws.cell(row=row + 1, column=k + 4).value = val
                        ws.cell(row=row + 1, column=k + 4).font = normal
                    row += 1
                ws.cell(row=row + 1, column=1).value = where[0]
                ws.cell(row=row + 1, column=1).font = normal
                ws.cell(row=row + 1, column=2).value = where[1]
                ws.cell(row=row + 1, column=2).font = normal
                ws.cell(row=row + 1, column=3).value = where[2]
                ws.cell(row=row + 1, column=3).font = normal
                valu = []
                for k in range(len(the_cols)):
                    valu.append(0)
                for k in range(12):
                    for l in range(len(the_cols)):
                        valu[l] += value[k][l]
                for k in range(len(valu)):
                    if drop_rainfall:
                        if k == the_cols.index('Rainfall'):
                            continue
                    val = round(valu[k] / (365 * the_hrs[k]), 1)
                    ws.cell(row=row + 1, column=k + 4).value = val
                    ws.cell(row=row + 1, column=k + 4).font = normal
                row += 1
            row = 1
            ws.cell(row=row + 1, column=3).value = 'Min.'
            ws.cell(row=row + 1, column=3).font = normal
            ws.cell(row=row + 2, column=3).value = 'Max.'
            ws.cell(row=row + 2, column=3).font = normal
            for j in range(len(col_min[0])):   # n values
                if drop_rainfall:
                    if j == the_cols.index('Rainfall'):
                        continue
                if col_min[0][j] == None:
                    ws.cell(row=row + 1, column=j + 4).value = 0.
                else:
                    ws.cell(row=row + 1, column=j + 4).value = round(col_min[0][j], 1)
                ws.cell(row=row + 1, column=j + 4).font = normal
                ws.cell(row=row + 2, column=j + 4).value = round(col_max[0][j], 1)
                ws.cell(row=row + 2, column=j + 4).font = normal
            lens = [8, 9, per_len + 1]
            for i in range(len(the_cols)):
                lens.append(len(the_cols[i]))
            for c in range(len(lens)):
                ws.column_dimensions[ssCol(c + 1)].width = lens[c]
            ws.freeze_panes = 'A4'
            wb.save(tgt_fil)
            wb.close()
            self.show_progress(f"Produced {tgt_fil[tgt_fil.rfind('/') + 1:]}")
            if self.hourly:
                for m in range(12):
                    wb = oxl.Workbook()
                    ws = wb.active
                    ws.title = self.src_year + '-{0:02d}'.format(m + 1)
                    ws.cell(row=1, column=1).value = 'Latitude'
                    ws.cell(row=1, column=1).font = normal
                    ws.cell(row=1, column=2).value = 'Longitude'
                    ws.cell(row=1, column=2).font = normal
                    ws.cell(row=1, column=3).value = 'Period'
                    ws.cell(row=1, column=3).font = normal
                    per_len = 13
                    for i in range(len(the_cols)):
                        if drop_rainfall:
                            if i == the_cols.index('Rainfall'):
                                continue
                        ws.cell(row=1, column=i + 4).value = the_cols[i]
                        ws.cell(row=1, column=i + 4).font = normal
                    row = 1
                    ws.cell(row=row + 1, column=3).value = 'Min.'
                    ws.cell(row=row + 1, column=3).font = normal
                    ws.cell(row=row + 2, column=3).value = 'Max.'
                    ws.cell(row=row + 2, column=3).font = normal
                    for j in range(len(col_min[1])):  # n values
                        if drop_rainfall:
                            if j == the_cols.index('Rainfall'):
                                continue
                        if col_min[1][j] == None:
                            ws.cell(row=row + 1, column=j + 4).value = 0.
                        else:
                            ws.cell(row=row + 1, column=j + 4).value = round(col_min[1][j], 1)
                        ws.cell(row=row + 1, column=j + 4).font = normal
                        ws.cell(row=row + 2, column=j + 4).value = round(col_max[1][j], 1)
                        ws.cell(row=row + 2, column=j + 4).font = normal
                    row += 2
                    for key in hrly_values:
                        where = key.split('_')
                        valueh = hrly_values[key]
                        for h in range(len(valueh)):  # 24 hours
                            ws.cell(row=row + 1, column=1).value = where[0]
                            ws.cell(row=row + 1, column=1).font = normal
                            ws.cell(row=row + 1, column=2).value = where[1]
                            ws.cell(row=row + 1, column=2).font = normal
                            ws.cell(row=row + 1, column=3).value = where[2] + '-' \
                                    + '{0:02d}'.format(m + 1) + '_{0:02d}'.format(h + 1) + ':00'
                            ws.cell(row=row + 1, column=3).font = normal
                            for j in range(len(valueh[0][m])):  # n values
                                if drop_rainfall:
                                    if j == the_cols.index('Rainfall'):
                                        continue
                                valu = round(valueh[h][m][j] / the_days[m], 1)  # need to multiply by 24
                                ws.cell(row=row + 1, column=j + 4).value = valu
                                ws.cell(row=row + 1, column=j + 4).font = normal
                            row += 1
                    lens = [8, 9, per_len + 1]
                    for i in range(len(the_cols)):
                        lens.append(len(the_cols[i]))
                    for c in range(len(lens)):
                        ws.column_dimensions[ssCol(c + 1)].width = lens[c]
                    ws.freeze_panes = 'A4'
                    tgt = tgt_fil[:yr_ndx] + '-{0:02d}'.format(m + 1) + tgt_fil[yr_ndx:]
                    wb.save(tgt)
                    wb.close()
                    self.show_progress(f"Produced {tgt[tgt.rfind('/') + 1:]}")
            if self.daily:
                mth = 0
                d = 0
                for dy in range(365):
                    wb = oxl.Workbook()
                    ws = wb.active
                    if ((dy + 1) * 24) > the_hour[mth + 1]:
                        mth += 1
                        d = 0
                    d += 1
                    ws.title = self.src_year + '-{0:02d}'.format(mth + 1) + '-{0:02d}'.format(d)
                    ws = wb.add_sheet(sht)
                    ws.cell(row=1, column=1).value = 'Latitude'
                    ws.cell(row=1, column=1).font = normal
                    ws.cell(row=1, column=2).value = 'Longitude'
                    ws.cell(row=1, column=2).font = normal
                    ws.cell(row=1, column=3).value = 'Period'
                    ws.cell(row=1, column=3).font = normal
                    per_len = 17
                    for i in range(len(the_cols)):
                        if drop_rainfall:
                            if i == the_cols.index('Rainfall'):
                                continue
                        ws.write(0, i + 3, the_cols[i])
                    row = 1
                    ws.cell(row=row + 1, column=3).value = 'Min.'
                    ws.cell(row=row + 1, column=3).font = normal
                    ws.cell(row=row + 2, column=3).value = 'Max.'
                    ws.cell(row=row + 2, column=3).font = normal
                    for j in range(len(col_min[1])):  # n values
                        if drop_rainfall:
                            if j == the_cols.index('Rainfall'):
                                continue
                        if col_min[1][j] == None:
                            ws.cell(row=row + 1, column=j + 4).value = 0.
                        else:
                            ws.cell(row=row + 1, column=j + 4).value = round(col_min[1][j], 1)
                        ws.cell(row=row + 1, column=j + 4).font = normal
                        ws.cell(row=row + 2, column=j + 4).value = round(col_max[1][j], 1)
                        ws.cell(row=row + 2, column=j + 4).font = normal
                    row += 2
                    for key in daily_values:
                        where = key.split('_')
                        valueh = daily_values[key]
                        for h in range(len(valueh)):  # 24 hours
                            ws.cell(row=row + 1, column=1).value = where[0]
                            ws.cell(row=row + 1, column=1).font = normal
                            ws.cell(row=row + 1, column=2).value = where[1]
                            ws.cell(row=row + 1, column=2).font = normal
                            ws.cell(row=row + 1, column=3).value = where[2] + '-{0:02d}'.format(mth + 1) \
                                     + '-{0:02d}'.format(d) + '_{0:02d}'.format(h + 1) + ':00'
                            ws.cell(row=row + 1, column=3).font = normal
                            for j in range(len(valueh[0][dy])):  # n values
                                if drop_rainfall:
                                    if j == the_cols.index('Rainfall'):
                                        continue
                                valu = round(valueh[h][dy][j], 1)
                                ws.cell(row=row + 1, column=j + 4).value = valu
                                ws.cell(row=row + 1, column=j + 4).font = normal
                            row += 1
                    lens = [8, 9, per_len + 1]
                    for i in range(len(the_cols)):
                        lens.append(len(the_cols[i]))
                    for c in range(len(lens)):
                        ws.column_dimensions[ssCol(c + 1)].width = lens[c]
                    ws.freeze_panes = 'A4'
                    tgt = tgt_fil[:yr_ndx] + '-{0:02d}'.format(mth + 1) + '-{0:02d}'.format(d) + tgt_fil[yr_ndx:]
                    wb.save(tgt)
                    self.show_progress(f"Produced {tgt[tgt.rfind('/') + 1:]}")
                    wb.close()
        elif tgt_fil[-4:] == '.xls':
            yr_ndx = tgt_fil.rfind(self.src_year) + 4
            if yr_ndx < 4:
                yr_ndx = tgt_fil.rfind('.') - 1
            wb = xlwt.Workbook()
            fnt = xlwt.Font()
            fnt.bold = True
            styleb = xlwt.XFStyle()
            styleb.font = fnt
            ws = wb.add_sheet(self.src_year)
            ws.write(0, 0, 'Latitude')
            ws.write(0, 1, 'Longitude')
            ws.write(0, 2, 'Period')
            per_len = 6
            for i in range(len(the_cols)):
                if drop_rainfall:
                    if i == the_cols.index('Rainfall'):
                        continue
                ws.write(0, i + 3, the_cols[i])
            row = 3   # allow two rows for min & max
            for key in all_values:
                where = key.split('_')
                value = all_values[key]
                for i in range(len(value)):
                    ws.write(row, 0, where[0])
                    ws.write(row, 1, where[1])
                    ws.write(row, 2, where[2] + '-' + '{0:02d}'.format(i + 1))
                    for j in range(len(value[i])):
                        if drop_rainfall:
                            if j == the_cols.index('Rainfall'):
                                continue
                        valu = round(value[i][j] / (the_days[i] * the_hrs[j]), 1)
                        ws.write(row, j + 3, valu)
                        if col_min[0][j] == None or valu < col_min[0][j]:
                            if (self.nonzero and valu > 0) or j == the_cols.index('Temperature'):
                                col_min[0][j] = valu
                        if valu > col_max[0][j]:
                            col_max[0][j] = valu
                    row += 1
                for j in range(len(seasons)):
                    ws.write(row, 0, where[0])
                    ws.write(row, 1, where[1])
                    ssn = where[2] + '-' + seasons[j][0]
                    per_len = max(per_len, len(ssn))
                    ws.write(row, 2, ssn)
                    valu = []
                    days = 0
                    for k in range(len(the_cols)):
                        valu.append(0)
                    for k in range(1, len(seasons[j])):
                        for l in range(len(the_cols)):
                            valu[l] += value[seasons[j][k]][l]
                        days += the_days[seasons[j][k]]
                    for k in range(len(valu)):
                        if drop_rainfall:
                            if k == the_cols.index('Rainfall'):
                                continue
                        val = round(valu[k] / (days * the_hrs[k]), 1)
                        ws.write(row, k + 3, val)
                    row += 1
                for j in range(len(periods)):
                    ws.write(row, 0, where[0])
                    ws.write(row, 1, where[1])
                    ssn = where[2] + '-' + periods[j][0]
                    per_len = max(per_len, len(ssn))
                    ws.write(row, 2, ssn)
                    valu = []
                    days = 0
                    for k in range(len(the_cols)):
                        valu.append(0)
                    for k in range(1, len(periods[j])):
                        for l in range(len(the_cols)):
                            valu[l] += value[periods[j][k]][l]
                        days += the_days[periods[j][k]]
                    for k in range(len(valu)):
                        if drop_rainfall:
                            if k == the_cols.index('Rainfall'):
                                continue
                        val = round(valu[k] / (days * the_hrs[k]), 1)
                        ws.write(row, k + 3, val)
                    row += 1
                ws.write(row, 0, where[0])
                ws.write(row, 1, where[1])
                ws.write(row, 2, where[2])
                valu = []
                for k in range(len(the_cols)):
                    valu.append(0)
                for k in range(12):
                    for l in range(len(the_cols)):
                        valu[l] += value[k][l]
                for k in range(len(valu)):
                    if drop_rainfall:
                        if k == the_cols.index('Rainfall'):
                            continue
                    val = round(valu[k] / (365 * the_hrs[k]), 1)
                    ws.write(row, k + 3, val)
                row += 1
            row = 1
            ws.write(row, 2, 'Min.')
            ws.write(row + 1, 2, 'Max.')
            for j in range(len(col_min[0])):   # n values
                if drop_rainfall:
                    if j == the_cols.index('Rainfall'):
                        continue
                if col_min[0][j] == None:
                    ws.write(row, j + 3, 0.)
                else:
                    ws.write(row, j + 3, round(col_min[0][j], 1))
                ws.write(row + 1, j + 3, round(col_max[0][j], 1))
            lens = [8, 9, per_len + 1]
            for i in range(len(the_cols)):
                lens.append(len(the_cols[i]))
            for c in range(len(lens)):
                if lens[c] * 275 > ws.col(c).width:
                    ws.col(c).width = lens[c] * 275
            ws.set_panes_frozen(True)  # frozen headings instead of split panes
            ws.set_horz_split_pos(3)  # in general, freeze after last heading row
            ws.set_remove_splits(True)  # if user does unfreeze, don't leave a split there
            wb.save(tgt_fil)
            self.show_progress(f"Produced {tgt_fil[tgt_fil.rfind('/') + 1:]}")
            if self.hourly:
                for m in range(12):
                    wb = xlwt.Workbook()
                    fnt = xlwt.Font()
                    fnt.bold = True
                    styleb = xlwt.XFStyle()
                    styleb.font = fnt
                    sht = self.src_year + '-{0:02d}'.format(m + 1)
                    ws = wb.add_sheet(sht)
                    ws.write(0, 0, 'Latitude')
                    ws.write(0, 1, 'Longitude')
                    ws.write(0, 2, 'Period')
                    per_len = 13
                    for i in range(len(the_cols)):
                        if drop_rainfall:
                            if i == the_cols.index('Rainfall'):
                                continue
                        ws.write(0, i + 3, the_cols[i])
                    row = 1
                    ws.write(row, 2, 'Min.')
                    ws.write(row + 1, 2, 'Max.')
                    for j in range(len(col_min[1])):  # n values
                        if drop_rainfall:
                            if j == the_cols.index('Rainfall'):
                                continue
                        if col_min[1][j] == None:
                            ws.write(row, j + 3, 0.)
                        else:
                            ws.write(row, j + 3, round(col_min[1][j], 1))
                        ws.write(row + 1, j + 3, round(col_max[1][j], 1))
                    row += 2
                    for key in hrly_values:
                        where = key.split('_')
                        valueh = hrly_values[key]
                        for h in range(len(valueh)):  # 24 hours
                            ws.write(row, 0, where[0])
                            ws.write(row, 1, where[1])
                            ws.write(row, 2, where[2] + '-' + '{0:02d}'.format(m + 1) +
                                     '_{0:02d}'.format(h + 1) + ':00')
                            for j in range(len(valueh[0][m])):  # n values
                                if drop_rainfall:
                                    if j == the_cols.index('Rainfall'):
                                        continue
                                valu = round(valueh[h][m][j] / the_days[m], 1)  # need to multiply by 24
                                ws.write(row, j + 3, valu)
                            row += 1
                    lens = [8, 9, per_len + 1]
                    for i in range(len(the_cols)):
                        lens.append(len(the_cols[i]))
                    for c in range(len(lens)):
                        if lens[c] * 275 > ws.col(c).width:
                            ws.col(c).width = lens[c] * 275
                    ws.set_panes_frozen(True)  # frozen headings instead of split panes
                    ws.set_horz_split_pos(3)  # in general, freeze after last heading row
                    ws.set_remove_splits(True)  # if user does unfreeze, don't leave a split there
                    tgt = tgt_fil[:yr_ndx] + '-{0:02d}'.format(m + 1) + tgt_fil[yr_ndx:]
                    wb.save(tgt)
                    self.show_progress(f"Produced {tgt[tgt.rfind('/') + 1:]}")
            if self.daily:
                mth = 0
                d = 0
                for dy in range(365):
                    wb = xlwt.Workbook()
                    fnt = xlwt.Font()
                    fnt.bold = True
                    styleb = xlwt.XFStyle()
                    styleb.font = fnt
                    if ((dy + 1) * 24) > the_hour[mth + 1]:
                        mth += 1
                        d = 0
                    d += 1
                    sht = self.src_year + '-{0:02d}'.format(mth + 1) + '-{0:02d}'.format(d)
                    ws = wb.add_sheet(sht)
                    ws.write(0, 0, 'Latitude')
                    ws.write(0, 1, 'Longitude')
                    ws.write(0, 2, 'Period')
                    per_len = 17
                    for i in range(len(the_cols)):
                        if drop_rainfall:
                            if i == the_cols.index('Rainfall'):
                                continue
                        ws.write(0, i + 3, the_cols[i])
                    row = 1
                    ws.write(row, 2, 'Min.')
                    ws.write(row + 1, 2, 'Max.')
                    for j in range(len(col_min[1])):  # n values
                        if drop_rainfall:
                            if j == the_cols.index('Rainfall'):
                                continue
                        if col_min[1][j] == None:
                            ws.write(row, j + 3, 0.)
                        else:
                            ws.write(row, j + 3, round(col_min[1][j], 1))
                        ws.write(row + 1, j + 3, round(col_max[1][j], 1))
                    row += 2
                    for key in daily_values:
                        where = key.split('_')
                        valueh = daily_values[key]
                        for h in range(len(valueh)):  # 24 hours
                            ws.write(row, 0, where[0])
                            ws.write(row, 1, where[1])
                            ws.write(row, 2, where[2] + '-{0:02d}'.format(mth + 1)
                                     + '-{0:02d}'.format(d) + '_{0:02d}'.format(h + 1) + ':00')
                            for j in range(len(valueh[0][dy])):  # n values
                                if drop_rainfall:
                                    if j == the_cols.index('Rainfall'):
                                        continue
                                valu = round(valueh[h][dy][j], 1)
                                ws.write(row, j + 3, valu)
                            row += 1
                    lens = [8, 9, per_len + 1]
                    for i in range(len(the_cols)):
                        lens.append(len(the_cols[i]))
                    for c in range(len(lens)):
                        if lens[c] * 275 > ws.col(c).width:
                            ws.col(c).width = lens[c] * 275
                    ws.set_panes_frozen(True)  # frozen headings instead of split panes
                    ws.set_horz_split_pos(3)  # in general, freeze after last heading row
                    ws.set_remove_splits(True)  # if user does unfreeze, don't leave a split there
                    tgt = tgt_fil[:yr_ndx] + '-{0:02d}'.format(mth + 1) + '-{0:02d}'.format(d) + tgt_fil[yr_ndx:]
                    wb.save(tgt)
                    self.show_progress(f"Produced {tgt[tgt.rfind('/') + 1:]}")
        else:
            tf = open(tgt_fil, 'w')
            hdr = 'Latitude,Longitude,Period'
            for i in range(len(the_cols)):
                hdr += ',' + the_cols[i]
            hdr += '\n'
            tf.write(hdr)
            for key in all_values:
                where = key.split('_')
                value = all_values[key]
                for i in range(len(value)):
                    line = '%s,%s,%s-%s' % (where[0], where[1], where[2], '{0:02d}'.format(i + 1))
                    for j in range(len(value[i])):
                        valu = round(value[i][j] / (the_days[i] * the_hrs[j]), 1)
                        line += ',' + str(valu)
                    tf.write(line + '\n')
                for j in range(len(seasons)):
                    line = '%s,%s,%s-%s' % (where[0], where[1], where[2], seasons[j][0])
                    valu = []
                    days = 0
                    for k in range(len(the_cols)):
                        valu.append(0)
                    for k in range(1, len(seasons[j])):
                        for l in range(len(the_cols)):
                            valu[l] += value[seasons[j][k]][l]
                        days += the_days[seasons[j][k]]
                    for k in range(len(valu)):
                        val = round(valu[k] / (days * the_hrs[k]), 1)
                        line += ',' + str(val)
                    tf.write(line + '\n')
                for j in range(len(periods)):
                    line = '%s,%s,%s-%s' % (where[0], where[1], where[2], periods[j][0])
                    valu = []
                    days = 0
                    for k in range(len(the_cols)):
                        valu.append(0)
                    for k in range(1, len(periods[j])):
                        for l in range(len(the_cols)):
                            valu[l] += value[periods[j][k]][l]
                        days += the_days[periods[j][k]]
                    for k in range(len(valu)):
                        val = round(valu[k] / (days * the_hrs[k]), 1)
                        line += ',' + str(val)
                    tf.write(line + '\n')
                line = '%s,%s,%s' % (where[0], where[1], str(where[2]))
                valu = []
                for k in range(len(the_cols)):
                    valu.append(0)
                for k in range(12):
                    for l in range(len(the_cols)):
                        valu[l] += value[k][l]
                for k in range(len(valu)):
                    val = round(valu[k] / (365 * the_hrs[k]), 1)
                    line += ',' + str(val)
                tf.write(line + '\n')
            tf.close()
            self.show_progress(f'Produced {tgt_fil}')
        self.log += '%s created' % tgt_fil[tgt_fil.rfind('/') + 1:]

class getParms(QtWidgets.QWidget):

    def __init__(self, help='help.html'):
        super(getParms, self).__init__()
        self.help = help
        self.initUI()

if "__main__" == __name__:
        sys.exit()
