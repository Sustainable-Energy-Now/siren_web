#!/usr/bin/python3
#
#  Copyright (C) 2016-2024 Sustainable Energy Now Inc., Angus King
#
#  indexweather.py - This file is part of SIREN.
#
#  SIREN is free software: you can redistribute it and/or modify
#  it under the terms of the GNU Affero General Public License as
#  published by the Free Software Foundation, either version 3 of
#  the License, or (at your option) any later version.
#
#  SIREN is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU Affero General Public License for more details.
#
#  You should have received a copy of the GNU Affero General
#  Public License along with SIREN.  If not, see
#  <http://www.gnu.org/licenses/>.
#

import openpyxl as oxl
import os
import sys
import xlwt

import displayobject
from credits import fileVersion
from getmodels import getModelFile
from senutils import getParents, getUser, ssCol


class makeIndex():

    def close(self):
        return

    def getLog(self):
        return self.log

    def __init__(self, what, src_dir, tgt_fil):
        if len(sys.argv) > 1:
            config_file = sys.argv[1]
        else:
            config_file = getModelFile('SIREN.ini')
        self.log = ''
        files = []
        fils = os.listdir(src_dir)
        for fil in fils:
            if (what[0].lower() == 's' and (fil[-4:] == '.csv' or fil[-4:] == '.smw')) \
              or (what[0].lower() == 'w' and fil[-4:] == '.srw'):
                tf = open(src_dir + '/' + fil, 'r')
                lines = tf.readlines()
                tf.close()
                if fil[-4:] == '.smw':
                    bits = lines[0].split(',')
                    src_lat = float(bits[4])
                    src_lon = float(bits[5])
                elif fil[-4:] == '.srw':
                    bits = lines[0].split(',')
                    src_lat = float(bits[5])
                    src_lon = float(bits[6])
                elif fil[-10:] == '(TMY2).csv' or fil[-10:] == '(TMY3).csv' \
                  or fil[-10:] == '(INTL).csv' or fil[-4:] == '.csv':
                    fst_row = len(lines) - 8760
                    if fst_row < 3:
                        bits = lines[0].split(',')
                        src_lat = float(bits[4])
                        src_lon = float(bits[5])
                    else:
                        cols = lines[fst_row - 3].split(',')
                        bits = lines[fst_row - 2].split(',')
                        for i in range(len(cols)):
                            if cols[i].lower() in ['latitude', 'lat']:
                                src_lat = float(bits[i])
                            elif cols[i].lower() in ['longitude', 'lon', 'long', 'lng']:
                                src_lon = float(bits[i])
                else:
                    continue
                files.append([src_lat, src_lon, fil])
        if tgt_fil[-5:] == '.xlsx':
            wb = oxl.Workbook()
            normal = oxl.styles.Font(name='Arial', size='10')
            bold = oxl.styles.Font(name='Arial', bold=True, size='10')
            ws = wb.active
            ws.title = 'Index'
            ws.cell(row=1, column=1).value = 'Latitude'
            ws.cell(row=1, column=1).font = normal
            ws.cell(row=1, column=2).value = 'Longitude'
            ws.cell(row=1, column=2).font = normal
            ws.cell(row=1, column=3).value = 'Filename'
            ws.cell(row=1, column=3).font = normal
            lens = [8, 9, 8]
            for i in range(len(files)):
                for c in range(3):
                    ws.cell(row=i + 2, column=c + 1).value = files[i][c]
                    ws.cell(row=i + 2, column=c + 1).font = normal
                    lens[c] = max(len(str(files[i][c])), lens[c])
            for c in range(len(lens)):
                ws.column_dimensions[ssCol(c + 1)].width = lens[c]
            ws.freeze_panes = 'A2'
            wb.save(tgt_fil)
        elif tgt_fil[-4:]:
            wb = xlwt.Workbook()
            fnt = xlwt.Font()
            fnt.bold = True
            styleb = xlwt.XFStyle()
            styleb.font = fnt
            ws = wb.add_sheet('Index')
            ws.write(0, 0, 'Latitude')
            ws.write(0, 1, 'Longitude')
            ws.write(0, 2, 'Filename')
            lens = [8, 9, 8]
            for i in range(len(files)):
                for c in range(3):
                    ws.write(i + 1, c, files[i][c])
                    lens[c] = max(len(str(files[i][c])), lens[c])
            for c in range(len(lens)):
                if lens[c] * 275 > ws.col(c).width:
                    ws.col(c).width = lens[c] * 275
            ws.set_panes_frozen(True)   # frozen headings instead of split panes
            ws.set_horz_split_pos(1)   # in general, freeze after last heading row
            ws.set_remove_splits(True)   # if user does unfreeze, don't leave a split there
            wb.save(tgt_fil)
        else:
            tf = open(tgt_fil, 'w')
            hdr = 'Latitude,Longitude,Filename\n'
            tf.write(hdr)
            for i in range(len(files)):
                line = '%s,%s,"%s"\n' % (files[i][0], files[i][1], files[i][2])
                tf.write(line)
            tf.close()
        self.log += '%s created' % tgt_fil[tgt_fil.rfind('/') + 1:]


class getParms():

    def __init__(self, help='help.html'):
        super(getParms, self).__init__()
        self.help = help
        self.initUI()

if "__main__" == __name__:
    app = QtWidgets.QApplication(sys.argv)
    if len(sys.argv) > 2:   # arguments
        src_dir_s = ''
        src_dir_w = ''
        tgt_fil = ''
        for i in range(1, len(sys.argv)):
            if sys.argv[i][:6] == 'solar=':
                src_dir_s = sys.argv[i][6:]
            elif sys.argv[i][:5] == 'wind=':
                src_dir_w = sys.argv[i][5:]
            elif sys.argv[i][:7] == 'target=' or sys.argv[i][:7] == 'tgtfil=':
                tgt_fil = sys.argv[i][7:]
        if src_dir_s != '':
            files = makeIndex('Solar', src_dir_s, tgt_fil)
        elif src_dir_w != '':
            files = makeIndex('Wind', src_dir_w, tgt_fil)
        else:
            print('No source directory specified')
    else:
        ex = getParms()
        app.exec_()
        app.deleteLater()
        sys.exit()
