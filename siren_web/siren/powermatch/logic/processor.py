from siren_web.siren.powermatch.logic.file_handler import FileHandler
from typing import Optional
from powermatchui.views.progress_handler import ProgressHandler
from siren_web.siren.powermatch.logic.excel import ExcelProcessor
from siren_web.siren.utilities.senutils import getUser, ssCol, techClean
import numpy as np
import openpyxl as oxl
import time
tech_names = ['Load', 'Onshore Wind', 'Offshore Wind', 'Rooftop PV', 'Fixed PV', 'Single Axis PV',
              'Dual Axis PV', 'Biomass', 'Geothermal', 'Other1', 'CST', 'Shortfall']
# initialise tech_names from .ini file
#            add dispatchable for re from [Grid] dispatchable?
# load data file. If not in data file then include in order and flag as RE
# tracking_pv is a synonym form dual_axis_pv
# phes is a synonym for pumped_hydro
# other1 is a synonym for other - or the other way around
# [Grid]
# dispatchable=pumped_hydro geothermal biomass solar_thermal cst
# consider: hydrogen bess
# [Power]
# technologies=backtrack_pv bess biomass cst fixed_pv geothermal offshore_wind rooftop_pv single_axis_pv solar_thermal tracking_pv wave wind other other_wave
#              add pumped_hydro hydrogen
#              maybe drop bess?
# fossil_technologies=fossil_ccgt fossil_coal fossil_cogen fossil_distillate fossil_gas fossil_mixed fossil_ocgt
target_keys = ['lcoe', 'load_pct', 'surplus_pct', 're_pct', 'cost', 'co2']
target_names = ['LCOE', 'Load%', 'Surplus%', 'RE%', 'Cost', 'CO2']
target_fmats = ['$%.2f', '%.1f%%', '%.1f%%', '%.1f%%', '$%.1fpwr_chr', '%.1fpwr_chr']
target_titles = ['LCOE ($)', 'Load met %', 'Surplus %', 'RE %', 'Total Cost ($)', 'tCO2e']
headers = ['Facility', 'Capacity\n(Gen, MW;\nStor, MWh)', 'To meet\nLoad (MWh)',
           'Subtotal\n(MWh)', 'CF', 'Cost ($/yr)', 'LCOG\nCost\n($/MWh)', 'LCOE\nCost\n($/MWh)',
           'Emissions\n(tCO2e)', 'Emissions\nCost', 'LCOE With\nCO2 Cost\n($/MWh)', 'Max.\nMWH',
           'Max.\nBalance', 'Capital\nCost', 'Lifetime\nCost', 'Lifetime\nEmissions',
           'Lifetime\nEmissions\nCost', 'Area (km^2)', 'Reference\nLCOE', 'Reference\nCF']
# set up columns for summary table. Hopefully to make it easier to add / alter columns
st_fac = 0 # Facility
st_cap = 1 # Capacity\n(Gen, MW;\nStor, MWh)
st_tml = 2 # To meet\nLoad (MWh)
st_sub = 3 # Subtotal\n(MWh)
st_cfa = 4 # CF
st_cst = 5 # Cost ($/yr)
st_lcg = 6 # LCOG\nCost\n($/MWh)
st_lco = 7 # LCOE\nCost\n($/MWh)
st_emi = 8 # Emissions\n(tCO2e)
st_emc = 9 # Emissions\nCost
st_lcc = 10 # LCOE With\nCO2 Cost\n($/MWh)
st_max = 11 # Max.\nMWH
st_bal = 12 # Max.\nBalance'
st_cac = 13 # Capital\nCost'
st_lic = 14 # Lifetime\nCost'
st_lie = 15 # Lifetime\nEmissions
st_lec = 16 # Lifetime\nEmissions\nCost
st_are = 17 # Area (km^2)
st_rlc = 18 # Reference\nLCOE
st_rcf = 19 # Reference\nCF

# same order as self.file_labels
C = 0 # Constraints - xls or xlsx
G = 1 # Generators - xls or xlsx
O = 2 # Optimisation - xls or xlsx
D = 3 # Data - xlsx
R = 4 # Results - xlsx
B = 5 # Batch input - xlsx
T = 6 # Transition input - xlsx
S = 'S' # Summary
O1 = 'O1'

class PowerMatchProcessor:
    def __init__(self, config, scenarios, generators, constraints, progress_handler: Optional[ProgressHandler] = None, 
                 event_callback=None, status_callback=None):
        self.config = config  # Access configuration details
        self.scenarios = scenarios
        self.generators =generators
        self.constraints = constraints
        self.listener = progress_handler
        self.event_callback = event_callback  # UI passes its event-processing function
        self.setStatus = status_callback or (lambda text: None)  # Default to no-op
        self.file_labels = ['Constraints', 'Generators', 'Optimisation', 'Data', 'Results', 'Batch']
        ifiles = [''] * len(self.file_labels)
        self.isheets = self.file_labels[:]
        del self.isheets[-2:]
        self.labels = [None] * len(self.file_labels)
        self.files = [None] * len(self.file_labels)
        self.sheets = self.file_labels[:]
        self.get_filename = FileHandler.get_filename
        self.adjusted_lcoe = True
        self.carbon_price = 0.
        self.carbon_price_max = 200.
        self.discount_rate = 0.
        self.load_folder = ''
        self.load_year = 'n/a'
        self.optimise_choice = 'LCOE'
        self.optimise_generations = 20
        self.optimise_mutation = 0.005
        self.optimise_population = 50
        self.optimise_stop = 0
        self.optimise_debug = False
        self.optimise_default = None
        self.optimise_multiplot = True
        self.optimise_multisurf = False
        self.optimise_multitable = False
        self.optimise_to_batch = True
        self.remove_cost = True
        self.results_prefix = ''
        self.dispatchable = ['Biomass', 'Geothermal', 'Pumped Hydro', 'Solar Thermal', 'CST'] # RE dispatchable
        self.save_tables = False
        self.show_multipliers = False
        self.show_correlation = False
        self.summary_sources = True
        self.surplus_sign = 1 # Note: Preferences file has it called shortfall_sign
        # it's easier for the user to understand while for the program logic surplus is easier
        self.underlying = ['Rooftop PV'] # technologies contributing to underlying (but not operational) load
        self.operational = []
        iorder = []
        self.targets = {}
        for t in range(len(target_keys)):
            if target_keys[t] in ['re_pct', 'surplus_pct']:
                self.targets[target_keys[t]] = [target_names[t], 0., -1, 0., 0, target_fmats[t],
                                                 target_titles[t]]
            else:
                self.targets[target_keys[t]] = [target_names[t], 0., 0., -1, 0, target_fmats[t],
                                                 target_titles[t]]
        try:
            dts = config.get('Grid', 'dispatchable').split(' ')
            dispatchable = []
            for dt in dts:
                dispatchable.append(techClean(dt.replace('_', ' ').title()))
            self.dispatchable = dispatchable
        except:
            pass
        try:
            adjust_cap = config.get('Power', 'adjust_cap')
            try:
                self.adjust_cap = float(adjust_cap)
            except:
                try:
                    self.adjust_cap = eval(adjust_cap)
                except:
                    pass
            if self.adjust_cap < 0:
                self.adjust_cap = pow(10, 12)
        except:
            pass
        try:
            items = config.items('Powermatch')
            for key, value in items:
                if key == 'batch_new_file':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.batch_new_file = True
                elif key == 'batch_prefix':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.batch_prefix = True
                elif key[:4] == 'tml_':
                    continue
                elif key[-5:] == '_file':
                    ndx = self.file_labels.index(key[:-5].title())
                    ifiles[ndx] = value.replace('$USER$', getUser())
                elif key[-6:] == '_sheet':
                    ndx = self.file_labels.index(key[:-6].title())
                    self.isheets[ndx] = value
                elif key == 'adjust_generators':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.adjust_gen = True
                elif key == 'adjusted_capacities':
                    self.adjustto = {}
                    bits = value.split(',')
                    for bit in bits:
                        bi = bit.split('=')
                        self.adjustto[bi[0]] = float(bi[1])
                elif key == 'carbon_price':
                    try:
                        self.carbon_price = float(value)
                    except:
                        pass
                elif key == 'carbon_price_max':
                    try:
                        self.carbon_price_max = float(value)
                    except:
                        pass
                elif key == 'adjusted_lcoe' or key == 'corrected_lcoe':
                    if value.lower() in ['false', 'no', 'off']:
                        self.adjusted_lcoe = False
                elif key == 'discount_rate':
                    try:
                        self.discount_rate = float(value)
                    except:
                        pass
                elif key == 'dispatch_order':
                    iorder = value.split(',')
                elif key == 'load':
                    try:
                        self.load_files = value
                        for ky, valu in parents:
                            self.load_files = self.load_files.replace(ky, valu)
                        self.load_files = self.load_files.replace('$USER$', getUser())
                    except:
                        pass
                elif key == 'load_year':
                    self.load_year = value
                elif key == 'log_status':
                    if value.lower() in ['false', 'no', 'off']:
                        self.log_status = False
                elif key == 'more_details':
                    if value.lower() in ['true', 'yes', 'on']:
                        self.more_details = True
                elif key == 'optimise_debug':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.optimise_debug = True
                elif key == 'optimise_default':
                    self.optimise_default = value
                elif key == 'optimise_choice':
                    self.optimise_choice = value
                elif key == 'optimise_generations':
                    try:
                        self.optimise_generations = int(value)
                    except:
                        pass
                elif key == 'optimise_multiplot':
                    if value.lower() in ['false', 'off', 'no']:
                        self.optimise_multiplot = False
                    elif value.lower() in ['surf', 'tri-surf', 'trisurf']:
                        self.optimise_multisurf = True
                elif key == 'optimise_multitable':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.optimise_multitable = True
                elif key == 'optimise_mutation':
                    try:
                        self.optimise_mutation = float(value)
                    except:
                        pass
                elif key == 'optimise_population':
                    try:
                        self.optimise_population = int(value)
                    except:
                        pass
                elif key == 'optimise_stop':
                    try:
                        self.optimise_stop = int(value)
                    except:
                        pass
                elif key == 'optimise_to_batch':
                    if value.lower() in ['false', 'off', 'no']:
                        self.optimise_to_batch = False
                elif key[:9] == 'optimise_':
                    try:
                        bits = value.split(',')
                        t = target_keys.index(key[9:])
                        # name, weight, minimum, maximum, widget index
                        self.targets[key[9:]] = [target_names[t], float(bits[0]), float(bits[1]),
                                                float(bits[2]), 0, target_fmats[t],
                                                 target_titles[t]]
                    except:
                        pass
                elif key == 'remove_cost':
                    if value.lower() in ['false', 'off', 'no']:
                        self.remove_cost = False
                elif key == 'results_prefix':
                    self.results_prefix = value
                elif key == 'save_tables':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.save_tables = True
                elif key == 'show_correlation':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.show_correlation = True
                elif key == 'show_multipliers':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.show_multipliers = True
                elif key == 'shortfall_sign':
                    if value[0] == '+' or value[0].lower() == 'p':
                        self.surplus_sign = -1
                elif key == 'summary_sources':
                    if value.lower() in ['false', 'off', 'no']:
                        self.summary_sources = False
                elif key == 'underlying':
                    self.underlying = value.split(',')
                elif key == 'operational':
                    self.operational = value.split(',')
        except:
            print('PME1: Error with', key)
            pass

    def data_sources(self, sheet, sheet_row, pm_data_file, option):
        normal = oxl.styles.Font(name='Arial')
        bold = oxl.styles.Font(name='Arial', bold=True)
        sheet.cell(row=sheet_row, column=1).value = 'Data sources'
        sheet.cell(row=sheet_row, column=1).font = bold
        sheet_row += 1
        sheet.cell(row=sheet_row, column=1).value = 'Scenarios folder'
        sheet.cell(row=sheet_row, column=1).font = normal
        sheet.cell(row=sheet_row, column=2).value = self.scenarios
        sheet.cell(row=sheet_row, column=2).font = normal
        sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
        sheet_row += 1
        sheet.cell(row=sheet_row, column=1).value = 'Powermatch data file'
        sheet.cell(row=sheet_row, column=1).font = normal
        if pm_data_file[: len(self.scenarios)] == self.scenarios:
            pm_data_file = pm_data_file[len(self.scenarios):]
        sheet.cell(row=sheet_row, column=2).value = pm_data_file
        sheet.cell(row=sheet_row, column=2).font = normal
        sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
        sheet_row += 1
        try:
            if self.loadCombo.currentText() != 'n/a':
                sheet.cell(row=sheet_row, column=1).value = 'Load file'
                sheet.cell(row=sheet_row, column=1).font = normal
                load_file = self.load_files.replace('$YEAR$', self.loadCombo.currentText())
                if load_file[: len(self.scenarios)] == self.scenarios:
                    load_file = load_file[len(self.scenarios):]
                sheet.cell(row=sheet_row, column=2).value = load_file
                sheet.cell(row=sheet_row, column=2).font = normal
                sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
                sheet_row += 1
        except:
            pass
        sheet.cell(row=sheet_row, column=1).value = 'Constraints worksheet'
        sheet.cell(row=sheet_row, column=1).font = normal
        sheet.cell(row=sheet_row, column=2).value = str(self.files[C].text()) \
               + '.' + str(self.sheets[C].currentText())
        sheet.cell(row=sheet_row, column=2).font = normal
        sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
        sheet_row += 1
        sheet.cell(row=sheet_row, column=1).value = 'Generators worksheet'
        sheet.cell(row=sheet_row, column=1).font = normal
        if option == T:
            sheet.cell(row=sheet_row, column=2).value = self.files[G].text()
        else:
            sheet.cell(row=sheet_row, column=2).value = self.files[G].text() \
                   + '.' + self.sheets[G].currentText()
        sheet.cell(row=sheet_row, column=2).font = normal
        sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
        return sheet_row

    def doDispatch(
        self, year, option, sender_name, pmss_details, pmss_data, re_order, dispatch_order, pm_data_file, 
        data_file, files, sheets, title=None
        ):
        def calcLCOE(annual_output, capital_cost, annual_operating_cost, discount_rate, lifetime):
            # Compute levelised cost of electricity
            if discount_rate > 0:
                annual_cost_capital = capital_cost * discount_rate * pow(1 + discount_rate, lifetime) / \
                                      (pow(1 + discount_rate, lifetime) - 1)
            else:
                annual_cost_capital = capital_cost / lifetime
            total_annual_cost = annual_cost_capital + annual_operating_cost
            try:
                return total_annual_cost / annual_output
            except:
                return total_annual_cost

        def format_period(per):
            hr = per % 24
            day = int((per - hr) / 24)
            mth = 0
            while day > the_days[mth] - 1:
                day -= the_days[mth]
                mth += 1
            return '{}-{:02d}-{:02d} {:02d}:00'.format(year, mth+1, day+1, hr)

        def summary_totals(title=''):
            sp_d = [' '] * len(headers)
            sp_d[st_fac] = title + 'Total'
            sp_d[st_cap] = cap_sum
            sp_d[st_tml] = tml_sum
            sp_d[st_sub] = gen_sum
            sp_d[st_cst] = cost_sum
            sp_d[st_lcg] = gs
            sp_d[st_lco] = gsw
            sp_d[st_emi] = co2_sum
            sp_d[st_emc] = co2_cost_sum
            sp_d[st_lcc] = gswc
            sp_d[st_cac] = capex_sum
            sp_d[st_lic] = lifetime_sum
            sp_d[st_lie] = lifetime_co2_sum
            sp_d[st_lec] = lifetime_co2_cost
            sp_d[st_are] = total_area
            sp_data.append(sp_d)
            if (self.carbon_price > 0 or option == B or option == T):
                sp_d = [' '] * len(headers)
                cc = co2_sum * self.carbon_price
                cl = cc * max_lifetime
                if self.adjusted_lcoe and tml_sum > 0:
                    cs = (cost_sum + cc) / tml_sum
                else:
                    if gen_sum > 0:
                        cs = (cost_sum + cc) / gen_sum
                    else:
                        cs = ''
                sp_d[st_fac] = title + 'Total incl. Carbon Cost'
                sp_d[st_cst] = cost_sum + cc
                sp_d[st_lic] = lifetime_sum + cl
                sp_data.append(sp_d)
            if tml_sum > 0:
                sp_d = [' '] * len(headers)
             #   sp_d[st_fac] = 'RE Direct Contribution to ' + title + 'Load'
                sp_d[st_fac] = 'RE %age'
                re_pct = (tml_sum - sto_sum - ff_sum) / tml_sum
                sp_d[st_cap] = '{:.1f}%'.format(re_pct * 100.)
                sp_d[st_tml] = tml_sum - ff_sum - sto_sum
                sp_data.append(sp_d)
                if sto_sum > 0:
                    sp_d = [' '] * len(headers)
                 #   sp_d[st_fac] = 'RE Contribution to ' + title + 'Load via Storage'
                    sp_d[st_fac] = 'Storage %age'
                    sp_d[st_cap] = '{:.1f}%'.format(sto_sum * 100. / tml_sum)
                    sp_d[st_tml] = sto_sum
                    sp_data.append(sp_d)
            sp_data.append([' '])
            sp_data.append([title + 'Load Analysis'])
            if sp_load != 0:
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = title + 'Load met'
                load_pct = (sp_load - sf_sums[0]) / sp_load
                sp_d[st_cap] = '{:.1f}%'.format(load_pct * 100)
                sp_d[st_tml] = sp_load - sf_sums[0]
                sp_data.append(sp_d)
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = 'Shortfall'
                sp_d[st_cap] = '{:.1f}%'.format(sf_sums[0] * 100 / sp_load)
                sp_d[st_tml] = sf_sums[0]
                sp_data.append(sp_d)
                if option == B or option == T:
                    sp_d = [' '] * len(headers)
                    sp_d[st_fac] = title + 'Total Load'
                    sp_d[st_tml] = sp_load
                    if title == '':
                        sp_d[st_max] = load_max
                    sp_data.append(sp_d)
                else:
                    load_mult = ''
                    try:
                        mult = round(pmss_details['Load'].multiplier, 3)
                        if mult != 1:
                            load_mult = ' x ' + str(mult)
                    except:
                        pass
                    sp_d = [' '] * len(headers)
                    sp_d[st_fac] = 'Total ' + title + 'Load - ' + year + load_mult
                    sp_d[st_tml] = sp_load
                    if title == '' or option == S:
                        sp_d[st_max] = load_max
                        sp_d[st_bal] = ' (' + format_period(load_hr)[5:] + ')'
                    sp_data.append(sp_d)
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = 'RE %age of Total ' + title + 'Load'
                sp_d[st_cap] = '{:.1f}%'.format((sp_load - sf_sums[0] - ff_sum) * 100. / sp_load)
                sp_data.append(sp_d)
                sp_data.append(' ')
                if tot_sto_loss != 0:
                    sp_d = [' '] * len(headers)
                    sp_d[st_fac] = 'Storage losses'
                    sp_d[st_sub] = tot_sto_loss
                    sp_data.append(sp_d)
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = title + 'Surplus'
                surp_pct = -sf_sums[1] / sp_load
                sp_d[st_cap] = '{:.1f}%'.format(surp_pct * 100)
                sp_d[st_sub] = -sf_sums[1]
                sp_data.append(sp_d)
            else:
                load_pct = 0
                surp_pct = 0
                re_pct = 0
            max_short = [0, 0]
            for h in range(len(shortfall)):
                if shortfall[h] > max_short[1]:
                    max_short[0] = h
                    max_short[1] = shortfall[h]
            if max_short[1] > 0:
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = 'Largest Shortfall'
                sp_d[st_sub] = round(max_short[1], 2)
                sp_d[st_cfa] = ' (' + format_period(max_short[0])[5:] + ')'
                sp_data.append(sp_d)
            if option == O or option == O1:
                return load_pct, surp_pct, re_pct

        def do_detail(fac, col, ss_row):
            if fac in self.generators.keys():
                gen = fac
            else:
                gen = pmss_details[fac].generator
            col += 1
            sp_cols.append(fac)
            sp_cap.append(pmss_details[fac].capacity * pmss_details[fac].multiplier)
            if do_zone and pmss_details[fac].zone != '':
                ns.cell(row=zone_row, column=col).value = pmss_details[fac].zone
                ns.cell(row=zone_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
            try:
                ns.cell(row=what_row, column=col).value = fac[fac.find('.') + 1:]
            except:
                ns.cell(row=what_row, column=col).value = fac # gen
            ns.cell(row=what_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
            ns.cell(row=cap_row, column=col).value = sp_cap[-1]
            ns.cell(row=cap_row, column=col).number_format = '#,##0.00'
            # capacity
            ns.cell(row=sum_row, column=col).value = '=SUM(' + ssCol(col) \
                    + str(hrows) + ':' + ssCol(col) + str(hrows + 8759) + ')'
            ns.cell(row=sum_row, column=col).number_format = '#,##0'
            # To meet load MWh
            ns.cell(row=tml_row, column=col).value = fac_tml[fac]
            ns.cell(row=tml_row, column=col).number_format = '#,##0'
            ns.cell(row=cf_row, column=col).value = '=IF(' + ssCol(col) + str(cap_row) + '>0,' + \
                    ssCol(col) + str(sum_row) + '/' + ssCol(col) + str(cap_row) + '/8760,"")'
            ns.cell(row=cf_row, column=col).number_format = '#,##0.0%'
            # subtotal MWh
            ns.cell(row=cf_row, column=col).value = '=IF(' + ssCol(col) + str(cap_row) + '>0,' + \
                    ssCol(col) + str(sum_row) +'/' + ssCol(col) + str(cap_row) + '/8760,"")'
            ns.cell(row=cf_row, column=col).number_format = '#,##0.0%'
            if gen not in self.generators.keys():
                return col
            if self.generators[gen].capex > 0 or self.generators[gen].fixed_om > 0 \
              or self.generators[gen].variable_om > 0 or self.generators[gen].fuel > 0:
                disc_rate = self.generators[gen].disc_rate
                if disc_rate == 0:
                    disc_rate = self.discount_rate
                if disc_rate == 0:
                    cst_calc = '/' + str(self.generators[gen].lifetime)
                else:
                    pwr_calc = 'POWER(1+' + str(disc_rate) + ',' + str(self.generators[gen].lifetime) + ')'
                    cst_calc = '*' + str(disc_rate) + '*' + pwr_calc + '/SUM(' + pwr_calc + ',-1)'
                ns.cell(row=cost_row, column=col).value = '=IF(' + ssCol(col) + str(cf_row) + \
                        '>0,' + ssCol(col) + str(cap_row) + '*' + str(self.generators[gen].capex) + \
                        cst_calc + '+' + ssCol(col) + str(cap_row) + '*' + \
                        str(self.generators[gen].fixed_om) + '+' + ssCol(col) + str(sum_row) + '*(' + \
                        str(self.generators[gen].variable_om) + '+' + str(self.generators[gen].fuel) + \
                        '),0)'
                ns.cell(row=cost_row, column=col).number_format = '$#,##0'
                ns.cell(row=lcoe_row, column=col).value = '=IF(AND(' + ssCol(col) + str(cf_row) + \
                        '>0,' + ssCol(col) + str(cap_row) + '>0),' + ssCol(col) + \
                        str(cost_row) + '/' + ssCol(col) + str(sum_row) + ',"")'
                ns.cell(row=lcoe_row, column=col).number_format = '$#,##0.00'
            elif self.generators[gen].lcoe > 0:
                if ss_row >= 0:
                    ns.cell(row=cost_row, column=col).value = '=IF(' + ssCol(col) + str(cf_row) + \
                            '>0,' + ssCol(col) + str(sum_row) + '*Summary!' + ssCol(st_rlc + 1) + str(ss_row) + \
                        '*Summary!' + ssCol(st_rcf + 1) + str(ss_row) + '/' + ssCol(col) + str(cf_row) + ',0)'
                    ns.cell(row=cost_row, column=col).number_format = '$#,##0'
                ns.cell(row=lcoe_row, column=col).value = '=IF(AND(' + ssCol(col) + str(cf_row) + '>0,' \
                        + ssCol(col) + str(cap_row) + '>0),' + ssCol(col) + str(cost_row) + '/8760/' \
                        + ssCol(col) + str(cf_row) +'/' + ssCol(col) + str(cap_row) + ',"")'
                ns.cell(row=lcoe_row, column=col).number_format = '$#,##0.00'
            elif self.generators[gen].lcoe_cf == 0: # no cost facility
                if ss_row >= 0:
                    ns.cell(row=cost_row, column=col).value = '=IF(' + ssCol(col) + str(cf_row) + \
                            '>0,' + ssCol(col) + str(sum_row) + '*Summary!' + ssCol(st_rlc + 1) + str(ss_row) + \
                        '*Summary!' + ssCol(st_rcf + 1) + str(ss_row) + '/' + ssCol(col) + str(cf_row) + ',0)'
                    ns.cell(row=cost_row, column=col).number_format = '$#,##0'
                ns.cell(row=lcoe_row, column=col).value = '=IF(AND(' + ssCol(col) + str(cf_row) + '>0,' \
                        + ssCol(col) + str(cap_row) + '>0),' + ssCol(col) + str(cost_row) + '/8760/' \
                        + ssCol(col) + str(cf_row) +'/' + ssCol(col) + str(cap_row) + ',"")'
                ns.cell(row=lcoe_row, column=col).number_format = '$#,##0.00'
            if self.generators[gen].emissions > 0:
                ns.cell(row=emi_row, column=col).value = '=' + ssCol(col) + str(sum_row) \
                        + '*' + str(self.generators[gen].emissions)
                ns.cell(row=emi_row, column=col).number_format = '#,##0'
            ns.cell(row=max_row, column=col).value = '=MAX(' + ssCol(col) + str(hrows) + \
                                           ':' + ssCol(col) + str(hrows + 8759) + ')'
            ns.cell(row=max_row, column=col).number_format = '#,##0.00'
            ns.cell(row=hrs_row, column=col).value = '=COUNTIF(' + ssCol(col) + str(hrows) + \
                                           ':' + ssCol(col) + str(hrows + 8759) + ',">0")'
            ns.cell(row=hrs_row, column=col).number_format = '#,##0'
            di = pmss_details[fac].col
            if pmss_details[fac].multiplier == 1:
                for row in range(hrows, 8760 + hrows):
                    ns.cell(row=row, column=col).value = pmss_data[di][row - hrows]
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            else:
                for row in range(hrows, 8760 + hrows):
                    ns.cell(row=row, column=col).value = pmss_data[di][row - hrows] * \
                                                         pmss_details[fac].multiplier
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            return col

        def do_detail_summary(fac, col, ss_row, dd_tml_sum, dd_re_sum):
            if do_zone and pmss_details[fac].zone != '':
                ss.cell(row=ss_row, column=st_fac+1).value = '=Detail!' + ssCol(col) + str(zone_row) + \
                                                      '&"."&Detail!' + ssCol(col) + str(what_row)
            else:
                ss.cell(row=ss_row, column=st_fac+1).value = '=Detail!' + ssCol(col) + str(what_row)
            if fac in self.generators.keys():
                gen = fac
            else:
                gen = pmss_details[fac].generator
            # capacity
            ss.cell(row=ss_row, column=st_cap+1).value = '=Detail!' + ssCol(col) + str(cap_row)
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.00'
            # To meet load MWh
            ss.cell(row=ss_row, column=st_tml+1).value = '=Detail!' + ssCol(col) + str(tml_row)
            ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
            dd_tml_sum += ssCol(st_tml+1) + str(ss_row) + '+'
            # subtotal MWh
            ss.cell(row=ss_row, column=st_sub+1).value = '=IF(Detail!' + ssCol(col) + str(sum_row) \
                                                  + '>0,Detail!' + ssCol(col) + str(sum_row) + ',"")'
            ss.cell(row=ss_row, column=st_sub+1).number_format = '#,##0'
            dd_re_sum += ssCol(st_sub+1) + str(ss_row) + '+'
            # CF
            ss.cell(row=ss_row, column=st_cfa+1).value = '=IF(Detail!' + ssCol(col) + str(cf_row) \
                                                  + '>0,Detail!' + ssCol(col) + str(cf_row) + ',"")'
            ss.cell(row=ss_row, column=st_cfa+1).number_format = '#,##0.0%'
            if gen not in self.generators.keys():
                return dd_tml_sum, dd_re_sum
            if self.generators[gen].capex > 0 or self.generators[gen].fixed_om > 0 \
              or self.generators[gen].variable_om > 0 or self.generators[gen].fuel > 0:
                disc_rate = self.generators[gen].disc_rate
                if disc_rate == 0:
                    disc_rate = self.discount_rate
                if disc_rate == 0:
                    cst_calc = '/' + str(self.generators[gen].lifetime)
                else:
                    pwr_calc = 'POWER(1+' + str(disc_rate) + ',' + str(self.generators[gen].lifetime) + ')'
                    cst_calc = '*' + str(disc_rate) + '*' + pwr_calc + '/SUM(' + pwr_calc + ',-1)'
                # cost / yr
                if self.remove_cost:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=IF(Detail!' + ssCol(col) + str(sum_row) \
                            + '>0,Detail!' + ssCol(col) + str(cost_row) + ',"")'
                else:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=Detail!' + ssCol(col) + str(cost_row)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                # lcog
                ss.cell(row=ss_row, column=st_lcg+1).value = '=IF(Detail!' + ssCol(col) + str(lcoe_row) \
                                                      + '>0,Detail!' + ssCol(col) + str(lcoe_row) + ',"")'
                ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
                # capital cost
                ss.cell(row=ss_row, column=st_cac+1).value = '=IF(Detail!' + ssCol(col) + str(cap_row) \
                                                        + '>0,Detail!' + ssCol(col) + str(cap_row) + '*'  \
                                                        + str(self.generators[gen].capex) + ',"")'
                ss.cell(row=ss_row, column=st_cac+1).number_format = '$#,##0'
            elif self.generators[gen].lcoe > 0:
                # cost / yr
                if self.remove_cost:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=IF(Detail!' + ssCol(col) + str(sum_row) \
                            + '>0,Detail!' + ssCol(col) + str(cost_row) + ',"")'
                else:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=Detail!' + ssCol(col) + str(cost_row)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                # lcog
                ss.cell(row=ss_row, column=st_lcg+1).value = '=Detail!' + ssCol(col) + str(lcoe_row)
                ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
                # ref lcoe
                ss.cell(row=ss_row, column=st_rlc+1).value = self.generators[gen].lcoe
                ss.cell(row=ss_row, column=st_rlc+1).number_format = '$#,##0.00'
                # ref cf
                ss.cell(row=ss_row, column=st_rcf+1).value = self.generators[gen].lcoe_cf
                ss.cell(row=ss_row, column=st_rcf+1).number_format = '#,##0.0%'
            elif self.generators[gen].lcoe_cf == 0: # no cost facility
                # cost / yr
                if self.remove_cost:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=IF(Detail!' + ssCol(col) + str(sum_row) \
                            + '>0,Detail!' + ssCol(col) + str(cost_row) + ',"")'
                else:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=Detail!' + ssCol(col) + str(cost_row)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                # lcog
                ss.cell(row=ss_row, column=st_lcg+1).value = '=Detail!' + ssCol(col) + str(lcoe_row)
                ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
                # ref lcoe
                ss.cell(row=ss_row, column=st_rlc+1).value = self.generators[gen].lcoe
                ss.cell(row=ss_row, column=st_rlc+1).number_format = '$#,##0.00'
                # ref cf
                ss.cell(row=ss_row, column=st_rcf+1).value = self.generators[gen].lcoe_cf
                ss.cell(row=ss_row, column=st_rcf+1).number_format = '#,##0.0%'
            # lifetime cost
            ss.cell(row=ss_row, column=st_lic+1).value = '=IF(Detail!' + ssCol(col) + str(sum_row) \
                                                    + '>0,Detail!' + ssCol(col) + str(cost_row) + '*lifetime,"")'
            ss.cell(row=ss_row, column=st_lic+1).number_format = '$#,##0'
            # max mwh
            ss.cell(row=ss_row, column=st_max+1).value = '=IF(Detail!' + ssCol(col) + str(sum_row) \
                                                   + '>0,Detail!' + ssCol(col) + str(max_row) + ',"")'
            ss.cell(row=ss_row, column=st_max+1).number_format = '#,##0.00'
            if self.generators[gen].emissions > 0:
                if self.remove_cost:
                    ss.cell(row=ss_row, column=st_emi+1).value = '=IF(Detail!' + ssCol(col) + str(sum_row) \
                            + '>0,Detail!' + ssCol(col) + str(emi_row) + ',"")'
                else:
                    ss.cell(row=ss_row, column=st_emi+1).value = '=Detail!' + ssCol(col) + str(emi_row)
                ss.cell(row=ss_row, column=st_emi+1).number_format = '#,##0'
                if self.carbon_price > 0:
                    ss.cell(row=ss_row, column=st_emc+1).value = '=IF(AND(' + ssCol(st_emi+1) + str(ss_row) + '<>"",' + \
                                                                 ssCol(st_emi+1) + str(ss_row) + '>0),' + \
                                                                 ssCol(st_emi+1) + str(ss_row) + '*carbon_price,"")'
                    ss.cell(row=ss_row, column=st_emc+1).number_format = '$#,##0'
            ss.cell(row=ss_row, column=st_lie+1).value = '=IF(AND(' + ssCol(st_emi+1) + str(ss_row) + '<>"",' + \
                                                         ssCol(st_emi+1) + str(ss_row) + '>0),' + \
                                                         ssCol(st_emi+1) + str(ss_row) + '*lifetime,"")'
            ss.cell(row=ss_row, column=st_lie+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_lec+1).value = '=IF(AND(' + ssCol(st_emi+1) + str(ss_row) + '<>"",' + \
                                                         ssCol(st_emi+1) + str(ss_row) + '>0),' + \
                                                         ssCol(st_emc+1) + str(ss_row) + '*lifetime,"")'
            ss.cell(row=ss_row, column=st_lec+1).number_format = '$#,##0'
            if self.generators[gen].area > 0:
                ss.cell(row=ss_row, column=st_are+1).value = '=Detail!' + ssCol(col) + str(cap_row) +\
                                                             '*' + str(self.generators[gen].area)
                ss.cell(row=ss_row, column=st_are+1).number_format = '#,##0.00'
            return dd_tml_sum, dd_re_sum

        def detail_summary_total(ss_row, title='', base_row='', back_row=''):
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = title + 'Total'
            for col in range(1, len(headers) + 1):
                ss.cell(row=3, column=col).font = bold
                ss.cell(row=ss_row, column=col).font = bold
            for col in [st_cap, st_tml, st_sub, st_cst, st_emi, st_emc, st_cac, st_lic, st_lie, st_lec, st_are]:
                if back_row != '':
                    strt = ssCol(col, base=0) + back_row + '+'
                else:
                    strt = ''
                ss.cell(row=ss_row, column=col+1).value = '=' + strt + 'SUM(' + ssCol(col, base=0) + \
                        base_row + ':' + ssCol(col, base=0) + str(ss_row - 1) + ')'
                if col in [st_cap, st_are]:
                    ss.cell(row=ss_row, column=col+1).number_format = '#,##0.00'
                elif col in [st_tml, st_sub, st_emi, st_lie]:
                    ss.cell(row=ss_row, column=col+1).number_format = '#,##0'
                else:
                    ss.cell(row=ss_row, column=col+1).number_format = '$#,##0'
            ss.cell(row=ss_row, column=st_lcg+1).value = '=' + ssCol(st_cst+1) + str(ss_row) + \
                                                         '/' + ssCol(st_sub+1) + str(ss_row)
            ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
            ss.cell(row=ss_row, column=st_lco+1).value = '=' + ssCol(st_cst+1) + str(ss_row) + \
                                                         '/' + ssCol(st_tml+1) + str(ss_row)
            ss.cell(row=ss_row, column=st_lco+1).number_format = '$#,##0.00'
            if self.carbon_price > 0:
                ss.cell(row=ss_row, column=st_lcc+1).value = '=(' + ssCol(st_cst+1) + str(ss_row) + \
                    '+' + ssCol(st_emc+1) + str(ss_row) + ')/' + ssCol(st_tml+1) + str(ss_row)
                ss.cell(row=ss_row, column=st_lcc+1).number_format = '$#,##0.00'
                ss.cell(row=ss_row, column=st_lcc+1).font = bold
            last_col = ssCol(ns.max_column)
            r = 1
            if self.carbon_price > 0:
                ss_row += 1
                ss.cell(row=ss_row, column=1).value = title + 'Total incl. Carbon Cost'
                ss.cell(row=ss_row, column=st_cst+1).value = '=' + ssCol(st_cst+1) + str(ss_row - 1) + \
                        '+' + ssCol(st_emc+1) + str(ss_row - 1)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                ss.cell(row=ss_row, column=st_lic+1).value = '=' + ssCol(st_lic+1) + str(ss_row - r) + \
                                                             '+' + ssCol(st_lec+1) + str(ss_row - 1)
                ss.cell(row=ss_row, column=st_lic+1).number_format = '$#,##0'
                r += 1
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = title + 'RE %age'
            ss.cell(row=ss_row, column=st_tml+1).value = ns_tml_sum[:-1] + ')'
            ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_cap+1).value = '=' + ssCol(st_tml+1) + str(ss_row) + '/' +\
                                                         ssCol(st_tml+1) + str(ss_row - r)
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.0%'
            ss_re_row = ss_row
            ss_sto_row = -1
            # if storage
            if ns_sto_sum != '':
                ss_row += 1
                ss.cell(row=ss_row, column=1).value = title + 'Storage %age'
                ss.cell(row=ss_row, column=st_tml+1).value = '=' + ns_sto_sum[1:]
                ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
                ss.cell(row=ss_row, column=st_cap+1).value = '=(' + ns_sto_sum[1:] + ')/' + ssCol(st_tml+1) + \
                                                             str(ss_row - r - 1)
                ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.0%'
                ss_sto_row = ss_row
            # now do the LCOE and LCOE with CO2 stuff
            if base_row == '4':
                base_col = 'C'
                if ss_sto_row >= 0:
                    for rw in range(ss_re_fst_row, ss_re_lst_row + 1):
                        ss.cell(row=rw, column=st_lco+1).value = '=IF(AND(' + ssCol(st_lcg+1) + str(rw) + '<>"",' + \
                                ssCol(st_lcg+1) + str(rw) + '>0),' + \
                                ssCol(st_cst+1) + str(rw) + '/(' + ssCol(st_tml+1) + str(rw) + '+(' + \
                                ssCol(st_tml+1) + '$' + str(ss_sto_row) + '*' + ssCol(st_tml+1) + str(rw) + \
                                ')/' + ssCol(st_tml+1) + '$' + str(ss_re_row) + '),"")'
                        ss.cell(row=rw, column=st_lco+1).number_format = '$#,##0.00'
                        if self.carbon_price > 0:
                            ss.cell(row=rw, column=st_lcc+1).value = '=IF(AND(' + ssCol(st_emc+1) + str(rw) + '<>"",' + \
                                    ssCol(st_emc+1) + str(rw) + '>0),(' + \
                                    ssCol(st_cst+1) + str(rw) + '+' + ssCol(st_emc+1) + str(rw) + ')/(' + \
                                    ssCol(st_tml+1) + str(rw) + '+(' + ssCol(st_tml+1) + '$' + str(ss_sto_row) + \
                                    '*' + ssCol(st_tml+1) + str(rw) + ')/' + ssCol(st_tml+1) + '$' + \
                                    str(ss_re_row) + '),"")'
                            ss.cell(row=rw, column=st_lcc+1).number_format = '$#,##0.00'
                else:
                    for rw in range(ss_re_fst_row, ss_re_lst_row):
                        ss.cell(row=rw, column=st_lco+1).value = '=IF(' + ssCol(st_lcg+1) + str(rw) + '>0,' + \
                                ssCol(st_cst+1) + str(rw) + '/' + ssCol(st_tml+1) + str(rw) + '),"")'
                        ss.cell(row=rw, column=st_lco+1).number_format = '$#,##0.00'
                        if self.carbon_price > 0:
                            ss.cell(row=rw, column=st_lcc+1).value = '=IF(AND(' + ssCol(st_emc+1) + str(rw) + '<>"",' + \
                                    ssCol(st_emc+1) + str(rw) + '>0),(' + \
                                    ssCol(st_cst+1) + str(rw) + ssCol(st_emc+1) + str(rw) + ')/' + \
                                    ssCol(st_tml+1) + str(rw) + '),"")'
                            ss.cell(row=rw, column=st_lcc+1).number_format = '$#,##0.00'
                for rw in range(ss_re_lst_row + 1, ss_lst_row + 1):
                    ss.cell(row=rw, column=st_lco+1).value = '=IF(AND(' + ssCol(st_tml+1) + str(rw) + '<>"",' + \
                                    ssCol(st_tml+1) + str(rw) + '>0),' + ssCol(st_cst+1) + str(rw) + \
                                                             '/' + ssCol(st_tml+1) + str(rw) + ',"")'
                    ss.cell(row=rw, column=st_lco+1).number_format = '$#,##0.00'
                    if self.carbon_price > 0:
                        ss.cell(row=rw, column=st_lcc+1).value = '=IF(AND(' + ssCol(st_emc+1) + str(rw) + '<>"",' + \
                                    ssCol(st_emc+1) + str(rw) + '>0),(' + \
                                ssCol(st_cst+1) + str(rw) + '+' + ssCol(st_emc+1) + str(rw) + ')/' + \
                                ssCol(st_tml+1) + str(rw) + ',"")'
                        ss.cell(row=rw, column=st_lcc+1).number_format = '$#,##0.00'
            else:
                base_col = ssCol(next_col)
                for rw in range(ul_fst_row, ul_lst_row + 1):
                    ss.cell(row=rw, column=st_lco+1).value = '=' + ssCol(st_cst+1) + str(rw) + \
                                                             '/' + ssCol(st_tml+1) + str(rw)
                    ss.cell(row=rw, column=st_lco+1).number_format = '$#,##0.00'
                    if self.carbon_price > 0:
                        ss.cell(row=rw, column=st_lcc+1).value = '=(' + ssCol(st_cst+1) + str(rw) + \
                            '+' + ssCol(st_emc+1) + str(rw) + ')/' + ssCol(st_tml+1) + str(rw)
                        ss.cell(row=rw, column=st_lcc+1).number_format = '$#,##0.00'
            ss_row += 2
            ss.cell(row=ss_row, column=1).value = title + 'Load Analysis'
            ss.cell(row=ss_row, column=1).font = bold
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = title + 'Load met'
      ##      lm_row = ss_row
      #      if self.surplus_sign < 0:
      #          addsub = ')+' + base_col
      #      else:
      #          addsub = ')-' + base_col
      #      ss.cell(row=ss_row, column=st_tml+1).value = '=SUMIF(Detail!' + last_col + str(hrows) + ':Detail!' \
      #          + last_col + str(hrows + 8759) + ',"' + sf_test[0] + '=0",Detail!C' + str(hrows) \
      #          + ':Detail!C' + str(hrows + 8759) + ')+SUMIF(Detail!' + last_col + str(hrows) + ':Detail!' \
      #          + last_col + str(hrows + 8759) + ',"' + sf_test[1] + '0",Detail!C' + str(hrows) + ':Detail!C' \
      #          + str(hrows + 8759) + addsub + str(ss_row + 1)
            ss.cell(row=ss_row, column=st_tml+1).value = '=Detail!' + base_col + str(sum_row) + '-' + base_col + str(ss_row + 1)
            ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_cap+1).value = '=' + ssCol(st_tml+1) + str(ss_row) + '/' + ssCol(st_tml+1) + \
                                                         str(ss_row + 2)
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.0%'
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = title + 'Shortfall'
            sf_text = 'SUMIF(Detail!' + last_col + str(hrows) + ':Detail!' + last_col \
                      + str(hrows + 8759) + ',"' + sf_test[0] + '0",Detail!' + last_col \
                      + str(hrows) + ':Detail!' + last_col + str(hrows + 8759) + ')'
            if self.surplus_sign > 0:
                ss.cell(row=ss_row, column=st_tml+1).value = '=-' + sf_text
            else:
                ss.cell(row=ss_row, column=st_tml+1).value = '=' + sf_text
            ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_cap+1).value = '=' + ssCol(st_tml+1) + str(ss_row) + '/' + ssCol(st_tml+1) + \
                                                         str(ss_row + 1)
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.0%'
            ss_row += 1
            ld_row = ss_row
            load_mult = ''
            try:
                mult = round(pmss_details['Load'].multiplier, 3)
                if mult != 1:
                    load_mult = ' x ' + str(mult)
            except:
                pass
            ss.cell(row=ss_row, column=1).value = 'Total ' + title + 'Load - ' + year + load_mult
            ss.cell(row=ss_row, column=1).font = bold
            ss.cell(row=ss_row, column=st_tml+1).value = '=SUM(' + ssCol(st_tml+1) + str(ss_row - 2) + ':' + \
                                                         ssCol(st_tml+1) + str(ss_row - 1) + ')'
            ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_tml+1).font = bold
            ss.cell(row=ss_row, column=st_max+1).value = '=Detail!' + base_col + str(max_row)
            ss.cell(row=ss_row, column=st_max+1).number_format = '#,##0.00'
            ss.cell(row=ss_row, column=st_max+1).font = bold
            ss.cell(row=ss_row, column=st_bal+1).value = '=" ("&OFFSET(Detail!B' + str(hrows - 1) + ',MATCH(Detail!' + \
                    base_col + str(max_row) + ',Detail!' + base_col + str(hrows) + ':Detail!' + base_col + \
                    str(hrows + 8759) + ',0),0)&")"'
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = 'RE %age of Total ' + title + 'Load'
            ss.cell(row=ss_row, column=1).font = bold
            if ns_sto_sum == '':
                ss.cell(row=ss_row, column=st_cap+1).value = ssCol(st_tml+1) + str(ss_re_row - 1) + \
                                                             '/' + ssCol(st_tml+1) + str(ss_row - 1)
            else:
                ss.cell(row=ss_row, column=st_cap+1).value = '=(' + ssCol(st_tml+1) + str(ss_re_row) + '+' + \
                                                             ssCol(st_tml+1) + str(ss_sto_row) + ')/' + \
                                                             ssCol(st_tml+1) + str(ss_row - 1)
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.0%'
            ss.cell(row=ss_row, column=st_cap+1).font = bold
            ss_row += 2
            if ns_loss_sum != '':
                ss.cell(row=ss_row, column=1).value = title + 'Storage Losses'
                ss.cell(row=ss_row, column=st_sub+1).value = '=' + ns_loss_sum[1:]
                ss.cell(row=ss_row, column=st_sub+1).number_format = '#,##0'
                ss_row += 1
            ss.cell(row=ss_row, column=1).value = title + 'Surplus'
            sf_text = 'SUMIF(Detail!' + last_col + str(hrows) + ':Detail!' + last_col \
                      + str(hrows + 8759) + ',"' + sf_test[1] + '0",Detail!' + last_col + str(hrows) \
                      + ':Detail!' + last_col + str(hrows + 8759) + ')'
            if self.surplus_sign < 0:
                ss.cell(row=ss_row, column=st_sub+1).value = '=-' + sf_text
            else:
                ss.cell(row=ss_row, column=st_sub+1).value = '=' + sf_text
            ss.cell(row=ss_row, column=st_sub+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_cap+1).value = '=' + ssCol(st_sub+1) + str(ss_row) + '/' + ssCol(st_tml+1) + str(ld_row)
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.0%'
            max_short = [0, 0]
            for h in range(len(shortfall)):
                if shortfall[h] > max_short[1]:
                    max_short[0] = h
                    max_short[1] = shortfall[h]
            if max_short[1] > 0:
                ss_row += 1
                ss.cell(row=ss_row, column=1).value = 'Largest ' + title + 'Shortfall:'
                ss.cell(row=ss_row, column=st_sub+1).value = '=Detail!' + last_col + str(hrows + max_short[0])
                ss.cell(row=ss_row, column=st_sub+1).number_format = '#,##0.00'
                ss.cell(row=ss_row, column=st_cfa+1).value = '=" ("&OFFSET(Detail!B' + str(hrows - 1) + \
                        ',MATCH(' + ssCol(st_sub+1) + str(ss_row) + ',Detail!' + last_col + str(hrows) + \
                        ':Detail!' + last_col + str(hrows + 8759) + ',0),0)&")"'
            return ss_row, ss_re_row

    # The "guts" of Powermatch processing. Have a single calculation algorithm
    # for Summary, Powermatch (detail), and Optimise. The detail makes it messy
    # Note: For Batch pmss_data is reused so don't update it in doDispatch
        self.files = files
        self.sheets = sheets
        the_days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        if self.surplus_sign < 0:
            sf_test = ['>', '<']
            sf_sign = ['+', '-']
        else:
            sf_test = ['<', '>']
            sf_sign = ['-', '+']
        sp_cols = []
        sp_cap = []
        shortfall = [0.] * 8760
        re_tml_sum = 0. # keep tabs on how much RE is used
        start_time = time.time()
        do_zone = False # could pass as a parameter
        max_lifetime = 0
        # find max. lifetime years for all technologies selected
        for key in pmss_details.keys():
            if key == 'Load'or key == 'Total':
                continue
            if pmss_details[key].capacity * pmss_details[key].multiplier > 0:
             #   gen = key.split('.')[-1]
                gen = pmss_details[key].generator
                max_lifetime = max(max_lifetime, self.generators[gen].lifetime)
        for key in pmss_details.keys():
            if key.find('.') > 0:
                do_zone = True
                break
        underlying_facs = []
        undercol = [] * len(self.underlying)
        operational_facs = []
        fac_tml = {}
        for fac in re_order:
            if fac == 'Load':
                continue
            fac_tml[fac] = 0.
            if fac in self.operational:
              #  operational_facs.append(fac)
                continue
            if fac.find('.') > 0:
                if fac[fac.find('.') + 1:] in self.underlying:
                    underlying_facs.append(fac)
                    continue
            elif fac in self.underlying:
                underlying_facs.append(fac)
                continue
        load_col = pmss_details['Load'].col
        for h in range(len(pmss_data[load_col])):
            load_h = pmss_data[load_col][h] * pmss_details['Load'].multiplier
            shortfall[h] = load_h
            for fac in fac_tml.keys():
                if fac in underlying_facs:
                    continue
                shortfall[h] -= pmss_data[pmss_details[fac].col][h] * pmss_details[fac].multiplier
            if shortfall[h] >= 0:
                alloc = 1.
            else:
                alloc = load_h / (load_h - shortfall[h])
            for fac in fac_tml.keys():
                if fac in underlying_facs:
                    fac_tml[fac] += pmss_data[pmss_details[fac].col][h] * pmss_details[fac].multiplier
                else:
                    fac_tml[fac] += pmss_data[pmss_details[fac].col][h] * pmss_details[fac].multiplier * alloc
            line = ''
        fac_tml_sum = 0
        for fac in fac_tml.keys():
            fac_tml_sum += fac_tml[fac]
        if self.show_correlation:
            col = pmss_details['Load'].col
            if pmss_details['Load'].multiplier == 1:
                df1 = pmss_data[col]
            else:
                tgt = []
                for h in range(len(pmss_data[col])):
                    tgt.append(pmss_data[col][h] * pmss_details['Load'].multiplier)
                df1 = tgt
            corr_src = []
            for h in range(len(shortfall)):
                if shortfall[h] < 0:
                    corr_src.append(pmss_data[col][h])
                else:
                    corr_src.append(pmss_data[col][h] - shortfall[h])
            try:
                corr = np.corrcoef(df1, corr_src)
                if np.isnan(corr.item((0, 1))):
                    corr = 0
                else:
                    corr = corr.item((0, 1))
            except:
                corr = 0
            corr_data = [['Correlation To Load']]
            corr_data.append(['RE Contribution', corr])
        else:
            corr_data = None
        if option == D:
            wb = oxl.Workbook()
            ns = wb.active
            ns.title = 'Detail'
            normal = oxl.styles.Font(name='Arial')
            bold = oxl.styles.Font(name='Arial', bold=True)
            ss = wb.create_sheet('Summary', 0)
            ns_re_sum = '=('
            ns_tml_sum = '=('
            ns_sto_sum = ''
            ns_loss_sum = ''
            ns_not_sum = ''
            cap_row = 1
            ns.cell(row=cap_row, column=2).value = 'Capacity (MW/MWh)' #headers[1].replace('\n', ' ')
            ss.row_dimensions[3].height = 40
            ss.cell(row=3, column=st_fac+1).value = headers[st_fac] # facility
            ss.cell(row=3, column=st_cap+1).value = headers[st_cap] # capacity
            ini_row = 2
            ns.cell(row=ini_row, column=2).value = 'Initial Capacity'
            tml_row = 3
            ns.cell(row=tml_row, column=2).value = headers[st_tml].replace('\n', ' ')
            ss.cell(row=3, column=st_tml+1).value = headers[st_tml] # to meet load
            sum_row = 4
            ns.cell(row=sum_row, column=2).value = headers[st_sub].replace('\n', ' ')
            ss.cell(row=3, column=st_sub+1).value = headers[st_sub] # subtotal MWh
            cf_row = 5
            ns.cell(row=cf_row, column=2).value = headers[st_cfa].replace('\n', ' ')
            ss.cell(row=3, column=st_cfa+1).value = headers[st_cfa] # CF
            cost_row = 6
            ns.cell(row=cost_row, column=2).value = headers[st_cst].replace('\n', ' ')
            ss.cell(row=3, column=st_cst+1).value = headers[st_cst] # Cost / yr
            lcoe_row = 7
            ns.cell(row=lcoe_row, column=2).value = headers[st_lcg].replace('\n', ' ')
            ss.cell(row=3, column=st_lcg+1).value = headers[st_lcg] # LCOG
            ss.cell(row=3, column=st_lco+1).value = headers[st_lco] # LCOE
            emi_row = 8
            ns.cell(row=emi_row, column=2).value = headers[st_emi].replace('\n', ' ')
            ss.cell(row=3, column=st_emi+1).value = headers[st_emi] # emissions
            ss.cell(row=3, column=st_emc+1).value = headers[st_emc] # emissions cost
            ss.cell(row=3, column=st_lcc+1).value = headers[st_lcc] # LCOE with CO2
            ss.cell(row=3, column=st_max+1).value = headers[st_max] # max. MWh
            ss.cell(row=3, column=st_bal+1).value = headers[st_bal] # max. balance
            ss.cell(row=3, column=st_cac+1).value = headers[st_cac] # capital cost
            ss.cell(row=3, column=st_lic+1).value = headers[st_lic] # lifetime cost
            ss.cell(row=3, column=st_lie+1).value = headers[st_lie] # lifetime emissions
            ss.cell(row=3, column=st_lec+1).value = headers[st_lec] # lifetime emissions cost
            ss.cell(row=3, column=st_are+1).value = headers[st_are] # area
            ss.cell(row=3, column=st_rlc+1).value = headers[st_rlc] # reference lcoe
            ss.cell(row=3, column=st_rcf+1).value = headers[st_rcf] # reference cf
            ss_row = 3
            ss_re_fst_row = 4
            fall_row = 9
            ns.cell(row=fall_row, column=2).value = 'Shortfall periods'
            max_row = 10
            ns.cell(row=max_row, column=2).value = 'Maximum (MW/MWh)'
            hrs_row = 11
            ns.cell(row=hrs_row, column=2).value = 'Hours of usage'
            if do_zone:
                zone_row = 12
                what_row = 13
                hrows = 14
                ns.cell(row=zone_row, column=1).value = 'Zone'
            else:
                what_row = 12
                hrows = 13
            ns.cell(row=what_row, column=1).value = 'Hour'
            ns.cell(row=what_row, column=2).value = 'Period'
            ns.cell(row=what_row, column=3).value = 'Load'
            ns.cell(row=sum_row, column=3).value = '=SUM(' + ssCol(3) + str(hrows) + \
                                                   ':' + ssCol(3) + str(hrows + 8759) + ')'
            ns.cell(row=sum_row, column=3).number_format = '#,##0'
            ns.cell(row=max_row, column=3).value = '=MAX(' + ssCol(3) + str(hrows) + \
                                                   ':' + ssCol(3) + str(hrows + 8759) + ')'
            ns.cell(row=max_row, column=3).number_format = '#,##0.00'
            o = 4
            col = 3
            # hour, period
            for row in range(hrows, 8760 + hrows):
                ns.cell(row=row, column=1).value = row - hrows + 1
                ns.cell(row=row, column=2).value = format_period(row - hrows)
            # and load
            load_col = pmss_details['Load'].col
            if pmss_details['Load'].multiplier == 1:
                for row in range(hrows, 8760 + hrows):
                    ns.cell(row=row, column=3).value = pmss_data[load_col][row - hrows]
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            else:
                for row in range(hrows, 8760 + hrows):
                    ns.cell(row=row, column=3).value = pmss_data[load_col][row - hrows] * \
                            pmss_details['Load'].multiplier
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            # here we're processing renewables (so no storage)
            for fac in re_order:
                if fac == 'Load':
                    continue
                if fac in underlying_facs:
                    continue
                if pmss_details[fac].col <= 0:
                    continue
                ss_row += 1
                col = do_detail(fac, col, ss_row)
                ns_tml_sum, ns_re_sum = do_detail_summary(fac, col, ss_row, ns_tml_sum, ns_re_sum)
            ss_re_lst_row = ss_row
            col += 1
            shrt_col = col
            ns.cell(row=fall_row, column=shrt_col).value = '=COUNTIF(' + ssCol(shrt_col) \
                            + str(hrows) + ':' + ssCol(shrt_col) + str(hrows + 8759) + \
                            ',"' + sf_test[0] + '0")'
            ns.cell(row=fall_row, column=shrt_col).number_format = '#,##0'
            ns.cell(row=what_row, column=shrt_col).value = 'Shortfall (' + sf_sign[0] \
                    + ') /\nSurplus (' + sf_sign[1] + ')'
            ns.cell(row=max_row, column=shrt_col).value = '=MAX(' + ssCol(shrt_col) + str(hrows) + \
                                           ':' + ssCol(shrt_col) + str(hrows + 8759) + ')'
            ns.cell(row=max_row, column=shrt_col).number_format = '#,##0.00'
            for col in range(3, shrt_col + 1):
                ns.cell(row=what_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                        vertical='bottom', horizontal='center')
                ns.cell(row=row, column=shrt_col).value = shortfall[row - hrows] * -self.surplus_sign
                for col in range(3, shrt_col + 1):
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            for row in range(hrows, 8760 + hrows):
                ns.cell(row=row, column=shrt_col).value = shortfall[row - hrows] * -self.surplus_sign
                ns.cell(row=row, column=col).number_format = '#,##0.00'
            col = shrt_col + 1
            ns.cell(row=tml_row, column=col).value = '=SUM(' + ssCol(col) + str(hrows) + \
                                                   ':' + ssCol(col) + str(hrows + 8759) + ')'
            ns.cell(row=tml_row, column=col).number_format = '#,##0'
            ns.cell(row=max_row, column=col).value = '=MAX(' + ssCol(col) + str(hrows) + \
                                           ':' + ssCol(col) + str(hrows + 8759) + ')'
            ns.cell(row=max_row, column=col).number_format = '#,##0.00'
            ns.cell(row=hrs_row, column=col).value = '=COUNTIF(' + ssCol(col) + str(hrows) + \
                                           ':' + ssCol(col) + str(hrows + 8759) + ',">0")'
            ns.cell(row=hrs_row, column=col).number_format = '#,##0'
            ns.cell(row=what_row, column=col).value = 'RE Contrib.\nto Load'
            ns.cell(row=what_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
            for row in range(hrows, 8760 + hrows):
                if shortfall[row - hrows] < 0:
                    if pmss_details['Load'].multiplier == 1:
                        rec = pmss_data[load_col][row - hrows]
                    else:
                        rec = pmss_data[load_col][row - hrows] * pmss_details['Load'].multiplier
                else:
                    if pmss_details['Load'].multiplier == 1:
                        rec = pmss_data[load_col][row - hrows] - shortfall[row - hrows]
                    else:
                        rec = pmss_data[load_col][row - hrows] * pmss_details['Load'].multiplier - \
                              shortfall[row - hrows]
                ns.cell(row=row, column=col).value = rec
               # the following formula will do the same computation
               # ns.cell(row=row, column=col).value = '=IF(' + ssCol(shrt_col) + str(row) + '>0,' + \
               #                            ssCol(3) + str(row) + ',' + ssCol(3) + str(row) + \
               #                            '+' + ssCol(shrt_col) + str(row) + ')'
                ns.cell(row=row, column=col).number_format = '#,##0.00'
          #  shrt_col += 1
           # col = shrt_col + 1
            ul_re_sum = ns_re_sum
            ul_tml_sum = ns_tml_sum
            nsul_sums = ['C']
            nsul_sum_cols = [3]
            for fac in underlying_facs:
                if pmss_details[fac].capacity * pmss_details[fac].multiplier == 0:
                    continue
                col = do_detail(fac, col, -1)
                nsul_sums.append(ssCol(col))
                nsul_sum_cols.append(col)
            if col > shrt_col + 1: # underlying
                col += 1
                ns.cell(row=what_row, column=col).value = 'Underlying\nLoad'
                ns.cell(row=what_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                        vertical='bottom', horizontal='center')
                ns.cell(row=sum_row, column=col).value = '=SUM(' + ssCol(col) + str(hrows) + \
                                                         ':' + ssCol(col) + str(hrows + 8759) + ')'
                ns.cell(row=sum_row, column=col).number_format = '#,##0'
                ns.cell(row=max_row, column=col).value = '=MAX(' + ssCol(col) + str(hrows) + \
                                                         ':' + ssCol(col) + str(hrows + 8759) + ')'
                ns.cell(row=max_row, column=col).number_format = '#,##0.00'
                for row in range(hrows, 8760 + hrows):
                    txt = '='
                    for c in nsul_sums:
                        txt += c + str(row) + '+'
                    ns.cell(row=row, column=col).value = txt[:-1]
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            next_col = col
            col += 1
        else: # O, O1, B, T
            sp_data = []
            sp_load = 0. # load from load curve
            hrows = 10
            load_max = 0
            load_hr = 0
            tml = 0.
            try:
                load_col = pmss_details['Load'].col
            except:
                load_col = 0
            if (option == B or option == T) and len(underlying_facs) > 0:
                load_facs = underlying_facs[:]
                load_facs.insert(0, 'Load')
                for h in range(len(pmss_data[load_col])):
                    amt = 0
                    for fac in load_facs:
                        amt += pmss_data[pmss_details[fac].col][h] * pmss_details[fac].multiplier
                    if amt > load_max:
                        load_max = amt
                        load_hr = h
                    sp_load += amt
                underlying_facs = []
            else:
                fac = 'Load'
                sp_load = sum(pmss_data[load_col]) * pmss_details[fac].multiplier
                for h in range(len(pmss_data[load_col])):
                    amt = pmss_data[load_col][h] * pmss_details[fac].multiplier
                    if amt > load_max:
                        load_max = amt
                        load_hr = h
            for fac in re_order:
                if fac == 'Load' or fac in underlying_facs:
                    continue
                if pmss_details[fac].capacity * pmss_details[fac].multiplier == 0:
                    continue
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = fac
                sp_d[st_cap] = pmss_details[fac].capacity * pmss_details[fac].multiplier
                try:
                    sp_d[st_tml] = fac_tml[fac]
                except:
                    pass
                sp_d[st_sub] = sum(pmss_data[pmss_details[fac].col]) * pmss_details[fac].multiplier
                sp_d[st_max] = max(pmss_data[pmss_details[fac].col]) * pmss_details[fac].multiplier
                sp_data.append(sp_d)
       #     for h in range(len(shortfall)):
        #        if shortfall[h] < 0:
         #           tml += pmss_data[load_col][h] * pmss_details['Load'].multiplier
          #      else:
           #         tml += pmss_data[load_col][h] * pmss_details['Load'].multiplier - shortfall[h]
        if option not in [O, O1, B, T]:
            self.listener.progress_bar.setValue(6)
            if self.event_callback:
                self.event_callback()
        storage_names = []
        # find any minimum generation for generators
        short_taken = {}
        short_taken_tot = 0
        for gen in dispatch_order:
            if pmss_details[gen].fac_type == 'G': # generators
                try:
                    const = self.generators[gen].constraint
                except:
                    try:
                        g2 = gen[gen.find('.') + 1:]
                        const = self.generators[g2].constraint
                    except:
                        continue
                if self.constraints[const].capacity_min != 0:
                    try:
                        short_taken[gen] = pmss_details[gen].capacity * pmss_details[gen].multiplier * \
                            self.constraints[const].capacity_min
                    except:
                        short_taken[gen] = pmss_details[gen].capacity * \
                            self.constraints[const].capacity_min
                    short_taken_tot += short_taken[gen]
                    for row in range(8760):
                        shortfall[row] = shortfall[row] - short_taken[gen]
        tot_sto_loss = 0.
        for gen in dispatch_order:
         #   min_after = [0, 0, -1, 0, 0, 0] # initial, low balance, period, final, low after, period
         #  Min_after is there to see if storage is as full at the end as at the beginning
            try:
                capacity = pmss_details[gen].capacity * pmss_details[gen].multiplier
            except:
                try:
                    capacity = pmss_details[gen].capacity
                except:
                    continue
            if gen not in self.generators.keys():
                continue
            if self.generators[gen].constraint in self.constraints and \
              self.constraints[self.generators[gen].constraint].category == 'Storage': # storage
                storage_names.append(gen)
                storage = [0., 0., 0., 0.] # capacity, initial, min level, max drain
                storage[0] = capacity
                if option == D:
                    ns.cell(row=cap_row, column=col + 2).value = capacity
                    ns.cell(row=cap_row, column=col + 2).number_format = '#,##0.00'
                try:
                    storage[1] = self.generators[gen].initial * pmss_details[gen].multiplier
                except:
                    storage[1] = self.generators[gen].initial
                if self.constraints[self.generators[gen].constraint].capacity_min > 0:
                    storage[2] = capacity * self.constraints[self.generators[gen].constraint].capacity_min
                if self.constraints[self.generators[gen].constraint].capacity_max > 0:
                    storage[3] = capacity * self.constraints[self.generators[gen].constraint].capacity_max
                else:
                    storage[3] = capacity
                recharge = [0., 0.] # cap, loss
                if self.constraints[self.generators[gen].constraint].recharge_max > 0:
                    recharge[0] = capacity * self.constraints[self.generators[gen].constraint].recharge_max
                else:
                    recharge[0] = capacity
                if self.constraints[self.generators[gen].constraint].recharge_loss > 0:
                    recharge[1] = self.constraints[self.generators[gen].constraint].recharge_loss
                discharge = [0., 0.] # cap, loss
                if self.constraints[self.generators[gen].constraint].discharge_max > 0:
                    discharge[0] = capacity * self.constraints[self.generators[gen].constraint].discharge_max
                if self.constraints[self.generators[gen].constraint].discharge_loss > 0:
                    discharge[1] = self.constraints[self.generators[gen].constraint].discharge_loss
                if self.constraints[self.generators[gen].constraint].parasitic_loss > 0:
                    parasite = self.constraints[self.generators[gen].constraint].parasitic_loss / 24.
                else:
                    parasite = 0.
                in_run = [False, False]
                min_runtime = self.constraints[self.generators[gen].constraint].min_runtime
                in_run[0] = True # start off in_run
                if min_runtime > 0 and self.generators[gen].initial == 0:
                    in_run[0] = False
                warm_time = self.constraints[self.generators[gen].constraint].warm_time
                storage_carry = storage[1] # self.generators[gen].initial
                if option == D:
                    ns.cell(row=ini_row, column=col + 2).value = storage_carry
                    ns.cell(row=ini_row, column=col + 2).number_format = '#,##0.00'
                storage_bal = []
                storage_can = 0.
                use_max = [0, None]
                sto_max = storage_carry
                for row in range(8760):
                    storage_loss = 0.
                    storage_losses = 0.
                    if storage_carry > 0:
                        loss = storage_carry * parasite
                        # for later: record parasitic loss
                        storage_carry = storage_carry - loss
                        storage_losses -= loss
                    if shortfall[row] < 0:  # excess generation
                        if min_runtime > 0:
                            in_run[0] = False
                        if warm_time > 0:
                            in_run[1] = False
                        can_use = - (storage[0] - storage_carry) * (1 / (1 - recharge[1]))
                        if can_use < 0: # can use some
                            if shortfall[row] > can_use:
                                can_use = shortfall[row]
                            if can_use < - recharge[0] * (1 / (1 - recharge[1])):
                                can_use = - recharge[0]
                        else:
                            can_use = 0.
                        # for later: record recharge loss
                        storage_losses += can_use * recharge[1]
                        storage_carry -= (can_use * (1 - recharge[1]))
                        shortfall[row] -= can_use
                        if corr_data is not None:
                            corr_src[row] += can_use
                    else: # shortfall
                        if min_runtime > 0 and shortfall[row] > 0:
                            if not in_run[0]:
                                if row + min_runtime <= 8759:
                                    for i in range(row + 1, row + min_runtime + 1):
                                        if shortfall[i] <= 0:
                                            break
                                    else:
                                        in_run[0] = True
                        if in_run[0]:
                            can_use = shortfall[row] * (1 / (1 - discharge[1]))
                            can_use = min(can_use, discharge[0])
                            if can_use > storage_carry - storage[2]:
                                can_use = storage_carry - storage[2]
                            if warm_time > 0 and not in_run[1]:
                                in_run[1] = True
                                can_use = can_use * (1 - warm_time)
                        else:
                            can_use = 0
                        if can_use > 0:
                            storage_loss = can_use * discharge[1]
                            storage_losses -= storage_loss
                            storage_carry -= can_use
                            can_use = can_use - storage_loss
                            shortfall[row] -= can_use
                            if corr_data is not None:
                                corr_src[row] += can_use
                            if storage_carry < 0:
                                storage_carry = 0
                        else:
                            can_use = 0.
                    if can_use < 0:
                        if use_max[1] is None or can_use < use_max[1]:
                            use_max[1] = can_use
                    elif can_use > use_max[0]:
                        use_max[0] = can_use
                    storage_bal.append(storage_carry)
                    if storage_bal[-1] > sto_max:
                        sto_max = storage_bal[-1]
                    if option == D:
                        if can_use > 0:
                            ns.cell(row=row + hrows, column=col).value = 0
                            ns.cell(row=row + hrows, column=col + 2).value = can_use * self.surplus_sign
                        else:
                            ns.cell(row=row + hrows, column=col).value = can_use * -self.surplus_sign
                            ns.cell(row=row + hrows, column=col + 2).value = 0
                        ns.cell(row=row + hrows, column=col + 1).value = storage_losses
                        ns.cell(row=row + hrows, column=col + 3).value = storage_carry
                        ns.cell(row=row + hrows, column=col + 4).value = (shortfall[row] + short_taken_tot) * -self.surplus_sign
                        for ac in range(5):
                            ns.cell(row=row + hrows, column=col + ac).number_format = '#,##0.00'
                            ns.cell(row=max_row, column=col + ac).value = '=MAX(' + ssCol(col + ac) + \
                                    str(hrows) + ':' + ssCol(col + ac) + str(hrows + 8759) + ')'
                            ns.cell(row=max_row, column=col + ac).number_format = '#,##0.00'
                    else:
                        tot_sto_loss += storage_losses
                        if can_use > 0:
                            storage_can += can_use
                if option == D:
                    ns.cell(row=sum_row, column=col).value = '=SUMIF(' + ssCol(col) + \
                            str(hrows) + ':' + ssCol(col) + str(hrows + 8759) + ',">0")'
                    ns.cell(row=sum_row, column=col).number_format = '#,##0'
                    ns.cell(row=sum_row, column=col + 1).value = '=SUMIF(' + ssCol(col + 1) + \
                            str(hrows) + ':' + ssCol(col + 1) + str(hrows + 8759) + ',"<0")'
                    ns.cell(row=sum_row, column=col + 1).number_format = '#,##0'
                    ns.cell(row=sum_row, column=col + 2).value = '=SUMIF(' + ssCol(col + 2) + \
                            str(hrows) + ':' + ssCol(col + 2) + str(hrows + 8759) + ',">0")'
                    ns.cell(row=sum_row, column=col + 2).number_format = '#,##0'
                    ns.cell(row=cf_row, column=col + 2).value = '=IF(' + ssCol(col + 2) + str(cap_row) + '>0,' + \
                            ssCol(col + 2) + str(sum_row) + '/' + ssCol(col + 2) + '1/8760,"")'
                    ns.cell(row=cf_row, column=col + 2).number_format = '#,##0.0%'
                    ns.cell(row=max_row, column=col).value = '=MAX(' + ssCol(col) + \
                            str(hrows) + ':' + ssCol(col) + str(hrows + 8759) + ')'
                    ns.cell(row=max_row, column=col).number_format = '#,##0.00'
                    ns.cell(row=hrs_row, column=col + 2).value = '=COUNTIF(' + ssCol(col + 2) + \
                            str(hrows) + ':' + ssCol(col + 2) + str(hrows + 8759) + ',">0")'
                    ns.cell(row=hrs_row, column=col + 2).number_format = '#,##0'
                    ns.cell(row=hrs_row, column=col + 3).value = '=' + ssCol(col + 2) + \
                            str(hrs_row) + '/8760'
                    ns.cell(row=hrs_row, column=col + 3).number_format = '#,##0.0%'
                    col += 5
                else:
                    if storage[0] == 0:
                        continue
               #     tml_tot += storage_can
                    sp_d = [' '] * len(headers)
                    sp_d[st_fac] = gen
                    sp_d[st_cap] = storage[0]
                    sp_d[st_tml] = storage_can
                    sp_d[st_max] = use_max[0]
                    sp_d[st_bal] = sto_max
                    sp_data.append(sp_d)
            else: # generator
                try:
                    if self.constraints[self.generators[gen].constraint].capacity_max > 0:
                        cap_capacity = capacity * self.constraints[self.generators[gen].constraint].capacity_max
                    else:
                        cap_capacity = capacity
                except:
                    cap_capacity = capacity
                if gen in short_taken.keys():
                    for row in range(8760):
                        shortfall[row] = shortfall[row] + short_taken[gen]
                    short_taken_tot -= short_taken[gen]
                    min_gen = short_taken[gen]
                else:
                    min_gen = 0
                if option == D:
                    ns.cell(row=cap_row, column=col).value = capacity
                    ns.cell(row=cap_row, column=col).number_format = '#,##0.00'
                    for row in range(8760):
                        if shortfall[row] >= 0: # shortfall?
                            if shortfall[row] >= cap_capacity:
                                shortfall[row] = shortfall[row] - cap_capacity
                                ns.cell(row=row + hrows, column=col).value = cap_capacity
                            elif shortfall[row] < min_gen:
                                ns.cell(row=row + hrows, column=col).value = min_gen
                                shortfall[row] -= min_gen
                            else:
                                ns.cell(row=row + hrows, column=col).value = shortfall[row]
                                shortfall[row] = 0
                        else:
                            shortfall[row] -= min_gen
                            ns.cell(row=row + hrows, column=col).value = min_gen
                        ns.cell(row=row + hrows, column=col + 1).value = (shortfall[row] + short_taken_tot) * -self.surplus_sign
                        ns.cell(row=row + hrows, column=col).number_format = '#,##0.00'
                        ns.cell(row=row + hrows, column=col + 1).number_format = '#,##0.00'
                    ns.cell(row=sum_row, column=col).value = '=SUM(' + ssCol(col) + str(hrows) + \
                            ':' + ssCol(col) + str(hrows + 8759) + ')'
                    ns.cell(row=sum_row, column=col).number_format = '#,##0'
                    ns.cell(row=cf_row, column=col).value = '=IF(' + ssCol(col) + str(cap_row) + '>0,' + \
                            ssCol(col) + str(sum_row) + '/' + ssCol(col) + str(cap_row) + '/8760,"")'
                    ns.cell(row=cf_row, column=col).number_format = '#,##0.0%'
                    ns.cell(row=max_row, column=col).value = '=MAX(' + ssCol(col) + \
                                str(hrows) + ':' + ssCol(col) + str(hrows + 8759) + ')'
                    ns.cell(row=max_row, column=col).number_format = '#,##0.00'
                    ns.cell(row=hrs_row, column=col).value = '=COUNTIF(' + ssCol(col) + \
                            str(hrows) + ':' + ssCol(col) + str(hrows + 8759) + ',">0")'
                    ns.cell(row=hrs_row, column=col).number_format = '#,##0'
                    ns.cell(row=hrs_row, column=col + 1).value = '=' + ssCol(col) + \
                            str(hrs_row) + '/8760'
                    ns.cell(row=hrs_row, column=col + 1).number_format = '#,##0.0%'
                    col += 2
                else:
                    gen_can = 0.
                    gen_max = 0
                    for row in range(8760):
                        if shortfall[row] >= 0: # shortfall?
                            if shortfall[row] >= cap_capacity:
                                shortfall[row] = shortfall[row] - cap_capacity
                                gen_can += cap_capacity
                                if cap_capacity > gen_max:
                                    gen_max = cap_capacity
                            elif shortfall[row] < min_gen:
                                gen_can += min_gen
                                if min_gen > gen_max:
                                    gen_max = min_gen
                                shortfall[row] -= min_gen
                            else:
                                gen_can += shortfall[row]
                                if shortfall[row] > gen_max:
                                    gen_max = shortfall[row]
                                shortfall[row] = 0
                        else:
                            if min_gen > gen_max:
                                gen_max = min_gen
                            gen_can += min_gen
                            shortfall[row] -= min_gen # ??
                    if capacity == 0:
                        continue
                    sp_d = [' '] * len(headers)
                    sp_d[st_fac] = gen
                    sp_d[st_cap] = capacity
                    sp_d[st_tml] = gen_can
                    sp_d[st_sub] = gen_can
                    sp_d[st_max] = gen_max
                    sp_data.append(sp_d)
#        if option == D: # Currently calculated elsewhere
#            if self.surplus_sign > 0:
#                maxmin = 'MIN'
#            else:
#                maxmin = 'MAX'
#            ns.cell(row=max_row, column=col-1).value = '=' + maxmin + '(' + \
#                    ssCol(col-1) + str(hrows) + ':' + ssCol(col - 1) + str(hrows + 8759) + ')'
#            ns.cell(row=max_row, column=col-1).number_format = '#,##0.00'
        if option not in [O, O1, B, T]:
            self.listener.progress_bar.setValue(8)
            if self.event_callback:
                self.event_callback()
        if corr_data is not None:
            try:
                corr = np.corrcoef(df1, corr_src)
                if np.isnan(corr.item((0, 1))):
                    corr = 0
                else:
                    corr = corr.item((0, 1))
            except:
                corr = 0
            corr_data.append(['RE plus Storage', corr])
            col = pmss_details['Load'].col
            corr_src = []
            for h in range(len(shortfall)):
                if shortfall[h] < 0:
                    corr_src.append(pmss_data[col][h])
                else:
                    corr_src.append(pmss_data[col][h] - shortfall[h])
            try:
                corr = np.corrcoef(df1, corr_src)
                if np.isnan(corr.item((0, 1))):
                    corr = 0
                else:
                    corr = corr.item((0, 1))
            except:
                corr = 0
            corr_data.append(['To Meet Load', corr])
            for c in range(1, len(corr_data)):
                if abs(corr_data[c][1]) < 0.1:
                    corr_data[c].append('None')
                elif abs(corr_data[c][1]) < 0.3:
                    corr_data[c].append('Little if any')
                elif abs(corr_data[c][1]) < 0.5:
                    corr_data[c].append('Low')
                elif abs(corr_data[c][1]) < 0.7:
                    corr_data[c].append('Moderate')
                elif abs(corr_data[c][1]) < 0.9:
                    corr_data[c].append('High')
                else:
                    corr_data[c].append('Very high')
        if option != D:
            load_col = pmss_details['Load'].col
            cap_sum = 0.
            gen_sum = 0.
            re_sum = 0.
            tml_sum = 0.
            ff_sum = 0.
            sto_sum = 0.
            cost_sum = 0.
            co2_sum = 0.
            co2_cost_sum = 0.
            capex_sum = 0.
            lifetime_sum = 0.
            lifetime_co2_sum = 0.
            lifetime_co2_cost = 0.
            total_area = 0.
            for sp in range(len(sp_data)):
                gen = sp_data[sp][st_fac]
                if gen in storage_names:
                    sto_sum += sp_data[sp][2]
                else:
                    try:
                        gen2 = gen[gen.find('.') + 1:]
                    except:
                        gen2 = gen
                    if gen in tech_names or gen2 in tech_names:
                        re_sum += sp_data[sp][st_sub]
            for sp in range(len(sp_data)):
                gen = sp_data[sp][st_fac]
                if gen in storage_names:
                    ndx = 2
                else:
                    if gen in self.generators.keys():
                        pass
                    else:
                        try:
                            gen = gen[gen.find('.') + 1:]
                        except:
                            pass
                    ndx = 3
                try:
                    if sp_data[sp][st_cap] > 0:
                        cap_sum += sp_data[sp][st_cap]
                        if self.generators[gen].lcoe > 0:
                            sp_data[sp][st_cfa] = sp_data[sp][ndx] / sp_data[sp][st_cap] / 8760 # need number for now
                        else:
                            sp_data[sp][st_cfa] = '{:.1f}%'.format(sp_data[sp][ndx] / sp_data[sp][st_cap] / 8760 * 100)
                    gen_sum += sp_data[sp][st_sub]
                except:
                    pass
                try:
                    tml_sum += sp_data[sp][st_tml]
                except:
                    pass
                if gen not in self.generators.keys():
                    continue
                ndx = 3
                if gen in storage_names:
                    ndx = 2
                else:
                    try:
                        gen2 = gen[gen.find('.') + 1:]
                    except:
                        gen2 = gen
                    if gen not in tech_names and gen2 not in tech_names:
                        ff_sum += sp_data[sp][ndx]
                if self.generators[gen].capex > 0 or self.generators[gen].fixed_om > 0 \
                  or self.generators[gen].variable_om > 0 or self.generators[gen].fuel > 0:
                    if option != T and self.remove_cost and sp_data[sp][ndx] == 0:
                        sp_data[sp][st_cst] = 0
                        continue
                    capex = sp_data[sp][st_cap] * self.generators[gen].capex
                    capex_sum += capex
                    opex = sp_data[sp][st_cap] * self.generators[gen].fixed_om \
                           + sp_data[sp][ndx] * self.generators[gen].variable_om \
                           + sp_data[sp][ndx] * self.generators[gen].fuel
                    disc_rate = self.generators[gen].disc_rate
                    if disc_rate == 0:
                        disc_rate = self.discount_rate
                    lifetime = self.generators[gen].lifetime
                    sp_data[sp][st_lcg] = calcLCOE(sp_data[sp][ndx], capex, opex, disc_rate, lifetime)
                    sp_data[sp][st_cst] = sp_data[sp][ndx] * sp_data[sp][st_lcg]
                    # To prevent ZeroDivisionError
                    if (gen in tech_names or gen2 in tech_names) and (fac_tml_sum > 0):
                        sp_data[sp][st_lco] = sp_data[sp][st_cst] / (sp_data[sp][st_tml] + (sto_sum * sp_data[sp][st_tml] / fac_tml_sum))
                    else:
                        sp_data[sp][st_lco] = sp_data[sp][st_lcg]
                    cost_sum += sp_data[sp][st_cst]
                    sp_data[sp][st_cac] = capex
                elif self.generators[gen].lcoe > 0:
                    if option != T and self.remove_cost and sp_data[sp][ndx] == 0:
                        sp_data[sp][st_cst] = 0
                        continue
                    if self.generators[gen].lcoe_cf > 0:
                        lcoe_cf = self.generators[gen].lcoe_cf
                    else:
                        lcoe_cf = sp_data[sp][st_cfa]
                    sp_data[sp][st_cst] = self.generators[gen].lcoe * lcoe_cf * 8760 * sp_data[sp][st_cap]
                    if sp_data[sp][st_cfa] > 0:
                        sp_data[sp][st_lcg] = sp_data[sp][st_cst] / sp_data[sp][ndx]
                        sp_data[sp][st_lco] = sp_data[sp][st_lcg]
                    sp_data[sp][st_cfa] = '{:.1f}%'.format(sp_data[sp][st_cfa] * 100.)
                    cost_sum += sp_data[sp][st_cst]
                    sp_data[sp][st_rlc] = self.generators[gen].lcoe
                    sp_data[sp][st_rcf] = '{:.1f}%'.format(lcoe_cf * 100.)
                elif self.generators[gen].lcoe_cf == 0: # no cost facility
                    if option != T and self.remove_cost and sp_data[sp][ndx] == 0:
                        sp_data[sp][st_cst] = 0
                        continue
                    lcoe_cf = sp_data[sp][st_cfa]
                    sp_data[sp][st_cst] = 0
                    cost_sum += sp_data[sp][st_cst]
                sp_data[sp][st_lic] = sp_data[sp][st_cst] * max_lifetime
                lifetime_sum += sp_data[sp][st_lic]
                if self.generators[gen].emissions > 0 and sp_data[sp][st_tml]> 0:
                    sp_data[sp][st_emi] = sp_data[sp][ndx] * self.generators[gen].emissions
                    co2_sum += sp_data[sp][st_emi]
                    sp_data[sp][st_emc] = sp_data[sp][st_emi] * self.carbon_price
                    if sp_data[sp][st_cst] == 0:
                        sp_data[sp][st_lcc] = sp_data[sp][st_emc] / sp_data[sp][st_tml]
                    else:
                        sp_data[sp][st_lcc] = sp_data[sp][st_lco] * ((sp_data[sp][st_cst] + sp_data[sp][st_emc]) / sp_data[sp][st_cst])
                    co2_cost_sum += sp_data[sp][st_emc]
                    sp_data[sp][st_lie] = sp_data[sp][st_emi] * max_lifetime
                    lifetime_co2_sum += sp_data[sp][st_lie]
                    sp_data[sp][st_lec] = sp_data[sp][st_lie] * self.carbon_price
                    lifetime_co2_cost += sp_data[sp][st_lec]
                else:
                    sp_data[sp][st_lcc] = sp_data[sp][st_lco]
                if self.generators[gen].area > 0:
                    sp_data[sp][st_are] = sp_data[sp][st_cap] * self.generators[gen].area
                    total_area += sp_data[sp][st_are]
            sf_sums = [0., 0., 0.]
            for sf in range(len(shortfall)):
                if shortfall[sf] > 0:
                    sf_sums[0] += shortfall[sf]
                    sf_sums[2] += pmss_data[load_col][sf] * pmss_details['Load'].multiplier
                else:
                    sf_sums[1] += shortfall[sf]
                    sf_sums[2] += pmss_data[load_col][sf] * pmss_details['Load'].multiplier
            if gen_sum > 0:
                gs = cost_sum / gen_sum
            else:
                gs = ''
            if tml_sum > 0:
                gsw = cost_sum / tml_sum # LCOE
                gswc = (cost_sum + co2_cost_sum) / tml_sum
            else:
                gsw = ''
                gswc = ''
            if option == O or option == O1:
                load_pct, surp_pct, re_pct = summary_totals()
            else:
                summary_totals()
            do_underlying = False
            if len(underlying_facs) > 0:
                for fac in underlying_facs:
                    if pmss_details[fac].capacity * pmss_details[fac].multiplier > 0:
                        do_underlying = True
                        break
            if do_underlying:
                sp_data.append(' ')
                sp_data.append('Additional Underlying Load')
                for fac in underlying_facs:
                    if pmss_details[fac].capacity * pmss_details[fac].multiplier == 0:
                        continue
                    if fac in self.generators.keys():
                        gen = fac
                    else:
                        gen = pmss_details[fac].generator
                    col = pmss_details[fac].col
                    sp_d = [' '] * len(headers)
                    sp_d[st_fac] = fac
                    sp_d[st_cap] = pmss_details[fac].capacity * pmss_details[fac].multiplier
                    cap_sum += sp_d[st_cap]
                    sp_d[st_tml] = sum(pmss_data[pmss_details[fac].col]) * pmss_details[fac].multiplier
                    tml_sum += sp_d[st_tml]
                    sp_d[st_sub] = sp_d[st_tml]
                    gen_sum += sp_d[st_tml]
                    sp_load += sp_d[st_tml]
                    sp_d[st_cfa] = '{:.1f}%'.format(sp_d[st_sub] / sp_d[st_cap] / 8760 * 100.)
                    sp_d[st_max] = max(pmss_data[pmss_details[fac].col]) * pmss_details[fac].multiplier
                    if self.generators[gen].capex > 0 or self.generators[gen].fixed_om > 0 \
                      or self.generators[gen].variable_om > 0 or self.generators[gen].fuel > 0:
                        capex = sp_d[st_cap] * self.generators[gen].capex
                        capex_sum += capex
                        opex = sp_d[st_cap] * self.generators[gen].fixed_om \
                               + sp_d[st_tml] * self.generators[gen].variable_om \
                               + sp_d[st_tml] * self.generators[gen].fuel
                        disc_rate = self.generators[gen].disc_rate
                        if disc_rate == 0:
                            disc_rate = self.discount_rate
                        lifetime = self.generators[gen].lifetime
                        sp_d[st_lcg] = calcLCOE(sp_d[st_tml], capex, opex, disc_rate, lifetime)
                        sp_d[st_cst] = sp_d[st_tml] * sp_d[st_lcg]
                        cost_sum += sp_d[st_cst]
                        sp_d[st_lco] = sp_d[st_lcg]
                        sp_d[st_cac] = capex
                    elif self.generators[gen].lcoe > 0:
                        if self.generators[gen].lcoe_cf > 0:
                            lcoe_cf = self.generators[gen].lcoe_cf
                        else:
                            lcoe_cf = sp_d[st_cfa]
                        sp_d[st_cst] = self.generators[gen].lcoe * lcoe_cf * 8760 * sp_d[st_cap]
                        cost_sum += sp_d[st_cst]
                        if sp_d[st_cfa] > 0:
                            sp_d[st_lcg] = sp_d[st_cst] / sp_d[st_tml]
                            sp_d[st_lco] = sp_d[st_lcg]
                        sp_d[st_cfa] = '{:.1f}%'.format(sp_d[st_cfa] * 100.)
                        sp_d[st_rlc] = self.generators[gen].lcoe
                        sp_d[st_rcf] = '{:.1f}%'.format(lcoe_cf * 100.)
                    elif self.generators[gen].lcoe_cf == 0: # no cost facility
                        sp_d[st_cst] = 0
                        sp_d[st_lcg] = 0
                        sp_d[st_lco] = 0
                        sp_d[st_rlc] = self.generators[gen].lcoe
                    sp_d[st_lic] = sp_d[st_cst] * max_lifetime
                    lifetime_sum += sp_d[st_lic]
                    if self.generators[gen].emissions > 0:
                        sp_d[st_emi] = sp_d[st_tml] * self.generators[gen].emissions
                        co2_sum += sp_d[st_emi]
                        sp_d[st_emc] = sp_d[st_emi] * self.carbon_price
                        if sp_d[st_cst] > 0:
                            sp_d[st_lcc] = sp_d[st_lco] * ((sp_d[st_cst] + sp_d[st_emc]) / sp_d[st_cst])
                        else:
                            sp_d[st_lcc] = sp_d[st_emc] / sp_d[st_tml]
                        co2_cost_sum += sp_d[st_emc]
                        sp_d[st_lie] = sp_d[st_emi] * max_lifetime
                        lifetime_co2_sum += sp_d[st_lie]
                        sp_d[st_lec] = sp_d[st_lie] * self.carbon_price
                        lifetime_co2_cost += sp_d[st_lec]
                    else:
                        sp_d[st_lcc] = sp_d[st_lco]
                    if self.generators[gen].area > 0:
                        sp_d[st_are] = sp_d[st_cap] * self.generators[gen].area
                    sp_data.append(sp_d)
                if gen_sum > 0:
                    gs = cost_sum / gen_sum
                else:
                    gs = ''
                if tml_sum > 0:
                    gsw = cost_sum / tml_sum # LCOE
                    gswc = (cost_sum + co2_cost_sum) / tml_sum
                else:
                    gsw = ''
                    gswc = ''
                # find maximum underlying load
                if option == S:
                    load_max = 0
                    load_hr = 0
                    load_col = pmss_details['Load'].col
                    for h in range(len(pmss_data[load_col])):
                        amt = pmss_data[load_col][h] * pmss_details['Load'].multiplier
                        for fac in underlying_facs:
                            amt += pmss_data[pmss_details[fac].col][h] * pmss_details[fac].multiplier
                        if amt > load_max:
                            load_max = amt
                            load_hr = h
                summary_totals('Underlying ')
            if corr_data is not None:
                sp_data.append(' ')
                sp_data = sp_data + corr_data
            sp_data.append(' ')
            sp_data.append(['Static Variables'])
            if self.carbon_price > 0:
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = 'Carbon Price ($/tCO2e)'
                sp_d[st_cap] = self.carbon_price
                sp_data.append(sp_d)
            sp_d = [' '] * len(headers)
            sp_d[st_fac] = 'Lifetime (years)'
            sp_d[st_cap] = max_lifetime
            sp_data.append(sp_d)
            sp_d = [' '] * len(headers)
            sp_d[st_fac] = 'Discount Rate'
            sp_d[st_cap] = '{:.2%}'.format(self.discount_rate)
            sp_data.append(sp_d)
            if option == B or option == T:
                if self.optimise_debug:
                    sp_pts = [0] * len(headers)
                    for p in [st_cap, st_lcg, st_lco, st_lcc, st_max, st_bal, st_rlc, st_are]:
                        sp_pts[p] = 2
                    if corr_data is not None:
                        sp_pts[st_cap] = 3 # compromise between capacity (2) and correlation (4)
                return sp_data, corr_data
            if option == O or option == O1:
                op_load_tot = pmss_details['Load'].capacity * pmss_details['Load'].multiplier
                if gswc != '':
                    lcoe = gswc
                elif self.adjusted_lcoe:
                    lcoe = gsw # target is lcoe
                else:
                    lcoe = gs
                if gen_sum == 0:
                    re_pct = 0
                    load_pct = 0
                    re_pct = 0
                multi_value = {'lcoe': lcoe, #lcoe. lower better
                    'load_pct': load_pct, #load met. 100% better
                    'surplus_pct': surp_pct, #surplus. lower better
                    're_pct': re_pct, # RE pct. higher better
                    'cost': cost_sum, # cost. lower better
                    'co2': co2_sum} # CO2. lower better
                if option == O:
                    if multi_value['lcoe'] == '':
                        multi_value['lcoe'] = 0
                    return multi_value, sp_data, None
                else:
                    extra = [gsw, op_load_tot, sto_sum, re_sum, re_pct, sf_sums]
                    return multi_value, sp_data, extra
        #    list(map(list, list(zip(*sp_data))))
            span = None
            sp_pts = [0] * len(headers)
            for p in [st_cap, st_lcg, st_lco, st_lcc, st_max, st_bal, st_rlc, st_are]:
                sp_pts[p] = 2
            if corr_data is not None:
                sp_pts[st_cap] = 3 # compromise between capacity (2) and correlation (4)
            self.setStatus(sender_name + ' completed')
            if title is not None:
                atitle = title
            elif self.results_prefix != '':
                atitle = self.results_prefix + '_' + sender_name
            else:
                atitle = sender_name
            return sp_data, corr_data # finish if not detailed spreadsheet
        col = next_col + 1
        is_storage = False
        ss_sto_rows = []
        ss_st_row = -1
        for gen in dispatch_order:
            ss_row += 1
            try:
                if self.constraints[self.generators[gen].constraint].category == 'Storage':
                    ss_sto_rows.append(ss_row)
                    nc = 2
                    ns.cell(row=what_row, column=col).value = 'Charge\n' + gen
                    ns.cell(row=what_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                            vertical='bottom', horizontal='center')
                    ns.cell(row=what_row, column=col + 1).value = gen + '\nLosses'
                    ns.cell(row=what_row, column=col + 1).alignment = oxl.styles.Alignment(wrap_text=True,
                            vertical='bottom', horizontal='center')
                    is_storage = True
                    ns_sto_sum += '+' + ssCol(st_tml+1) + str(ss_row)
                    ns_loss_sum += '+Detail!' + ssCol(col + 1) + str(sum_row)
                else:
                    nc = 0
                    is_storage = False
                    ns_not_sum += '-' + ssCol(st_tml+1) + str(ss_row)
            except KeyError as err:
                msg = 'Key Error: No Constraint for ' + gen
                if title is not None:
                    msg += ' (model ' + title + ')'
                self.setStatus(msg)
                nc = 0
                is_storage = False
                ns_not_sum += '-' + ssCol(st_tml+1) + str(ss_row)
            ns.cell(row=what_row, column=col + nc).value = gen
            ss.cell(row=ss_row, column=st_fac+1).value = '=Detail!' + ssCol(col + nc) + str(what_row)
            # facility
            ss.cell(row=ss_row, column=st_cap+1).value = '=Detail!' + ssCol(col + nc) + str(cap_row)
            # capacity
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.00'
            # tml
            ss.cell(row=ss_row, column=st_tml+1).value = '=Detail!' + ssCol(col + nc) + str(sum_row)
            ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
            # subtotal
            try:
                if self.constraints[self.generators[gen].constraint].category != 'Storage':
                    ss.cell(row=ss_row, column=st_sub+1).value = '=Detail!' + ssCol(col + nc) + str(sum_row)
                    ss.cell(row=ss_row, column=st_sub+1).number_format = '#,##0'
            except KeyError as err:
                ss.cell(row=ss_row, column=st_sub+1).value = '=Detail!' + ssCol(col + nc) + str(sum_row)
                ss.cell(row=ss_row, column=st_sub+1).number_format = '#,##0'
            # cf
            ss.cell(row=ss_row, column=st_cfa+1).value = '=Detail!' + ssCol(col + nc) + str(cf_row)
            ss.cell(row=ss_row, column=st_cfa+1).number_format = '#,##0.0%'
            if self.generators[gen].capex > 0 or self.generators[gen].fixed_om > 0 \
              or self.generators[gen].variable_om > 0 or self.generators[gen].fuel > 0:
                disc_rate = self.generators[gen].disc_rate
                if disc_rate == 0:
                    disc_rate = self.discount_rate
                if disc_rate == 0:
                    cst_calc = '/' + str(self.generators[gen].lifetime)
                else:
                    pwr_calc = 'POWER(1+' + str(disc_rate) + ',' + str(self.generators[gen].lifetime) + ')'
                    cst_calc = '*' + str(disc_rate) + '*' + pwr_calc + '/SUM(' + pwr_calc + ',-1)'
                ns.cell(row=cost_row, column=col + nc).value = '=IF(' + ssCol(col + nc) + str(cf_row) + \
                        '>0,' + ssCol(col + nc) + str(cap_row) + '*' + str(self.generators[gen].capex) + \
                        cst_calc + '+' + ssCol(col + nc) + str(cap_row) + '*' + \
                        str(self.generators[gen].fixed_om) + '+' + ssCol(col + nc) + str(sum_row) + '*(' + \
                        str(self.generators[gen].variable_om) + '+' + str(self.generators[gen].fuel) + \
                        '),0)'
                ns.cell(row=cost_row, column=col + nc).number_format = '$#,##0'
                # cost / yr
                if self.remove_cost:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=IF(Detail!' + ssCol(col + nc) + str(sum_row) \
                            + '>0,Detail!' + ssCol(col + nc) + str(cost_row) + ',"")'
                else:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=Detail!' + ssCol(col + nc) + str(cost_row)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                ns.cell(row=lcoe_row, column=col + nc).value = '=IF(AND(' + ssCol(col + nc) + str(cf_row) + \
                        '>0,' + ssCol(col + nc) + str(cap_row) + '>0),' + ssCol(col + nc) + \
                        str(cost_row) + '/' + ssCol(col + nc) + str(sum_row) + ',"")'
                ns.cell(row=lcoe_row, column=col + nc).number_format = '$#,##0.00'
                # lcog
                ss.cell(row=ss_row, column=st_lcg+1).value = '=Detail!' + ssCol(col + nc) + str(lcoe_row)
                ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
                # lcoe
                ss.cell(row=ss_row, column=st_lco+1).value = '=Detail!' + ssCol(col + nc) + str(lcoe_row)
                ss.cell(row=ss_row, column=st_lco+1).number_format = '$#,##0.00'
                # capital cost
                ss.cell(row=ss_row, column=st_cac+1).value = '=IF(Detail!' + ssCol(col + nc) + str(cap_row) \
                                                            + '>0,Detail!' + ssCol(col + nc) + str(cap_row) + '*'  \
                                                            + str(self.generators[gen].capex) + ',"")'
                ss.cell(row=ss_row, column=st_cac+1).number_format = '$#,##0'
            elif self.generators[gen].lcoe > 0:
                ns.cell(row=cost_row, column=col + nc).value = '=IF(' + ssCol(col + nc) + str(cf_row) + \
                        '>0,' + ssCol(col + nc) + str(sum_row) + '*Summary!' + ssCol(st_rlc + 1) + str(ss_row) + \
                        '*Summary!' + ssCol(st_rcf + 1) + str(ss_row) + '/' + ssCol(col + nc) + str(cf_row) + ',0)'
                ns.cell(row=cost_row, column=col + nc).number_format = '$#,##0'
                # cost / yr
                if self.remove_cost:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=IF(Detail!' + ssCol(col + nc) + str(sum_row) \
                            + '>0,Detail!' + ssCol(col + nc) + str(cost_row) + ',"")'
                else:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=Detail!' + ssCol(col + nc) + str(cost_row)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                ns.cell(row=lcoe_row, column=col + nc).value = '=IF(AND(' + ssCol(col + nc) + str(cf_row) + '>0,' \
                            + ssCol(col + nc) + str(cap_row) + '>0),' + ssCol(col + nc) + str(cost_row) + '/8760/' \
                            + ssCol(col + nc) + str(cf_row) + '/' + ssCol(col + nc) + str(cap_row)+  ',"")'
                ns.cell(row=lcoe_row, column=col + nc).number_format = '$#,##0.00'
                # lcog
                ss.cell(row=ss_row, column=st_lcg+1).value = '=Detail!' + ssCol(col + nc) + str(lcoe_row)
                ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
                # lcoe
                ss.cell(row=ss_row, column=st_lco+1).value = '=Detail!' + ssCol(col + nc) + str(lcoe_row)
                ss.cell(row=ss_row, column=st_lco+1).number_format = '$#,##0.00'
                # ref lcoe
                ss.cell(row=ss_row, column=st_rlc+1).value = self.generators[gen].lcoe
                ss.cell(row=ss_row, column=st_rlc+1).number_format = '$#,##0.00'
                # ref cf
                if self.generators[gen].lcoe_cf == 0:
                    ss.cell(row=ss_row, column=st_rcf+1).value = '=' + ssCol(st_cfa+1) + str(ss_row)
                else:
                    ss.cell(row=ss_row, column=st_rcf+1).value = self.generators[gen].lcoe_cf
                ss.cell(row=ss_row, column=st_rcf+1).number_format = '#,##0.0%'
            elif self.generators[gen].lcoe_cf == 0: # no cost facility
                ns.cell(row=cost_row, column=col + nc).value = '=IF(' + ssCol(col + nc) + str(cf_row) + \
                        '>0,' + ssCol(col + nc) + str(sum_row) + '*Summary!' + ssCol(st_rlc + 1) + str(ss_row) + \
                        '*Summary!' + ssCol(st_rcf + 1) + str(ss_row) + '/' + ssCol(col + nc) + str(cf_row) + ',0)'
                ns.cell(row=cost_row, column=col + nc).number_format = '$#,##0'
                # cost / yr
                if self.remove_cost:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=IF(Detail!' + ssCol(col + nc) + str(sum_row) \
                            + '>0,Detail!' + ssCol(col + nc) + str(cost_row) + ',"")'
                else:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=Detail!' + ssCol(col + nc) + str(cost_row)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                ns.cell(row=lcoe_row, column=col + nc).value = '=IF(AND(' + ssCol(col + nc) + str(cf_row) + '>0,' \
                            + ssCol(col + nc) + str(cap_row) + '>0),' + ssCol(col + nc) + str(cost_row) + '/8760/' \
                            + ssCol(col + nc) + str(cf_row) + '/' + ssCol(col + nc) + str(cap_row)+  ',"")'
                ns.cell(row=lcoe_row, column=col + nc).number_format = '$#,##0.00'
                # lcog
                ss.cell(row=ss_row, column=st_lcg+1).value = '=Detail!' + ssCol(col + nc) + str(lcoe_row)
                ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
                # lcoe
                ss.cell(row=ss_row, column=st_lco+1).value = '=Detail!' + ssCol(col + nc) + str(lcoe_row)
                ss.cell(row=ss_row, column=st_lco+1).number_format = '$#,##0.00'
                # ref lcoe
                ss.cell(row=ss_row, column=st_rlc+1).value = self.generators[gen].lcoe
                ss.cell(row=ss_row, column=st_rlc+1).number_format = '$#,##0.00'
                # ref cf
                if self.generators[gen].lcoe_cf == 0:
                    ss.cell(row=ss_row, column=st_rcf+1).value = '=' + ssCol(st_cfa+1) + str(ss_row)
                else:
                    ss.cell(row=ss_row, column=st_rcf+1).value = self.generators[gen].lcoe_cf
                ss.cell(row=ss_row, column=st_rcf+1).number_format = '#,##0.0%'
            if self.generators[gen].emissions > 0:
                ns.cell(row=emi_row, column=col + nc).value = '=' + ssCol(col + nc) + str(sum_row) \
                        + '*' + str(self.generators[gen].emissions)
                ns.cell(row=emi_row, column=col + nc).number_format = '#,##0'
                # emissions
                if self.remove_cost:
                    ss.cell(row=ss_row, column=st_emi+1).value = '=IF(Detail!' + ssCol(col + nc) + str(sum_row) \
                            + '>0,Detail!' + ssCol(col + nc) + str(emi_row) + ',"")'
                else:
                    ss.cell(row=ss_row, column=st_emi+1).value = '=Detail!' + ssCol(col + nc) + str(emi_row)
                ss.cell(row=ss_row, column=st_emi+1).number_format = '#,##0'
                if self.carbon_price > 0:
                    ss.cell(row=ss_row, column=st_emc+1).value = '=IF(AND(' + ssCol(st_emi+1) + str(ss_row) + '<>"",' + \
                                                                 ssCol(st_emi+1) + str(ss_row) + '>0),' + \
                                                                 ssCol(st_emi+1) + str(ss_row) + '*carbon_price,"")'
                    ss.cell(row=ss_row, column=st_emc+1).number_format = '$#,##0'
            # max mwh
            ss.cell(row=ss_row, column=st_max+1).value = '=Detail!' + ssCol(col + nc) + str(max_row)
            ss.cell(row=ss_row, column=st_max+1).number_format = '#,##0.00'
            # max balance
            if nc > 0: # storage
                ss.cell(row=ss_row, column=st_bal+1).value = '=Detail!' + ssCol(col + nc + 1) + str(max_row)
                ss.cell(row=ss_row, column=st_bal+1).number_format = '#,##0.00'
            ns.cell(row=what_row, column=col + nc).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
            ns.cell(row=what_row, column=col + nc + 1).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
            if is_storage:
                # lifetime cost
                ss.cell(row=ss_row, column=st_lic+1).value = '=IF(Detail!' + ssCol(col + 2) + str(sum_row) \
                                                        + '>0,Detail!' + ssCol(col + 2) + str(cost_row) + '*lifetime,"")'
                ss.cell(row=ss_row, column=st_lic+1).number_format = '$#,##0'
                # ns.cell(row=what_row, column=col + 1).value = gen
                ns.cell(row=what_row, column=col + 3).value = gen + '\nBalance'
                ns.cell(row=what_row, column=col + 3).alignment = oxl.styles.Alignment(wrap_text=True,
                        vertical='bottom', horizontal='center')
                ns.cell(row=what_row, column=col + 4).value = 'After\n' + gen
                ns.cell(row=what_row, column=col + 4).alignment = oxl.styles.Alignment(wrap_text=True,
                        vertical='bottom', horizontal='center')
                ns.cell(row=fall_row, column=col + 4).value = '=COUNTIF(' + ssCol(col + 4) \
                        + str(hrows) + ':' + ssCol(col + 4) + str(hrows + 8759) + \
                        ',"' + sf_test[0] + '0")'
                ns.cell(row=fall_row, column=col + 4).number_format = '#,##0'
                col += 5
            else:
                # lifetime cost
                ss.cell(row=ss_row, column=st_lic+1).value = '=IF(Detail!' + ssCol(col) + str(sum_row) \
                                                        + '>0,Detail!' + ssCol(col) + str(cost_row) + '*lifetime,"")'
                ss.cell(row=ss_row, column=st_lic+1).number_format = '$#,##0'
                ns.cell(row=what_row, column=col + 1).value = 'After\n' + gen
                ns.cell(row=fall_row, column=col + 1).value = '=COUNTIF(' + ssCol(col + 1) \
                        + str(hrows) + ':' + ssCol(col + 1) + str(hrows + 8759) + \
                        ',"' + sf_test[0] + '0")'
                ns.cell(row=fall_row, column=col + 1).number_format = '#,##0'
                col += 2
            ss.cell(row=ss_row, column=st_lie+1).value = '=IF(AND(' + ssCol(st_emi+1) + str(ss_row) + '<>"",' + \
                                                         ssCol(st_emi+1) + str(ss_row) + '>0),' + \
                                                         ssCol(st_emi+1) + str(ss_row) + '*lifetime,"")'
            ss.cell(row=ss_row, column=st_lie+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_lec+1).value = '=IF(AND(' + ssCol(st_emi+1) + str(ss_row) + '<>"",' + \
                                                         ssCol(st_emi+1) + str(ss_row) + '>0),' + \
                                                         ssCol(st_emc+1) + str(ss_row) + '*lifetime,"")'
            ss.cell(row=ss_row, column=st_lec+1).number_format = '$#,##0'
        if is_storage:
            ns.cell(row=emi_row, column=col - 2).value = '=MIN(' + ssCol(col - 2) + str(hrows) + \
                    ':' + ssCol(col - 2) + str(hrows + 8759) + ')'
            ns.cell(row=emi_row, column=col - 2).number_format = '#,##0.00'
        for column_cells in ns.columns:
            length = 0
            value = ''
            row = 0
            sum_value = 0
            do_sum = False
            do_cost = False
            for cell in column_cells:
                if cell.row >= hrows:
                    if do_sum:
                        try:
                            sum_value += abs(cell.value)
                        except:
                            pass
                    else:
                        try:
                            value = str(round(cell.value, 2))
                            if len(value) > length:
                                length = len(value) + 2
                        except:
                            pass
                elif cell.row > 0:
                    if str(cell.value)[0] != '=':
                        values = str(cell.value).split('\n')
                        for value in values:
                            if cell.row == cost_row:
                                valf = value.split('.')
                                alen = int(len(valf[0]) * 1.6)
                            else:
                                alen = len(value) + 2
                            if alen > length:
                                length = alen
                    else:
                        if cell.row == cost_row:
                            do_cost = True
                        if cell.value[1:4] == 'SUM':
                            do_sum = True
            if sum_value > 0:
                alen = len(str(int(sum_value))) * 1.5
                if do_cost:
                    alen = int(alen * 1.5)
                if alen > length:
                    length = alen
            if isinstance(cell.column, int):
                cel = ssCol(cell.column)
            else:
                cel = cell.column
            ns.column_dimensions[cel].width = max(length, 10)
        ns.column_dimensions['A'].width = 6
        ns.column_dimensions['B'].width = 21
        st_row = hrows + 8760
        st_col = col
        for row in range(1, st_row):
            for col in range(1, st_col):
                try:
                    ns.cell(row=row, column=col).font = normal
                except:
                    pass
        self.listener.progress_bar.setValue(12)
        if self.event_callback:
            self.event_callback()
        ns.row_dimensions[what_row].height = 30
        ns.freeze_panes = 'C' + str(hrows)
        ns.activeCell = 'C' + str(hrows)
        if self.results_prefix != '':
            ss.cell(row=1, column=1).value = 'Powermatch - ' + self.results_prefix + ' Summary'
        else:
            ss.cell(row=1, column=1).value = 'Powermatch - Summary'
        ss.cell(row=1, column=1).font = bold
        ss_lst_row = ss_row + 1
        ss_row, ss_re_row = detail_summary_total(ss_row, base_row='4')
        if len(nsul_sum_cols) > 1: # if we have underlying there'll be more than one column
            ss_row += 2
            ss.cell(row=ss_row, column=1).value = 'Additional Underlying Load'
            ss.cell(row=ss_row, column=1).font = bold
            base_row = str(ss_row + 1)
            for col in nsul_sum_cols[1:]:
                ss_row += 1
                ul_tml_sum, ul_re_sum = do_detail_summary(fac, col, ss_row, ul_tml_sum, ul_re_sum)
            ul_fst_row = int(base_row)
            ul_lst_row = ss_row
            ns_re_sum = ul_re_sum
            ns_tml_sum = ul_tml_sum
            ss_row, ss_re_row = detail_summary_total(ss_row, title='Underlying ', base_row=base_row,
                                          back_row=str(ss_lst_row))
        wider = [ssCol(st_cac + 1), ssCol(st_lic + 1)]
        for column_cells in ss.columns:
            length = 0
            value = ''
            for cell in column_cells:
                if str(cell.value)[0] != '=':
                    values = str(cell.value).split('\n')
                    for value in values:
                        if len(value) + 1 > length:
                            length = len(value) + 1
            if isinstance(cell.column, int):
                cel = ssCol(cell.column)
            else:
                cel = cell.column
            if cel in wider:
                ss.column_dimensions[cel].width = max(length, 10) * 1.5
            else:
                ss.column_dimensions[cel].width = max(length, 10) * 1.2

        if corr_data is not None:
            ss_row += 2
            for corr in corr_data:
                ss.cell(row=ss_row, column=1).value = corr[0]
                if len(corr) > 1:
                    ss.cell(row=ss_row, column=2).value = corr[1]
                    ss.cell(row=ss_row, column=2).number_format = '#0.0000'
                    ss.cell(row=ss_row, column=3).value = corr[2]
                ss_row += 1
        ss_row += 2
        ss.cell(row=ss_row, column=1).value = 'Static Variables'
        ss.cell(row=ss_row, column=1).font = bold
        if self.carbon_price > 0:
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = 'Carbon Price ($/tCO2e)'
            ss.cell(row=ss_row, column=st_cap+1).value = self.carbon_price
            ss.cell(row=ss_row, column=st_cap+1).number_format = '$#,##0.00'
            attr_text = 'Summary!$' + ssCol(st_cap+1) + '$' + str(ss_row)
            carbon_cell = oxl.workbook.defined_name.DefinedName('carbon_price', attr_text=attr_text)
            wb.defined_names.append(carbon_cell)
        ss_row += 1
        attr_text = 'Summary!$' + ssCol(st_cap+1) + '$' + str(ss_row)
        lifetime_cell = oxl.workbook.defined_name.DefinedName('lifetime', attr_text=attr_text)
        wb.defined_names.append(lifetime_cell)
        ss.cell(row=ss_row, column=1).value = 'Lifetime (years)'
        ss.cell(row=ss_row, column=st_cap+1).value = max_lifetime
        ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0'
        ss_row += 1
        ss.cell(row=ss_row, column=1).value = 'Discount Rate'
        ss.cell(row=ss_row, column=st_cap+1).value = self.discount_rate
        ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.00%'
        ss_row += 2
        ss_row = self.data_sources(ss, ss_row, pm_data_file, option)
        self.listener.progress_bar.setValue(14)
        if self.event_callback:
            self.event_callback()
        for row in range(1, ss_row + 1):
            for col in range(1, len(headers) + 1):
                try:
                    if ss.cell(row=row, column=col).font.name != 'Arial':
                        ss.cell(row=row, column=col).font = normal
                except:
                    pass
        ss.freeze_panes = 'B4'
        ss.activeCell = 'B4'
        if self.save_tables:
            gens = []
            cons = []
            for fac in re_order:
                if fac == 'Load':
                    continue
                if pmss_details[fac].multiplier <= 0:
                    continue
                if fac.find('.') > 0:
                    gens.append(fac[fac.find('.') + 1:])
                else:
                    gens.append(fac)
                cons.append(self.generators[pmss_details[fac].generator].constraint)
            for gen in dispatch_order:
                gens.append(gen)
                cons.append(self.generators[gen].constraint)
            gs = wb.create_sheet(self.sheets[G].currentText())
            fields = []
            col = 1
            row = 1
            if hasattr(self.generators[list(self.generators.keys())[0]], 'name'):
                fields.append('name')
                gs.cell(row=row, column=col).value = 'Name'
                col += 1
            for prop in dir(self.generators[list(self.generators.keys())[0]]):
                if prop[:2] != '__' and prop[-2:] != '__':
                    if prop != 'name':
                        fields.append(prop)
                        txt = prop.replace('_', ' ').title()
                        txt = txt.replace('Cf', 'CF')
                        txt = txt.replace('Lcoe', 'LCOE')
                        txt = txt.replace('Om', 'OM')
                        gs.cell(row=row, column=col).value = txt
                        if prop == 'capex':
                            txt = txt + txt
                        gs.column_dimensions[ssCol(col)].width = max(len(txt) * 1.4, 10)
                        col += 1
            nme_width = 4
            con_width = 4
            for key, value in self.generators.items():
                if key in gens:
                    row += 1
                    col = 1
                    for field in fields:
                        gs.cell(row=row, column=col).value = getattr(value, field)
                        if field in ['name', 'constraint']:
                            txt = getattr(value, field)
                            if field == 'name':
                                if len(txt) > nme_width:
                                    nme_width = len(txt)
                                    gs.column_dimensions[ssCol(col)].width = nme_width * 1.4
                            else:
                                if len(txt) > con_width:
                                    con_width = len(txt)
                                    gs.column_dimensions[ssCol(col)].width = con_width * 1.4
                        elif field in ['capex', 'fixed_om']:
                            gs.cell(row=row, column=col).number_format = '$#,##0'
                        elif field in ['lcoe', 'variable_om', 'fuel']:
                            gs.cell(row=row, column=col).number_format = '$#,##0.00'
                        elif field in ['disc_rate']:
                            gs.cell(row=row, column=col).number_format = '#,##0.00%'
                        elif field in ['capacity', 'lcoe_cf', 'initial']:
                            gs.cell(row=row, column=col).number_format = '#,##0.00'
                        elif field in ['emissions']:
                            gs.cell(row=row, column=col).number_format = '#,##0.000'
                        elif field in ['lifetime', 'order']:
                            gs.cell(row=row, column=col).number_format = '#,##0'
                        col += 1
            for row in range(1, row + 1):
                for col in range(1, len(fields) + 1):
                    gs.cell(row=row, column=col).font = normal
            gs.freeze_panes = 'B2'
            gs.activeCell = 'B2'
            fields = []
            col = 1
            row = 1
            cs = wb.create_sheet(self.sheets[C].currentText())
            if hasattr(self.constraints[list(self.constraints.keys())[0]], 'name'):
                fields.append('name')
                cs.cell(row=row, column=col).value = 'Name'
                col += 1
            for prop in dir(self.constraints[list(self.constraints.keys())[0]]):
                if prop[:2] != '__' and prop[-2:] != '__':
                    if prop != 'name':
                        fields.append(prop)
                        if prop == 'warm_time':
                            cs.cell(row=row, column=col).value = 'Warmup Time'
                        else:
                            cs.cell(row=row, column=col).value = prop.replace('_', ' ').title()
                        cs.column_dimensions[ssCol(col)].width = max(len(prop) * 1.4, 10)
                        col += 1
            nme_width = 4
            cat_width = 4
            for key, value in self.constraints.items():
                if key in cons:
                    row += 1
                    col = 1
                    for field in fields:
                        cs.cell(row=row, column=col).value = getattr(value, field)
                        if field in ['name', 'category']:
                            txt = getattr(value, field)
                            if field == 'name':
                                if len(txt) > nme_width:
                                    nme_width = len(txt)
                                    cs.column_dimensions[ssCol(col)].width = nme_width * 1.4
                            else:
                                if len(txt) > cat_width:
                                    cat_width = len(txt)
                                    cs.column_dimensions[ssCol(col)].width = cat_width * 1.4
                        elif field == 'warm_time':
                            cs.cell(row=row, column=col).number_format = '#0.00'
                        elif field != 'category':
                            cs.cell(row=row, column=col).number_format = '#,##0%'
                        col += 1
            for row in range(1, row + 1):
                for col in range(1, len(fields) + 1):
                    try:
                        cs.cell(row=row, column=col).font = normal
                    except:
                        pass
            cs.freeze_panes = 'B2'
            cs.activeCell = 'B2'
        wb.save(data_file)
        self.listener.progress_bar.setValue(20)
        if self.event_callback:
            self.event_callback()
        j = data_file.rfind('/')
        data_file = data_file[j + 1:]
        msg = '%s created (%.2f seconds)' % (data_file, time.time() - start_time)
        msg = '%s created.' % data_file
        self.setStatus(msg)
        self.listener.progress_bar.setHidden(True)
        self.listener.progress_bar.setValue(0)

    def run_optimise(self, in_year, in_option, in_pmss_details, in_pmss_data, in_re_order,
                   in_dispatch_order, pm_data_file, data_file):

        def create_starting_population(individuals, chromosome_length):
            # Set up an initial array of all zeros
            population = np.zeros((individuals, chromosome_length))
            # Loop through each row (individual)
            for i in range(individuals):
                # Choose a random number of ones to create but at least one 1
                ones = random.randint(1, chromosome_length)
                # Change the required number of zeros to ones
                population[i, 0:ones] = 1
                # Sfuffle row
                np.random.shuffle(population[i])
            return population

        def select_individual_by_tournament(population, *argv):
            # Get population size
            population_size = len(population)

            # Pick individuals for tournament
            fighter = [0, 0]
            fighter_fitness = [0, 0]
            fighter[0] = random.randint(0, population_size - 1)
            fighter[1] = random.randint(0, population_size - 1)

            # Get fitness score for each
            if len(argv) == 1:
                fighter_fitness[0] = argv[0][fighter[0]]
                fighter_fitness[1] = argv[0][fighter[1]]
            else:
                for arg in argv:
                    min1 = min(arg)
                    max1 = max(arg)
                    for f in range(len(fighter)):
                        try:
                            fighter_fitness[f] += (arg[f] - min1) / (max1 - min1)
                        except:
                            pass
            # Identify individual with lowest score
            # Fighter 1 will win if score are equal
            if fighter_fitness[0] <= fighter_fitness[1]:
                winner = fighter[0]
            else:
                winner = fighter[1]

            # Return the chromsome of the winner
            return population[winner, :]

        def breed_by_crossover(parent_1, parent_2, points=2):
            # Get length of chromosome
            chromosome_length = len(parent_1)

            # Pick crossover point, avoiding ends of chromsome
            if points == 1:
                crossover_point = random.randint(1,chromosome_length-1)
            # Create children. np.hstack joins two arrays
                child_1 = np.hstack((parent_1[0:crossover_point],
                                     parent_2[crossover_point:]))
                child_2 = np.hstack((parent_2[0:crossover_point],
                                     parent_1[crossover_point:]))
            else: # only do 2 at this    stage
                crossover_point_1 = random.randint(1, chromosome_length - 2)
                crossover_point_2 = random.randint(crossover_point_1 + 1, chromosome_length - 1)
                child_1 = np.hstack((parent_1[0:crossover_point_1],
                                     parent_2[crossover_point_1:crossover_point_2],
                                     parent_1[crossover_point_2:]))
                child_2 = np.hstack((parent_2[0:crossover_point_1],
                                     parent_1[crossover_point_1:crossover_point_2],
                                     parent_2[crossover_point_2:]))
            # Return children
            return child_1, child_2

        def randomly_mutate_population(population, mutation_probability):
            # Apply random mutation
            random_mutation_array = np.random.random(size=(population.shape))
            random_mutation_boolean = random_mutation_array <= mutation_probability
        #    random_mutation_boolean[0][:] = False # keep the best multi and lcoe
       #     random_mutation_boolean[1][:] = False
            population[random_mutation_boolean] = np.logical_not(population[random_mutation_boolean])
            # Return mutation population
            return population

        def calculate_fitness(population):
            lcoe_fitness_scores = [] # scores = LCOE values
            multi_fitness_scores = [] # scores = multi-variable weight
            multi_values = [] # values for each of the six variables
            if len(population) == 1:
                option = O1
            else:
                option = O
            if self.debug:
                self.popn += 1
                self.chrom = 0
            for chromosome in population:
                # now get random amount of generation per technology (both RE and non-RE)
                for fac, value in opt_order.items():
                    capacity = value[2]
                    for c in range(value[0], value[1]):
                        if chromosome[c]:
                            capacity = capacity + capacities[c]
                    try:
                        pmss_details[fac].multiplier = capacity / pmss_details[fac].capacity
                    except:
                        print('PME2:', gen, capacity, pmss_details[fac].capacity)
                multi_value, op_data, extra = self.doDispatch(year, option, pmss_details, pmss_data, re_order,
                    dispatch_order, pm_data_file, data_file, files=None, sheets=None)
                if multi_value['load_pct'] < self.targets['load_pct'][3]:
                    if multi_value['load_pct'] == 0:
                        print('PME3:', multi_value['lcoe'], self.targets['load_pct'][3], multi_value['load_pct'])
                        lcoe_fitness_scores.append(1)
                    else:
                        try:
                            lcoe_fitness_scores.append(pow(multi_value['lcoe'],
                                self.targets['load_pct'][3] / multi_value['load_pct']))
                        except OverflowError as err:
                            self.setStatus(f"Overflow error: {err}; POW({multi_value['lcoe']:,}, " \
                                         + f"{self.targets['load_pct'][3] / multi_value['load_pct']:,}) " \
                                         + f"({self.targets['load_pct'][3]:,} / {multi_value['load_pct']:,} )")
                        except:
                            pass
                else:
                    lcoe_fitness_scores.append(multi_value['lcoe'])
                multi_values.append(multi_value)
                multi_fitness_scores.append(calc_weight(multi_value))
                if self.debug:
                    self.chrom += 1
                    line = str(self.popn) + ',' + str(self.chrom) + ','
                    for fac, value in opt_order.items():
                        try:
                            line += str(pmss_details[fac].capacity * pmss_details[fac].multiplier) + ','
                        except:
                            line += ','
                    for key in self.targets.keys():
                        try:
                            line += '{:.3f},'.format(multi_value[key])
                        except:
                            line += multi_value[key] + ','
                    line += '{:.5f},'.format(multi_fitness_scores[-1])
                    self.db_file.write(line + '\n')
            # alternative approach to calculating fitness
            multi_fitness_scores1 = []
            maxs = {}
            mins = {}
            tgts = {}
            for key in multi_value.keys():
                if key[-4:] == '_pct':
                    tgts[key] = abs(self.targets[key][2] - self.targets[key][3])
                else:
                    maxs[key] = 0
                    mins[key] = -1
                    for popn in multi_values:
                        try:
                            tgt = abs(self.targets[key][2] - popn[key])
                        except:
                            continue
                        if tgt > maxs[key]:
                            maxs[key] = tgt
                        if mins[key] < 0 or tgt < mins[key]:
                            mins[key] = tgt
            for popn in multi_values:
                weight = 0
                for key, value in multi_value.items():
                    if self.targets[key][1] <= 0:
                        continue
                    try:
                        tgt = abs(self.targets[key][2] - popn[key])
                    except:
                        continue
                    if key[-4:] == '_pct':
                        if tgts[key] != 0:
                            if tgt > tgts[key]:
                                weight += 1 * self.targets[key][1]
                            else:
                                try:
                                    weight += 1 - ((tgt / tgts[key]) * self.targets[key][1])
                                except:
                                    pass
                    else:
                        try:
                            weight += 1 - (((maxs[key] - tgt) / (maxs[key] - mins[key])) \
                                      * self.targets[key][1])
                        except:
                            pass
                multi_fitness_scores1.append(weight)
            if len(population) == 1: # return the table for best chromosome
                return op_data, multi_values
            else:
                return lcoe_fitness_scores, multi_fitness_scores, multi_values

        def optQuitClicked(event):
            self.optExit = True
            optDialog.close()

        def chooseClicked(event):
            self.opt_choice = sender_name
            chooseDialog.close()

        def calc_weight(multi_value, calc=0):
            weight = [0., 0.]
            if calc == 0:
                for key, value in self.targets.items():
                    if multi_value[key] == '':
                        continue
                    if value[1] <= 0:
                        continue
                    if value[2] == value[3]: # wants specific target
                        if multi_value[key] == value[2]:
                            w = 0
                        else:
                            w = 1
                    elif value[2] > value[3]: # wants higher target
                        if multi_value[key] > value[2]: # high no weight
                            w = 0.
                        elif multi_value[key] < value[3]: # low maximum weight
                            w = 2.
                        else:
                            w = 1 - (multi_value[key] - value[3]) / (value[2] - value[3])
                    else: # lower target
                        if multi_value[key] == -1 or multi_value[key] > value[3]: # high maximum weight
                            w = 2.
                        elif multi_value[key] < value[2]: # low no weight
                            w = 0.
                        else:
                            w = multi_value[key] / (value[3] - value[2])
                    weight[0] += w * value[1]
            elif calc == 1:
                for key, value in self.targets.items():
                    if multi_value[key] == '':
                        continue
                    if value[1] <= 0:
                        continue
                    if multi_value[key] < 0:
                        w = 1
                    elif value[2] == value[3]: # wants specific target
                        if multi_value[key] == value[2]:
                            w = 0
                        else:
                            w = 1
                    else: # target range
                        w = min(abs(value[2] - multi_value[key]) / abs(value[2] - value[3]), 1)
                    weight[1] += w * value[1]
            return weight[calc]

        def plot_multi(multi_scores, multi_best, multi_order, title):
            data = [[], [], []]
            max_amt = [0., 0.]
            for multi in multi_best:
                max_amt[0] = max(max_amt[0], multi['cost'])
                max_amt[1] = max(max_amt[1], multi['co2'])
            pwr_chr = ['', '']
            divisor = [1., 1.]
            pwr_chrs = ' KMBTPEZY'
            for m in range(2):
                for pwr in range(len(pwr_chrs) - 1, -1, -1):
                    if max_amt[m] > pow(10, pwr * 3):
                        pwr_chr[m] = pwr_chrs[pwr]
                        divisor[m] = 1. * pow(10, pwr * 3)
                        break
            self.targets['cost'][5] = self.targets['cost'][5].replace('pwr_chr', pwr_chr[0])
            self.targets['co2'][5] = self.targets['co2'][5].replace('pwr_chr', pwr_chr[1])
            for multi in multi_best:
                for axis in range(3): # only three axes in plot
                    if multi_order[axis] == 'cost':
                        data[axis].append(multi[multi_order[axis]] / divisor[0]) # cost
                    elif multi_order[axis] == 'co2':
                        data[axis].append(multi[multi_order[axis]] / divisor[1]) # co2
                    elif multi_order[axis][-4:] == '_pct': # percentage
                        data[axis].append(multi[multi_order[axis]] * 100.)
                    else:
                        data[axis].append(multi[multi_order[axis]])
            # create colour map
            colours = multi_scores[:]
            cmax = max(colours)
            cmin = min(colours)
            if cmin == cmax:
                return
            for c in range(len(colours)):
                colours[c] = (colours[c] - cmin) / (cmax - cmin)
            scolours = sorted(colours)
            cvals  = [-1., 0, 1]
            colors = ['green' ,'orange', 'red']
            norm = plt.Normalize(min(cvals), max(cvals))
            tuples = list(zip(map(norm,cvals), colors))
            cmap = matplotlib.colors.LinearSegmentedColormap.from_list('', tuples)
            fig = plt.figure(title + QtCore.QDateTime.toString(QtCore.QDateTime.currentDateTime(), '_yyyy-MM-dd_hhmm'))
            mx = plt.axes(projection='3d')
            plt.title('\n' + title.title() + '\n')
            try:
                for i in range(len(data[0])):
                    mx.scatter3D(data[2][i], data[1][i], data[0][i], picker=True, color=cmap(colours[i]), cmap=cmap)
                    if title[:5] == 'start':
                        mx.text(data[2][i], data[1][i], data[0][i], '%s' % str(i+1))
                    else:
                        j = scolours.index(colours[i])
                        if j < 10:
                            mx.text(data[2][i], data[1][i], data[0][i], '%s' % str(j+1))
            except:
                return
            if self.optimise_multisurf:
                cvals_r  = [-1., 0, 1]
                colors_r = ['red' ,'orange', 'green']
                norm_r = plt.Normalize(min(cvals_r), max(cvals_r))
                tuples_r = list(zip(map(norm_r, cvals_r), colors_r))
                cmap_r = matplotlib.colors.LinearSegmentedColormap.from_list('', tuples_r)
                # https://www.fabrizioguerrieri.com/blog/surface-graphs-with-irregular-dataset/
                triang = mtri.Triangulation(data[2], data[1])
                mx.plot_trisurf(triang, data[0], cmap=cmap_r)
            mx.xaxis.set_major_formatter(FormatStrFormatter(self.targets[multi_order[2]][5]))
            mx.yaxis.set_major_formatter(FormatStrFormatter(self.targets[multi_order[1]][5]))
            mx.zaxis.set_major_formatter(FormatStrFormatter(self.targets[multi_order[0]][5]))
            mx.set_xlabel(self.targets[multi_order[2]][6])
            mx.set_ylabel(self.targets[multi_order[1]][6])
            mx.set_zlabel(self.targets[multi_order[0]][6])
            zp = ZoomPanX()
            f = zp.zoom_pan(mx, base_scale=1.2, annotate=True)
            plt.show()
            if zp.datapoint is not None: # user picked a point
                if zp.datapoint[0][0] < 0: # handle problem in matplotlib sometime after version 3.0.3
                    best = [0, 999]
                    for p in range(len(multi_best)):
                        diff = 0
                        for v in range(3):
                            key = multi_order[v]
                            valu = multi_best[p][key]
                            if key[-4:] == '_pct':
                                valu = valu * 100.
                            diff += abs((valu - zp.datapoint[0][v + 1]) / valu)
                        if diff < best[1]:
                            best = [p, diff]
                    zp.datapoint = [[best[0]]]
                    for v in range(3):
                        key = multi_order[v]
                        zp.datapoint[0].append(multi_best[p][key])
                if self.more_details:
                    for p in zp.datapoint:
                        msg = 'iteration ' + str(p[0]) + ': '
                        mult = []
                        for i in range(3):
                            if self.targets[multi_order[i]][6].find('%') > -1:
                                mult.append(100.)
                            else:
                                mult.append(1.)
                            msg += self.targets[multi_order[i]][6].replace('%', '%%') + ': ' + \
                                   self.targets[multi_order[i]][5] + '; '
                        msg = msg % (p[1] * mult[0], p[2] * mult[1], p[3] * mult[2])
                        self.setStatus(msg)
            return zp.datapoint

        def show_multitable(best_score_progress, multi_best, multi_order, title):
            def pwr_chr(amt):
                pwr_chrs = ' KMBTPEZY'
                pchr = ''
                divisor = 1.
                for pwr in range(len(pwr_chrs) - 1, -1, -1):
                    if amt > pow(10, pwr * 3):
                        pchr = pwr_chrs[pwr]
                        divisor = 1. * pow(10, pwr * 3)
                        break
                return amt / divisor, pchr

            def opt_fmat(amt, fmat):
                tail = ' '
                p = fmat.find('pwr_chr')
                if p > 0:
                    amt, tail = pwr_chr(amt)
                    fmat = fmat.replace('pwr_chr', '')
              #  d = fmat.find('d')
             #   if d > 0:
              #      amt = amt * 100.
                fmat = fmat.replace('.1f', '.2f')
                i = fmat.find('%')
                fmt = fmat[:i] + '{:> 8,' + fmat[i:].replace('%', '').replace('d', '.2%') + '}' + tail
                return fmt.format(amt)

            best_table = []
            best_fmate = []
            for b in range(len(multi_best)):
                bl = list(multi_best[b].values())
                bl.insert(0, best_score_progress[b])
                bl.insert(0, b + 1)
                best_table.append(bl)
                best_fmate.append([b + 1])
                best_fmate[-1].insert(1, best_score_progress[b])
                for f in range(2, len(best_table[-1])):
                    if target_keys[f - 2] in ['load_pct', 'surplus_pct', 're_pct']:
                        best_fmate[-1].append(opt_fmat(bl[f], '%d%%'))
                    else:
                        best_fmate[-1].append(opt_fmat(bl[f], target_fmats[f - 2]))
            fields = target_names[:]
            fields.insert(0, 'weight')
            fields.insert(0, 'iteration')
            dialog = displaytable.Table(best_fmate, fields=fields, txt_align='R', decpts=[0, 4],
                     title=title, sortby='weight')
            dialog.exec_()
            b = int(dialog.getItem(0)) - 1
            del dialog
            pick = [b]
            for fld in multi_order[:3]:
                i = target_keys.index(fld)
                pick.append(best_table[b][i + 2])
            return [pick]

#       optClicked mainline starts here
        year = in_year
        option = in_option
        pmss_details = dict(in_pmss_details)
        pmss_data = in_pmss_data[:]
        re_order = in_re_order[:]
        dispatch_order = in_dispatch_order[:]
        if self.optimise_debug:
            self.debug = True
        else:
            self.debug = False
        missing = []
        for fac in re_order:
            if fac == 'Load':
                continue
            if fac not in self.optimisation.keys():
                missing.append(fac)
        for gen in dispatch_order:
            if gen not in self.optimisation.keys():
                missing.append(gen)
        if len(missing) > 0:
            bad = False
            if self.optimise_default is not None:
                defaults = self.optimise_default.split(',')
                if len(defaults) < 1 or len(defaults) > 3:
                    bad = True
                else:
                    try:
                        for miss in missing:
                            if len(defaults) == 2:
                                minn = 0
                            else:
                                if defaults[0][-1] == 'd':
                                    minn = pmss_details[miss].capacity * float(defaults[0][:-1])
                                elif defaults[0][-1] == 'c':
                                    minn = pmss_details[miss].capacity * pmss_details[miss].multiplier * float(defaults[0][:-1])
                                else:
                                    minn = float(defaults[0])
                            if defaults[-2][-1] == 'd':
                                maxx = pmss_details[miss].capacity * float(defaults[-2][:-1])
                            elif defaults[-2][-1] == 'c':
                                maxx = pmss_details[miss].capacity * pmss_details[miss].multiplier * float(defaults[-2][:-1])
                            else:
                                maxx = float(defaults[-2])
                            if len(defaults) == 3:
                                if defaults[-1][-1].lower() == 'd':
                                    step = capacity * float(defaults[-1][:-1])
                                elif defaults[-1][-1].lower() == 'c':
                                    step = capacity * multiplier * float(defaults[-1][:-1])
                                else:
                                    step = float(defaults[-1])
                            else:
                                step = (maxx - minn) / float(defaults[-1])
                            self.optimisation[miss] =  Optimisation(miss, 'None', None)
                            self.optimisation[miss].approach = 'Range'
                            self.optimisation[miss].capacity_min = minn
                            self.optimisation[miss].capacity_max = maxx
                            self.optimisation[miss].capacity_step = step
                    except:
                        bad = True
                check = ''
                for miss in missing:
                    check += miss + ', '
                check = check[:-2]
                if bad:
                    self.setStatus('Key Error: Missing Optimisation entries for: ' + check)
                    return
                self.setStatus('Missing Optimisation entries added for: ' + check)
        self.optExit = False
        self.setStatus('Optimise processing started')
        err_msg = ''
        optDialog = QDialog()
        grid = QGridLayout()
        grid.addWidget(QLabel('Adjust load'), 0, 0)
        self.optLoad = QDoubleSpinBox()
        self.optLoad.setRange(-1, self.adjust_cap)
        self.optLoad.setDecimals(4)
        self.optLoad.setSingleStep(.1)
        rw = 0
        grid.addWidget(self.optLoad, rw, 1)
        grid.addWidget(QLabel('Multiplier for input Load'), rw, 2, 1, 3)
        self.optLoad.setValue(pmss_details['Load'].multiplier)
        rw += 1
        grid.addWidget(QLabel('Population size'), rw, 0)
        optPopn = QSpinBox()
        optPopn.setRange(10, 500)
        optPopn.setSingleStep(10)
        optPopn.setValue(self.optimise_population)
        optPopn.valueChanged.connect(self.changes)
        grid.addWidget(optPopn, rw, 1)
        grid.addWidget(QLabel('Size of population'), rw, 2, 1, 3)
        rw += 1
        grid.addWidget(QLabel('No. of iterations'), rw, 0, 1, 3)
        optGenn = QSpinBox()
        optGenn.setRange(10, 500)
        optGenn.setSingleStep(10)
        optGenn.setValue(self.optimise_generations)
        optGenn.valueChanged.connect(self.changes)
        grid.addWidget(optGenn, rw, 1)
        grid.addWidget(QLabel('Number of iterations (generations)'), rw, 2, 1, 3)
        rw += 1
        grid.addWidget(QLabel('Mutation probability'), rw, 0)
        optMutn = QDoubleSpinBox()
        optMutn.setRange(0, 1)
        optMutn.setDecimals(4)
        optMutn.setSingleStep(0.001)
        optMutn.setValue(self.optimise_mutation)
        optMutn.valueChanged.connect(self.changes)
        grid.addWidget(optMutn, rw, 1)
        grid.addWidget(QLabel('Add in mutation'), rw, 2, 1, 3)
        rw += 1
        grid.addWidget(QLabel('Exit if stable'), rw, 0)
        optStop = QSpinBox()
        optStop.setRange(0, 50)
        optStop.setSingleStep(10)
        optStop.setValue(self.optimise_stop)
        optStop.valueChanged.connect(self.changes)
        grid.addWidget(optStop, rw, 1)
        grid.addWidget(QLabel('Exit if LCOE/weight remains the same after this many iterations'),
                       rw, 2, 1, 3)
        rw += 1
        grid.addWidget(QLabel('Optimisation choice'), rw, 0)
        optCombo = QComboBox()
        choices = ['LCOE', 'Multi', 'Both']
        for choice in choices:
            optCombo.addItem(choice)
            if choice == self.optimise_choice:
                optCombo.setCurrentIndex(optCombo.count() - 1)
        grid.addWidget(optCombo, rw, 1)
        grid.addWidget(QLabel('Choose type of optimisation'),
                       rw, 2, 1, 3)
        rw += 1
        # for each variable name
        grid.addWidget(QLabel('Variable'), rw, 0)
        grid.addWidget(QLabel('Weight'), rw, 1)
        grid.addWidget(QLabel('Better'), rw, 2)
        grid.addWidget(QLabel('Worse'), rw, 3)
        rw += 1
        ndx = grid.count()
        for key in self.targets.keys():
            self.targets[key][4] = ndx
            ndx += 4
        for key, value in self.targets.items():
            if value[2] == value[3]:
                ud = '(=)'
            elif value[2] < 0:
                ud = '(<html>&uarr;</html>)'
            elif value[3] < 0 or value[3] > value[2]:
                ud = '(<html>&darr;</html>)'
            else:
                ud = '(<html>&uarr;</html>)'
            grid.addWidget(QLabel(value[0] + ': ' + ud), rw, 0)
            weight = QDoubleSpinBox()
            weight.setRange(0, 1)
            weight.setDecimals(2)
            weight.setSingleStep(0.05)
            weight.setValue(value[1])
            grid.addWidget(weight, rw, 1)
            if key[-4:] == '_pct':
                minim = QDoubleSpinBox()
                minim.setRange(-.1, 1.)
                minim.setDecimals(2)
                minim.setSingleStep(0.1)
                minim.setValue(value[2])
                grid.addWidget(minim, rw, 2)
                maxim = QDoubleSpinBox()
                maxim.setRange(-.1, 1.)
                maxim.setDecimals(2)
                maxim.setSingleStep(0.1)
                maxim.setValue(value[3])
                grid.addWidget(maxim, rw, 3)
            else:
                minim = QLineEdit()
                minim.setValidator(QtGui.QDoubleValidator())
                minim.validator().setDecimals(2)
                minim.setText(str(value[2]))
                grid.addWidget(minim, rw, 2)
                maxim = QLineEdit()
                maxim.setValidator(QtGui.QDoubleValidator())
                maxim.validator().setDecimals(2)
                maxim.setText(str(value[3]))
                grid.addWidget(maxim, rw, 3)
            rw += 1
        quit = QPushButton('Quit', self)
        grid.addWidget(quit, rw, 0)
        quit.clicked.connect(optQuitClicked)
        show = QPushButton('Proceed', self)
        grid.addWidget(show, rw, 1)
        show.clicked.connect(optDialog.close)
        optDialog.setLayout(grid)
        optDialog.setWindowTitle('Choose Optimisation Parameters')
        optDialog.setWindowIcon(QtGui.QIcon('resources/resources/sen_icon32.ico'))
        optDialog.exec_()
        if self.optExit: # a fudge to exit
            self.setStatus('Execution aborted.')
            return
        # check we have optimisation entries for generators and storage
        # update any changes to targets
        self.optimise_choice = optCombo.currentText()
        for key in self.targets.keys():
            weight = grid.itemAt(self.targets[key][4] + 1).widget()
            self.targets[key][1] = weight.value()
            minim = grid.itemAt(self.targets[key][4] + 2).widget()
            try:
                self.targets[key][2] = minim.value()
            except:
                self.targets[key][2] = float(minim.text())
            maxim = grid.itemAt(self.targets[key][4] + 3).widget()
            try:
                self.targets[key][3] = maxim.value()
            except:
                self.targets[key][3] = float(maxim.text())
        # might want to save load value if changed
        pmss_details['Load'].multiplier = self.optLoad.value()
        updates = {}
        lines = []
        lines.append('optimise_choice=' + self.optimise_choice)
        lines.append('optimise_generations=' + str(self.optimise_generations))
        lines.append('optimise_mutation=' + str(self.optimise_mutation))
        lines.append('optimise_population=' + str(self.optimise_population))
        lines.append('optimise_stop=' + str(self.optimise_stop))
        if self.optimise_choice == 'LCOE':
            do_lcoe = True
            do_multi = False
        elif self.optimise_choice == 'Multi':
            do_lcoe = False
            do_multi = True
        else:
            do_lcoe = True
            do_multi = True
        multi_order = []
        for key, value in self.targets.items():
            line = 'optimise_{}={:.2f},{:.2f},{:.2f}'.format(key, value[1], value[2], value[3])
            lines.append(line)
            multi_order.append('{:.2f}{}'.format(value[1], key))
        updates['Powermatch'] = lines
        SaveIni(updates)
        multi_order.sort(reverse=True)
    #    multi_order = multi_order[:3] # get top three weighted variables - but I want them all
        multi_order = [o[4:] for o in multi_order]
        self.adjust_gen = True
        orig_load = []
        load_col = -1
        orig_tech = {}
        orig_capacity = {}
        opt_order = {} # rely on it being processed in added order
        # each entry = [first entry in chrom, last entry, minimum capacity]
        # first get original renewables generation from data sheet
        for fac in re_order:
            if fac == 'Load':
                continue
            opt_order[fac] = [0, 0, 0]
        # now add scheduled generation
        for gen in dispatch_order:
            opt_order[gen] = [0, 0, 0]
        capacities = []
        for gen in opt_order.keys():
            opt_order[gen][0] = len(capacities) # first entry
            try:
                if self.optimisation[gen].approach == 'Discrete':
                    capacities.extend(self.optimisation[gen].capacities)
                    opt_order[gen][1] = len(capacities) # last entry
                elif self.optimisation[gen].approach == 'Range':
                    if self.optimisation[gen].capacity_max == self.optimisation[gen].capacity_min:
                        capacities.extend([0])
                        opt_order[gen][1] = len(capacities)
                        opt_order[gen][2] = self.optimisation[gen].capacity_min
                        continue
                    ctr = int((self.optimisation[gen].capacity_max - self.optimisation[gen].capacity_min) / \
                              self.optimisation[gen].capacity_step)
                    if ctr < 1:
                        self.setStatus("Error with Optimisation table entry for '" + gen + "'")
                        return
                    capacities.extend([self.optimisation[gen].capacity_step] * ctr)
                    tot = self.optimisation[gen].capacity_step * ctr + self.optimisation[gen].capacity_min
                    if tot < self.optimisation[gen].capacity_max:
                        capacities.append(self.optimisation[gen].capacity_max - tot)
                    opt_order[gen][1] = len(capacities)
                    opt_order[gen][2] = self.optimisation[gen].capacity_min
                else:
                    opt_order[gen][1] = len(capacities)
            except KeyError as err:
                self.setStatus('Key Error: No Optimisation entry for ' + str(err))
                opt_order[gen] = [len(capacities), len(capacities) + 5, 0]
                capacities.extend([pmss_details[gen].capacity / 5.] * 5)
            except ZeroDivisionError as err:
                self.setStatus('Zero capacity: ' + gen + ' ignored')
            except:
                err = str(sys.exc_info()[0]) + ',' + str(sys.exc_info()[1]) + ',' + gen + ',' \
                      + str(opt_order[gen])
                self.setStatus('Error: ' + str(err))
                return
        # chromosome = [1] * int(len(capacities) / 2) + [0] * (len(capacities) - int(len(capacities) / 2))
        # we have the original data - from here down we can do our multiple optimisations
        # Set general parameters
        self.setStatus('Optimisation choice is ' + self.optimise_choice)
        chromosome_length = len(capacities)
        self.setStatus(f'Chromosome length: {chromosome_length}; {pow(2, chromosome_length):,} permutations')
        self.optimise_population = optPopn.value()
        population_size = self.optimise_population
        self.optimise_generations = optGenn.value()
        maximum_generation = self.optimise_generations
        self.optimise_mutation = optMutn.value()
        self.optimise_stop = optStop.value()
        lcoe_scores = []
        multi_scores = []
        multi_values = []
     #   if do_lcoe:
      #      lcoe_target = 0. # aim for this LCOE
        if do_multi:
            multi_best = [] # list of six variables for best weight
            multi_best_popn = [] # list of chromosomes for best weight
        self.show_ProgressBar(maximum=optGenn.value(), msg='Process iterations', title='SIREN - Powermatch Progress')
        self.opt_progressbar.setVisible(True)
        start_time = time.time()
        # Create starting population
        self.opt_progressbar.barProgress(1, 'Processing iteration 1')
        QtCore.QCoreApplication.processEvents()
        population = create_starting_population(population_size, chromosome_length)
        # calculate best score(s) in starting population
        # if do_lcoe best_score = lowest non-zero lcoe
        # if do_multi best_multi = lowest weight and if not do_lcoe best_score also = best_weight
        if self.debug:
            filename = self.scenarios + 'opt_debug_' + \
                       QtCore.QDateTime.toString(QtCore.QDateTime.currentDateTime(),
                       'yyyy-MM-dd_hhmm') + '.csv'
            self.db_file = open(filename, 'w')
            line0 = 'Popn,Chrom,'
            line1 = 'Weights,,'
            line2 = 'Targets,,'
            line3 = 'Range,' + str(population_size) + ','
            for gen, value in opt_order.items():
                 line0 += gen + ','
                 line1 += ','
                 line2 += ','
                 line3 += ','
            for key in self.targets.keys():
                 line0 += key + ','
                 line1 += str(self.targets[key][1]) + ','
                 line2 += str(self.targets[key][2]) + ','
                 if key[-4:] == '_pct':
                     line3 += str(abs(self.targets[key][2] - self.targets[key][3])) + ','
                 else:
                     line3 += ','
            line0 += 'Score'
            self.db_file.write(line0 + '\n' + line1 + '\n' + line2 + '\n' + line3 + '\n')
            self.popn = 0
            self.chrom = 0
        lcoe_scores, multi_scores, multi_values = calculate_fitness(population)
        if do_lcoe:
            try:
                best_score = np.min(lcoe_scores)
            except:
                print('PME4:', lcoe_scores)
            best_ndx = lcoe_scores.index(best_score)
            lowest_chrom = population[best_ndx]
            self.setStatus('Starting LCOE: $%.2f' % best_score)
        if do_multi:
            if self.more_details: # display starting population ?
                pick = plot_multi(multi_scores, multi_values, multi_order, 'starting population')
            # want maximum from first round to set base upper limit
            for key in self.targets.keys():
                if self.targets[key][2] < 0: # want a maximum from first round
                    setit = 0
                    for multi in multi_values:
                        setit = max(multi[key], setit)
                    self.targets[key][2] = setit
                if self.targets[key][3] < 0: # want a maximum from first round
                    setit = 0
                    for multi in multi_values:
                        setit = max(multi[key], setit)
                    self.targets[key][3] = setit
            # now we can find the best weighted result - lowest is best
            best_multi = np.min(multi_scores)
            best_mndx = multi_scores.index(best_multi)
            multi_lowest_chrom = population[best_mndx]
            multi_best_popn.append(multi_lowest_chrom)
            multi_best.append(multi_values[best_mndx])
            self.setStatus('Starting Weight: %.4f' % best_multi)
            multi_best_weight = best_multi
            best_multi_progress = [best_multi]
            if not do_lcoe:
                best_score = best_multi
            last_multi_score = best_multi
            lowest_multi_score = best_multi
            mud = '='
        # Add starting best score to progress tracker
        best_score_progress = [best_score]
        best_ctr = 1
        last_score = best_score
        lowest_score = best_score
        lud = '='
        # Now we'll go through the generations of genetic algorithm
        for generation in range(1, maximum_generation):
            lcoe_status = ''
            multi_status = ''
            if do_lcoe:
                lcoe_status = ' %s $%.2f ;' % (lud, best_score)
            if do_multi:
                multi_status = ' %s %.4f ;' % (mud, best_multi)
            tim = (time.time() - start_time)
            if tim < 60:
                tim = ' (%s%s %.1f secs)' % (lcoe_status, multi_status, tim)
            else:
                tim = ' (%s%s %.2f mins)' % (lcoe_status, multi_status, tim / 60.)
            self.opt_progressbar.barProgress(generation + 1,
                'Processing iteration ' + str(generation + 1) + tim)
            if self.event_callback:
                self.event_callback()
            if not self.opt_progressbar.be_open:
                break
        # Create an empty list for new population
            new_population = []
        # Using elitism approach include best individual
            if do_lcoe:
                new_population.append(lowest_chrom)
            if do_multi:
                new_population.append(multi_lowest_chrom)
            # Create new population generating two children at a time
            if do_lcoe:
                if do_multi:
                    for i in range(int(population_size/2)):
                        parent_1 = select_individual_by_tournament(population, lcoe_scores,
                                                                   multi_scores)
                        parent_2 = select_individual_by_tournament(population, lcoe_scores,
                                                                   multi_scores)
                        child_1, child_2 = breed_by_crossover(parent_1, parent_2)
                        new_population.append(child_1)
                        new_population.append(child_2)
                else:
                    for i in range(int(population_size/2)):
                        parent_1 = select_individual_by_tournament(population, lcoe_scores)
                        parent_2 = select_individual_by_tournament(population, lcoe_scores)
                        child_1, child_2 = breed_by_crossover(parent_1, parent_2)
                        new_population.append(child_1)
                        new_population.append(child_2)
            else:
                for i in range(int(population_size/2)):
                    parent_1 = select_individual_by_tournament(population, multi_scores)
                    parent_2 = select_individual_by_tournament(population, multi_scores)
                    child_1, child_2 = breed_by_crossover(parent_1, parent_2)
                    new_population.append(child_1)
                    new_population.append(child_2)
            # get back to original size (after elitism adds)
            if do_lcoe:
                new_population.pop()
            if do_multi:
                new_population.pop()
            # Replace the old population with the new one
            population = np.array(new_population)
            if self.optimise_mutation > 0:
                population = randomly_mutate_population(population, self.optimise_mutation)
            # Score best solution, and add to tracker
            lcoe_scores, multi_scores, multi_values = calculate_fitness(population)
            if do_lcoe:
                best_lcoe = np.min(lcoe_scores)
                best_ndx = lcoe_scores.index(best_lcoe)
                best_score = best_lcoe
            # now we can find the best weighted result - lowest is best
            if do_multi:
                best_multi = np.min(multi_scores)
                best_mndx = multi_scores.index(best_multi)
                multi_lowest_chrom = population[best_mndx]
                multi_best_popn.append(multi_lowest_chrom)
                multi_best.append(multi_values[best_mndx])
           #     if multi_best_weight > best_multi:
                multi_best_weight = best_multi
                if not do_lcoe:
                    best_score = best_multi
                best_multi_progress.append(best_multi)
            best_score_progress.append(best_score)
            if best_score < lowest_score:
                lowest_score = best_score
                if do_lcoe:
                    lowest_chrom = population[best_ndx]
                else: #(do_multi only)
                    multi_lowest_chrom = population[best_mndx]
            if self.optimise_stop > 0:
                if best_score == last_score:
                    best_ctr += 1
                    if best_ctr >= self.optimise_stop:
                        break
                else:
                    last_score = best_score
                    best_ctr = 1
            last_score = best_score
            if do_lcoe:
                if best_score == best_score_progress[-2]:
                    lud = '='
                elif best_score < best_score_progress[-2]:
                    lud = '<html>&darr;</html>'
                else:
                    lud = '<html>&uarr;</html>'
            if do_multi:
                if best_multi == last_multi_score:
                    mud = '='
                elif best_multi < last_multi_score:
                    mud = '<html>&darr;</html>'
                else:
                    mud = '<html>&uarr;</html>'
                last_multi_score = best_multi
        if self.debug:
            try:
                self.db_file.close()
                optimiseDebug(self.db_file.name)
                os.remove(self.db_file.name)
            except:
                pass
            self.debug = False
        self.opt_progressbar.setVisible(False)
        self.opt_progressbar.close()
        tim = (time.time() - start_time)
        if tim < 60:
            tim = '%.1f secs)' % tim
        else:
            tim = '%.2f mins)' % (tim / 60.)
        msg = 'Optimise completed (%0d iterations; %s' % (generation + 1, tim)
        if best_score > lowest_score:
            msg += ' Try more iterations.'
        # we'll keep two or three to save re-calculating_fitness
        op_data = [[], [], [], [], []]
        score_data = [None, None, None, None, None]
        if do_lcoe:
            op_data[0], score_data[0] = calculate_fitness([lowest_chrom])
        if do_multi:
            op_data[1], score_data[1] = calculate_fitness([multi_lowest_chrom])
        self.setStatus(msg)
        if self.event_callback:
            self.event_callback()
        self.listener.progress_bar.setHidden(True)
        self.listener.progress_bar.setValue(0)
        # GA has completed required generation
        if do_lcoe:
            self.setStatus('Final LCOE: $%.2f' % best_score)
            fig = 'optimise_lcoe'
            titl = 'Optimise LCOE using Genetic Algorithm'
            ylbl = 'Best LCOE ($/MWh)'
        else:
            fig = 'optimise_multi'
            titl = 'Optimise Multi using Genetic Algorithm'
            ylbl = 'Best Weight'
        if do_multi:
            self.setStatus('Final Weight: %.4f' % multi_best_weight)
        # Plot progress
        x = list(range(1, len(best_score_progress)+ 1))
        matplotlib.rcParams['savefig.directory'] = self.scenarios
        plt.figure(fig + QtCore.QDateTime.toString(QtCore.QDateTime.currentDateTime(),
                   '_yyyy-MM-dd_hhmm'))
        lx = plt.subplot(111)
        plt.title(titl)
        lx.plot(x, best_score_progress)
        lx.set_xlabel('Optimise Cycle (' + str(len(best_score_progress)) + ' iterations)')
        lx.set_ylabel(ylbl)
        zp = ZoomPanX()
        f = zp.zoom_pan(lx, base_scale=1.2, annotate=True)
        plt.show()
        pick = None
        pickf = None
        if do_multi:
            if self.optimise_multiplot:
                pick = plot_multi(best_multi_progress, multi_best, multi_order, 'best of each iteration')
                if self.more_details:
                    pickf = plot_multi(multi_scores, multi_values, multi_order, 'final iteration')
            if self.optimise_multitable:
                pick2 = show_multitable(best_multi_progress, multi_best, multi_order, 'best of each iteration')
                try:
                    pick = pick + pick2
                except:
                    pick = pick2
                if self.more_details:
                    pick2 = show_multitable(multi_scores, multi_values, multi_order, 'final iteration')
                    try:
                        pickf = pickf + pick2
                    except:
                        pickf = pick2
        op_pts = [0] * len(headers)
        for p in [st_lcg, st_lco, st_lcc, st_max, st_bal, st_rlc, st_are]:
            op_pts[p] = 2
        op_pts[st_cap] = 3
        if self.more_details:
            if do_lcoe:
                list(map(list, list(zip(*op_data[0]))))
                dialog = displaytable.Table(op_data[0], title=sender_name, fields=headers,
                         save_folder=self.scenarios, sortby='', decpts=op_pts)
                dialog.exec_()
                del dialog
            if do_multi:
                list(map(list, list(zip(*op_data[1]))))
                dialog = displaytable.Table(op_data[1], title='Multi_' + sender_name, fields=headers,
                         save_folder=self.scenarios, sortby='', decpts=op_pts)
                dialog.exec_()
                del dialog
        # now I'll display the resulting capacities for LCOE, lowest weight, picked
        # now get random amount of generation per technology (both RE and non-RE)
        its = {}
        for fac, value in opt_order.items():
            its[fac] = []
        chrom_hdrs = []
        chroms = []
        ndxes = []
        if do_lcoe:
            chrom_hdrs = ['Lowest LCOE']
            chroms = [lowest_chrom]
            ndxes = [0]
        if do_multi:
            chrom_hdrs.append('Lowest Weight')
            chroms.append(multi_lowest_chrom)
            ndxes.append(1)
        if pickf is not None:
            for p in range(len(pickf)):
                if pick is None:
                    pick = [pickf[f]]
                else:
                    pick.append(pickf[p][:])
                pick[-1][0] = len(multi_best_popn)
                multi_best_popn.append(population[pickf[p][0]])
        if pick is not None:
            # at present I'll calculate the best weight for the chosen picks. Could actually present all for user choice
            if len(pick) <= 3:
                multi_lowest_chrom = multi_best_popn[pick[0][0]]
                op_data[2], score_data[2] = calculate_fitness([multi_lowest_chrom])
                if self.more_details:
                    list(map(list, list(zip(*op_data[2]))))
                    dialog = displaytable.Table(op_data[2], title='Pick_' + sender_name, fields=headers,
                             save_folder=self.scenarios, sortby='', decpts=op_pts)
                    dialog.exec_()
                    del dialog
                chrom_hdrs.append('Your pick')
                chroms.append(multi_lowest_chrom)
                ndxes.append(2)
                if len(pick) >= 2:
                    multi_lowest_chrom = multi_best_popn[pick[1][0]]
                    op_data[3], score_data[3] = calculate_fitness([multi_lowest_chrom])
                    if self.more_details:
                        list(map(list, list(zip(*op_data[3]))))
                        dialog = displaytable.Table(op_data[3], title='Pick_' + sender_name, fields=headers,
                                 save_folder=self.scenarios, sortby='', decpts=op_pts)
                        dialog.exec_()
                        del dialog
                    chrom_hdrs.append('Your 2nd pick')
                    chroms.append(multi_lowest_chrom)
                    ndxes.append(3)
                if len(pick) == 3:
                    multi_lowest_chrom = multi_best_popn[pick[2][0]]
                    op_data[4], score_data[4] = calculate_fitness([multi_lowest_chrom])
                    if self.more_details:
                        list(map(list, list(zip(*op_data[4]))))
                        dialog = displaytable.Table(op_data[4], title='Pick_' + sender_name, fields=headers,
                                 save_folder=self.scenarios, sortby='', decpts=op_pts)
                        dialog.exec_()
                        del dialog
                    chrom_hdrs.append('Your 3rd pick')
                    chroms.append(multi_lowest_chrom)
                    ndxes.append(4)
            else:
                picks = []
                for pck in pick:
                    picks.append(multi_best_popn[pck[0]])
                a, b, c = calculate_fitness(picks)
                best_multi = np.min(b)
                best_mndx = b.index(best_multi)
                multi_lowest_chrom = picks[best_mndx]
                op_data[2], score_data[2] = calculate_fitness([multi_lowest_chrom])
                if self.more_details:
                    list(map(list, list(zip(*op_data[2]))))
                    dialog = displaytable.Table(op_data[2], title='Pick_' + sender_name, fields=headers,
                             save_folder=self.scenarios, sortby='', decpts=op_pts)
                    dialog.exec_()
                    del dialog
                chrom_hdrs.append('Your pick')
                chroms.append(multi_lowest_chrom)
                ndxes.append(2)
        for chromosome in chroms:
            for fac, value in opt_order.items():
                capacity = opt_order[fac][2]
                for c in range(value[0], value[1]):
                    if chromosome[c]:
                        capacity = capacity + capacities[c]
                its[fac].append(capacity / pmss_details[fac].capacity)
        chooseDialog = QDialog()
        hbox = QHBoxLayout()
        grid = [QGridLayout()]
        label = QLabel('<b>Facility</b>')
        label.setAlignment(QtCore.Qt.AlignCenter)
        grid[0].addWidget(label, 0, 0)
        for h in range(len(chrom_hdrs)):
            grid.append(QGridLayout())
            label = QLabel('<b>' + chrom_hdrs[h] + '</b>')
            label.setAlignment(QtCore.Qt.AlignCenter)
            grid[-1].addWidget(label, 0, 0, 1, 3)
        rw = 1
        for key, value in its.items():
            grid[0].addWidget(QLabel(key), rw, 0)
            if pmss_details[key].fac_type == 'S':
                typ = ' MWh'
            else:
                typ = ' MW'
            if self.show_multipliers:
                for h in range(len(chrom_hdrs)):
                    label = QLabel('{:,.1f}'.format(value[h] * pmss_details[key].capacity))
                    label.setAlignment(QtCore.Qt.AlignRight)
                    grid[h + 1].addWidget(label, rw, 0)
                    label = QLabel(typ)
                    label.setAlignment(QtCore.Qt.AlignLeft)
                    grid[h + 1].addWidget(label, rw, 1)
                    label = QLabel('({:.2f})'.format(value[h]))
                    label.setAlignment(QtCore.Qt.AlignRight)
                    grid[h + 1].addWidget(label, rw, 2)
            else:
                for h in range(len(chrom_hdrs)):
                    label = QLabel('{:,.1f}'.format(value[h] * pmss_details[key].capacity))
                    label.setAlignment(QtCore.Qt.AlignRight)
                    grid[h + 1].addWidget(label, rw, 0, 1, 2)
                    label = QLabel(typ)
                    label.setAlignment(QtCore.Qt.AlignLeft)
                    grid[h + 1].addWidget(label, rw, 2)
            rw += 1
        max_amt = [0., 0.]
        if do_lcoe:
            max_amt[0] = score_data[0][0]['cost']
            max_amt[1] = score_data[0][0]['co2']
        if do_multi:
            for multi in multi_best:
                max_amt[0] = max(max_amt[0], multi['cost'])
                max_amt[1] = max(max_amt[1], multi['co2'])
        pwr_chr = ['', '']
        divisor = [1., 1.]
        pwr_chrs = ' KMBTPEZY'
        for m in range(2):
            for pwr in range(len(pwr_chrs) - 1, -1, -1):
                if max_amt[m] > pow(10, pwr * 3):
                    pwr_chr[m] = pwr_chrs[pwr]
                    divisor[m] = 1. * pow(10, pwr * 3)
                    break
        self.targets['cost'][5] = self.targets['cost'][5].replace('pwr_chr', pwr_chr[0])
        self.targets['co2'][5] = self.targets['co2'][5].replace('pwr_chr', pwr_chr[1])
        for key in multi_order:
            lbl = QLabel('<i>' + self.targets[key][0] + '</i>')
            lbl.setAlignment(QtCore.Qt.AlignCenter)
            grid[0].addWidget(lbl, rw, 0)
            for h in range(len(chrom_hdrs)):
                if key == 'cost':
                    amt = score_data[ndxes[h]][0][key] / divisor[0] # cost
                elif key == 'co2':
                    amt = score_data[ndxes[h]][0][key] / divisor[1] # co2
                elif key[-4:] == '_pct': # percentage
                    amt = score_data[ndxes[h]][0][key] * 100.
                else:
                    amt = score_data[ndxes[h]][0][key]
                txt = '<i>' + self.targets[key][5] + '</i>'
                try:
                    label = QLabel(txt % amt)
                except:
                    label = QLabel('?')
                    print('PME5:', key, txt, amt)
                label.setAlignment(QtCore.Qt.AlignCenter)
                grid[h + 1].addWidget(label, rw, 0, 1, 3)
            rw += 1
        cshow = QPushButton('Quit', self)
        grid[0].addWidget(cshow)
        cshow.clicked.connect(chooseDialog.close)
        for h in range(len(chrom_hdrs)):
            button = QPushButton(chrom_hdrs[h], self)
            grid[h + 1].addWidget(button, rw, 0, 1, 3)
            button.clicked.connect(chooseClicked) #(chrom_hdrs[h]))
        for gri in grid:
            frame = QFrame()
            frame.setFrameStyle(QFrame.Box)
            frame.setLineWidth(1)
            frame.setLayout(gri)
            hbox.addWidget(frame)
   #     grid.addWidget(show, rw, 1)
    #    show.clicked.connect(optDialog.close)
        chooseDialog.setLayout(hbox)
        chooseDialog.setWindowTitle('Choose Optimal Generator Mix')
        chooseDialog.setWindowIcon(QtGui.QIcon('resources/resources/sen_icon32.ico'))
     #  this is a big of a kluge but I couldn't get it to behave
        self.opt_choice = ''
        chooseDialog.exec_()
        del chooseDialog
        try:
            h = chrom_hdrs.index(self.opt_choice)
        except:
            return
        op_data[h], score_data[h] = calculate_fitness([chroms[h]]) # make it current
        msg = chrom_hdrs[h] + ': '
        for key in multi_order[:3]:
            msg += self.targets[key][0] + ': '
            if key == 'cost':
                amt = score_data[h][0][key] / divisor[0] # cost
            elif key == 'co2':
                amt = score_data[h][0][key] / divisor[1] # co2
            elif key[-4:] == '_pct': # percentage
                amt = score_data[h][0][key] * 100.
            else:
                amt = score_data[h][0][key]
            txt = self.targets[key][5]
            txt = txt % amt
            msg += txt + '; '
        self.setStatus(msg)
        list(map(list, list(zip(*op_data[h]))))
        op_data[h].append(' ')
        op_data[h].append('Optimisation Parameters')
        op_op_prm = len(op_data[h])
        op_data[h].append(['Population size', str(self.optimise_population)])
        op_data[h].append(['No. of iterations', str(self.optimise_generations)])
        op_data[h].append(['Mutation probability', '%0.4f' % self.optimise_mutation])
        op_data[h].append(['Exit if stable', str(self.optimise_stop)])
        op_data[h].append(['Optimisation choice', self.optimise_choice])
        op_data[h].append(['Variable', 'Weight', 'Better', 'Worse'])
        for i in range(len(target_keys)):
            op_data[h].append([])
            for j in range(4):
                if j == 0:
                    op_data[h][-1].append(self.targets[target_keys[i]][j])
                else:
                    op_data[h][-1].append('{:.2f}'.format(self.targets[target_keys[i]][j]))
        op_max_row = len(op_data[h])
        for key in self.optimisation.keys():
            op_data[h].append(['Max. ' + key, self.optimisation[key].capacity_max])
        dialog = displaytable.Table(op_data[h], title='Chosen_' + sender_name, fields=headers,
                 save_folder=self.scenarios, sortby='', decpts=op_pts)
        dialog.exec_()
        del dialog
        if self.optimise_to_batch:
            msgbox = QMessageBox()
            msgbox.setWindowTitle('SIREN - Add to Batch')
            msgbox.setText("Press 'Yes' to add to Batch file")
            msgbox.setIcon(QMessageBox.Question)
            msgbox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            reply = msgbox.exec_()
            if reply == QMessageBox.Yes:
                check_list = []
                tot_row = 0
                save_opt_rows = False
                for o_r in range(len(op_data[h])):
                    if op_data[h][o_r][0] == 'Total':
                        break
                    if op_data[h][o_r][0] != 'RE Contribution To Load':
                        check_list.append(o_r)
                if self.files[B].text() != '':
                    newfile = self.get_filename(self.files[B].text())
                else:
                    curfile = self.scenarios[:-1]
                    newfile = QFileDialog.getSaveFileName(None, 'Create and save ' + self.file_labels[B] + ' file',
                              curfile, 'Excel Files (*.xlsx)')[0]
                    if newfile == '':
                        return
                if os.path.exists(newfile):
                    wb = oxl.load_workbook(newfile)
                elif self.batch_template == '':
                    return
                else:
                    wb = oxl.load_workbook(self.batch_template)   #copy batch
                    if newfile[: len(self.scenarios)] == self.scenarios:
                        self.files[B].setText(newfile[len(self.scenarios):])
                    else:
                        if newfile.rfind('/') > 0:
                            that_len = len(commonprefix([self.scenarios, newfile]))
                            if that_len > 0:
                                bits = self.scenarios[that_len:].split('/')
                                pfx = ('..' + '/') * (len(bits) - 1)
                                newfile = pfx + newfile[that_len + 1:]
                        if newfile[-5:] != '.xlsx':
                            newfile += '.xlsx'
                        self.files[B].setText(newfile)
                if wb.worksheets[0].max_column > 1024:
                    self.clean_batch_sheet()
                    ds = oxl.load_workbook(self.get_filename(self.files[B].text()))
                batch_input_sheet = wb.worksheets[0]
                batch_input_sheet.protection.sheet = False
                normal = oxl.styles.Font(name='Arial')
                bold = oxl.styles.Font(name='Arial', bold=True)
                col = batch_input_sheet.max_column + 1
                fst_row = -1
                if col == 4: # possibly only chart stuff in columns 2 and 3
                    get_out = False
                    for col in range(3, 1, -1):
                        for row in range(1, batch_input_sheet.max_row + 1):
                            if batch_input_sheet.cell(row=row, column=col).value is not None:
                                col += 1
                                get_out = True
                                break
                            if batch_input_sheet.cell(row=row, column=1).value == 'Total':
                                break
                        if get_out:
                            break
                for row in range(1, batch_input_sheet.max_row + 1):
                    if batch_input_sheet.cell(row=row, column=1).value is None:
                        continue
                    if batch_input_sheet.cell(row=row, column=1).value in ['Model', 'Model Label', 'Technology']:
                        new_cell = batch_input_sheet.cell(row=row, column=col)
                        new_cell.value = QtCore.QDateTime.toString(QtCore.QDateTime.currentDateTime(), 'MM-dd hh:mm')
                        msg += " Added to batch as '" + new_cell.value + "' (column " + ssCol(col) + ')'
                        continue
                    if batch_input_sheet.cell(row=row, column=1).value == 'Capacity (MW)':
                        fst_row = row + 1
                        cell = batch_input_sheet.cell(row=row, column=col - 1)
                        new_cell = batch_input_sheet.cell(row=row, column=col)
                        new_cell.value = 'MW'
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.protection = copy(cell.protection)
                            new_cell.alignment = copy(cell.alignment)
                        continue
                    if batch_input_sheet.cell(row=row, column=1).value == 'Optimisation Parameters':
                        save_opt_rows = True
                        break
                    for o_r in range(len(op_data[h])):
                        if op_data[h][o_r][0] == batch_input_sheet.cell(row=row, column=1).value:
                            if op_data[h][o_r][0] == 'Total' and col > 2:
                                cell = batch_input_sheet.cell(row=row, column=2)
                            else:
                                cell = batch_input_sheet.cell(row=row, column=col - 1)
                            new_cell = batch_input_sheet.cell(row=row, column=col)
                            try:
                                new_cell.value = float(op_data[h][o_r][1])
                            except:
                                try:
                                    new_cell.value = op_data[h][o_r][1]
                                except:
                                    pass
                            if cell.has_style:
                                new_cell.font = copy(cell.font)
                                new_cell.border = copy(cell.border)
                                new_cell.fill = copy(cell.fill)
                                new_cell.protection = copy(cell.protection)
                                new_cell.alignment = copy(cell.alignment)
                                if col == 2:
                                    new_cell.font = normal
                                    new_cell.number_format = '#0.00'
                                else:
                                    new_cell.number_format = copy(cell.number_format)
                            elif col == 2:
                                new_cell.font = normal
                                new_cell.number_format = '#0.00'
                            try:
                                i = check_list.index(o_r)
                                del check_list[i]
                            except:
                                pass
                    if batch_input_sheet.cell(row=row, column=1).value == 'Total':
                        tot_row = row
                if save_opt_rows: # want optimisation?
                    for o_r in range(op_op_prm, len(op_data[h])):
                        row += 1
                        new_cell = batch_input_sheet.cell(row=row, column=1)
                        new_cell.value = op_data[h][o_r][0]
                        new_cell = batch_input_sheet.cell(row=row, column=col)
                        try:
                            new_cell.value = float(op_data[h][o_r][1])
                        except:
                            new_cell.value = op_data[h][o_r][1]
                if len(check_list) > 0:
                    check_list.reverse()
                    if col > 2:
                        cell = batch_input_sheet.cell(row=fst_row, column=2)
                    else:
                        cell = batch_input_sheet.cell(row=fst_row, column=col)
                    for o_r in check_list:
                        batch_input_sheet.insert_rows(tot_row)
                        new_cell = batch_input_sheet.cell(row=tot_row, column=1)
                        new_cell.value = op_data[h][o_r][0]
                        new_cell = batch_input_sheet.cell(row=tot_row, column=col)
                        try:
                            new_cell.value = float(op_data[h][o_r][1])
                        except:
                            new_cell.value = op_data[h][o_r][1]
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.protection = copy(cell.protection)
                            new_cell.alignment = copy(cell.alignment)
                wb.save(self.get_filename(self.files[B].text()))
                self.setStatus(msg)
        if self.adjust.isChecked():
            self.adjustto = {}
            for fac, value in sorted(pmss_details.items()):
                self.adjustto[fac] = value.capacity * value.multiplier
        return
