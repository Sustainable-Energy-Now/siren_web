import configparser  # decode .ini file
from copy import copy
from powermatch.ui.credits import fileVersion
import powermatch.ui.displaytable as displaytable
from siren.ui.editini import EdtDialog, SaveIni
from powermatch.ui.floaters import ProgressBar, FloatStatus
import glob
from logic.config_manager import ConfigManager
from logic.file_handler import FileHandler
from logic.excel import ExcelProcessor
from logic.logic import Constraint, Facility, PM_Facility, Optimisation
from logic.processor import PowerMatchProcessor, ProgressHandler
import openpyxl as oxl
from openpyxl.chart import (
    LineChart,
    Reference,
    Series
)
import os
from math import log10
from utilities.senutils import ClickableQLabel, getParents, getUser, ListWidget, ssCol, techClean, WorkBook
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QComboBox, QCheckBox, QDesktopWidget, QDialog, QDoubleSpinBox,  \
    QFileDialog, QFrame, QGridLayout, QLabel, QLineEdit, QMainWindow, QMessageBox, QProgressBar,   \
    QPushButton, QScrollArea, QShortcut,  QVBoxLayout, QWidget, QDesktopWidget
from PyQt5.QtCore import pyqtSignal, QSize
import subprocess
import sys
import time

tech_names = ['Load', 'Onshore Wind', 'Offshore Wind', 'Rooftop PV', 'Fixed PV', 'Single Axis PV',
              'Dual Axis PV', 'Biomass', 'Geothermal', 'Other1', 'CST', 'Shortfall']
# initialise tech_names from .ini file
#            add dispatchable for re from [Grid] dispatchable?
# load data file. If not in data file then include in order and flag as RE
# tracking_pv is a synonym form dual_axis_pv
# phes is a synonym for pumped_hydro
# other1 is a synonym for other - or the other way around
# [Grid]
# dispatchable=pumped_hydro geothermal biomass solar_thermal cst
# consider: hydrogen bess
# [Power]
# technologies=backtrack_pv bess biomass cst fixed_pv geothermal offshore_wind rooftop_pv single_axis_pv solar_thermal tracking_pv wave wind other other_wave
#              add pumped_hydro hydrogen
#              maybe drop bess?
# fossil_technologies=fossil_ccgt fossil_coal fossil_cogen fossil_distillate fossil_gas fossil_mixed fossil_ocgt
target_keys = ['lcoe', 'load_pct', 'surplus_pct', 're_pct', 'cost', 'co2']
target_names = ['LCOE', 'Load%', 'Surplus%', 'RE%', 'Cost', 'CO2']
target_fmats = ['$%.2f', '%.1f%%', '%.1f%%', '%.1f%%', '$%.1fpwr_chr', '%.1fpwr_chr']
target_titles = ['LCOE ($)', 'Load met %', 'Surplus %', 'RE %', 'Total Cost ($)', 'tCO2e']
headers = ['Facility', 'Capacity\n(Gen, MW;\nStor, MWh)', 'To meet\nLoad (MWh)',
           'Subtotal\n(MWh)', 'CF', 'Cost ($/yr)', 'LCOG\nCost\n($/MWh)', 'LCOE\nCost\n($/MWh)',
           'Emissions\n(tCO2e)', 'Emissions\nCost', 'LCOE With\nCO2 Cost\n($/MWh)', 'Max.\nMWH',
           'Max.\nBalance', 'Capital\nCost', 'Lifetime\nCost', 'Lifetime\nEmissions',
           'Lifetime\nEmissions\nCost', 'Area (km^2)', 'Reference\nLCOE', 'Reference\nCF']
# set up columns for summary table. Hopefully to make it easier to add / alter columns
st_fac = 0 # Facility
st_cap = 1 # Capacity\n(Gen, MW;\nStor, MWh)
st_tml = 2 # To meet\nLoad (MWh)
st_sub = 3 # Subtotal\n(MWh)
st_cfa = 4 # CF
st_cst = 5 # Cost ($/yr)
st_lcg = 6 # LCOG\nCost\n($/MWh)
st_lco = 7 # LCOE\nCost\n($/MWh)
st_emi = 8 # Emissions\n(tCO2e)
st_emc = 9 # Emissions\nCost
st_lcc = 10 # LCOE With\nCO2 Cost\n($/MWh)
st_max = 11 # Max.\nMWH
st_bal = 12 # Max.\nBalance'
st_cac = 13 # Capital\nCost'
st_lic = 14 # Lifetime\nCost'
st_lie = 15 # Lifetime\nEmissions
st_lec = 16 # Lifetime\nEmissions\nCost
st_are = 17 # Area (km^2)
st_rlc = 18 # Reference\nLCOE
st_rcf = 19 # Reference\nCF

# same order as self.file_labels
C = 0 # Constraints - xls or xlsx
G = 1 # Generators - xls or xlsx
O = 2 # Optimisation - xls or xlsx
D = 3 # Data - xlsx
R = 4 # Results - xlsx
B = 5 # Batch input - xlsx
T = 6 # Transition input - xlsx
S = 'S' # Summary
O1 = 'O1'

def get_value(ws, row, col):
    def get_range(text, alphabet=None, base=1):
        if len(text) < 1:
            return None
        if alphabet is None:
            alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        if alphabet[0] == ' ':
            alphabet = alphabet[1:]
        alphabet = alphabet.upper()
        bits = ['', '']
        b = 0
        in_char = True
        for char in text:
            if char.isdigit():
                if in_char:
                    in_char = False
                    b += 1
            else:
                if alphabet.find(char.upper()) < 0:
                    continue
                if not in_char:
                    in_char = True
                    b += 1
            if b >= len(bits):
                break
            bits[b] += char.upper()
        try:
            bits[1] = int(bits[1]) - (1 - base)
        except:
            pass
        row = 0
        ndx = 1
        for c in range(len(bits[0]) -1, -1, -1):
            ndx1 = alphabet.index(bits[0][c]) + 1
            row = row + ndx1 * ndx
            ndx = ndx * len(alphabet)
        bits[0] = row - (1 - base)
        for c in bits:
            if c == '':
                return None
    try:
        while ws.cell(row=row, column=col).value[0] == '=':
            row, col = get_range(ws.cell(row=row, column=col).value)
    except:
        return ''
    return ws.cell(row=row, column=col).value

class MainWindow(QMainWindow):
    # Signals can be connected to slots for interactivity
    log_signal = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Powermatch UI")
        # Initialize UI components (grid layout, buttons, etc.)
        # Connect UI events to logic or API

    def display_log(self, message: str):
        # Logic to display logs in the UI
        pass

class Adjustments(QDialog):
    def setAdjValueUnits(self, key, typ, capacity):
        if key != 'Load':
            mw = capacity
            if typ == 'S':
                unit = 'MWh'
            else:
                unit = 'MW'
            dp = self._decpts
            div = 0
        else:
            dimen = log10(capacity)
            unit = 'MWh'
            if dimen > 11:
                unit = 'PWh'
                div = 9
            elif dimen > 8:
                unit = 'TWh'
                div = 6
            elif dimen > 5:
                unit = 'GWh'
                div = 3
            else:
                div = 0
            mw = capacity / pow(10, div)
            dp = None
        mwtxt = unit
        mwcty = round(mw, dp)
        return mw, mwtxt, mwcty, div

    def niceSize(window, ctr): # works for Adjustments window (probably because less that 640*480)
        height = window.frameSize().height() / 1.07
        height = 65 + ctr * 32
        width = window.frameSize().width()
        screen = QDesktopWidget().availableGeometry()
        if height > (screen.height() - 70):
            height = screen.height() - 70
        if width > (screen.width() - 70):
            width = screen.width() - 70
        size = QtCore.QSize(QtCore.QSize(int(width), int(height)))
        window.resize(size)

    def __init__(self, parent, data, adjustin, adjust_cap, prefix, show_multipliers=False, save_folder=None,
                 batch_file=None):
        super(Adjustments, self).__init__()
        self.ignoreEnter = False
        self._adjust_typ = {} # facility type = G, S or L
        self._adjust_cty = {} # (actual) adjust capacity
        self.show_multipliers = show_multipliers
        if self.show_multipliers:
            self._adjust_mul = {} # (actual) adjust multiplier
            self._adjust_rnd = {} # multiplier widget (rounded to 4 digits)
        self._adjust_txt = {} # string with capacity units
        self._save_folder = save_folder
        self._batch_file = None
        if batch_file is not None:
            if os.path.isfile(batch_file):
                self._batch_file = batch_file
        self._ignore = False
        self._results = None
        self.grid = QGridLayout()
        self._data = {}
        ctr = 0
        self._decpts = 1
        for key, typ, capacity in data:
            if key == 'Load' or capacity is None:
                continue
            dimen = log10(capacity)
            if dimen < 2.:
                if dimen < 1.:
                    self._decpts = 2
                elif self._decpts != 2:
                    self._decpts = 1
        if prefix is not None:
            self.grid.addWidget(QLabel('Results Prefix:'), ctr, 0)
            self.pfx_fld = QLineEdit()
            self.pfx_fld.setText(prefix)
            self.grid.addWidget(self.pfx_fld, ctr, 1, 1, 2)
            ctr += 1
        # Note: relies on Load being first entry
        for key, typ, capacity in data:
            self._adjust_typ[key] = typ
            if key != 'Load' and capacity is None:
                continue
       #     if key not in adjustin.keys():
       #         continue
            try:
                mw, mwtxt, mwcty, div = self.setAdjValueUnits(key, typ, adjustin[key])
            except:
                mw = 0
                mwtxt = 'MW'
                mwcty = 0
                div = 0
            self._data[key] = [capacity / pow(10, div), div]
            self._adjust_cty[key] = QDoubleSpinBox()
            self._adjust_cty[key].setRange(0, capacity / pow(10, div) * adjust_cap)
            self._adjust_cty[key].setDecimals(self._decpts)
            if self.show_multipliers:
                self._adjust_rnd[key] = QDoubleSpinBox()
                self._adjust_rnd[key].setRange(0, adjust_cap)
                self._adjust_rnd[key].setDecimals(4)
            if key in adjustin.keys():
                self._adjust_cty[key].setValue(mwcty)
                if self.show_multipliers:
                    try:
                        self._adjust_mul[key] = adjustin[key] / capacity
                        self._adjust_rnd[key].setValue(round(self._adjust_mul[key], 4))
                    except:
                        self._adjust_mul[key] = 1.
                        self._adjust_rnd[key].setValue(1.)
            else:
                self._adjust_cty[key].setValue(0)
                if self.show_multipliers:
                    self._adjust_mul[key] = 0.
                    self._adjust_rnd[key].setValue(0.)
            self._adjust_cty[key].setObjectName(key)
            self.grid.addWidget(QLabel(key), ctr, 0)
            self.grid.addWidget(self._adjust_cty[key], ctr, 1)
            self._adjust_txt[key] = QLabel('')
            self._adjust_txt[key].setObjectName(key + 'label')
            self._adjust_txt[key].setText(mwtxt)
            self.grid.addWidget(self._adjust_txt[key], ctr, 2)
            if self.show_multipliers:
                self._adjust_cty[key].valueChanged.connect(self.adjustCap)
                self._adjust_rnd[key].setSingleStep(.1)
                self._adjust_rnd[key].setObjectName(key)
                self.grid.addWidget(self._adjust_rnd[key], ctr, 3)
                self._adjust_rnd[key].valueChanged.connect(self.adjustMult)
            ctr += 1
            if key == 'Load' and len(data) > 1:
                self.grid.addWidget(QLabel('Facility'), ctr, 0)
                self.grid.addWidget(QLabel('Capacity'), ctr, 1)
                if self.show_multipliers:
                    self.grid.addWidget(QLabel('Multiplier'), ctr, 3)
                ctr += 1
        quit = QPushButton('Quit', self)
        self.grid.addWidget(quit, ctr, 0)
        quit.clicked.connect(self.quitClicked)
        show = QPushButton('Proceed', self)
        self.grid.addWidget(show, ctr, 1)
        show.clicked.connect(self.showClicked)
        if prefix is not None:
            reset = QPushButton('Reset', self)
            self.grid.addWidget(reset, ctr, 2)
            reset.clicked.connect(self.resetClicked)
        resetload = QPushButton('Reset Load', self)
        self.grid.addWidget(resetload, ctr, 3)
        resetload.clicked.connect(self.resetloadClicked)
        if save_folder is not None:
            ctr += 1
            save = QPushButton('Save', self)
            self.grid.addWidget(save, ctr, 0)
            save.clicked.connect(self.saveClicked)
            restore = QPushButton('Restore', self)
            self.grid.addWidget(restore, ctr, 1)
            restore.clicked.connect(self.restoreClicked)
            listi = QPushButton('List', self)
            self.grid.addWidget(listi, ctr, 2)
            listi.clicked.connect(self.listClicked)
            if self._batch_file is not None:
                batch = QPushButton('Add to Batch', self)
                self.grid.addWidget(batch, ctr, 3)
                batch.clicked.connect(self.addtoBatch)
        frame = QFrame()
        frame.setLayout(self.grid)
        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setWidget(frame)
        self.layout = QVBoxLayout(self)
        self.layout.addWidget(self.scroll)
        self.niceSize(ctr)
        self.setWindowTitle('SIREN - Powermatch - Adjust generators')
        self.setWindowIcon(QtGui.QIcon('resources/resources/sen_icon32.ico'))
        QShortcut(QtGui.QKeySequence('q'), self, self.quitClicked)
        self.show()

    def adjustMult(self):
        key = self.sender().objectName()
        if not self._ignore:
            self._adjust_mul[key] = self._adjust_rnd[key].value()
            self._adjust_cty[key].setValue(self._data[key][0] * self._adjust_rnd[key].value())
        mw, mwtxt, mwstr, div = self.setAdjValueUnits(key, self._adjust_typ[key], self._data[key][0])
        self._adjust_txt[key].setText(mwtxt)
     #   if not self._ignore:
      #      self._adjust_val[key].setText(mwstr)
        self._ignore = False

    def adjustCap(self):
        if self._ignore:
            return
        key = self.sender().objectName()
        if key != 'Load':
            adj = self._adjust_cty[key].value() / self._data[key][0]
         #   self._adjust_rnd[key].setValue(adj)
        else:
            dimen = log10(self._data[key][0])
            if dimen > 11:
                mul = 9
            elif dimen > 8:
                mul = 6
            elif dimen > 5:
                mul = 3
            else:
                mul = 0
            adj = (self._adjust_cty[key].value() * pow(10, mul)) / self._data[key][0]
        self._adjust_mul[key] = adj
      #  self._adjust_cty[key] = self._data[key] * adj
        self._ignore = True
        self._adjust_rnd[key].setValue(round(adj, 4))
        self._ignore = False

    def quitClicked(self):
        self.ignoreEnter = False
        self.close()

    def resetClicked(self, to):
        if to is None:
            to = 0.
        else:
            to = 1.
        if self.show_multipliers:
            for key in self._adjust_rnd.keys():
                self._adjust_rnd[key].setValue(to)
        else:
            if to == 0:
                for key in self._adjust_cty.keys():
                    self._adjust_cty[key].setValue(0.)
            else:
                for key in self._adjust_cty.keys():
                    self._adjust_cty[key].setValue(self._data[key][0])
        self.pfx_fld.setText('')

    def resetloadClicked(self, to):
        if isinstance(to, bool):
            to = 1.
        if self.show_multipliers:
            self._adjust_rnd['Load'].setValue(to)
        else:
            self._adjust_cty['Load'].setValue(self._data['Load'][0])

    def restoreClicked(self):
        ini_file = QFileDialog.getOpenFileName(self, 'Open Adjustments file',
                   self._save_folder, 'Preferences Files (*.ini)')[0]
        if ini_file != '':
            self._ignore = True
            reshow = False
            config = configparser.RawConfigParser()
            config.read(ini_file)
            try:
                prefix = ini_file[ini_file.rfind('/') + 1: - 4]
            except:
                prefix = ''
            self.getIt(config, prefix)

    def getIt(self, config, prefix=''):
        try:
            adjustto = self.config.get('Powermatch', 'adjusted_capacities')
        except:
            return
        self.resetClicked(to=None)
        bits = adjustto.split(',')
        for bit in bits:
            bi = bit.split('=')
            key = bi[0]
            try:
                mw, mwtxt, mwcty, div = self.setAdjValueUnits(key, self._adjust_typ[key],
                                        float(bi[1]))
                self._adjust_cty[key].setValue(mwcty)
                if self.show_multipliers:
                    self._adjust_mul[key] = float(bi[1]) / (self._data[key][0] * pow(10, self._data[key][1]))
                    self._adjust_rnd[key].setValue(round(self._adjust_mul[key], 4))
            except:
                pass
        self._ignore = False
        self.pfx_fld.setText(prefix)

    def listClicked(self):
        if os.path.exists(self._save_folder):
            names = {}
            techs = []
            ini_files = os.listdir(self._save_folder)
            ini_files.sort()
            for ini_file in ini_files:
                if ini_file[-4:] == '.ini':
                    config = configparser.RawConfigParser()
                    try:
                        config.read(self._save_folder + ini_file)
                    except:
                        continue
                    try:
                        adjustto = self.config.get('Powermatch', 'adjusted_capacities')
                    except:
                        continue
                    names[ini_file[:-4]] = [0] * len(techs)
                    bits = adjustto.split(',')
                    for bit in bits:
                        bi = bit.split('=')
                        key = bi[0]
                        try:
                            mw, mwtxt, mwcty, div = self.setAdjValueUnits(key, self._adjust_typ[key],
                                                                          float(bi[1]))
                        except:
                            mwcty = 0
                        if mwcty == 0:
                            continue
                        if key not in techs:
                            techs.append(key)
                            names[ini_file[:-4]].append(0)
                        ndx = techs.index(key)
                        names[ini_file[:-4]][ndx] = mwcty
            techs.insert(0, 'Preference File')
            decpts = [1] * len(techs)
            dialog = displaytable.Table(names, title=self.sender().text(), decpts=decpts, fields=techs,
                                        save_folder=self._save_folder)
            dialog.exec_()
            chosen = dialog.getItem(0)
            self._ignore = True
            reshow = False
            config = configparser.RawConfigParser()
            config.read(self._save_folder + chosen + '.ini')
            self.getIt(config, chosen)
            del dialog

    def saveClicked(self):
        line = ''
        for key, value in self._adjust_cty.items():
            if self._decpts == 2:
                line += '{}={:.2f},'.format(key, value.value() * pow(10, self._data[key][1]))
            else:
                line += '{}={:.1f},'.format(key, value.value() * pow(10, self._data[key][1]))
        if line != '':
            line = 'adjusted_capacities=' + line[:-1]
            updates = {'Powermatch': ['adjustments=', line]}
            save_file = self._save_folder
            if self.pfx_fld.text() != '':
                save_file += '/' + self.pfx_fld.text()
            inifile = QFileDialog.getSaveFileName(None, 'Save Adjustments to file',
                      save_file, 'Preferences Files (*.ini)')[0]
            if inifile != '':
                if inifile[-4:] != '.ini':
                    inifile = inifile + '.ini'
                SaveIni(updates, ini_file=inifile)

    def showClicked(self):
        self.ignoreEnter = False
        self._results = {}
        for key in list(self._adjust_cty.keys()):
            self._results[key] = self._adjust_cty[key].value() * pow(10, self._data[key][1])
        self.close()

    def getValues(self):
        return self._results

    def getPrefix(self):
        return self.pfx_fld.text()

    def addtoBatch(self):
        check_list = list(self._adjust_cty.keys())[1:]
        # wb = oxl.load_workbook(self._batch_file)
        wb = ExcelProcessor(self._batch_file)
        batch_input_sheet = wb.get_worksheet(0)
        batch_input_sheet.protection.sheet = False
        normal = wb.get_font(name='Arial')
        bold = wb.get_font(name='Arial', bold=True)
        col = batch_input_sheet.max_column + 1
        tot_row = -1
        fst_row = -1
        if col == 4: # possibly only chart stuff in columns 2 and 3
            get_out = False
            for col in range(3, 1, -1):
                for row in range(1, batch_input_sheet.max_row + 1):
                    if batch_input_sheet.cell(row=row, column=col).value is not None:
                        col += 1
                        get_out = True
                        break
                    if batch_input_sheet.cell(row=row, column=1).value == 'Total':
                        break
                if get_out:
                    break
        for row in range(1, batch_input_sheet.max_row + 1):
            if batch_input_sheet.cell(row=row, column=1).value is None:
                continue
            if batch_input_sheet.cell(row=row, column=1).value in ['Model', 'Model Label', 'Technology']:
                new_cell = batch_input_sheet.cell(row=row, column=col)
                new_cell.value = QtCore.QDateTime.toString(QtCore.QDateTime.currentDateTime(), 'MM-dd hh:mm')
                add_msg = new_cell.value
            if batch_input_sheet.cell(row=row, column=1).value == 'Capacity (MW)':
                fst_row = row + 1
                cell = batch_input_sheet.cell(row=row, column=col - 1)
                new_cell = batch_input_sheet.cell(row=row, column=col)
                new_cell.value = 'MW'
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
                continue
            for key in self._adjust_cty.keys():
                if key == batch_input_sheet.cell(row=row, column=1).value:
                    cell = batch_input_sheet.cell(row=fst_row, column=col - 1)
                    new_cell = batch_input_sheet.cell(row=row, column=col)
                    new_cell.value = self._adjust_cty[key].value()
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)
                        if col == 2:
                            new_cell.font = normal
                            new_cell.number_format = '#0.00'
                        else:
                            new_cell.number_format = copy(cell.number_format)
                    elif col == 2:
                        new_cell.font = normal
                        new_cell.number_format = '#0.00'
                    try:
                        i = check_list.index(key)
                        del check_list[i]
                    except:
                        pass
            if batch_input_sheet.cell(row=row, column=1).value == 'Total':
                tot_row = row
           #     if len(check_list) > 0:
           #         tot_row = row
        if len(check_list) > 0:
            check_list.reverse()
            cell = batch_input_sheet.cell(row=fst_row, column=col)
            for key in check_list:
                if self._adjust_cty[key].value() == 0:
                    continue
                batch_input_sheet.insert_rows(tot_row)
                new_cell = batch_input_sheet.cell(row=tot_row, column=1)
                new_cell.value = key
                new_cell = batch_input_sheet.cell(row=tot_row, column=col)
                new_cell.value = self._adjust_cty[key].value()
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
                tot_row += 1
        if fst_row > 0 and tot_row > 0:
            new_cell = batch_input_sheet.cell(row=tot_row, column=col)
            new_cell.value = '=SUM(' + ssCol(col) + str(fst_row) + ':' + ssCol(col) + str(tot_row - 1) + ')'
            if col > 2:
                cell = batch_input_sheet.cell(row=tot_row, column=2)
            else:
                cell = batch_input_sheet.cell(row=tot_row, column=col)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
        wb.save()
        QMessageBox.about(self, 'SIREN - Add to Batch', "Added to batch as '" + add_msg + "' (column " + ssCol(col) + ')')

class setTransition(QDialog):
    def niceSize(window, ctr): # works for Adjustments window (probably because less that 640*480)
        height = window.frameSize().height() / 1.07
        height = 70 + ctr * 32
        width = window.frameSize().width()
        screen = QDesktopWidget().availableGeometry()
        if height > (screen.height() - 70):
            height = screen.height() - 70
        if width > (screen.width() - 70):
            width = screen.width() - 70
        size = QtCore.QSize(QtCore.QSize(int(width), int(height)))
        window.resize(size)

    def __init__(self, parent, label, generators, sheet, year):
        super(setTransition, self).__init__()
        self._results = None
        i = generators.rfind('/')
        generator_file = generators[i + 1:]
        self.grid = QGridLayout()
        r = 0
        self.grid.addWidget(QLabel(label + ' File:'), r, 0)
        file_name = QLabel(generator_file)
        file_name.setStyleSheet("border: 1px inset grey; min-height: 22px; border-radius: 4px;")
        self.grid.addWidget(file_name, r, 1, 1, 5)
        r += 1
        self.grid.addWidget(QLabel(label + ' Sheet:'), r, 0)
        self.sheet = QLineEdit()
        if sheet[-4:].isnumeric():
            sheet = sheet[:-4] + '$YEAR$'
        else:
            sheet = sheet.replace(year, '$YEAR$')
        self.sheet.setText(sheet)
        self.grid.addWidget(self.sheet, r, 1, 1, 2)
        r += 1
        quit = QPushButton('Quit', self)
        self.grid.addWidget(quit, r, 0)
        quit.clicked.connect(self.quitClicked)
        show = QPushButton('Proceed', self)
        self.grid.addWidget(show, r, 1)
        show.clicked.connect(self.showClicked)
        frame = QFrame()
        frame.setLayout(self.grid)
        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setWidget(frame)
        self.layout = QVBoxLayout(self)
        self.layout.addWidget(self.scroll)
        self.niceSize(r)
        self.setWindowTitle('SIREN - Powermatch - Transition files')
        self.setWindowIcon(QtGui.QIcon('resources/resources/sen_icon32.ico'))
        QShortcut(QtGui.QKeySequence('q'), self, self.quitClicked)
        self.show()

    def quitClicked(self):
        self.ignoreEnter = False
        self.close()

    def showClicked(self):
        self.ignoreEnter = False
        self._results = self.sheet.text()
        self.close()

    def getValues(self):
        return self._results
    
class PyQtProgressHandler(ProgressHandler):
    def __init__(self, progress_bar, log_output):
        self.progress_bar = progress_bar
        self.log_output = log_output

    def update_progress(self, value: int):
        """Update the PyQt5 progress bar."""
        self.progress_bar.setValue(value)

    def log_message(self, message: str):
        """Log messages in the PyQt5 interface."""
        print(message)  # Replace with a UI log widget if needed

class PyQtProgressListener(ProgressHandler):
    def __init__(self, status_label: QLabel, progress_bar: QProgressBar):
        self.status_label = status_label
        self.progress_bar = progress_bar

    def on_status_update(self, message: str):
        """Update the status label."""
        self.status_label.setText(message)

    def on_progress_update(self, progress: int):
        """Update the progress bar."""
        self.progress_bar.setValue(progress)
        
class powerMatchUI(QWidget):
    log = pyqtSignal(str)
    progress = pyqtSignal(str)
    
    def get_filename(self, filename):
        if filename.find('/') == 0: # full directory in non-Windows
            return filename
        elif (sys.platform == 'win32' or sys.platform == 'cygwin') \
          and filename[1:2] == ':/': # full directory for Windows
            return filename
        elif filename[:3] == '../': # directory upwards of scenarios
            ups = filename.split('../')
            scens = self.scenarios.split('/')
            scens = scens[: -(len(ups) - 1)]
            scens.append(ups[-1])
            return '/'.join(scens)
        else: # subdirectory of scenarios
            return self.scenarios + filename
        
    def get_load_years(self):
        load_years = ['n/a']
        i = self.load_files.find('$YEAR$')
        if i < 0:
            return load_years
        j = len(self.load_files) - i - 6
        files = glob.glob(self.load_files[:i] + '*' + self.load_files[i + 6:])
        for fil in files:
            load_years.append(fil[i:len(fil) - j])
        return sorted(load_years, reverse=True)
    
    def __init__(self, config, settings, help='help.html'):
        super().__init__()
        self.help = help
        self.config = config
        self.settings = settings
        self.file_handler = FileHandler()
        self.setup_ui(config, settings)
        self.initialize_processor(self.config)
        
    def initialize_processor(self, config):
        """Initialize the processor with a PyQt5-specific listener."""
        listener = PyQtProgressListener(self.status_label, self.progress_bar)
        self.processor = PowerMatchProcessor(
            config, self.scenarios, self.generators, self.constraints, listener,
            event_callback=QApplication.processEvents,
            status_callback=self.setStatus  # Pass setStatus as a callback
            )
    def setup_ui(self, config, settings):
        self.status_label = QLabel(self)
        self.progress_bar = QProgressBar(self)
        parents = settings['parents']
        self.batch_template = settings['batch_template']
        self.scenarios = settings['scenarios']
        self.load_files = settings['load_files']
        self._load_folder = settings['_load_folder']
        self.log_status = settings['log_status']
        self.log_status = True
        self.file_labels = ['Constraints', 'Generators', 'Optimisation', 'Data', 'Results', 'Batch']
        ifiles = [''] * len(self.file_labels)
        self.isheets = self.file_labels[:]
        del self.isheets[-2:]
        self.batch_new_file = False
        self.batch_prefix = False
        self.more_details = False
        self.constraints = None
        self.generators = None
        self.optimisation = None
        self.adjustto = None # adjust capacity to this
        self.adjust_cap = 25
        self.adjust_gen = False
        self.change_res = True
        self.adjusted_lcoe = True
        self.carbon_price = 0.
        self.carbon_price_max = 200.
        self.discount_rate = 0.
        self.load_folder = ''
        self.load_year = 'n/a'
        self.optimise_choice = 'LCOE'
        self.optimise_generations = 20
        self.optimise_mutation = 0.005
        self.optimise_population = 50
        self.optimise_stop = 0
        self.optimise_debug = False
        self.optimise_default = None
        self.optimise_multiplot = True
        self.optimise_multisurf = False
        self.optimise_multitable = False
        self.optimise_to_batch = True
        self.remove_cost = True
        self.results_prefix = ''
        self.dispatchable = ['Biomass', 'Geothermal', 'Pumped Hydro', 'Solar Thermal', 'CST'] # RE dispatchable
        self.save_tables = False
        self.show_multipliers = False
        self.show_correlation = False
        self.summary_sources = True
        self.surplus_sign = 1 # Note: Preferences file has it called shortfall_sign
        # it's easier for the user to understand while for the program logic surplus is easier
        self.underlying = ['Rooftop PV'] # technologies contributing to underlying (but not operational) load
        self.operational = []
        iorder = []
        self.targets = {}
        for t in range(len(target_keys)):
            if target_keys[t] in ['re_pct', 'surplus_pct']:
                self.targets[target_keys[t]] = [target_names[t], 0., -1, 0., 0, target_fmats[t],
                                                 target_titles[t]]
            else:
                self.targets[target_keys[t]] = [target_names[t], 0., 0., -1, 0, target_fmats[t],
                                                 target_titles[t]]
        try:
            dts = self.config.get('Grid', 'dispatchable').split(' ')
            dispatchable = []
            for dt in dts:
                dispatchable.append(techClean(dt.replace('_', ' ').title()))
            self.dispatchable = dispatchable
        except:
            pass
        try:
            adjust_cap = self.config.get('Power', 'adjust_cap')
            try:
                self.adjust_cap = float(adjust_cap)
            except:
                try:
                    self.adjust_cap = eval(adjust_cap)
                except:
                    pass
            if self.adjust_cap < 0:
                self.adjust_cap = pow(10, 12)
        except:
            pass
        try:
            items = self.config.items('Powermatch')
            for key, value in items:
                if key == 'batch_new_file':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.batch_new_file = True
                elif key == 'batch_prefix':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.batch_prefix = True
                elif key[:4] == 'tml_':
                    continue
                elif key[-5:] == '_file':
                    ndx = self.file_labels.index(key[:-5].title())
                    ifiles[ndx] = value.replace('$USER$', getUser())
                elif key[-6:] == '_sheet':
                    ndx = self.file_labels.index(key[:-6].title())
                    self.isheets[ndx] = value
                elif key == 'adjust_generators':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.adjust_gen = True
                elif key == 'adjusted_capacities':
                    self.adjustto = {}
                    bits = value.split(',')
                    for bit in bits:
                        bi = bit.split('=')
                        self.adjustto[bi[0]] = float(bi[1])
                elif key == 'carbon_price':
                    try:
                        self.carbon_price = float(value)
                    except:
                        pass
                elif key == 'carbon_price_max':
                    try:
                        self.carbon_price_max = float(value)
                    except:
                        pass
                elif key == 'change_results':
                    if value.lower() in ['false', 'off', 'no']:
                        self.change_res = False
                elif key == 'adjusted_lcoe' or key == 'corrected_lcoe':
                    if value.lower() in ['false', 'no', 'off']:
                        self.adjusted_lcoe = False
                elif key == 'discount_rate':
                    try:
                        self.discount_rate = float(value)
                    except:
                        pass
                elif key == 'dispatch_order':
                    iorder = value.split(',')
                elif key == 'load':
                    try:
                        self.load_files = value
                        for ky, valu in parents:
                            self.load_files = self.load_files.replace(ky, valu)
                        self.load_files = self.load_files.replace('$USER$', getUser())
                    except:
                        pass
                elif key == 'load_year':
                    self.load_year = value
                elif key == 'log_status':
                    if value.lower() in ['false', 'no', 'off']:
                        self.log_status = False
                elif key == 'more_details':
                    if value.lower() in ['true', 'yes', 'on']:
                        self.more_details = True
                elif key == 'optimise_debug':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.optimise_debug = True
                elif key == 'optimise_default':
                    self.optimise_default = value
                elif key == 'optimise_choice':
                    self.optimise_choice = value
                elif key == 'optimise_generations':
                    try:
                        self.optimise_generations = int(value)
                    except:
                        pass
                elif key == 'optimise_multiplot':
                    if value.lower() in ['false', 'off', 'no']:
                        self.optimise_multiplot = False
                    elif value.lower() in ['surf', 'tri-surf', 'trisurf']:
                        self.optimise_multisurf = True
                elif key == 'optimise_multitable':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.optimise_multitable = True
                elif key == 'optimise_mutation':
                    try:
                        self.optimise_mutation = float(value)
                    except:
                        pass
                elif key == 'optimise_population':
                    try:
                        self.optimise_population = int(value)
                    except:
                        pass
                elif key == 'optimise_stop':
                    try:
                        self.optimise_stop = int(value)
                    except:
                        pass
                elif key == 'optimise_to_batch':
                    if value.lower() in ['false', 'off', 'no']:
                        self.optimise_to_batch = False
                elif key[:9] == 'optimise_':
                    try:
                        bits = value.split(',')
                        t = target_keys.index(key[9:])
                        # name, weight, minimum, maximum, widget index
                        self.targets[key[9:]] = [target_names[t], float(bits[0]), float(bits[1]),
                                                float(bits[2]), 0, target_fmats[t],
                                                 target_titles[t]]
                    except:
                        pass
                elif key == 'remove_cost':
                    if value.lower() in ['false', 'off', 'no']:
                        self.remove_cost = False
                elif key == 'results_prefix':
                    self.results_prefix = value
                elif key == 'save_tables':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.save_tables = True
                elif key == 'show_correlation':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.show_correlation = True
                elif key == 'show_multipliers':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.show_multipliers = True
                elif key == 'shortfall_sign':
                    if value[0] == '+' or value[0].lower() == 'p':
                        self.surplus_sign = -1
                elif key == 'summary_sources':
                    if value.lower() in ['false', 'off', 'no']:
                        self.summary_sources = False
                elif key == 'underlying':
                    self.underlying = value.split(',')
                elif key == 'operational':
                    self.operational = value.split(',')
        except:
            print('PME1: Error with', key)
            pass
        self.restorewindows = False
        try:
            rw = self.config.get('Windows', 'restorewindows')
            if rw.lower() in ['true', 'yes', 'on']:
                self.restorewindows = True
        except:
            pass
        self.opt_progressbar = None
        self.floatstatus = None # status window
        self.grid = QGridLayout()
        self.labels = [None] * len(self.file_labels)
        self.files = [None] * len(self.file_labels)
        self.sheets = self.file_labels[:]
        del self.sheets[-2:]
        self.updated = False
        edit = [None] * D
        r = 0
        for i in range(len(self.file_labels)):
            if i == R:
                self.grid.addWidget(QLabel('Results Prefix:'), r, 0)
                self.results_pfx_fld = QLineEdit()
                self.results_pfx_fld.setText(self.results_prefix)
                self.results_pfx_fld.textChanged.connect(self.pfxChanged)
                self.grid.addWidget(self.results_pfx_fld, r, 1, 1, 2)
                r += 1
            self.labels[i] = QLabel(self.file_labels[i] + ' File:')
            self.grid.addWidget(self.labels[i], r, 0)
            self.files[i] = ClickableQLabel()
            self.files[i].setStyleSheet("background-color: white; border: 1px inset grey; min-height: 22px; border-radius: 4px;")
            self.files[i].setText(ifiles[i])
            self.files[i].clicked.connect(self.fileChanged)
            self.grid.addWidget(self.files[i], r, 1, 1, 5)
            button = QPushButton(f'Open {self.file_labels[i]} file', self)
            self.grid.addWidget(button, r, 6)
            button.clicked.connect(self.openClicked)
            if i < D:
                r += 1
                self.grid.addWidget(QLabel(self.file_labels[i] + ' Sheet:'), r, 0)
                self.sheets[i] = QComboBox()
                try:
                    curfile = self.get_filename(ifiles[i])
                    ts = WorkBook()
                    ts.open_workbook(curfile)
                    ndx = 0
                    j = -1
                    for sht in ts.sheet_names():
                        j += 1
                        self.sheets[i].addItem(sht)
                        if sht == self.isheets[i]:
                            ndx = j
                    self.sheets[i].setCurrentIndex(ndx)
                    ws = ts.sheet_by_index(ndx)
                    if i == G:
                        self.getGenerators(ws)
                    elif i == O:
                        self.getOptimisation(ws)
                    ts.close()
                    del ts
                except:
                    self.sheets[i].addItem(self.isheets[i])
                self.grid.addWidget(self.sheets[i], r, 1, 1, 3)
                self.sheets[i].currentIndexChanged.connect(self.sheetChanged)
                edit[i] = QPushButton(self.file_labels[i], self)
                self.grid.addWidget(edit[i], r, 4, 1, 2)
                edit[i].clicked.connect(self.editClicked)
            elif i == D and self.load_files != '':
                r += 1
                self.grid.addWidget(QLabel('Load Folder:'), r, 0)
                self.load_dir = ClickableQLabel()
                try:
                    self.load_dir.setText(self.load_files[:self.load_files.rfind('/')])
                except:
                    self.load_dir.setText('')
                self.load_dir.setStyleSheet("background-color: white; border: 1px inset grey; min-height: 22px; border-radius: 4px;")
                self.load_dir.clicked.connect(self.loaddirChanged)
                self.grid.addWidget(self.load_dir, r, 1, 1, 5)
                r += 1
                self.grid.addWidget(QLabel('Load Year:'), r, 0)
                self.load_years = self.get_load_years()
                self.loadCombo = QComboBox()
                for choice in self.load_years:
                    self.loadCombo.addItem(choice)
                    if choice == self.load_year:
                        self.loadCombo.setCurrentIndex(self.loadCombo.count() - 1)
                self.loadCombo.currentIndexChanged.connect(self.changes)
                self.grid.addWidget(self.loadCombo, r, 1)
                self.grid.addWidget(QLabel("(To to use a different load year to the data file. Otherwise choose 'n/a')"), r, 2, 1, 4)
            r += 1
      #  wdth = edit[1].fontMetrics().boundingRect(edit[1].text()).width() + 9
        self.grid.addWidget(QLabel('Replace Last:'), r, 0)
        if self.batch_new_file:
            msg = '(check to replace an existing Results workbook)'
        else:
            msg = '(check to replace last Results worksheet in Batch spreadsheet)'
        self.replace_last = QCheckBox(msg, self)
        self.replace_last.setCheckState(QtCore.Qt.Unchecked)
        self.grid.addWidget(self.replace_last, r, 1, 1, 3)
        self.grid.addWidget(QLabel('Prefix facility names in Batch report:'), r, 4)
        self.batch_prefix_check = QCheckBox('', self)
        if self.batch_prefix:
            self.batch_prefix_check.setCheckState(QtCore.Qt.Checked)
        else:
            self.batch_prefix_check.setCheckState(QtCore.Qt.Unchecked)
        self.grid.addWidget(self.batch_prefix_check, r, 5)
        self.batch_prefix_check.stateChanged.connect(self.bpcchanged)
        r += 1
        self.grid.addWidget(QLabel('Discount Rate:'), r, 0)
        self.discount = QDoubleSpinBox()
        self.discount.setRange(0, 100)
        self.discount.setDecimals(2)
        self.discount.setSingleStep(.5)
        try:
            self.discount.setValue(self.discount_rate * 100.)
        except:
            self.discount.setValue(0.)
        self.grid.addWidget(self.discount, r, 1)
        self.discount.valueChanged.connect(self.drchanged)
        self.grid.addWidget(QLabel('(%. Only required if using input costs rather than reference LCOE)'), r, 2, 1, 4)
        r += 1
        self.grid.addWidget(QLabel('Carbon Price:'), r, 0)
        self.carbon = QDoubleSpinBox()
        self.carbon.setRange(0, self.carbon_price_max)
        self.carbon.setDecimals(2)
        try:
            self.carbon.setValue(self.carbon_price)
        except:
            self.carbon.setValue(0.)
        self.grid.addWidget(self.carbon, r, 1)
        self.carbon.valueChanged.connect(self.cpchanged)
        self.grid.addWidget(QLabel('($/tCO2e. Use only if LCOE excludes carbon price)'), r, 2, 1, 4)
        r += 1
        self.grid.addWidget(QLabel('Adjust Generators:'), r, 0)
        self.adjust = QCheckBox('(check to adjust generators capacity data)', self)
        if self.adjust_gen:
            self.adjust.setCheckState(QtCore.Qt.Checked)
        self.grid.addWidget(self.adjust, r, 1, 1, 4)
        r += 1
        self.grid.addWidget(QLabel('Dispatch Order:\n(move to right\nto exclude)'), r, 0)
        self.order = ListWidget(self) #QListWidget()
      #  self.order.setDragDropMode(QtGui.QAbstractItemView.InternalMove)
        self.grid.addWidget(self.order, r, 1, 1, 3)
        self.ignore = ListWidget(self) # QListWidget()
      #  self.ignore.setDragDropMode(QtGui.QAbstractItemView.InternalMove)
        self.grid.addWidget(self.ignore, r, 4, 1, 3)
        r += 1
        self.log = QLabel('')
        msg_palette = QtGui.QPalette()
        msg_palette.setColor(QtGui.QPalette.Foreground, QtCore.Qt.red)
        self.log.setPalette(msg_palette)
        self.grid.addWidget(self.log, r, 1, 1, 6)
        r += 1
        self.progressbar = QProgressBar()
        self.progressbar.setMinimum(0)
        self.progressbar.setMaximum(20) # was 10 set to 20 to get 5% steps
        self.progressbar.setValue(0)
        self.progressbar.setStyleSheet('QProgressBar {border: 1px solid grey; border-radius: 2px; text-align: center;}' \
                                       + 'QProgressBar::chunk { background-color: #06A9D6;}')
        self.grid.addWidget(self.progressbar, r, 1, 1, 6)
        self.progressbar.setHidden(True)
        r += 1
        r += 1
        quit = QPushButton('Done', self)
        self.grid.addWidget(quit, r, 0)
        quit.clicked.connect(self.quitClicked)
        QShortcut(QtGui.QKeySequence('q'), self, self.quitClicked)
        pms = QPushButton('Summary', self)
        self.grid.addWidget(pms, r, 1)
        pms.clicked.connect(self.pmClicked)
        pm = QPushButton('Detail', self)
     #   pm.setMaximumWidth(wdth)
        self.grid.addWidget(pm, r, 2)
        pm.clicked.connect(self.pmClicked)
        btch = QPushButton('Batch', self)
        self.grid.addWidget(btch, r, 3)
        btch.clicked.connect(self.pmClicked)
        trns = QPushButton('Transition', self)
        self.grid.addWidget(trns, r, 4)
        trns.clicked.connect(self.pmClicked)
        opt = QPushButton('Optimise', self)
        self.grid.addWidget(opt, r, 5)
        opt.clicked.connect(self.pmClicked)
        help = QPushButton('Help', self)
     #   help.setMaximumWidth(wdth)
      #  quit.setMaximumWidth(wdth)
        self.grid.addWidget(help, r, 6)
        help.clicked.connect(self.helpClicked)
        QShortcut(QtGui.QKeySequence('F1'), self, self.helpClicked)
     #   self.grid.setColumnStretch(0, 2)
        r += 1
        editini = QPushButton('Preferences', self)
     #   editini.setMaximumWidth(wdth)
        self.grid.addWidget(editini, r, 0)
        editini.clicked.connect(self.editIniFile)
        do_tml = False
        if sys.platform == 'win32' or sys.platform == 'cygwin':
            if os.path.exists('pmtmldetail.exe'):
                do_tml = True
        elif os.path.exists('pmtmldetail.py'):
            do_tml = True
        if do_tml:
            tmld = QPushButton('TML Detail', self)
            self.grid.addWidget(tmld, r, 1)
            tmld.clicked.connect(self.tmlClicked)
        self.setOrder()
        if len(iorder) > 0:
            self.order.clear()
            self.ignore.clear()
            for gen in iorder:
                self.order.addItem(gen)
            try:
                for gen in self.generators.keys():
                    if (gen in tech_names and gen not in self.dispatchable) or gen in iorder:
                        continue
                    try:
                        chk = gen[gen.find('.') + 1:]
                        if chk in tech_names and chk not in self.dispatchable:
                            continue
                    except:
                        pass
                    self.ignore.addItem(gen)
            except:
                pass
        if self.adjust_gen and self.adjustto is None:
           self.adjustto = {}
           self.adjustto['Load'] = 0
           for gen in tech_names:
               try:
                   if self.generators[gen].capacity > 0:
                       self.adjustto[gen] = self.generators[gen].capacity
               except:
                   pass
           for gen in iorder:
               try:
                   if self.generators[gen].capacity > 0:
                       self.adjustto[gen] = self.generators[gen].capacity
               except:
                   pass
        frame = QFrame()
        frame.setLayout(self.grid)
        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setWidget(frame)
        self.layout = QVBoxLayout(self)
        self.layout.addWidget(self.scroll)
        self.setWindowTitle('SIREN - powermatch (' + fileVersion() + ') - Powermatch')
        self.setWindowIcon(QtGui.QIcon('resources/resources/sen_icon32.ico'))
        if self.restorewindows:
            try:
                rw = self.config.get('Windows', 'powermatch_size').split(',')
                self.resize(int(rw[0]), int(rw[1]))
                mp = self.config.get('Windows', 'powermatch_pos').split(',')
                self.move(int(mp[0]), int(mp[1]))
            except:
                pass
        else:
            self.center()
            self.resize(int(self.sizeHint().width() * 1.2), int(self.sizeHint().height() * 1.2))
        self.show_FloatStatus() # status window
        self.show()

    def center(self):
        frameGm = self.frameGeometry()
        screen = QApplication.desktop().screenNumber(QApplication.desktop().cursor().pos())
        centerPoint = QApplication.desktop().availableGeometry(screen).center()
        frameGm.moveCenter(centerPoint)
        self.move(frameGm.topLeft())

    def fileChanged(self):
        self.setStatus('')
        for i in range(len(self.file_labels)):
            if self.files[i].hasFocus():
                break
        if self.files[i].text() == '':
            curfile = self.scenarios[:-1]
        else:
            curfile = self.get_filename(self.files[i].text())
        if i == R:
            if self.files[i].text() == '':
                curfile = self.get_filename(self.files[D].text())
                curfile = curfile.replace('data', 'results')
                curfile = curfile.replace('Data', 'Results')
                if curfile == self.scenarios + self.files[D].text():
                    j = curfile.find(' ')
                    if j > 0:
                        jnr = ' '
                    else:
                        jnr = '_'
                    j = curfile.rfind('.')
                    curfile = curfile[:j] + jnr + 'Results' + curfile[j:]
            else:
                curfile = self.get_filename(self.files[R].text())
            newfile = QFileDialog.getSaveFileName(None, 'Save ' + self.file_labels[i] + ' file',
                      curfile, 'Excel Files (*.xlsx)')[0]
        elif i == B and not self.batch_new_file:
            options = QFileDialog.Options()
            # options |= QFileDialog.DontUseNativeDialog
            newfile = QFileDialog.getSaveFileName(None, 'Open/Create and save ' + self.file_labels[i] + ' file',
                      curfile, 'Excel Files (*.xlsx)', options=options)[0]
        else:
            newfile = QFileDialog.getOpenFileName(self, 'Open ' + self.file_labels[i] + ' file',
                      curfile)[0]
        if newfile != '':
            if i < D:
                if i == C:
                    self.constraints = None
                elif i == G:
                    self.generators = None
                elif i == O:
                    self.optimisation = None
                ts = WorkBook()
                ts.open_workbook(newfile)
                ndx = 0
                self.sheets[i].clear()
                j = -1
                for sht in ts.sheet_names():
                    j += 1
                    self.sheets[i].addItem(sht)
                    if len(sht) >= len(self.file_labels[i]):
                        if sht[:len(self.file_labels[i])].lower() == self.file_labels[i].lower():
                            ndx = j
                self.sheets[i].setCurrentIndex(ndx)
                if i == G:
                    ws = ts.sheet_by_index(ndx)
                    self.getGenerators(ws)
                    self.setOrder()
                ts.close()
                del ts
            if newfile[: len(self.scenarios)] == self.scenarios:
                self.files[i].setText(newfile[len(self.scenarios):])
            else:
                if newfile.rfind('/') > 0:
                    that_len = len(commonprefix([self.scenarios, newfile]))
                    if that_len > 0:
                        bits = self.scenarios[that_len:].split('/')
                        pfx = ('..' + '/') * (len(bits) - 1)
                        newfile = pfx + newfile[that_len + 1:]
                self.files[i].setText(newfile)
            if i == D and self.change_res:
                newfile = self.files[D].text()
                newfile = newfile.replace('data', 'results')
                newfile = newfile.replace('Data', 'Results')
                if newfile != self.files[D].text():
                    self.files[R].setText(newfile)
            self.updated = True

    def pfxChanged(self):
        self.results_prefix = self.results_pfx_fld.text()
     #   self.setStatus('Results filename will be ' + self.results_pfx_fld.text() + '_' + self.files[R].text())
        self.updated = True

    def sheetChanged(self, i):
        try:
            for i in range(3):
                if self.sheets[i].hasFocus():
                    break
            else:
                return
        except:
            return # probably file changed
        self.setStatus('')
        newfile = self.get_filename(self.files[i].text())
        ts = WorkBook()
        ts.open_workbook(newfile)
        ws = ts.sheet_by_name(self.sheets[i].currentText())
        self.setStatus('Sheet ' + self.sheets[i].currentText() + ' loaded')
        if i == G:
            self.getGenerators(ws)
            self.setOrder()
        elif i == O:
            self.getOptimisation(ws)
        ts.close()
        del ts

    def loaddirChanged(self):
        curdir = self.load_dir.text()
        newdir = QFileDialog.getExistingDirectory(self, 'Choose Load File Folder',
                 curdir, QFileDialog.ShowDirsOnly)
        if newdir != '':
            try:
                self.load_files = newdir + self.load_files[self.load_files.rfind('/'):]
            except:
                self.load_files = newdir + self.load_files
            if newdir[: len(self.scenarios)] == self.scenarios:
                self.load_dir.setText(newdir[len(self.scenarios):])
            else:
                if newdir.rfind('/') > 0:
                    that_len = len(commonprefix([self.scenarios, newdir]))
                    if that_len > 0:
                        bits = self.scenarios[that_len:].split('/')
                        pfx = ('..' + '/') * (len(bits) - 1)
                        newdir = pfx + newdir[that_len + 1:]
                self.load_dir.setText(newdir)
            self.load_years = self.get_load_years()
            self.loadCombo.clear()
            for choice in self.load_years:
                self.loadCombo.addItem(choice)
                if choice == self.load_year:
                    self.loadCombo.setCurrentIndex(self.loadCombo.count() - 1)
            self.updated = True

    def helpClicked(self):
        dialog = displayobject.AnObject(QDialog(), self.help,
                 title='Help for powermatch (' + fileVersion() + ')', section='powermatch')
        dialog.exec_()

    def drchanged(self):
        self.updated = True
        self.discount_rate = self.discount.value() / 100.

    def cpchanged(self):
        self.updated = True
        self.carbon_price = self.carbon.value()

    def changes(self):
        self.updated = True

    def bpcchanged(self):
        if self.batch_prefix_check.isChecked():
            self.batch_prefix = True
        else:
            self.batch_prefix = False
        self.updated = True

    def openClicked(self):
        bit = self.sender().text().split()
        fnr = self.file_labels.index(bit[1])
        curfile = self.get_filename(self.files[fnr].text())
        if not os.path.exists(curfile):
            if fnr == R and self.results_pfx_fld.text() != '':
                i = curfile.rfind('/')
                curfile = curfile[:i + 1] + self.results_pfx_fld.text() + '_' + curfile[i+1:]
                print(curfile)
                if not os.path.exists(curfile):
                    self.setStatus(self.file_labels[fnr] + ' not found.')
                    return
            else:
                self.setStatus(self.file_labels[fnr] + ' not found.')
                return
        if sys.platform == 'win32' or sys.platform == 'cygwin':
            os.startfile(curfile)
        elif sys.platform == 'darwin':
            subprocess.call('open', curfile)
        elif sys.platform == 'linux2' or sys.platform == 'linux':
            subprocess.call(('xdg-open', curfile))
        self.setStatus(self.file_labels[fnr] + ' file "launched".')

    def quitClicked(self):
        if self.updated or self.order.updated or self.ignore.updated:
            updates = {}
            lines = []
            lines.append('adjust_generators=' + str(self.adjust.isChecked()))
            lines.append('adjustments=') # clean up the old way
            if self.adjustto is not None:
                line = ''
                for key, value in self.adjustto.items():
                    line += '{}={:.1f},'.format(key, value)
                if line != '':
                    lines.append('adjusted_capacities=' + line[:-1])
            line = 'batch_prefix='
            if self.batch_prefix:
                line += 'True'
            lines.append(line)
            lines.append('carbon_price=' + str(self.carbon_price))
            lines.append('discount_rate=' + str(self.discount_rate))
            line = ''
            for itm in range(self.order.count()):
                line += self.order.item(itm).text() + ','
            lines.append('dispatch_order=' + line[:-1])
            for i in range(len(self.file_labels)):
                lines.append(self.file_labels[i].lower() + '_file=' + self.files[i].text().replace(getUser(), '$USER$'))
            for i in range(D):
                lines.append(self.file_labels[i].lower() + '_sheet=' + self.sheets[i].currentText())
            line = 'load='
            if self.load_dir.text() != self._load_folder:
                line += self.load_files.replace(getUser(), '$USER$')
            lines.append(line)
            line = 'load_year='
            if self.loadCombo.currentText() != 'n/a':
                line += self.loadCombo.currentText()
            lines.append(line)
            lines.append('optimise_choice=' + self.optimise_choice)
            lines.append('optimise_generations=' + str(self.optimise_generations))
            lines.append('optimise_mutation=' + str(self.optimise_mutation))
            lines.append('optimise_population=' + str(self.optimise_population))
            lines.append('optimise_stop=' + str(self.optimise_stop))
            for key, value in self.targets.items():
                line = 'optimise_{}={:.2f},{:.2f},{:.2f}'.format(key, value[1], value[2], value[3])
                lines.append(line)
            lines.append('results_prefix=' + self.results_prefix)
            updates['Powermatch'] = lines
            SaveIni(updates)
        self.close()

    def closeEvent(self, event):
        if self.floatstatus is not None:
            self.floatstatus.exit()
        if self.restorewindows:
            updates = {}
            lines = []
            add = int((self.frameSize().width() - self.size().width()) / 2)   # need to account for border
            lines.append('powermatch_pos=%s,%s' % (str(self.pos().x() + add), str(self.pos().y() + add)))
            lines.append('powermatch_size=%s,%s' % (str(self.width()), str(self.height())))
            updates['Windows'] = lines
            SaveIni(updates)
        event.accept()

    def tmlClicked(self):
        if len(sys.argv) > 1:
            config_file = sys.argv[1]
        else:
            config_file = getModelFile('SIREN.ini')
        if sys.platform == 'win32' or sys.platform == 'cygwin':
            if os.path.exists('pmtmldetail.exe'):
                pid = subprocess.Popen(['pmtmldetail.exe', config_file]).pid
            elif os.path.exists('pmtmldetail.py'):
                pid = subprocess.Popen(['pmtmldetail.py', config_file], shell=True).pid
        else:
            pid = subprocess.Popen(['python3', 'pmtmldetail.py', config_file]).pid # python3
        return

    def editIniFile(self):
        if len(sys.argv) > 1:
            config_file = sys.argv[1]
        else:
            config_file = getModelFile('SIREN.ini')
        dialr = EdtDialog(config_file, section='[Powermatch]')
        dialr.exec_()
     #   self.get_config()   # refresh config values
        config = configparser.RawConfigParser()
        config.read(config_file)
        try:
            st = self.config.get('Powermatch', 'save_tables')
        except:
            st = 'False'
        if st.lower() in ['true', 'yes', 'on']:
            self.save_tables = True
        else:
            self.save_tables = False
        try:
            st = self.config.get('Powermatch', 'more_details')
        except:
            st = 'False'
        if st.lower() in ['true', 'yes', 'on']:
            self.more_details = True
        else:
            self.more_details = False
        try:
            st = self.config.get('Powermatch', 'optimise_to_batch')
        except:
            st = 'True'
        if st.lower() in ['true', 'yes', 'on']:
            self.optimise_to_batch = True
        else:
            self.optimise_to_batch = False
        try:
            st = self.config.get('Powermatch', 'show_multipliers')
        except:
            st = 'False'
        if st.lower() in ['true', 'yes', 'on']:
            self.show_multipliers = True
        else:
            self.show_multipliers = False
        try:
            st = self.config.get('Powermatch', 'batch_new_file')
        except:
            st = 'False'
        if st.lower() in ['true', 'yes', 'on']:
            self.batch_new_file = True
            msg = '(check to replace an existing Results workbook)'
        else:
            self.batch_new_file = False
            msg = '(check to replace last Results worksheet in Batch spreadsheet)'
        self.replace_last = QCheckBox(msg, self)
        try:
            st = self.config.get('Powermatch', 'batch_prefix')
        except:
            st = 'False'
        if st.lower() in ['true', 'yes', 'on']:
            self.batch_prefix = True
        else:
            self.batch_prefix = False
        QApplication.processEvents()
        self.setStatus(config_file + ' edited. Reload may be required.')

    def editClicked(self):
        def update_dictionary(it, source):
            new_keys = list(source.keys())
            # first we delete and add keys to match updated dictionary
            if it == C:
                old_keys = list(self.constraints.keys())
                for key in old_keys:
                    if key in new_keys:
                        del new_keys[new_keys.index(key)]
                    else:
                        del self.constraints[key]
                for key in new_keys:
                    self.constraints[key] = Constraint(key, '<category>', 0., 1., 1., 1., 1., 0.,
                                                       1., 0., 0., 0, 0)
                target = self.constraints
            elif it == G:
                old_keys = list(self.generators.keys())
                for key in old_keys:
                    if key in new_keys:
                        del new_keys[new_keys.index(key)]
                    else:
                        del self.generators[key]
                for key in new_keys:
                    self.generators[key] = Facility(name=key, constraint='<constraint>')
                target = self.generators
            elif it == O:
                old_keys = list(self.optimisation.keys())
                for key in old_keys:
                    if key in new_keys:
                        del new_keys[new_keys.index(key)]
                    else:
                        del self.optimisation[key]
                for key in new_keys:
                    self.optimisation[key] = Optimisation(key, 'None', None)
                target = self.optimisation
            # now update the data
            for key in list(target.keys()):
                for prop in dir(target[key]):
                    if prop[:2] != '__' and prop[-2:] != '__':
                        try:
                            if prop == 'lifetime' and source[key][prop] == 0:
                                setattr(target[key], prop, 20)
                            else:
                                setattr(target[key], prop, source[key][prop])
                        except:
                            pass

        self.setStatus('')
        msg = ''
        ts = None
        it = self.file_labels.index(self.sender().text())
        if it == C and self.constraints is not None:
            pass
        elif it == G and self.generators is not None:
            pass
        elif it == O and self.optimisation is not None:
            pass
        else:
            try:
                ts = WorkBook()
                ts.open_workbook(self.get_filename(self.files[it].text()))
                try:
                    sht = self.sheets[it].currentText()
                except:
                    self.setStatus(self.sheets[it].currentText() + ' not found in ' \
                                     + self.file_labels[it] + ' spreadsheet.')
                    return
                ws = ts.sheet_by_name(sht)
            except:
                ws = None
        if it == G: # generators
            if self.generators is None:
                try:
                    self.getGenerators(ws)
                except:
                    return
            sp_pts = []
            for key in self.generators.keys():
                break
            for prop in dir(self.generators[key]):
                if prop[:2] != '__' and prop[-2:] != '__':
                    if prop == 'name':
                        sp_pts.insert(0, 0)
                    elif prop in ['capex', 'constraint', 'fixed_om', 'order']:
                        sp_pts.append(0)
                    elif prop == 'disc_rate' or prop == 'emissions':
                        sp_pts.append(3)
                    elif prop == 'area':
                        sp_pts.append(5)
                    else:
                        sp_pts.append(2)
            dialog = displaytable.Table(self.generators, title=self.sender().text(),
                 save_folder=self.scenarios, edit=True, decpts=sp_pts, abbr=False)
            dialog.exec_()
            if dialog.getValues() is not None:
                update_dictionary(it, dialog.getValues())
                self.setOrder()
                msg = ' table updated'
        elif it == O: # self.optimisation
            if self.optimisation is None:
                try:
                    self.getOptimisation(ws)
                except:
                    return
            dialog = displaytable.Table(self.optimisation, title=self.sender().text(),
                     save_folder=self.scenarios, edit=True)
            dialog.exec_()
            if dialog.getValues() is not None:
                update_dictionary(it, dialog.getValues())
                for key in self.optimisation.keys():
                    if self.optimisation[key].approach == 'Discrete':
                        caps = self.optimisation[key].capacities.split()
                        self.optimisation[key].capacities = []
                        cap_max = 0.
                        for cap in caps:
                            try:
                                self.optimisation[key].capacities.append(float(cap))
                                cap_max += float(cap)
                            except:
                                pass
                        self.optimisation[key].capacity_min = 0
                        self.optimisation[key].capacity_max = round(cap_max, 3)
                        self.optimisation[key].capacity_step = None
                msg = ' table updated'
        if ts is not None:
            ts.close()
            del ts
        newfile = dialog.savedfile
        if newfile is not None:
            if newfile[: len(self.scenarios)] == self.scenarios:
                self.files[it].setText(newfile[len(self.scenarios):])
            else:
                self.files[it].setText(newfile)
            if msg == '':
                msg = ' table exported'
            else:
                msg += ' and exported'
        if msg != '':
            self.setStatus(self.file_labels[it] + msg)

    def getGenerators(self, ws):
        if ws is None:
            self.generators = {}
            args = {'name': '<name>', 'constraint': '<constraint>'}
            self.generators['<name>'] = Facility(**args)
            return
        if ws.cell_value(0, 0) != 'Name':
            self.setStatus('Not a ' + self.file_labels[G] + ' worksheet.')
            return
        args = ['name', 'order', 'constraint', 'capacity', 'lcoe', 'lcoe_cf', 'emissions', 'initial',
                'capex', 'fixed_om', 'variable_om', 'fuel', 'disc_rate', 'lifetime', 'area']
        possibles = {'name': 0}
        for col in range(ws.ncols):
            try:
                arg = ws.cell_value(0, col).lower()
            except:
                continue
            if arg in args:
                possibles[arg] = col
            elif ws.cell_value(0, col)[:9] == 'Capital':
                possibles['capex'] = col
            elif ws.cell_value(0, col)[:8] == 'Discount':
                possibles['disc_rate'] = col
            elif ws.cell_value(0, col)[:8] == 'Dispatch':
                possibles['order'] = col
            elif ws.cell_value(0, col)[:9] == 'Emissions':
                possibles['emissions'] = col
            elif ws.cell_value(0, col) == 'FOM':
                possibles['fixed_om'] = col
            elif ws.cell_value(0, col) == 'LCOE CF':
                possibles['lcoe_cf'] = col
            elif ws.cell_value(0, col)[:4] == 'LCOE':
                possibles['lcoe'] = col
            elif ws.cell_value(0, col) == 'VOM':
                possibles['variable_om'] = col
        self.generators = {}
        for row in range(1, ws.nrows):
            if ws.cell_value(row, 0) is None:
                continue
            in_args = {}
            for key, value in possibles.items():
                in_args[key] = ws.cell_value(row, value)
            self.generators[str(ws.cell_value(row, 0))] = Facility(**in_args)
        return

    def getOptimisation(self, ws):
        if ws is None:
            self.optimisation = {}
            self.optimisation['<name>'] = Optimisation('<name>', 'None', None)
            return
        if ws.cell_value(0, 0) != 'Name':
            self.setStatus('Not an ' + self.file_labels[O] + ' worksheet.')
            return
        cols = ['Name', 'Approach', 'Values', 'Capacity Max', 'Capacity Min',
                'Capacity Step', 'Capacities']
        coln = [-1] * len(cols)
        for col in range(ws.ncols):
            try:
                i = cols.index(ws.cell_value(0, col))
                coln[i] = col
            except:
                pass
        if coln[0] < 0:
            self.setStatus('Not an ' + self.file_labels[O] + ' worksheet.')
            return
        self.optimisation = {}
        for row in range(1, ws.nrows):
            tech = ws.cell_value(row, 0)
            if tech is None:
                continue
            if coln[2] > 0: # values format
                self.optimisation[tech] = Optimisation(tech,
                                     ws.cell_value(row, coln[1]),
                                     ws.cell_value(row, coln[2]))
            else:
                if ws.cell_value(row, coln[1]) == 'Discrete': # fudge values format
                    self.optimisation[tech] = Optimisation(tech,
                                         ws.cell_value(row, coln[1]),
                                         ws.cell_value(row, coln[-1]))
                else:
                    self.optimisation[tech] = Optimisation(tech, '', '')
                    for col in range(1, len(coln)):
                        if coln[col] > 0:
                            attr = cols[col].lower().replace(' ', '_')
                            setattr(self.optimisation[tech], attr,
                                    ws.cell_value(row, coln[col]))
            try:
                self.optimisation[tech].capacity = self.generators[tech].capacity
            except:
                pass
        return

    def getBatch(self, ws, option):
        global columns, rows, values
        def recurse(lvl):
            if lvl >= len(rows) - 1:
                return
            for i in range(len(values[lvl])):
                columns[lvl] = columns[lvl] + [values[lvl][i]] * cols[lvl+1]
                recurse(lvl + 1)

        def step_split(steps):
            bits = steps.split(',')
            if len(bits) == 1:
                bits = steps.split(';')
            try:
                strt = int(bits[0])
            except:
                return 0, 0, 0, -1
            try:
                stop = int(bits[1])
                step = int(bits[2])
                try:
                    frst = int(bits[3])
                except:
                    frst = -1
            except:
                return strt, strt, strt, frst
            return strt, stop, step, frst

        if ws is None:
            self.setStatus(self.file_labels[B] + ' worksheet missing.')
            return False
        istrt = 0
        year_row = -1
        for row in range(3):
            if ws.cell_value(row, 0) in ['Model', 'Model Label', 'Technology']:
                istrt = row + 1
                break
        else:
            self.setStatus('Not a ' + self.file_labels[B] + ' worksheet.')
            return False
        self.batch_models = [{}] # cater for a range of capacities
        self.batch_report = [['Capacity (MW/MWh)', 1]]
        self.batch_tech = []
        istop = ws.nrows
        inrows = False
        for row in range(istrt, ws.nrows):
            tech = ws.cell_value(row, 0)
            if tech is not None and tech != '':
                if year_row < 0 and tech[:4].lower() == 'year':
                    year_row = row
                    continue
                inrows = True
                if tech[:8].lower() != 'capacity':
                    if tech.find('.') > 0:
                        tech = tech[tech.find('.') + 1:]
                    if tech != 'Total' and tech not in self.generators.keys():
                        self.setStatus('Unknown technology - ' + tech + ' - in batch file.')
                        return False
                    self.batch_tech.append(ws.cell_value(row, 0))
                else:
                    self.batch_report[0][1] = row + 1
            elif inrows:
                istop = row
                break
            if tech[:5] == 'Total':
                istop = row + 1
                break
        if len(self.batch_tech) == 0:
            self.setStatus('No input technologies found in ' + self.file_labels[B] + ' worksheet (try opening and re-saving the workbook).')
            return False
        carbon_row = -1
        discount_row = -1
        for row in range(istop, ws.nrows):
            if ws.cell_value(row, 0) is not None and ws.cell_value(row, 0) != '':
                if ws.cell_value(row, 0).lower() in ['chart', 'graph', 'plot']:
                    self.batch_report.append(['Chart', row + 1])
                    break
                if ws.cell_value(row, 0).lower() in ['carbon price', 'carbon price ($/tco2e)']:
                    carbon_row = row
                if ws.cell_value(row, 0).lower() == 'discount rate' or ws.cell_value(row, 0).lower() == 'wacc':
                    discount_row = row
                self.batch_report.append([techClean(ws.cell_value(row, 0), full=True), row + 1])
        range_rows = {}
        for col in range(1, ws.ncols):
            model = ws.cell_value(istrt - 1, col)
            if model is None:
                break
            self.batch_models[0][col] = {'name': model}
            if option == T and year_row < 0:
                self.batch_models[0][col]['year'] = str(model)
            for row in range(istrt, istop):
                if row == year_row:
                    if ws.cell_value(row, col) is not None and ws.cell_value(row, col) != '':
                        self.batch_models[0][col]['year'] = str(ws.cell_value(row, col))
                    continue
                tech = ws.cell_value(row, 0)
                try:
                    if ws.cell_value(row, col) > 0:
                        self.batch_models[0][col][tech] = ws.cell_value(row, col)
                except:
                    if ws.cell_value(row, col) is None:
                        pass
                    elif ws.cell_value(row, col).find(',') >= 0 or ws.cell_value(row, col).find(';') >= 0:
                        try:
                            range_rows[col].append(row)
                        except:
                            range_rows[col] = [row]
                        try:
                            strt, stop, step, frst = step_split(ws.cell_value(row, col))
                            self.batch_models[0][col][tech] = strt
                            if frst >= 0 and len(range_rows[col]) > 1:
                                del range_rows[col][-1]
                                range_rows[col].insert(0, row)
                        except:
                            pass
                    pass
            if carbon_row >= 0:
                if isinstance(ws.cell_value(carbon_row, col), float):
                    self.batch_models[0][col]['Carbon Price'] = ws.cell_value(carbon_row, col)
                elif isinstance(ws.cell_value(carbon_row, col), int):
                    self.batch_models[0][col]['Carbon Price'] = float(ws.cell_value(carbon_row, col))
            if discount_row >= 0:
                if isinstance(ws.cell_value(discount_row, col), float):
                    self.batch_models[0][col]['Discount Rate'] = ws.cell_value(discount_row, col)
                elif isinstance(ws.cell_value(discount_row, col), int):
                    self.batch_models[0][col]['Discount Rate'] = float(ws.cell_value(discount_row, col))
        if len(self.batch_models[0]) == 0:
            self.setStatus('No models found in ' + self.file_labels[B] + ' worksheet (try opening and re-saving the workbook).')
            return False
        if len(range_rows) == 0:
            return True
        # cater for ranges - so multiple batch_models lists
        for rcol, ranges in range_rows.items():
            rows = {}
            for rw in ranges:
                rows[rw] = ws.cell_value(rw, rcol)
            if len(ranges) > 1: # create sheet for each range else one sheet
                values = []
                cols = [1]
                for i in range(len(ranges) -1, 0, -1):
                    strt, stop, step, frst = step_split(rows[ranges[i]])
                    values.insert(0, [])
                    for stp in range(strt, stop + step, step):
                        values[0].append(stp)
                    cols.insert(0, cols[0] * len(values[0]))
                columns = [[]] * len(rows)
                recurse(0)
                my_tech = ws.cell_value(ranges[0], 0)
                tech_2 = ws.cell_value(ranges[1], 0)
              # produce new batch_models entry for first range tech
                techs = {}
                for c in range(1, len(ranges)):
                    techs[ws.cell_value(ranges[c], 0)] = c - 1
                bits = my_tech.split('.')
                strt, stop, step, frst = step_split(rows[ranges[0]])
                for sht in range(strt, stop + step, step):
                    self.batch_models.append({})
                    for c2 in range(len(columns[0])):
                        self.batch_models[-1][c2] = {}
                        for key, value in self.batch_models[0][rcol].items():
                            self.batch_models[-1][c2][key] = value
                        self.batch_models[-1][c2][my_tech] = sht
                        for key, value in techs.items():
                            self.batch_models[-1][c2][key] = columns[value][c2]
                        self.batch_models[-1][c2]['name'] = f'{bits[-1]}_{sht}_{tech_2}'
            else:
                my_tech = ws.cell_value(ranges[0], 0)
                self.batch_models.append({})
                strt, stop, step, frst = step_split(rows[ranges[0]])
                c2 = -1
                for ctr in range(strt, stop + step, step):
                    c2 += 1
                    self.batch_models[-1][c2] = {}
                    if c2 == 0:
                        self.batch_models[-1][c2]['hdr'] = ws.cell_value(ranges[0], 0) # fudge to get header name
                    for key, value in self.batch_models[0][rcol].items():
                        self.batch_models[-1][c2][key] = value
                    self.batch_models[-1][c2][my_tech] = ctr
                #    for key, value in techs.items():
                 #       self.batch_models[-1][c2][key] = columns[value][c2]
                    self.batch_models[-1][c2]['name'] = f'Model {c2 + 1}'
        return True

    def setOrder(self):
        self.order.clear()
        self.ignore.clear()
        self.re_capacity = {}
        if self.generators is None:
            order = ['Storage', 'Biomass', 'PHES', 'Gas', 'CCG1', 'Other', 'Coal']
            for stn in order:
                self.order.addItem(stn)
        else:
            order = []
            zero = []
            for key, value in self.generators.items():
            #    if value.capacity == 0:
            #        continue
                if key in tech_names and key not in self.dispatchable:
                    self.re_capacity[key] = value.capacity
                    continue
                try:
                    gen = key[key.find('.') + 1:]
                    if gen in tech_names and gen not in self.dispatchable:
                        self.re_capacity[key] = value.capacity
                        continue
                except:
                    pass
                try:
                    o = int(value.order)
                    if o > 0:
                        while len(order) <= o:
                            order.append([])
                        order[o - 1].append(key)
                    elif o == 0:
                        zero.append(key)
                except:
                    pass
            order.append(zero)
            for cat in order:
                for stn in cat:
                    self.order.addItem(stn)

    def data_sources(self, sheet, sheet_row, pm_data_file, option):
        normal = oxl.styles.Font(name='Arial')
        bold = oxl.styles.Font(name='Arial', bold=True)
        sheet.cell(row=sheet_row, column=1).value = 'Data sources'
        sheet.cell(row=sheet_row, column=1).font = bold
        sheet_row += 1
        sheet.cell(row=sheet_row, column=1).value = 'Scenarios folder'
        sheet.cell(row=sheet_row, column=1).font = normal
        sheet.cell(row=sheet_row, column=2).value = self.scenarios
        sheet.cell(row=sheet_row, column=2).font = normal
        sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
        sheet_row += 1
        sheet.cell(row=sheet_row, column=1).value = 'Powermatch data file'
        sheet.cell(row=sheet_row, column=1).font = normal
        if pm_data_file[: len(self.scenarios)] == self.scenarios:
            pm_data_file = pm_data_file[len(self.scenarios):]
        sheet.cell(row=sheet_row, column=2).value = pm_data_file
        sheet.cell(row=sheet_row, column=2).font = normal
        sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
        sheet_row += 1
        try:
            if self.loadCombo.currentText() != 'n/a':
                sheet.cell(row=sheet_row, column=1).value = 'Load file'
                sheet.cell(row=sheet_row, column=1).font = normal
                load_file = self.load_files.replace('$YEAR$', self.loadCombo.currentText())
                if load_file[: len(self.scenarios)] == self.scenarios:
                    load_file = load_file[len(self.scenarios):]
                sheet.cell(row=sheet_row, column=2).value = load_file
                sheet.cell(row=sheet_row, column=2).font = normal
                sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
                sheet_row += 1
        except:
            pass
        sheet.cell(row=sheet_row, column=1).value = 'Constraints worksheet'
        sheet.cell(row=sheet_row, column=1).font = normal
        sheet.cell(row=sheet_row, column=2).value = str(self.files[C].text()) \
               + '.' + str(self.sheets[C].currentText())
        sheet.cell(row=sheet_row, column=2).font = normal
        sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
        sheet_row += 1
        sheet.cell(row=sheet_row, column=1).value = 'Generators worksheet'
        sheet.cell(row=sheet_row, column=1).font = normal
        if option == T:
            sheet.cell(row=sheet_row, column=2).value = self.files[G].text()
        else:
            sheet.cell(row=sheet_row, column=2).value = self.files[G].text() \
                   + '.' + self.sheets[G].currentText()
        sheet.cell(row=sheet_row, column=2).font = normal
        sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
        return sheet_row

    def clean_batch_sheet(self):
        msgbox = QMessageBox()
        msgbox.setWindowTitle('SIREN - Powermatch Batch')
        msgbox.setText("Batch worksheet has more that 1,024 columns.\nSome may be invalid/empty. Would you like these to be removed")
        msgbox.setIcon(QMessageBox.Warning)
        msgbox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        reply = msgbox.exec_()
        if reply != QMessageBox.Yes:
            return
        batch_report_file = self.get_filename(self.files[B].text())
        if os.path.exists(batch_report_file + '~'):
            os.remove(batch_report_file + '~')
        shutil.copy2(batch_report_file, batch_report_file + '~')
        wb = oxl.load_workbook(batch_report_file)
        ws = wb.worksheets[0]
        for row in range(1, 4):
            try:
                if ws.cell(row=row, column=1).value.lower() in ['model', 'model label', 'technology']:
                    break
            except:
                pass
        else:
            return # bail out
        for col in range(ws.max_column, 1, -1):
            if ws.cell(row=row, column=col).value is None:
               ws.delete_cols(col, 1)
        wb.save(batch_report_file)

    def pmClicked(self):
        sender_name = self.sender().text()
        def get_load_data(load_file):
            try:
                tf = open(load_file, 'r')
                lines = tf.readlines()
                tf.close()
            except:
                return None
            load_data = []
            bit = lines[0].rstrip().split(',')
            if len(bit) > 0: # multiple columns
                for b in range(len(bit)):
                    if bit[b][:4].lower() == 'load':
                        if bit[b].lower().find('kwh') > 0: # kWh not MWh
                            for i in range(1, len(lines)):
                                bit = lines[i].rstrip().split(',')
                                load_data.append(float(bit[b]) * 0.001)
                        else:
                            for i in range(1, len(lines)):
                                bit = lines[i].rstrip().split(',')
                                load_data.append(float(bit[b]))
            else:
                for i in range(1, len(lines)):
                    load_data.append(float(lines[i].rstrip()))
            return load_data

        def get_batch_prefix(report_group):
            if report_group == 'Lifetime Emissions':
                return 'LES_'
            if report_group in ['Correlation To Load', 'Static Variables']:
                return ''
            bits = report_group.split(' ')
            for i in range(len(bits) -1, -1, -1):
                if bits[i][0] == '(' and bits[i][-1] == ')':
                    del bits[i]
            if len(bits) == 1:
                abr = bits[0][0] + bits[0][-1]
            else:
                abr = ''
                for bit in bits:
                    abr += bit[0]
            return abr.upper() + '_'

        col_letters = ' ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        self.setStatus(self.sender().text() + ' processing started')
        if self.sender().text() == 'Detail': # detailed spreadsheet?
            option = D
        elif self.sender().text() == 'Optimise': # do optimisation?
            option = O
            self.optExit = False #??
        elif self.sender().text() == 'Batch': # do batch processsing
            option = B
        elif self.sender().text() == 'Transition': # do transition processsing
            option = T
        else:
            option = S
        if option != O:
            self.progressbar.setMinimum(0)
            self.progressbar.setMaximum(20)
            self.progressbar.setHidden(False)
            QtWidgets.QApplication.processEvents()
        err_msg = ''
        if self.generators is None:
            try:
                ts = WorkBook()
                ts.open_workbook(self.get_filename(self.files[G].text()))
                ws = ts.sheet_by_name(self.sheets[G].currentText())
                self.getGenerators(ws)
                ts.close()
                del ts
            except FileNotFoundError:
                if err_msg != '':
                    err_msg += ' nor Generators - ' + self.files[G].text()
                else:
                    err_msg = 'Generators file not found - ' + self.files[G].text()
                self.getGenerators(None)
            except:
                if err_msg != '':
                    err_msg += ' and Generators'
                else:
                    err_msg = 'Error accessing Generators'
                self.getGenerators(None)
        if option == B or option == T: # has to be xlsx workbook
            try:
                ts = WorkBook()
                bwbopen_start = time.time()
                ts.open_workbook(self.get_filename(self.files[B].text()))
                ws = ts.sheet_by_index(0)
                if ws.ncols > 1024:
                    ts.close()
                    self.clean_batch_sheet()
                    ts = WorkBook()
                    ts.open_workbook(self.get_filename(self.files[B].text()))
                    ws = ts.sheet_by_index(0)
                tim = time.time() - bwbopen_start
                if tim < 60:
                    tim = '%.1f secs' % tim
                else:
                    hhmm = tim / 60.
                    tim = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
                self.setStatus(f'{self.file_labels[B]} workbook opened ({tim})')
                ok = self.getBatch(ws, option)
                ts.close()
                del ts
                if not ok:
                    return
            except FileNotFoundError:
                err_msg = 'Batch file not found - ' + self.files[B].text()
            except Exception as e:
                err_msg = 'Error accessing Batch file ' + str(e)
        if option == O and self.optimisation is None:
            try:
                ts = WorkBook()
                ts.open_workbook(self.get_filename(self.files[O].text()))
                ws = ts.sheet_by_name(self.sheets[O].currentText())
                self.getOptimisation(ws)
                ts.close()
                del ts
                if self.optimisation is None:
                    if err_msg != '':
                        err_msg += ' not an Optimisation worksheet'
                    else:
                        err_msg = 'Not an optimisation worksheet'
            except FileNotFoundError:
                if err_msg != '':
                    err_msg += ' nor Optimisation - ' + self.files[O].text()
                else:
                    err_msg = 'Optimisation file not found - ' + self.files[O].text()
            except:
                if err_msg != '':
                    err_msg += ' and Optimisation'
                else:
                    err_msg = 'Error accessing Optimisation'
            if self.optimisation is None:
                self.getOptimisation(None)
        if err_msg != '':
            self.setStatus(err_msg)
            return
        pm_data_file = self.get_filename(self.files[D].text())
        if pm_data_file[-5:] != '.xlsx': #xlsx format only
            self.setStatus('Not a Powermatch data spreadsheet (1)')
            self.progressbar.setHidden(True)
            return
        try:
            ts = oxl.load_workbook(pm_data_file)
        except FileNotFoundError:
            self.setStatus('Data file not found - ' + self.files[D].text())
            return
        except Exception as e:
            self.setStatus('Error accessing Data file - ' + self.files[D].text() + ' ' + str(e))
            return
        ws = ts.worksheets[0]
        top_row = ws.max_row - 8760
        if top_row < 1 or (ws.cell(row=top_row, column=1).value != 'Hour' \
                           or ws.cell(row=top_row, column=2).value != 'Period'):
            self.setStatus(f'Not a Powermatch data spreadsheet (2; {top_row})')
            self.progressbar.setHidden(True)
            return
        typ_row = top_row - 1
        gen_row = typ_row
        while typ_row > 0:
            if ws.cell(row=typ_row, column=1).value[:9] == 'Generated':
                gen_row = typ_row
            if ws.cell(row=typ_row, column=3).value in tech_names:
                break
            typ_row -= 1
        else:
            self.setStatus('no suitable data')
            return
        do_zone = False
        zone_row = typ_row - 1
        try:
            if ws.cell(row=zone_row, column=1).value.lower() == 'zone':
                do_zone = True
                zone_techs = []
        except:
            pass
        icap_row = typ_row + 1
        while icap_row < top_row:
            if ws.cell(row=icap_row, column=1).value[:8] == 'Capacity':
                break
            icap_row += 1
        else:
            self.setStatus('no capacity data')
            return
        year = ws.cell(row=top_row + 1, column=2).value[:4]
        pmss_details = {} # contains name, generator, capacity, tech_type, col, multiplier
        pmss_data = []
        re_order = [] # order for re technology
        dispatch_order = [] # order for dispatchable technology
        load_columns = {}
        load_col = -1
        strt_col = 3
        try:
            if self.loadCombo.currentText() != 'n/a':
                year = self.loadCombo.currentText()
                strt_col = 4
                load_col = len(pmss_data)
                typ = 'L'
                capacity = 0
                fctr = 1
                pmss_details['Load'] = PM_Facility('Load', 'Load', 0, 'L', len(pmss_data), 1)
                load_columns[self.loadCombo.currentText()] = len(pmss_data)
                pmss_data.append([])
                load_file = self.load_files.replace('$YEAR$', self.loadCombo.currentText())
                pmss_data[-1] = get_load_data(load_file)
                re_order.append('Load')
        except:
            pass
        zone = ''
        for col in range(strt_col, ws.max_column + 1):
            try:
                valu = ws.cell(row=typ_row, column=col).value.replace('-','')
                i = tech_names.index(valu)
            except:
                continue
            key = tech_names[i]
            if key == 'Load':
                load_col = len(pmss_data)
                typ = 'L'
                capacity = 0
                fctr = 1
            else:
                if do_zone:
                    cell = ws.cell(row=zone_row, column=col)
                    if type(cell).__name__ == 'MergedCell':
                        pass
                    else:
                        zone = ws.cell(row=zone_row, column=col).value
                    if zone is None or zone == '':
                        zone_tech = valu
                    else:
                        zone_tech = zone + '.' + valu
                    key = zone_tech
                    zone_techs.append(key)
                else: # temp
                    if len(self.re_capacity) > 0 and tech_names[i] not in self.re_capacity.keys():
                        continue
                try:
                    capacity = float(ws.cell(row=icap_row, column=col).value)
                except:
                    continue
                if capacity <= 0:
                    continue
                typ = 'R'
                if do_zone:
                    fctr = 1
                elif tech_names[i] in self.re_capacity and capacity > 0:
                    fctr = self.re_capacity[tech_names[i]] / capacity
                else:
                    fctr = 1
            pmss_details[key] = PM_Facility(key, tech_names[i], capacity, typ, len(pmss_data), fctr)
            if key == 'Load':
                load_columns[year] = len(pmss_data)
            pmss_data.append([])
            re_order.append(key)
            for row in range(top_row + 1, ws.max_row + 1):
                pmss_data[-1].append(ws.cell(row=row, column=col).value)
        pmss_details['Load'].capacity = sum(pmss_data[load_col])
        do_adjust = False
        if option == O:
            for itm in range(self.order.count()):
                gen = self.order.item(itm).text()
                try:
                    if self.generators[gen].capacity <= 0:
                        continue
                except KeyError as err:
                    self.setStatus('Key Error: No Generator entry for ' + str(err))
                    continue
                try:
                    if self.generators[gen].constraint in self.constraints and \
                      self.constraints[self.generators[gen].constraint].category == 'Generator':
                        typ = 'G'
                    else:
                        typ = 'S'
                except:
                    continue
                dispatch_order.append(gen)
                pmss_details[gen] = PM_Facility(gen, gen, self.generators[gen].capacity, typ, -1, 1)
            if self.adjust.isChecked():
                 pmss_details['Load'].multiplier = self.adjustto['Load'] / pmss_details['Load'].capacity
            self.optClicked(year, option, pmss_details, pmss_data, re_order, dispatch_order,
                            None, None)
            return
        if self.adjust.isChecked() and option != B and option != T:
            if self.adjustto is None:
                self.adjustto = {}
                self.adjustto['Load'] = 0
                if do_zone:
                    tns = zone_techs[:]
                else:
                    tns = tech_names[:]
                for gen in tns:
                    try:
                        if self.generators[gen].capacity > 0:
                            self.adjustto[gen] = self.generators[gen].capacity
                    except:
                        pass
                for i in range(self.order.count()):
                    gen = self.order.item(i).text()
                    try:
                        if self.generators[gen].capacity > 0:
                            self.adjustto[gen] = self.generators[gen].capacity
                    except:
                       pass
            generated = sum(pmss_data[load_col])
            datain = []
            datain.append(['Load', 'L', generated])
            if self.adjustto['Load'] == 0:
                self.adjustto['Load'] = generated
            for col in range(4, ws.max_column + 1):
                try:
                    valu = ws.cell(row=typ_row, column=col).value.replace('-','')
                    i = tech_names.index(valu)
                except:
                    continue
                key = tech_names[i]
                if do_zone:
                    cell = ws.cell(row=zone_row, column=col)
                    if type(cell).__name__ == 'MergedCell':
                        pass
                    else:
                        zone = ws.cell(row=zone_row, column=col).value
                    if zone is None or zone == '':
                        zone_tech = valu
                    else:
                        zone_tech = zone + '.' + valu
                    key = zone_tech
                try:
                    typ = self.constraints[tech_names[i]].category[0]
                    if typ == '':
                        typ = 'R'
                    datain.append([key, typ, float(ws.cell(row=icap_row, column=col).value)])
                except:
                    try:
                        datain.append([key, 'R', float(ws.cell(row=icap_row, column=col).value)])
                    except:
                        pass
            for i in range(self.order.count()):
                try:
                    if self.generators[self.order.item(i).text()].capacity > 0:
                        gen = self.order.item(i).text()
                        try:
                            if self.generators[gen].constraint in self.constraints and \
                               self.constraints[self.generators[gen].constraint].category == 'Generator':
                                typ = 'G'
                            else:
                                typ = 'S'
                        except:
                            continue
                        datain.append([gen, typ, self.generators[gen].capacity])
                except:
                    pass
            adjust = Adjustments(self, datain, self.adjustto, self.adjust_cap, self.results_prefix,
                                 show_multipliers=self.show_multipliers, save_folder=self.scenarios,
                                 batch_file=self.get_filename(self.files[B].text()))
            adjust.exec_()
            if adjust.getValues() is None:
                self.setStatus('Execution aborted.')
                self.progressbar.setHidden(True)
                return
            self.adjustto = adjust.getValues()
            results_prefix = adjust.getPrefix()
            if results_prefix != self.results_prefix:
                self.results_prefix = results_prefix
                self.results_pfx_fld.setText(self.results_prefix)
            self.updated = True
            do_adjust = True
        ts.close()
        self.progressbar.setValue(0) # was 1
        QtWidgets.QApplication.processEvents()
        if self.files[R].text() == '':
            i = pm_data_file.rfind('/')
            if i >= 0:
                data_file = pm_data_file[i + 1:]
            else:
                data_file = pm_data_file
            data_file = data_file.replace('data', 'results')
            data_file = data_file.replace('Data', 'Results')
            if data_file == pm_data_file[i + 1:]:
                j = data_file.find(' ')
                if j > 0:
                    jnr = ' '
                else:
                    jnr = '_'
                j = data_file.rfind('.')
                data_file = data_file[:j] + jnr + 'Results' + data_file[j:]
            self.files[R].setText(data_file)
        else:
            data_file = self.get_filename(self.files[R].text())
        if self.results_prefix != '':
            j = data_file.rfind('/')
            data_file = data_file[: j + 1] + self.results_prefix + '_' + data_file[j + 1:]
        for itm in range(self.order.count()):
            gen = self.order.item(itm).text()
            try:
                if self.generators[gen].capacity <= 0:
                    continue
            except KeyError as err:
                self.setStatus('Key Error: No Generator entry for ' + str(err))
                continue
            except:
                continue
            if do_adjust:
                try:
                    if self.adjustto[gen] <= 0:
                        continue
                except:
                    pass
            try:
                if self.generators[gen].constraint in self.constraints and \
                  self.constraints[self.generators[gen].constraint].category == 'Generator':
                    typ = 'G'
                else:
                    typ = 'S'
            except:
                continue
            dispatch_order.append(gen)
            pmss_details[gen] = PM_Facility(gen, gen, self.generators[gen].capacity, typ, -1, 1)
        if option == B or option == T:
            if option == T:
                files = setTransition(self, self.file_labels[G], self.get_filename(self.files[G].text()),
                                      self.sheets[G].currentText(), self.loadCombo.currentText())
                files.exec_()
                if files.getValues() is None:
                    self.setStatus('Execution aborted.')
                    self.progressbar.setHidden(True)
                    return
                gen_sheet = files.getValues()
                trn_year = ''
                newfile = self.get_filename(self.files[G].text())
                gen_book = WorkBook()
                gen_book.open_workbook(newfile)
                pmss_details['Load'].multiplier = 1
            elif self.adjust.isChecked():
                generated = sum(pmss_data[load_col])
                datain = [['Load', 'L', generated]]
                adjustto = {'Load': generated}
                adjust = Adjustments(self, datain, adjustto, self.adjust_cap, None,
                                     show_multipliers=self.show_multipliers)
                adjust.exec_()
                if adjust.getValues() is None:
                    self.setStatus('Execution aborted.')
                    self.progressbar.setHidden(True)
                    return
                adjustto = adjust.getValues()
                pmss_details['Load'].multiplier = adjustto['Load'] / pmss_details['Load'].capacity
       #     start_time = time.time() # just for fun
            batch_details = {'Capacity (MW/MWh)': [st_cap, '#,##0.00'],
                             'To Meet Load (MWh)': [st_tml, '#,##0'],
                             'Generation (MWh)': [st_sub, '#,##0'],
                             'Capacity Factor': [st_cfa, '#,##0.0%'],
                             'Cost ($/Yr)': [st_cst, '#,##0'],
                             'LCOG ($/MWh)': [st_lcg, '#,##0.00'],
                             'LCOE ($/MWh)': [st_lco, '#,##0.00'],
                             'Emissions (tCO2e)': [st_emi, '#,##0'],
                             'Emissions Cost': [st_emc, '#,##0'],
                             'LCOE With CO2 ($/MWh)': [st_lcc, '#,##0.00'],
                             'Max MWh': [st_max, '#,##0'],
                             'Capital Cost': [st_cac, '#,##0'],
                             'Lifetime Cost': [st_lic, '#,##0'],
                             'Lifetime Emissions': [st_lie, '#,##0'],
                             'Lifetime Emissions Cost': [st_lec, '#,##0'],
                             'Area': [st_are, '#,###0.00']}
            batch_extra = {'RE': ['#,##0.00', ['RE %age', st_cap], ['Storage %age', st_cap], ['RE %age of Total Load', st_cap]],
                           'Load Analysis': ['#,##0', ['Load met', st_tml], ['Load met %age', st_cap], ['Shortfall', st_tml], ['Total Load', st_tml],
                           ['Largest Shortfall', st_cap], ['Storage losses', st_sub], ['Surplus', st_sub], ['Surplus %age', st_cap]],
                           'Carbon': ['#,##0.00', ['Carbon Price', st_cap], ['Carbon Cost', st_emc], ['LCOE incl. Carbon Cost', st_lcc],
                           ['Lifetime Emissions Cost', st_lec]],
                           'Correlation To Load': ['0.0000', ['RE Contribution', st_cap], ['RE plus Storage', st_cap],
                           ['To Meet Load', st_cap]],
                           'Static Variables': ['#,##0.00', ['Carbon Price', st_cap], ['Lifetime', st_cap],
                           ['Discount Rate', st_cap]],
                           'Optimisation Parameters': ['#,##0.00', ['Population size', 1], ['No. of iterations', 1],
                           ['Mutation probability', 1], ['Exit if stable', 1], ['Optimisation choice', 1],
                           ['Variable', 1], ['LCOE', 1], ['Load%', 1], ['Surplus%', 1], ['RE%', 1],
                           ['Cost', 1], ['CO2', 1]]}
                           # LCOE (incl. CO2)
         #   batch_extra['Optimisation Parameters'] = []
            batch_extra['LCOE ($/MWh)'] = ['#,##0.00']
            for tech in self.batch_tech:
                if tech == 'Total':
                    batch_extra['LCOE ($/MWh)'].append([tech + ' LCOE ($/MWh)'])
                else:
                    batch_extra['LCOE ($/MWh)'].append([tech])
            batch_extra['LCOE ($/MWh)'].append(['LCOE', st_lco])
            batch_extra['LCOE With CO2 ($/MWh)'] = ['#,##0.00']
            for tech in self.batch_tech:
                batch_extra['LCOE With CO2 ($/MWh)'].append([tech])
            batch_extra['LCOE With CO2 ($/MWh)'].append(['LCOE incl. Carbon Cost', st_lcc])
         #   batch_extra['To Meet Load (MWh)'] = ['#,##0.00', ['Total', st_tml]]
            wbopen_start = time.time()
            wb = oxl.load_workbook(self.get_filename(self.files[B].text()))
            tim = time.time() - wbopen_start
            if tim < 60:
                tim = '%.1f secs' % tim
            else:
                hhmm = tim / 60.
                tim = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
            self.setStatus(f'{self.file_labels[B]} workbook re-opened for update ({tim})')
            batch_input_sheet = wb.worksheets[0]
            rpt_time = QtCore.QDateTime.toString(QtCore.QDateTime.currentDateTime(), 'yyyy-MM-dd_hhmm')
            if self.batch_new_file:
                wb.close()
                i = self.files[B].text().rfind('.')
                suffix = '_report_' + rpt_time
                batch_report_file = self.get_filename(self.files[B].text()[:i] + suffix + self.files[B].text()[i:])
                batch_report_file = QtWidgets.QFileDialog.getSaveFileName(None, 'Save Batch Report file',
                                    batch_report_file, 'Excel Files (*.xlsx)')[0]
                if batch_report_file == '':
                    self.setStatus(self.sender().text() + ' aborted')
                    return
                if batch_report_file[-5:] != '.xlsx':
                    batch_report_file += '.xlsx'
                if os.path.exists(batch_report_file) and not self.replace_last.isChecked():
                    wb = oxl.load_workbook(batch_report_file)
                    bs = wb.create_sheet('Results_' + rpt_time)
                else:
                    wb = oxl.Workbook()
                    bs = wb.active
                    bs.title = 'Results_' + rpt_time
            else:
                batch_report_file = self.get_filename(self.files[B].text())
                if self.replace_last.isChecked():
                    del_sht = ''
                    for sht in wb.sheetnames:
                        if sht[:8] == 'Results_' and sht > del_sht:
                            del_sht = sht
                    if del_sht != '':
                        del wb[del_sht]
                        del_sht = del_sht.replace('Results', 'Charts')
                        if del_sht in wb.sheetnames:
                            del wb[del_sht]
                bs = wb.create_sheet('Results_' + rpt_time)
            start_time = time.time() # just for fun
            normal = oxl.styles.Font(name='Arial')
            bold = oxl.styles.Font(name='Arial', bold=True)
            grey = oxl.styles.colors.Color(rgb='00f2f2f2')
            grey_fill = oxl.styles.fills.PatternFill(patternType='solid', fgColor=grey)
            total_models = 0
            for sht in range(len(self.batch_models)):
                total_models = total_models + len(self.batch_models[sht])
            try:
                incr = 20 / total_models
            except:
                incr = .05
            prgv = incr
            prgv_int = 0
            model_row = False
            model_row_no = 0
            sht_nam_len = max(len(str(len(self.batch_models))), 2)
            for sht in range(len(self.batch_models)):
                sheet_start = time.time()
                if sht == 0: # normal case
                   # copy header rows to new worksheet
                   merged_cells = []
                   merge_cells = None
                   model_row = False
                   model_cols = len(self.batch_models[sht])
                   for row in range(1, self.batch_report[0][1] + 2):
                       if batch_input_sheet.cell(row=row, column=1).value in ['Model', 'Model Label', 'Technology']:
                           model_row = True
                           model_row_no = row
                       else:
                           model_row = False
                       for col in range(1, model_cols + 2):
                           cell = batch_input_sheet.cell(row=row, column=col)
                           if type(cell).__name__ == 'MergedCell':
                               if merge_cells is None:
                                   merge_cells = [row, col - 1, col]
                               else:
                                   merge_cells[2] = col
                               continue
                           if model_row and col > 1:
                               new_cell = bs.cell(row=row, column=col, value=self.batch_models[sht][col - 1]['name'])
                           else:
                               new_cell = bs.cell(row=row, column=col, value=cell.value)
                           if cell.has_style:
                               new_cell.font = copy(cell.font)
                               new_cell.border = copy(cell.border)
                               new_cell.fill = copy(cell.fill)
                               new_cell.number_format = copy(cell.number_format)
                               new_cell.protection = copy(cell.protection)
                               new_cell.alignment = copy(cell.alignment)
                           if merge_cells is not None:
                               bs.merge_cells(start_row=row, start_column=merge_cells[1], end_row=row, end_column=merge_cells[2])
                               merged_cells.append(merge_cells)
                               merge_cells = None
                       if merge_cells is not None:
                           bs.merge_cells(start_row=row, start_column=merge_cells[1], end_row=row, end_column=merge_cells[2])
                           merged_cells.append(merge_cells)
                           merge_cells = None
                   try:
                       normal = oxl.styles.Font(name=cell.font.name, sz=cell.font.sz)
                       bold = oxl.styles.Font(name=cell.font.name, sz=cell.font.sz, bold=True)
                   except:
                       pass
                else:
                    sheet_name = f'{sht:0{sht_nam_len}}'
                    if sheet_name in wb.sheetnames:
                        del wb[sheet_name]
                        if 'Charts_' + sheet_name in wb.sheetnames:
                            del wb['Charts_' + sheet_name]
                    bs = wb.create_sheet(sheet_name)
                    if model_row_no > 1:
                        title = self.batch_models[sht][0]['name']
                        tech_2 = title.split('_')
                        if len(tech_2) > 1:
                            tech_2 = tech_2[-1]
                            bits_2 = tech_2.split('.')[-1]
                            title = title.replace(tech_2, bits_2)
                            cap_2 = self.batch_models[sht][0][tech_2]
                            fst_col = 2
                            bs.cell(row=1, column=2).value = f'{title}_{cap_2}'
                            bs.cell(row=1, column=2).font = normal
                            bs.cell(row=1, column=2).alignment = oxl.styles.Alignment(wrap_text=True, vertical='bottom', horizontal='center')
                            g = 1
                            for i in range(1, len(self.batch_models[sht])):
                                if self.batch_models[sht][i][tech_2] != cap_2:
                                    bs.merge_cells(start_row=1, start_column=fst_col, end_row=1, end_column=i + 1)
                                    fst_col = i + 2
                                    cap_2 = self.batch_models[sht][i][tech_2]
                                    bs.cell(row=1, column=fst_col).value = f'{title}_{cap_2}'
                                    if g == 0:
                                        g = 1
                                    else:
                                        bs.cell(row=1, column=fst_col).fill = grey_fill
                                        g = 0
                                    bs.cell(row=1, column=fst_col).font = normal
                                    bs.cell(row=1, column=fst_col).alignment = oxl.styles.Alignment(wrap_text=True, vertical='bottom', horizontal='center')
                            bs.merge_cells(start_row=1, start_column=fst_col, end_row=1, end_column=i + 2)
                        else:
                            try:
                                title = self.batch_models[sht][0]['hdr'].split('.')[-1]
                                del self.batch_models[sht][0]['hdr']
                            except:
                                pass
                            bs.cell(row=1, column=2).value = f'{title}'
                            bs.cell(row=1, column=2).font = normal
                            bs.cell(row=1, column=2).alignment = oxl.styles.Alignment(wrap_text=True, vertical='bottom', horizontal='center')
                            bs.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(self.batch_models[sht]) + 1)
                column = 1
                gndx = self.batch_report[0][1] # Capacity group starting row
                do_opt_parms = [False, 0, 0, 0]
                total_load_row = 0
                if self.discount_rate > 0:
                    batch_disc_row = 0
                else:
                    batch_disc_row = -1
                if self.carbon_price > 0:
                    batch_carbon_row = 0
                else:
                    batch_carbon_row = -1
                batch_lifetime = False
                batch_data_sources_row = 0
                re_tml_row = 0
                max_load_row = -1
                report_keys = []
                for g in range(len(self.batch_report)):
                    report_keys.append(self.batch_report[g][0])
                if 'Lifetime Cost' in report_keys:
                    batch_lifetime = True
                for g in range(len(self.batch_report)):
                    if self.batch_report[g][0] == 'Chart':
                        continue
                    elif self.batch_report[g][0] == 'Carbon Price':
                        batch_carbon_row = self.batch_report[g][1]
                        continue
                    elif self.batch_report[g][0] == 'Discount Rate' or self.batch_report[g][0].lower() == 'wacc':
                        batch_disc_row = self.batch_report[g][1]
                        continue
                    elif self.batch_report[g][0].lower() == 'data sources':
                        batch_data_sources_row = gndx
                        gndx += 6
                        try:
                            if self.loadCombo.currentText() != 'n/a':
                                gndx += 1
                        except:
                            pass
                        continue
                    if self.batch_report[g][0] not in batch_details.keys() and self.batch_report[g][0] not in batch_extra.keys():
                        continue
                    self.batch_report[g][1] = gndx
                    if self.batch_prefix:
                        batch_pfx = get_batch_prefix(self.batch_report[g][0])
                    else:
                        batch_pfx = ''
                    bs.cell(row=gndx, column=1).value = self.batch_report[g][0]
                    bs.cell(row=gndx, column=1).font = bold
                    if self.batch_report[g][0] in batch_extra.keys():
                        key = self.batch_report[g][0]
                        if self.batch_report[g][0] == 'Optimisation Parameters':
                            for row in range(1, batch_input_sheet.max_row + 1):
                                if batch_input_sheet.cell(row=row, column=1).value == 'Optimisation Parameters':
                                    do_opt_parms[0] = True
                                    do_opt_parms[1] = gndx
                                    do_opt_parms[2] = row
                                    break
                            for row in range(row, batch_input_sheet.max_row + 1):
                                gndx += 1
                                if batch_input_sheet.cell(row=row, column=1).value == '':
                                    break
                            do_opt_parms[3] = row
                            continue
                        for sp in range(1, len(batch_extra[key])):
                            if batch_extra[key][sp][0] == 'Total Load':
                                total_load_row = gndx + sp
                            elif batch_extra[key][sp][0] == 'Carbon Price':
                                bs.cell(row=gndx + sp, column=1).value = batch_pfx + batch_extra[key][sp][0] + ' ($/tCO2e)'
                            elif batch_extra[key][sp][0] == 'Lifetime':
                                bs.cell(row=gndx + sp, column=1).value = batch_pfx + batch_extra[key][sp][0] + ' (years)'
                            elif batch_extra[key][sp][0] == 'Total incl. Carbon Cost':
                                bs.cell(row=gndx + sp, column=1).value = batch_pfx + 'LCOE incl. Carbon Cost'
                            else:
                                bs.cell(row=gndx + sp, column=1).value = batch_pfx + batch_extra[key][sp][0]
                            if batch_extra[key][sp][0] in ['RE %age of Total Load', 'Total incl. Carbon Cost'] or \
                              batch_extra[key][sp][0].find('LCOE') >= 0 and batch_extra[key][sp][0].find('Total LCOE') < 0:
                                bs.cell(row=gndx + sp, column=1).font = bold
                            else:
                                bs.cell(row=gndx + sp, column=1).font = normal
                        gndx += len(batch_extra[key]) + 1
                        if key == 'Carbon':
                            if not batch_lifetime:
                                gndx -= 1
                                tot_carb_row = gndx - 3
                            else:
                                tot_carb_row = gndx - 4
                        elif key == 'LCOE ($/MWh)':
                            tot_lco_row = gndx - 2
                        elif key == 'LCOE With CO2 ($/MWh)':
                            tot_lcc_row = gndx - 2
                    else:
                        if self.batch_report[g][0] not in batch_details.keys():
                            continue
                        if self.batch_prefix:
                            batch_pfx = get_batch_prefix(self.batch_report[g][0])
                        else:
                            batch_pfx = ''
                        for sp in range(len(self.batch_tech)):
                        #    if self.batch_report[g][0] == 'To Meet Load (MWh)' and sp == 0:
                         #       bs.cell(row=gndx + sp + 1, column=1).value = 'RE Contribution To Load'
                            if self.batch_report[g][0] != 'Capacity Factor' or self.batch_tech[sp] != 'Total':
                                bs.cell(row=gndx + sp + 1, column=1).value = batch_pfx + self.batch_tech[sp]
                            if self.batch_report[g][0] == 'Max MWh' and self.batch_tech[sp] == 'Total':
                                max_load_row = gndx + sp + 1
                                bs.cell(row=max_load_row, column=1).value = batch_pfx + 'Max Load'
                            elif self.batch_tech[sp] == 'Total' and self.batch_report[g][0] != 'Capacity Factor':
                                bs.cell(row=gndx + sp + 1, column=1).value = batch_pfx + self.batch_tech[sp] + ' ' + self.batch_report[g][0]
                            bs.cell(row=gndx + sp + 1, column=1).font = normal
                        if self.batch_report[g][0] == 'Cost ($/Yr)' and batch_disc_row >= 0:
                            batch_disc_row = gndx + sp + 2
                            bs.cell(row=batch_disc_row, column=1).value = batch_pfx + 'Discount Rate'
                            bs.cell(row=batch_disc_row, column=1).font = normal
                        if self.batch_report[g][0] == 'Capacity Factor' and self.batch_tech[-1] == 'Total':
                            gndx += len(self.batch_tech) + 1
                        else:
                            gndx += len(self.batch_tech) + 2
                        if self.batch_report[g][0] == 'Cost ($/Yr)' and batch_disc_row >= 0:
                            gndx += 1
                        if self.batch_report[g][0] == 'To Meet Load (MWh)':
                            re_tml_row = gndx - 1
                            bs.cell(row=re_tml_row, column=1).value = batch_pfx + 'RE Contribution To Load'
                            bs.cell(row=re_tml_row, column=1).font = normal
                            bs.cell(row=re_tml_row + 1, column=1).value = batch_pfx + 'Storage Contribution To Load'
                            bs.cell(row=re_tml_row + 1, column=1).font = normal
                            gndx += 2
                merge_col = 1
                last_name = ''
                # find first varying capacity to create model name
                model_key = ''
                model_nme = ''
                if sht > 0:
                    for key in self.batch_models[sht][0].keys():
                        if key == 'name':
                            continue
                        try:
                            if self.batch_models[sht][0][key] != self.batch_models[sht][1][key]:
                                model_key = key
                                bits = key.split('.')[-1].split(' ')
                                for bit in bits:
                                    model_nme += bit.strip('()')[0]
                                model_nme += '-'
                                break
                        except:
                            pass
                if option == T:
                    capex_table = {}
                    for fac in pmss_details.keys():
                        capex_table[fac] = {'cum': 0}
                for model, capacities in self.batch_models[sht].items():
                    if option == T:
                        if capacities['year'] != trn_year:
                            # get generators and load for new year
                            trn_year = capacities['year']
                            year = str(trn_year)
                            ws = gen_book.sheet_by_name(gen_sheet.replace('$YEAR$', year))
                            self.getGenerators(ws)
                            if year not in load_columns.keys():
                                load_columns[year] = len(pmss_data)
                                pmss_data.append([])
                                load_file = self.load_files.replace('$YEAR$', year)
                                pmss_data[-1] = get_load_data(load_file)
                    for fac in pmss_details.keys():
                        if fac == 'Load':
                            pmss_details['Load'].capacity = sum(pmss_data[load_columns[year]])
                            pmss_details['Load'].col = load_columns[year]
                            continue
                        pmss_details[fac].multiplier = 0
                    if int(prgv) > prgv_int:
                        prgv_int = int(prgv)
                        self.progressbar.setValue(int(prgv))
                        QtWidgets.QApplication.processEvents()
                    prgv += incr
                    column += 1
                    dispatch_order = []
                    for key, capacity in capacities.items(): # cater for zones
                        if key in ['Carbon Price', 'Discount Rate', 'Total']:
                            continue
                        if key == 'name' and model_row_no > 0:
                            if model_key != '':
                                bs.cell(row=model_row_no, column=column).value = f'{model_nme}{capacities[model_key]}'
                            elif option == T:
                                bs.cell(row=model_row_no, column=column).value = f'{capacity}'
                            else:
                                bs.cell(row=model_row_no, column=column).value = f'Model {model + 1}'
                            bs.cell(row=model_row_no, column=column).font = normal
                            bs.cell(row=model_row_no, column=column).alignment = oxl.styles.Alignment(wrap_text=True,
                                    vertical='bottom', horizontal='center')
                            continue
                        if key == 'year':
                            if capacity in load_columns.keys():
                                pmss_details['Load'].col = load_columns[capacity]
                            else:
                                load_columns[capacity] = len(pmss_data)
                                pmss_data.append([])
                                load_file = self.load_files.replace('$YEAR$', capacity)
                                pmss_data[-1] = get_load_data(load_file)
                                pmss_details['Load'].col = load_columns[capacity]
                            pmss_details['Load'].capacity = sum(pmss_data[pmss_details['Load'].col])
                            continue
                        if key not in re_order:
                            dispatch_order.append(key)
                        if key not in pmss_details.keys():
                            gen = key[key.find('.') + 1:]
                            if gen in re_order:
                                typ = 'R'
                            elif self.generators[gen].constraint in self.constraints and \
                              self.constraints[self.generators[gen].constraint].category == 'Generator':
                                typ = 'G'
                            else:
                                typ = 'S'
                            pmss_details[key] = PM_Facility(key, gen, capacity, typ, -1, 1)
                    for fac in pmss_details.keys():
                        if fac == 'Load':
                            continue
                        gen = pmss_details[fac].generator
                        try:
                            pmss_details[fac].multiplier = capacities[fac] * 1.0 / pmss_details[fac].capacity
                        except:
                            pass
                        if option == T:
                            if fac not in capex_table.keys():
                                capex_table[fac] = {'cum': 0}
                            if year not in capex_table[fac].keys():
                                try:
                                    capex_table[fac][year] = [self.generators[fac].capex, 0]
                                except:
                                    capex_table[fac][year] = [self.generators[fac[fac.find('.') + 1:]].capex, 0]
                            capx = pmss_details[fac].multiplier * pmss_details[fac].capacity
                            capex_table[fac][year][1] = capx - capex_table[fac]['cum']
                            capex_table[fac]['cum'] = capx
                    if option == T:
                        for fac in capex_table.keys():
                            if capex_table[fac]['cum'] == 0:
                                continue
                            capx = 0
                            for key, detail in capex_table[fac].items():
                                if key == 'cum':
                                    continue
                                capx = capx + detail[0] * detail[1]
                            capx = capx / capex_table[fac]['cum']
                            try:
                                self.generators[fac].capex = round(capx)
                            except:
                                self.generators[fac[fac.find('.') + 1:]].capex = round(capx)
                    save_carbon_price = None
                    if 'Carbon Price' in capacities.keys():
                        save_carbon_price = self.carbon_price
                        self.carbon_price = capacities['Carbon Price']
                    if 'Discount Rate' in capacities.keys():
                        save_discount_rate = self.discount_rate
                        self.discount_rate = capacities['Discount Rate']
                    sp_data, corr_data = self.processor.doDispatch(year, option, sender_name, pmss_details, pmss_data, re_order, 
                        dispatch_order, pm_data_file, data_file, files=None,sheets=None)
                    if 'Carbon Price' in capacities.keys():
                        self.carbon_price = save_carbon_price
                    # first the Facility/technology table at the top of sp_data
                    for sp in range(len(self.batch_tech) + 1):
                        if sp_data[sp][st_fac] in self.batch_tech:
                            tndx = self.batch_tech.index(sp_data[sp][st_fac]) + 1
                            for group in self.batch_report:
                                if group[0] in batch_details.keys():
                                    gndx = group[1]
                                    col = batch_details[group[0]][0]
                                    if group[0] == 'Capacity Factor' and sp_data[sp][0] == 'Total':
                                        continue
                                    if group[0] == 'Capacity Factor' and isinstance(sp_data[sp][col], str):
                                        bs.cell(row=gndx + tndx, column=column).value = float(sp_data[sp][col].strip('%')) / 100.
                                    else:
                                        bs.cell(row=gndx + tndx, column=column).value = sp_data[sp][col]
                                    bs.cell(row=gndx + tndx, column=column).number_format = batch_details[group[0]][1]
                                    bs.cell(row=gndx + tndx, column=column).font = normal
                        if sp_data[sp][st_fac] == 'Total':
                            break
                    if batch_disc_row > 1:
                         bs.cell(row=batch_disc_row, column=column).value = self.discount_rate
                         bs.cell(row=batch_disc_row, column=column).number_format = '#0.00%'
                         bs.cell(row=batch_disc_row, column=column).font = normal
                    # save details from Total row
                    for group in self.batch_report:
                        if group[0] == 'LCOE ($/MWh)':
                            try:
                                col = batch_details['LCOE ($/MWh)'][0]
                                bs.cell(row=tot_lco_row, column=column).value = sp_data[sp][col]
                                bs.cell(row=tot_lco_row, column=column).number_format = batch_details['LCOE ($/MWh)'][1]
                                bs.cell(row=tot_lco_row, column=column).font = bold
                            except:
                                pass
                        elif group[0] == 'LCOE With CO2 ($/MWh)':
                            try:
                                col = batch_details['LCOE With CO2 ($/MWh)'][0]
                                bs.cell(row=tot_lcc_row, column=column).value = sp_data[sp][col]
                                bs.cell(row=tot_lcc_row, column=column).number_format = batch_details['LCOE With CO2 ($/MWh)'][1]
                                bs.cell(row=tot_lcc_row, column=column).font = bold
                            except:
                                pass
                        elif group[0] == 'Carbon':
                            try:
                                bs.cell(row=tot_carb_row, column=column).value = sp_data[sp][st_emc]
                                bs.cell(row=tot_carb_row, column=column).number_format = '#,##0'
                                bs.cell(row=tot_carb_row, column=column).font = normal
                                bs.cell(row=tot_carb_row + 1, column=column).value = sp_data[sp][st_lcc]
                                bs.cell(row=tot_carb_row + 1, column=column).number_format = '#,##0.00'
                                bs.cell(row=tot_carb_row + 1, column=column).font = bold
                                bs.cell(row=tot_carb_row + 2, column=column).value = sp_data[sp][st_lec]
                                bs.cell(row=tot_carb_row + 2, column=column).number_format = '#,##0'
                                bs.cell(row=tot_carb_row + 2, column=column).font = normal
                            except:
                                pass
                    if 'Discount Rate' in capacities.keys():
                        self.discount_rate = save_discount_rate
                    # now the other stuff in sp_data
                    for sp in range(sp + 1, len(sp_data)):
                        if sp_data[sp][st_fac] == '':
                            continue
                        i = sp_data[sp][st_fac].find(' (')
                        if i >= 0:
                            tgt = sp_data[sp][st_fac][: i]
                        else:
                            tgt = sp_data[sp][st_fac]
                        if tgt == 'RE %age':
                            for group in self.batch_report:
                                if group[0] == 'To Meet Load (MWh)':
                                    try:
                                        col = batch_details['To Meet Load (MWh)'][0]
                                        bs.cell(row=re_tml_row, column=column).value = sp_data[sp][col]
                                        bs.cell(row=re_tml_row, column=column).number_format = batch_details['To Meet Load (MWh)'][1]
                                        bs.cell(row=re_tml_row, column=column).font = normal
                                    except:
                                        pass
                                    break
                        elif tgt == 'Storage %age':
                            for group in self.batch_report:
                                if group[0] == 'To Meet Load (MWh)':
                                    try:
                                        col = batch_details['To Meet Load (MWh)'][0]
                                        bs.cell(row=re_tml_row + 1, column=column).value = sp_data[sp][col]
                                        bs.cell(row=re_tml_row + 1, column=column).number_format = batch_details['To Meet Load (MWh)'][1]
                                        bs.cell(row=re_tml_row + 1, column=column).font = normal
                                    except:
                                        pass
                                    break
                        elif tgt == 'LCOE':
                            for group in self.batch_report:
                                if group[0] == 'LCOE ($/MWh)':
                                    try:
                                        col = batch_details['LCOE ($/MWh)'][0]
                                        bs.cell(row=re_tml_row + 1, column=column).value = sp_data[sp][col]
                                        bs.cell(row=re_tml_row + 1, column=column).number_format = batch_details['LCOE ($/MWh)'][1]
                                        bs.cell(row=re_tml_row + 1, column=column).font = normal
                                    except:
                                        pass
                                    break
                        elif tgt == 'Carbon Price':
                            for group in batch_extra['Carbon'][1:]:
                                if group[0] == 'Carbon Price':
                                    try:
                                        col = group[1]
                                        bs.cell(row=tot_carb_row - 1, column=column).value = sp_data[sp][col]
                                        bs.cell(row=tot_carb_row - 1, column=column).number_format = batch_extra['Carbon'][0]
                                        bs.cell(row=tot_carb_row - 1, column=column).font = normal
                                    except:
                                        pass
                                    break
                        elif tgt[:10] == 'Total Load':
                            for group in self.batch_report:
                                if group[0] == 'Max MWh':
                                    try:
                                        col = batch_details['Max MWh'][0]
                                        bs.cell(row=max_load_row, column=column).value = sp_data[sp][col]
                                        bs.cell(row=max_load_row, column=column).number_format = batch_extra['Max MWh'][0]
                                        bs.cell(row=max_load_row, column=column).font = normal
                                    except:
                                        pass
                                    break
                        for key, details in batch_extra.items():
                            try:
                                x = [x for x in details if tgt in x][0]
                                for group in self.batch_report:
                                    if group[0] == key:
                                        gndx = group[1]
                                        break
                                else:
                                    continue
                                tndx = details.index(x)
                                col = x[1]
                                bs.cell(row=gndx + tndx, column=column).value = sp_data[sp][col]
                                if key == 'RE' or (key == 'Static Variables' and x[0] == 'Discount Rate'):
                                    pct = float(sp_data[sp][col].strip('%')) / 100.
                                    bs.cell(row=gndx + tndx, column=column).value = pct
                                    bs.cell(row=gndx + tndx, column=column).number_format = '0.0%'
                                else:
                                    bs.cell(row=gndx + tndx, column=column).value = sp_data[sp][col]
                                    bs.cell(row=gndx + tndx, column=column).number_format = details[0]
                                bs.cell(row=gndx + tndx, column=column).font = normal
                                if sp_data[sp][st_fac] == 'RE %age of Total Load' or \
                                  sp_data[sp][st_fac].find('LCOE') >= 0 or \
                                  sp_data[sp][st_fac].find('incl.') >= 0:
                                    bs.cell(row=gndx + tndx, column=column).font = bold
                                else:
                                    bs.cell(row=gndx + tndx, column=column).font = normal
                                if key == 'Load Analysis':
                                    if x[0] in ['Load met', 'Surplus']:
                                        tndx += 1
                                        col = batch_extra['Load Analysis'][tndx][1]
                                        pct = float(sp_data[sp][col].strip('%')) / 100.
                                        bs.cell(row=gndx + tndx, column=column).value = pct
                                        bs.cell(row=gndx + tndx, column=column).number_format = '0.0%'
                                        bs.cell(row=gndx + tndx, column=column).font = normal
                            except:
                                pass
                tim = (time.time() - sheet_start)
                if tim < 60:
                    tim = '%.1f secs' % tim
                else:
                    hhmm = tim / 60.
                    tim = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
                timt = (time.time() - start_time)
                if timt < 60:
                    timt = '%.1f secs' % timt
                else:
                    hhmm = timt / 60.
                    timt = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
                self.setStatus(f'Processed sheet {sht + 1} of {len(self.batch_models)}; ({len(self.batch_models[sht])} models; {tim}. Total {timt})')
                QtWidgets.QApplication.processEvents()
                if total_load_row > 0:
                    if self.batch_prefix:
                        batch_pfx = get_batch_prefix('Load Analysis')
                    if option == T:
                        bs.cell(row=total_load_row, column=1).value = batch_pfx + 'Total Load'
                    else:
                        load_mult = ''
                        try:
                            mult = round(pmss_details['Load'].multiplier, 3)
                            if mult != 1:
                                load_mult = ' x ' + str(mult)
                        except:
                            pass
                        bs.cell(row=total_load_row, column=1).value = batch_pfx + 'Total Load - ' + year + load_mult
                if do_opt_parms[0]:
                    t_row = do_opt_parms[1]
                    for row in range(do_opt_parms[2], do_opt_parms[3] + 1):
                        for col in range(1, batch_input_sheet.max_column + 1):
                            cell = batch_input_sheet.cell(row=row, column=col)
                            new_cell = bs.cell(row=t_row, column=col, value=cell.value)
                            if cell.has_style:
                                new_cell.font = copy(cell.font)
                                new_cell.border = copy(cell.border)
                                new_cell.fill = copy(cell.fill)
                                new_cell.number_format = copy(cell.number_format)
                                new_cell.protection = copy(cell.protection)
                                new_cell.alignment = copy(cell.alignment)
                        t_row += 1
                del_rows = []
                for group in self.batch_report:
                    if group[0] in ['Generation (MWh)']:
                        # remove storage or RE
                        gndx = group[1]
                        if group[0] == 'Generation (MWh)':
                            tst = 'S'
                        else:
                            tst = 'R' # probably redundant
                        for row in range(gndx, gndx + len(self.batch_tech)):
                            try:
                                if pmss_details[bs.cell(row=row, column=1).value].tech_type == tst:
                                    del_rows.append(row)
                            except:
                                pass
                for row in sorted(del_rows, reverse=True):
                    bs.delete_rows(row, 1)
                for column_cells in bs.columns:
                    length = 0
                    for cell in column_cells:
                        if cell.row < self.batch_report[0][1] - 1:
                            continue
                        try:
                            value = str(round(cell.value, 2))
                        except:
                            value = cell.value
                        if value is None:
                            continue
                        if len(value) > length:
                            length = len(value)
                    if isinstance(cell.column, int):
                        cel = ssCol(cell.column)
                    else:
                        cel = cell.column
                    bs.column_dimensions[cel].width = max(length * 1.05, 10)
                if batch_data_sources_row > 0:
                    i = self.data_sources(bs, batch_data_sources_row - len(del_rows), pm_data_file, option)
                bs.freeze_panes = 'B' + str(self.batch_report[0][1])
                bs.activeCell = 'B' + str(self.batch_report[0][1])
                for sheet in wb:
                    wb[sheet.title].views.sheetView[0].tabSelected = False
                wb.active = bs
                # check if any charts/graphs
                if self.batch_report[-1][0] == 'Chart':
                    bold = oxl.styles.Font(name='Arial', bold=True)
                    min_col = 2
                    max_col = len(self.batch_models[sht]) + 1
                    chs = None
                    in_chart = False
                    cht_cells = ['N', 'B']
                    cht_row = -27
                    tndx_rows = max(9, len(self.batch_tech) + 4)
                    cats = None
                    chart_group = ''
                    for row in range(self.batch_report[-1][1], batch_input_sheet.max_row + 1):
                        if batch_input_sheet.cell(row=row, column=1).value is None:
                            continue
                        if batch_input_sheet.cell(row=row, column=1).value.lower() in ['chart', 'graph', 'plot']:
                            if in_chart:
                                charts[-1].width = 20
                                charts[-1].height = 12
                                for s in range(len(charts[-1].series)):
                                    ser = charts[-1].series[s]
                                    ser.marker.symbol = 'circle' #'dot', 'plus', 'triangle', 'x', 'picture', 'star', 'diamond', 'square', 'circle', 'dash', 'auto'
                              #      ser.graphicalProperties.line.solidFill = "00AAAA"
                                if charts2[-1] is not None:
                                    for s in range(len(charts2[-1].series)):
                                        ser = charts2[-1].series[s]
                                        ser.marker.symbol = 'triangle'
                               #         ser.graphicalProperties.line.solidFill = "00AAAA"
                                    charts2[-1].y_axis.crosses = 'max'
                                    charts[-1] += charts2[-1]
                                if cats is not None:
                                    charts[-1].set_categories(cats)
                                if len(charts) % 2:
                                    cht_row += 30
                                if chart_group != '':
                                    cht_col = col_letters.index(cht_cells[len(charts) % 2])
                                    chs.cell(row=cht_row - 1, column=cht_col).value = chart_group
                                    chs.cell(row=cht_row - 1, column=cht_col).font = bold
                                chs.add_chart(charts[-1], cht_cells[len(charts) % 2] + str(cht_row))
                            in_chart = True
                            if chs is None:
                                if bs.title.find('Results') >= 0:
                                    txt = bs.title.replace('Results', 'Charts')
                                else:
                                    txt = 'Charts_' + bs.title
                                chs = wb.create_sheet(txt)
                                charts = []
                                charts2 = []
                            charts.append(LineChart())
                            charts2.append(None)
                            if batch_input_sheet.cell(row=row, column=2).value is None or len(merged_cells) == 0:
                                min_col = 2
                                max_col = len(self.batch_models[sht]) + 1
                                chart_group = ''
                            else:
                                merge_group = get_value(batch_input_sheet, row, 2)
                                for i in range(len(merged_cells) -1, -1, -1):
                                    merge_value = get_value(batch_input_sheet, merged_cells[i][0], merged_cells[i][1])
                                    if merge_value == merge_group:
                                        min_col = merged_cells[i][1]
                                        max_col = merged_cells[i][2]
                                        chart_group = merge_group
                                        break
                        elif not in_chart:
                            continue
                        elif batch_input_sheet.cell(row=row, column=1).value.lower() == 'title':
                            charts[-1].title = batch_input_sheet.cell(row=row, column=2).value
                        elif batch_input_sheet.cell(row=row, column=1).value.lower() == 'x-title':
                            charts[-1].x_axis.title = get_value(batch_input_sheet, row, 2)
                        elif batch_input_sheet.cell(row=row, column=1).value.lower() == 'y-title':
                            charts[-1].y_axis.title = batch_input_sheet.cell(row=row, column=2).value
                        elif batch_input_sheet.cell(row=row, column=1).value.lower() == 'y-title2':
                            if charts2[-1] is None:
                                charts2[-1] = LineChart()
                                charts2[-1].x_axis.title = None
                            charts2[-1].y_axis.axId = 200
                            charts2[-1].y_axis.title = batch_input_sheet.cell(row=row, column=2).value
                        elif batch_input_sheet.cell(row=row, column=1).value.lower() in ['categories', 'y-labels', 'data', 'data2']:
                            dgrp = get_value(batch_input_sheet, row, 2)
                            if batch_input_sheet.cell(row=row, column=1).value.lower() == 'categories' \
                              and dgrp.lower() in ['model', 'model label', 'technology']: # models as categories
                                rw = self.batch_report[0][1] - 1
                                cats = Reference(bs, min_col=min_col, min_row=rw, max_col=max_col, max_row=rw)
                                continue
                            if dgrp.lower() in ['capacity (mw)', 'capacity (mw/mwh)']:
                                gndx = self.batch_report[0][1]
                            else:
                                for group in self.batch_report:
                                    if group[0].lower() == dgrp.lower():
                                        gndx = group[1]
                                        break
                                else:
                                     continue
                                # backup a bit in case rows deleted
                                for r in range(len(del_rows)):
                                    try:
                                        if bs.cell(row=gndx, column=1).value.lower() == group[0].lower():
                                            break
                                    except:
                                        pass
                                    gndx -= 1
                            ditm = get_value(batch_input_sheet, row, 3)
                            for tndx in range(tndx_rows):
                                if bs.cell(row=gndx + tndx, column=1).value is None:
                                    break
                                if bs.cell(row=gndx + tndx, column=1).value.lower() == ditm.lower():
                                    if batch_input_sheet.cell(row=row, column=1).value.lower() == 'data':
                                        values = Reference(bs, min_col=min_col, min_row=gndx + tndx, max_col=max_col, max_row=gndx + tndx)
                                        series = Series(values)
                                        series.title = oxl.chart.series.SeriesLabel(oxl.chart.data_source.StrRef("'" + bs.title + "'!A" + str(gndx + tndx)))
                                        charts[-1].append(series)
                                    elif batch_input_sheet.cell(row=row, column=1).value.lower() == 'data2':
                                        if charts2[-1] is None:
                                            charts2[-1] = LineChart()
                                        values = Reference(bs, min_col=min_col, min_row=gndx + tndx, max_col=max_col, max_row=gndx + tndx)
                                        series = Series(values)
                                        series.title = oxl.chart.series.SeriesLabel(oxl.chart.data_source.StrRef("'" + bs.title + "'!A" + str(gndx + tndx)))
                                        charts2[-1].append(series)
                                    else:
                                        cats = Reference(bs, min_col=min_col, min_row=gndx + tndx, max_col=max_col, max_row=gndx + tndx)
                                    break
                    if in_chart:
                        charts[-1].width = 20
                        charts[-1].height = 12
                        for s in range(len(charts[-1].series)):
                            ser = charts[-1].series[s]
                            ser.marker.symbol = 'circle' #'dot', 'plus', 'triangle', 'x', 'picture', 'star', 'diamond', 'square', 'circle', 'dash', 'auto'
                        if charts2[-1] is not None:
                            for s in range(len(charts2[-1].series)):
                                ser = charts2[-1].series[s]
                                ser.marker.symbol = 'triangle'
                            charts2[-1].y_axis.crosses = 'max'
                            charts[-1] += charts2[-1]
                        if cats is not None:
                            charts[-1].set_categories(cats)
                        if len(charts) % 2:
                            cht_row += 30
                        if chart_group != '':
                            cht_col = col_letters.index(cht_cells[len(charts) % 2])
                            chs.cell(row=cht_row - 1, column=cht_col).value = chart_group
                            chs.cell(row=cht_row - 1, column=cht_col).font = bold
                        chs.add_chart(charts[-1], cht_cells[len(charts) % 2] + str(cht_row))
            if len(self.batch_models) > 1 and len(self.batch_models[0]) == 1:
                try:
                    del wb['Results_' + rpt_time]
                    del wb['Charts_' + rpt_time]
                except:
                    pass
            tim = (time.time() - start_time)
            if tim < 60:
                tim = '%.1f secs' % tim
            else:
                hhmm = tim / 60.
                tim = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
            self.setStatus(f'Saving {self.sender().text()} report ({total_models:,} models; {tim})')
           #     self.setStatus('Saving %s report' % (self.sender().text()))
            self.progressbar.setValue(20)
            QtWidgets.QApplication.processEvents()
            wb.save(batch_report_file)
            tim = (time.time() - start_time)
            if tim < 60:
                tim = '%.1f secs' % tim
            else:
                hhmm = tim / 60.
                tim = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
            self.setStatus(f'{self.sender().text()} completed ({len(self.batch_models)} sheets, {total_models:,} models; {tim}). You may need to open and save the workbook to reprocess it.')
            return
        if do_adjust:
            if self.adjustto is not None:
                for fac, value in self.adjustto.items():
                    try:
                        pmss_details[fac].multiplier = value / pmss_details[fac].capacity
                    except:
                        pass
        if option == D:
            self.processor.doDispatch(year, option, sender_name, pmss_details, pmss_data, re_order, dispatch_order,
                pm_data_file, data_file, self.files, self.sheets)
        else:
            sp_data, corr_data = self.processor.doDispatch(year, option, sender_name, pmss_details, pmss_data, re_order, dispatch_order,
                pm_data_file, data_file, files = None, sheets = None)
        title = None
        if option == B or option == T:
            dialog = displaytable.Table(sp_data, title='Debug', fields=headers,
            save_folder=self.scenarios, sortby='', decpts=sp_pts)
            dialog.exec_()
        span = None
        if self.summary_sources and option != D: # want data sources
            sp_data.append(' ')
            sp_data.append('Data sources')
            span = 'Data sources'
            sp_data.append(['Scenarios folder', self.scenarios])
            if pm_data_file[: len(self.scenarios)] == self.scenarios:
                pm_data_file = pm_data_file[len(self.scenarios):]
            sp_data.append(['Powermatch data file', pm_data_file])
            load_file = self.load_files.replace('$YEAR$', self.loadCombo.currentText())
            if load_file[: len(self.scenarios)] == self.scenarios:
                load_file = load_file[len(self.scenarios):]
            sp_data.append(['Load file', load_file])
            sp_data.append(['Constraints worksheet', str(self.files[C].text()) \
                            + '.' + str(self.sheets[C].currentText())])
            sp_data.append(['Generators worksheet', str(self.files[G].text()) \
                            + '.' + str(self.sheets[G].currentText())])
            sp_pts = [0] * len(headers)
            for p in [st_cap, st_lcg, st_lco, st_lcc, st_max, st_bal, st_rlc, st_are]:
                sp_pts[p] = 2
            if corr_data is not None:
                sp_pts[st_cap] = 3 # compromise between capacity (2) and correlation (4)
            self.setStatus(self.sender().text() + ' completed')
            if title is not None:
                atitle = title
            elif self.results_prefix != '':
                atitle = self.results_prefix + '_' + self.sender().text()
            else:
                atitle = self.sender().text()
            dialog = displaytable.Table(sp_data, title=atitle, fields=headers,
                     save_folder=self.scenarios, sortby='', decpts=sp_pts,
                     span=span)
            dialog.exec_()
            self.progressbar.setValue(20)
            self.progressbar.setHidden(True)
            self.progressbar.setValue(0)

# Detailed Processing
    def show_ProgressBar(self, maximum, msg, title):
        if self.opt_progressbar is None:
            self.opt_progressbar = ProgressBar(maximum=maximum, msg=msg, title=title)
            self.opt_progressbar.setWindowModality(QtCore.Qt.WindowModal)
            self.opt_progressbar.show()
            self.opt_progressbar.setVisible(False)
            self.activateWindow()
        else:
            self.opt_progressbar.barRange(0, maximum, msg=msg)

    def show_FloatStatus(self):
        if not self.log_status:
            return
        if self.floatstatus is None:
            self.floatstatus = FloatStatus(self, self.scenarios, None, program='Powermatch')
            self.floatstatus.setWindowModality(QtCore.Qt.WindowModal)
            self.floatstatus.setWindowFlags(self.floatstatus.windowFlags() |
                         QtCore.Qt.WindowSystemMenuHint |
                         QtCore.Qt.WindowMinMaxButtonsHint)
            self.floatstatus.procStart.connect(self.getStatus)
            self.floatstatus.show()
            self.activateWindow()

    def setStatus(self, text):
        if self.log.text() == text:
            return
        self.log.setText(text)
        if text == '':
            return
        if self.floatstatus and self.log_status:
            self.floatstatus.log(text)
            QApplication.processEvents()

    @QtCore.pyqtSlot(str)
    def getStatus(self, text):
        if text == 'goodbye':
            self.floatstatus = None

    def exit(self):
        self.updated = False
        self.order.updated = False
        self.ignore.updated = False
        if self.floatstatus is not None:
            self.floatstatus.exit()
        self.close()

    def optClicked(self, in_year, in_option, in_pmss_details, in_pmss_data, in_re_order,
                   in_dispatch_order, pm_data_file, data_file):
        sp_data = self.processor.run_optimise(self, in_year, in_option, in_pmss_details, in_pmss_data, in_re_order,
                   in_dispatch_order, pm_data_file, data_file)
        span = None
        dialog = displaytable.Table(sp_data, title=atitle, fields=headers,
            save_folder=self.scenarios, sortby='', decpts=sp_pts,
            span=span)
        dialog.exec_()
