# powerplotui/services/esoo_workbook_parser.py
"""
FR-F02 — structured-source-first extraction. Preferred over PDF parsing
(esoo_pdf_parser.py) wherever a structured workbook is available for a
vintage.

AEMO publishes a 'Data Register' workbook (.xlsx) alongside each WEM ESOO
report: one sheet per report figure, each carrying the exact numeric series
the report chart was drawn from — including the Low/Expected/High scenario
and full POE10/50/90 band values that the report PDF itself only shows as
charts (see esoo_pdf_parser.py's docstring). This is the intended FR-F02
"structured source" for those figures.

Findings from inspecting the 2026 Data Register (worth checking against
future vintages, since sheet names/layout can shift year to year):
  - 88 sheets, one per report figure, named e.g. 'Ch 2_F.7' (chapter 2,
    Figure 7). Each has a title row, blank rows, then a header row buried
    partway down (row ~23) with columns: [blank, label, 'Unit', year
    columns '20XX-YY', ...growth-rate columns].
  - Data rows are labelled either '{vintage_year} {Scenario}' (e.g.
    '2026 Low') or '{vintage_year} {N}% POE' (e.g. '2026 10% POE'),
    interleaved with the same rows from the *prior* vintage (e.g.
    '2025 Low') and an 'Actual' row — only rows whose label starts with
    the current vintage's year are genuine forecast figures.
  - Figure 7 (energy, 3 scenarios), Figure 12 (summer peak POE10, 3
    scenarios), Figure 13 (summer peak POE10/50/90, Expected), Figure 17
    (winter peak POE10, 3 scenarios), Figure 18 (minimum POE90, 3
    scenarios) and Figure 19 (minimum POE10/50/90, Expected) together give
    near-complete Low/Expected/High x POE coverage for operational demand
    and underlying energy — see _extract_2026() below.
"""
import logging
import re
from datetime import date
from pathlib import Path
from typing import Callable, Dict, List

import openpyxl
import pandas as pd
from django.conf import settings

logger = logging.getLogger(__name__)


def read_workbook_sheets(path) -> Dict[str, "pd.DataFrame"]:
    """
    Generic flat-table read (header assumed on row 1) — a convenience for
    simple single-table structured exports (e.g. a Portal CSV/XLSX). NOT
    suitable for the AEMO 'Data Register' format described in the module
    docstring, which has a buried header row; use open_workbook() for that.
    """
    path = Path(path)
    sheets = pd.read_excel(path, sheet_name=None, engine='openpyxl')
    logger.info(f"Read {len(sheets)} sheet(s) from {path.name}: {list(sheets.keys())}")
    return sheets


def open_workbook(path):
    """
    Open a workbook for raw row-level access (values only — formulas are
    pre-evaluated by data_only=True, reading Excel's cached result).
    """
    return openpyxl.load_workbook(Path(path), read_only=True, data_only=True)


_YEAR_RE = re.compile(r'^20\d\d-\d\d$')
_ROW_LABEL_RE = re.compile(r'^(\d{4})\s+(.+)$')
# search (not fullmatch): some vintages label rows '{year} WEM ESOO N% POE'
# rather than plain '{year} N% POE' — the extra words shouldn't matter.
_POE_LABEL_RE = re.compile(r'(\d+)%\s*POE', re.IGNORECASE)


def _extract_scenario_series(
    workbook, sheet_name, vintage_year, metric, *,
    domain='demand', demand_basis='operational', unit='MW', value_scale=1.0,
    label_kind='scenario', fixed_poe_level=None, fixed_scenario=None,
) -> List[dict]:
    """
    Pull one vintage's rows out of a Data-Register-style sheet (see module
    docstring for the layout). label_kind='scenario' reads rows labelled
    '{year} Low/Expected/High' (poe_level comes from fixed_poe_level);
    label_kind='poe' reads rows labelled '{year} N% POE' (scenario comes
    from fixed_scenario).
    """
    if sheet_name not in workbook.sheetnames:
        logger.warning(f"esoo_workbook_parser: sheet '{sheet_name}' not found")
        return []

    rows = list(workbook[sheet_name].iter_rows(values_only=True))

    # Locate the header row by finding a run of 4+ consecutive year-pattern
    # cells ('20XX-YY'), rather than anchoring on header text — confirmed
    # across vintages that the exact column layout shifts (a 'Unit' column
    # before the years in 2023-2026; no 'Unit' column at all in 2022, just
    # 'Scenario'/'POE' — or sometimes nothing — directly before them). The
    # label column is whatever sits immediately to the run's left, skipping
    # one more column left if that cell literally says 'Unit'.
    header_idx = None
    label_col = None
    year_cols: List[tuple] = []
    for i, row in enumerate(rows):
        run_start = None
        run: List[tuple] = []
        for col_idx, cell in enumerate(row):
            if cell is not None and _YEAR_RE.match(str(cell).strip()):
                if run_start is None:
                    run_start = col_idx
                run.append((col_idx, str(cell).strip()))
            else:
                if len(run) >= 4:
                    break
                run_start = None
                run = []
        if len(run) >= 4:
            before_idx = run_start - 1
            is_unit_col = before_idx >= 0 and row[before_idx] is not None and str(row[before_idx]).strip().lower() == 'unit'
            if not is_unit_col and before_idx >= 0:
                # The header row itself may leave this column blank while
                # every *data* row underneath repeats a literal unit value
                # there instead (seen in 2017: no 'Unit' header text, but
                # each row still carries 'MW'/'GWh' at that position).
                for peek in rows[i + 1:i + 4]:
                    if before_idx < len(peek) and peek[before_idx] is not None:
                        if str(peek[before_idx]).strip().lower() in ('mw', 'gwh', 'mwh', 'twh', 'kwh'):
                            is_unit_col = True
                        break
            label_col = before_idx - 1 if is_unit_col else before_idx
            year_cols = run
            header_idx = i
            break

    if header_idx is None or not year_cols or label_col is None or label_col < 0:
        logger.warning(f"esoo_workbook_parser: no usable header row found in '{sheet_name}'")
        return []

    figures = []
    for row in rows[header_idx + 1:]:
        if len(row) <= label_col or not row[label_col]:
            continue
        label = str(row[label_col]).strip()
        match = _ROW_LABEL_RE.match(label)
        if match:
            row_year = int(match.group(1))
            if row_year != vintage_year:
                continue  # skip the prior-vintage comparison rows and 'Actual'
            rest = match.group(2).strip()
        else:
            # Some vintages label single-scenario-set sheets without a year
            # prefix at all (just 'Low'/'Expected'/'High' or 'N% POE') —
            # only every other vintage we've seen prefixes rows with a
            # year, so an un-prefixed row is never mistakeable for a prior
            # vintage's; accept it as this vintage's if (and only if) it's
            # an unambiguous bare scenario/POE label.
            rest = label
            if label.lower() not in ('low', 'expected', 'high') and not _POE_LABEL_RE.search(label):
                continue
        if label_kind == 'scenario':
            scenario = rest.lower()
            if scenario not in ('low', 'expected', 'high'):
                continue
            poe_level = fixed_poe_level
        else:
            poe_match = _POE_LABEL_RE.search(rest)
            if not poe_match:
                continue
            poe_level = int(poe_match.group(1))
            scenario = fixed_scenario

        for col_idx, year_label in year_cols:
            if col_idx >= len(row) or row[col_idx] is None:
                continue
            try:
                num = float(row[col_idx])
            except (TypeError, ValueError):
                continue
            figures.append({
                'domain': domain, 'metric': metric,
                'forecast_year': int(year_label[:4]),
                'demand_growth_scenario': scenario, 'poe_level': poe_level,
                'demand_basis': demand_basis, 'value': num * value_scale, 'unit': unit,
                'table_ref': sheet_name, 'page_ref': '',
                'cell_ref': f"{label} / {year_label}",
            })

    return figures


def _run_sheet_plan(vintage, workbook, plan: List[dict]) -> List[dict]:
    """
    Shared extraction engine: run _extract_scenario_series() once per sheet
    entry in `plan`. This is the reusable part across vintages — verified
    identical between the 2025 and 2026 Data Registers even though the
    *sheet names* shifted (AEMO renumbered figures between editions). Only
    SHEET_PLANS (below) needs a new entry per vintage; this function
    shouldn't need to change just because a year's figure numbering moved.
    """
    figures = []
    for entry in plan:
        entry = dict(entry)
        sheet_name = entry.pop('sheet_name')
        metric = entry.pop('metric')
        figures += _extract_scenario_series(workbook, sheet_name, vintage.year, metric, **entry)
    return figures


# Per-vintage sheet plans: which sheet backs each metric/scenario/POE
# combination, since AEMO renumbers figures (and occasionally reorganises
# them) between editions — confirmed by comparing the 2025 and 2026 Data
# Registers directly. Each entry's kwargs are passed to
# _extract_scenario_series() via _run_sheet_plan() above.
SHEET_PLANS: Dict[int, List[dict]] = {
    2026: [
        dict(sheet_name='Ch 2_F.7', metric='energy',
             domain='demand', demand_basis='underlying', unit='GWh', label_kind='scenario'),
        dict(sheet_name='Ch 2_F.12', metric='peak_summer',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='scenario', fixed_poe_level=10),
        dict(sheet_name='Ch 2_F.13', metric='peak_summer',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='poe', fixed_scenario='expected'),
        dict(sheet_name='Ch 2_F.17', metric='peak_winter',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='scenario', fixed_poe_level=10),
        dict(sheet_name='Ch 2_F.18', metric='minimum',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='scenario', fixed_poe_level=90),
        dict(sheet_name='Ch 2_F.19', metric='minimum',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='poe', fixed_scenario='expected'),
    ],
    2025: [
        # F.8/F.15/F.16/F.17/F.18 — figure numbers shifted vs. 2026 but the
        # row/column template (header row + '{year} Low/Expected/High' or
        # '{year} N% POE' labels) is identical. Unlike 2026, 2025 has no
        # separate "90% POE minimum, 3 scenarios" sheet — F.18 covers only
        # the Expected-scenario POE10/50/90 band, so Low/High minimum
        # figures simply aren't available for this vintage.
        dict(sheet_name='Ch 2_F.8', metric='energy',
             domain='demand', demand_basis='underlying', unit='GWh', label_kind='scenario'),
        dict(sheet_name='Ch 2_F.15', metric='peak_summer',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='scenario', fixed_poe_level=10),
        dict(sheet_name='Ch 2_F.16', metric='peak_summer',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='poe', fixed_scenario='expected'),
        dict(sheet_name='Ch 2_F.17', metric='peak_winter',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='scenario', fixed_poe_level=10),
        dict(sheet_name='Ch 2_F.18', metric='minimum',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='poe', fixed_scenario='expected'),
    ],
    2024: [
        # Same template again; POE rows here are labelled '{year} WEM ESOO
        # N% POE' rather than '{year} N% POE' — handled generically by
        # _POE_LABEL_RE using search() instead of a full match.
        dict(sheet_name='Ch 2_F.6', metric='energy',
             domain='demand', demand_basis='underlying', unit='GWh', label_kind='scenario'),
        dict(sheet_name='Ch 2_F.15', metric='peak_summer',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='scenario', fixed_poe_level=10),
        dict(sheet_name='Ch 2_F.16', metric='peak_summer',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='poe', fixed_scenario='expected'),
        dict(sheet_name='Ch 2_F.17', metric='peak_winter',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='scenario', fixed_poe_level=10),
        dict(sheet_name='Ch 2_F.18', metric='minimum',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='poe', fixed_scenario='expected'),
    ],
    2023: [
        # Different sheet-naming convention ('F. 17' not 'Ch 2_F.17') and an
        # extra leading blank column shifts 'Unit' to column 3 instead of 2
        # — handled generically now (see _extract_scenario_series' header
        # search). F.17 reports energy on the OPERATIONAL basis (unlike
        # 2024-2026, which only give underlying-basis energy by scenario).
        dict(sheet_name='F. 17', metric='energy',
             domain='demand', demand_basis='operational', unit='GWh', label_kind='scenario'),
        dict(sheet_name='F. 22', metric='peak_summer',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='scenario', fixed_poe_level=10),
        dict(sheet_name='F. 23', metric='peak_summer',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='poe', fixed_scenario='expected'),
        dict(sheet_name='F. 24', metric='peak_winter',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='scenario', fixed_poe_level=10),
        dict(sheet_name='F. 25', metric='minimum',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='poe', fixed_scenario='expected'),
    ],
    2022: [
        # No 'Unit' header column at all this year (just 'Scenario'/'POE',
        # or nothing, directly before the years) — handled generically now.
        # No winter-peak figure was published in 2022 (peak/minimum aren't
        # split by season at all this year); F.16/F.17 are the general
        # (summer/annual) peak. F.23's basis isn't labelled explicitly, but
        # its magnitude (~15,800-18,900 GWh) matches 2023's confirmed
        # operational figures far more closely than 2024-2026's underlying
        # ones (~19,900-20,600 GWh), and the report text contrasts it
        # against "underlying consumption" — inferred as operational.
        dict(sheet_name='F.23', metric='energy',
             domain='demand', demand_basis='operational', unit='GWh', label_kind='scenario'),
        dict(sheet_name='F.16', metric='peak_summer',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='scenario', fixed_poe_level=10),
        dict(sheet_name='F.17', metric='peak_summer',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='poe', fixed_scenario='expected'),
        dict(sheet_name='F.18', metric='minimum',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='poe', fixed_scenario='expected'),
    ],
    2021: [
        # Sheet names are plain 'Figure N' this year. Figure 32/33/35 label
        # their rows with no year prefix at all ('Low'/'Expected'/'High'
        # or 'N% POE' bare) — handled generically now (see the bare-label
        # fallback above). No winter-peak figure published in 2021 either.
        dict(sheet_name='Figure 36', metric='energy',
             domain='demand', demand_basis='operational', unit='GWh', label_kind='scenario'),
        dict(sheet_name='Figure 32', metric='peak_summer',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='scenario', fixed_poe_level=10),
        dict(sheet_name='Figure 33', metric='peak_summer',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='poe', fixed_scenario='expected'),
        dict(sheet_name='Figure 35', metric='minimum',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='poe', fixed_scenario='expected'),
    ],
    2020: [
        # File is the 'Figures' half of a Figures/Tables split; sheet names
        # again plain 'Figure N', bare (non-prefixed) row labels again.
        # Note: this vintage's scenario-split minimum-demand figure (32) is
        # anchored at 50% POE, not 90% as in every other vintage — trust
        # the sheet's own stated POE level rather than assuming 90.
        dict(sheet_name='Figure 34', metric='energy',
             domain='demand', demand_basis='operational', unit='GWh', label_kind='scenario'),
        dict(sheet_name='Figure 30', metric='peak_summer',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='scenario', fixed_poe_level=10),
        dict(sheet_name='Figure 31', metric='peak_summer',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='poe', fixed_scenario='expected'),
        dict(sheet_name='Figure 32', metric='minimum',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='scenario', fixed_poe_level=50),
        dict(sheet_name='Figure 33', metric='minimum',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='poe', fixed_scenario='expected'),
    ],
    2019: [
        # No minimum-demand forecast figure was published at all this year
        # (rooftop-PV-driven minimum-demand concern hadn't yet become a
        # headline topic) — energy and summer peak only. Missing-data cells
        # are the literal string '-' rather than blank; already handled
        # since float('-') raises and is caught like any other bad cell.
        dict(sheet_name='Figure 31', metric='energy',
             domain='demand', demand_basis='operational', unit='GWh', label_kind='scenario'),
        dict(sheet_name='Figure 25', metric='peak_summer',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='scenario', fixed_poe_level=10),
        dict(sheet_name='Figure 26', metric='peak_summer',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='poe', fixed_scenario='expected'),
    ],
    2018: [
        # Sheet names are lowercase 'figN' this year; no minimum-demand
        # figure published (same as 2019).
        dict(sheet_name='fig25', metric='energy',
             domain='demand', demand_basis='operational', unit='GWh', label_kind='scenario'),
        dict(sheet_name='fig20', metric='peak_summer',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='scenario', fixed_poe_level=10),
        dict(sheet_name='fig21', metric='peak_summer',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='poe', fixed_scenario='expected'),
    ],
    2017: [
        # Oldest modern-comparable vintage (2016 was deferred — see the
        # Foundation outcomes doc). Header row leaves the label/unit
        # columns blank; each data row repeats its own unit ('MW'/'GWh')
        # instead — handled generically now (see the unit-column peek
        # above). No minimum-demand figure published. fig29 has an extra
        # '10% POE adjusted historical' row alongside Low/Expected/High —
        # correctly ignored since it doesn't match the bare scenario check.
        dict(sheet_name='fig31', metric='energy',
             domain='demand', demand_basis='operational', unit='GWh', label_kind='scenario'),
        dict(sheet_name='fig29', metric='peak_summer',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='scenario', fixed_poe_level=10),
        dict(sheet_name='fig28', metric='peak_summer',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='poe', fixed_scenario='expected'),
        dict(sheet_name='fig30', metric='peak_winter',
             domain='demand', demand_basis='operational', unit='MW',
             label_kind='poe', fixed_scenario='expected'),
    ],
}


# Registered per-vintage workbook -> figure mappers. Signature:
#
#   extractor(vintage: EsooVintage, workbook: openpyxl.Workbook) -> List[dict]
#
# Each returned dict carries the same keys as esoo_pdf_parser.py's
# FIGURE_EXTRACTORS (domain, metric, forecast_year, demand_growth_scenario,
# poe_level, demand_basis, value, unit, table_ref, page_ref, cell_ref).
# parse_esoo_workbook_to_figures() fills in the remaining common provenance.
# All years currently share _run_sheet_plan; add a genuinely different
# extractor function only if some future vintage's workbook breaks the
# scenario/POE row-label template this engine assumes.
FIGURE_EXTRACTORS: Dict[int, Callable] = {
    year: (lambda vintage, workbook, _plan=plan: _run_sheet_plan(vintage, workbook, _plan))
    for year, plan in SHEET_PLANS.items()
}


def parse_esoo_workbook_to_figures(vintage, doc_type='data_register') -> List[dict]:
    """
    FR-F02 structured-source entry point. Looks up the vintage's
    SourceDocument for doc_type (default 'data_register'), opens it,
    and hands it to the registered FIGURE_EXTRACTORS mapper.

    Raises NotImplementedError if no mapper is registered for this vintage,
    and ValueError if the source document hasn't been fetched yet — same
    "surface, don't fabricate" rationale as esoo_pdf_parser.py.
    """
    doc = vintage.source_documents.filter(doc_type=doc_type).first()
    if doc is None or not doc.local_file_path:
        raise ValueError(
            f"EsooVintage {vintage.year} has no retrieved '{doc_type}' document; "
            f"fetch it first via `ingest_esoo_vintage --year {vintage.year} --doc-type {doc_type}`."
        )

    extractor = FIGURE_EXTRACTORS.get(vintage.year)
    if extractor is None:
        raise NotImplementedError(
            f"No FIGURE_EXTRACTORS mapping registered for ESOO {vintage.year}. "
            f"Register a mapper in esoo_workbook_parser.FIGURE_EXTRACTORS once "
            f"the sheet layout is confirmed."
        )

    # Opening/scanning the workbook below is slow enough on large files
    # that an idle DB connection can be dropped server-side (MySQL
    # wait_timeout) before the caller's next write — close it explicitly so
    # Django reconnects fresh afterward instead of erroring on a stale one.
    from django.db import connection
    connection.close()

    workbook = open_workbook(Path(settings.ESOO_ARCHIVE_DIR) / doc.local_file_path)
    figures = extractor(vintage, workbook)

    today = date.today()
    for figure in figures:
        figure.setdefault('source_document', doc.local_file_path)
        figure.setdefault('source_version', str(vintage.year))
        figure.setdefault('extraction_date', today)
        figure.setdefault('extraction_method', 'structured')

    return figures
