import base64
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.generic import TemplateView
from django.shortcuts import render
from ..forms import PlotForm
import io
import logging
import matplotlib.pyplot as plt
import re
from siren_web.models import Analysis, Scenarios, variations, Technologies
from siren_web.database_operations import fetch_analysis_scenario
import openpyxl
import pandas as pd

class VariantsView(TemplateView):
    template_name = 'variants.html'

    # Function to fetch data from the database
    def get_analysis_data(self, analysis_queryset):
        analysis_data = []
        for obj in analysis_queryset:
            obj_data = {}
            for field in obj._meta.fields:
                if field.name == 'idanalysis':
                    continue
                value = getattr(obj, field.name)
                if field.name == 'idscenarios':
                    obj_data['Scenario'] = value.title
                else:
                    obj_data[field.name.capitalize()] = value
            analysis_data.append(obj_data)
        return analysis_data
    
    def export_to_excel(self, request):
        # Get the selected parameters from the request.POST
        idscenarios = request.POST.get('scenario')
        idvariant = request.POST.get('variant')

        # Filter the Analysis data based on the selected parameters
        analysis_queryset = Analysis.objects.filter(
            idscenarios_id=idscenarios,
            variation__in=[variations.objects.get(pk=idvariant).variation_name, 'Baseline'],
        ).order_by('idanalysis')
        stages = Analysis.objects.filter(
            idscenarios_id=idscenarios,
            variation__in=[variations.objects.get(pk=idvariant).variation_name, 'Baseline'],
        ).values_list('stage', flat=True).distinct().order_by('stage')
        # Create a new workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Write the column headers
        headers = [field.name for field in Analysis._meta.fields if field.name not in ['idanalysis', 'idscenarios']]
        # stages = analysis_queryset.values_list('stage', flat=True).distinct()
        headers = ['Statistic', 'Component']
        headers.extend(['Stage ' + str(stage) for stage in stages])
        worksheet.append(headers)

        # Write the data rows
        last_column = 0
        for analysis in analysis_queryset:
            column = analysis.stage + 3
            if column != last_column:
                row = 2
                last_column = column
            if column == 3:
                cell = worksheet.cell(row=row, column=1)
                cell.value = analysis.heading
                cell = worksheet.cell(row=row, column=2)
                cell.value = analysis.component
            cell = worksheet.cell(row=row, column=column)
            cell.value = analysis.quantity  # Set the value of the new column
            row = row + 1

        # Set the response headers
        file_name = Scenarios.objects.get(pk=idscenarios).title + '_' + variations.objects.get(pk=idvariant).variation_name
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename={file_name}.xlsx"

        # Save the workbook to the response
        workbook.save(response)
        return response
        
    def get(self, request):
        analysis_queryset = fetch_analysis_scenario('Current')
        analysis_data = self.get_analysis_data(analysis_queryset)
            
        plotform = PlotForm()
        context = {
            'analysis_data' : analysis_data,
            'plotform': plotform,
        }
        return render(request, 'variants.html', context)

    def post(self, request):
        # Check if this is a back navigation from a chart
        if request.POST.get('back_from_chart') == 'true':
            return self.handle_back_navigation(request)
        
        plotform = PlotForm(request.POST or None, selected_scenario=request.POST.get('scenario'))
        idscenarios = request.POST.get('scenario')
        idvariant = request.POST.get('variant')
        
        # Get form data
        series_1 = request.POST.get('series_1')
        series_2 = request.POST.get('series_2')
        series_1_component = request.POST.get('series_1_component')
        series_2_component = request.POST.get('series_2_component')
        plot_type = request.POST.get('plot_type', '')
        if plotform.is_valid():
            # Handle different form actions
            if plot_type:
                return self.handle_plot_action(
                    request, plot_type, idscenarios, idvariant, 
                    series_1, series_2, series_1_component, series_2_component
                )
            
            elif 'export' in request.POST:
                return self.handle_export_action(request, idscenarios, idvariant)
        else:
            context = {'plotform': plotform}
            return render(request, 'variants.html', context)
        
        # If no specific action, return to form
        return self.render_form_with_error(request, plotform)

    def handle_back_navigation(self, request):
        """Handle when user returns from chart page with preserved parameters"""
        # Extract parameters from POST data
        scenario = request.POST.get('scenario')
        variant = request.POST.get('variant') 
        series_1 = request.POST.get('series_1')
        series_2 = request.POST.get('series_2')
        series_1_component = request.POST.get('series_1_component')
        series_2_component = request.POST.get('series_2_component')
        chart_type = request.POST.get('chart_type')
        chart_specialization = request.POST.get('chart_specialization')
        
        # Create form with preserved data
        form_data = {
            'scenario': scenario,
            'variant': variant,
            'series_1': series_1,
            'series_2': series_2,
            'series_1_component': series_1_component,
            'series_2_component': series_2_component,
            'chart_type': chart_type,
            'chart_specialization': chart_specialization
        }
        
        # Create form instance with the preserved data
        plotform = PlotForm(form_data=form_data, selected_scenario=scenario)
        
        # Get analysis data for display
        analysis_queryset = Analysis.objects.filter(
        idscenarios=scenario).all()[:8]
        analysis_data = self.get_analysis_data(analysis_queryset)
        
        context = {
            'analysis_data': analysis_data,
            'plotform': plotform,
        }
        request.session['chartback'] = variant  # Set session variable to indicate back navigation
        return render(request, 'variants.html', context)

    def render_form_with_error(self, request, plotform=None):
        """Render the form with any error messages"""
        if plotform is None:
            plotform = PlotForm()
        
        # Get some default analysis data to display
        analysis_queryset = Analysis.objects.all()[:20]
        analysis_data = self.get_analysis_data(analysis_queryset)
        
        context = {
            'plotform': plotform,
            'analysis_data': analysis_data,
        }
        return render(request, 'variants.html', context)

    def handle_plot_action(self, request, plot_type, idscenarios, idvariant, series_1, series_2, series_1_component, series_2_component):
        """Handle plot generation actions"""
        
        if plot_type == 'Echart':
            chart_type = request.POST.get('chart_type')
            chart_specialization = request.POST.get('chart_specialization')
            
            context = {
                'series_1': series_1,
                'series_2': series_2,
                'series_1_component': series_1_component,
                'series_2_component': series_2_component,
                'scenario': idscenarios,
                'variant': idvariant,
                'chart_type': chart_type,
                'chart_specialization': chart_specialization,
            }
            return render(request, 'echarts.html', context)
        
        else:
            messages.error(request, f'Unknown plot type: {plot_type}')
            return self.render_form_with_error(request)

    def handle_export_action(self, request, idscenarios, idvariant):
        """Handle export to Excel action"""
        try:
            # Get scenario and variant names for validation
            scenario = Scenarios.objects.get(pk=idscenarios)
            variant = variations.objects.get(pk=idvariant)
            
            # Generate and return the file response directly
            return self.export_to_excel(request)
            
        except Scenarios.DoesNotExist:
            messages.error(request, 'Selected scenario does not exist.')
            return self.render_form_with_error(request)
        except variations.DoesNotExist:
            messages.error(request, 'Selected variant does not exist.')
            return self.render_form_with_error(request)
        except Exception as e:
            messages.error(request, f'Error exporting to Excel: {str(e)}')
            return self.render_form_with_error(request)

    @staticmethod
    def get_valid_choices(request):
        scenario_id = request.GET.get('scenario')
        variant_id = request.GET.get('variant')
        series_1_heading = request.GET.get('series_1_heading')
        series_2_heading = request.GET.get('series_2_heading')
        series_2_component = request.GET.get('series_2_component')
        update_type = request.GET.get('update_type')
        # Handle back navigation reset without removing variants
        if update_type == 'updateVariants':
            chartback= request.session.get('chartback', '')
            if chartback:
                request.session['chartback'] = ''  # Reset after reading
                response_data = {'chartback': 'true'}
                return JsonResponse(response_data)
        request.session['chartback'] = ''
        response_data = {
            'variants': [],
            'series_1_headings': [],
            'series_1_components': [],
            'series_2_headings': [],
            'series_2_components': []
        }
        
        try:
            # Handle specific update types for series 2 bidirectional dependencies
            if update_type == 'series_2_headings_for_component':
                return VariantsView.handle_series_2_headings_for_component(request, scenario_id, variant_id, series_2_component)
            elif update_type == 'series_2_components_for_heading':
                return VariantsView.handle_series_2_components_for_heading(request, scenario_id, variant_id, series_2_heading)
            
            if scenario_id:
                variants_queryset = variations.objects.filter(idscenarios=scenario_id)
                variant_count = variants_queryset.count()
                
                if variant_count == 0:
                    # No variants available
                    response_data['variants'] = [
                        {'id': '', 'name': 'No variants available for this scenario'}
                    ]
                elif variant_count == 1:
                    # Only one variant - auto-select it
                    variant = variants_queryset.first()
                    variant_id = variant.idvariations # type: ignore
                    response_data['variants'] = [
                        {'id': variant.idvariations, 'name': variant.variation_name, 'selected': True} # type: ignore
                    ]
                else:
                    # Multiple variants - show selection prompt
                    response_data['variants'] = [
                        {'id': '', 'name': 'Select a Variant'}
                    ] + [
                        {'id': variant.idvariations, 'name': variant.variation_name} 
                        for variant in variants_queryset
                    ]
                
                # Filter Analysis based on scenario
                analysis_filter = {'idscenarios': scenario_id}
                
                # If a specific variant is selected, apply technology and dimension constraints
                if variant_id:
                    try:
                        selected_variant = variations.objects.get(idvariations=variant_id)
                        variation_name = selected_variant.variation_name
                        analysis_filter['variation'] = variation_name
                        
                        # Parse variation_name to extract technology_signature and dimension
                        technology_signature, dimension = VariantsView.parse_variation_name(variation_name)
                        
                        if technology_signature and dimension:
                            # Apply constraints based on parsed variation name
                            response_data = VariantsView.apply_variation_constraints(
                                response_data, scenario_id, technology_signature, 
                                dimension, series_1_heading, series_2_heading, analysis_filter, request
                            )
                        else:
                            # Display error message if parsing fails
                            messages.error(request, f"Invalid variation name format: '{variation_name}'. Expected format: technology_dimension[digits]")
                            response_data['series_1_headings'] = []
                            response_data['series_1_components'] = []
                            response_data['series_2_headings'] = []
                            response_data['series_2_components'] = []
                            
                    except variations.DoesNotExist:
                        # Display error message if variant doesn't exist
                        messages.error(request, f"Selected variant with ID '{variant_id}' does not exist")
                        response_data['series_1_headings'] = []
                        response_data['series_1_components'] = []
                        response_data['series_2_headings'] = []
                        response_data['series_2_components'] = []
                else:
                    if update_type != 'updateVariants':
                        messages.info(request, f"No variant selected. Please select a variant to see valid choices.")
                    response_data['series_1_headings'] = []
                    response_data['series_1_components'] = []
                    response_data['series_2_headings'] = []
                    response_data['series_2_components'] = []
        
        except Exception as e:
            # Log the error and display error message
            error_msg = f"Error in get_valid_choices: {str(e)}"
            messages.error(request, error_msg)
            response_data['series_1_headings'] = []
            response_data['series_1_components'] = []
            response_data['series_2_headings'] = []
            response_data['series_2_components'] = []
        
        return JsonResponse(response_data)

    @staticmethod
    def handle_series_2_headings_for_component(request, scenario_id, variant_id, series_2_component):
        """
        Handle updating series 2 headings when series 2 component changes.
        If component is null, return all headings for the scenario/variant.
        """
        response_data = {
            'series_2_headings': []
        }
        
        try:
            if not scenario_id or not variant_id:
                return JsonResponse(response_data)
            
            # Get the selected variant to build analysis filter
            try:
                selected_variant = variations.objects.get(idvariations=variant_id)
                variation_name = selected_variant.variation_name
                
                analysis_filter = {
                    'idscenarios': scenario_id,
                    'variation': variation_name
                }
                
                # If component is specified, add it to the filter
                if series_2_component:
                    analysis_filter['component'] = series_2_component
                
                # Query available headings
                headings = Analysis.objects.filter(**analysis_filter).values_list('heading', flat=True).distinct().order_by('heading')
                response_data['series_2_headings'] = list(headings)
                
            except variations.DoesNotExist:
                messages.error(request, f"Selected variant with ID '{variant_id}' does not exist")
                
        except Exception as e:
            error_msg = f"Error getting series 2 headings for component: {str(e)}"
            messages.error(request, error_msg)
        
        return JsonResponse(response_data)

    @staticmethod
    def handle_series_2_components_for_heading(request, scenario_id, variant_id, series_2_heading):
        """
        Handle updating series 2 components when series 2 heading changes.
        If heading is null, return all components for the scenario/variant.
        """
        response_data = {
            'series_2_components': []
        }
        
        try:
            if not scenario_id or not variant_id:
                return JsonResponse(response_data)
            
            # Get the selected variant to build analysis filter
            try:
                selected_variant = variations.objects.get(idvariations=variant_id)
                variation_name = selected_variant.variation_name
                
                analysis_filter = {
                    'idscenarios': scenario_id,
                    'variation': variation_name
                }
                
                # If heading is specified, add it to the filter
                if series_2_heading:
                    analysis_filter['heading'] = series_2_heading
                
                # Query available components
                components = Analysis.objects.filter(**analysis_filter).values_list('component', flat=True).distinct().order_by('component')
                response_data['series_2_components'] = list(components)
                
            except variations.DoesNotExist:
                messages.error(request, f"Selected variant with ID '{variant_id}' does not exist")
                
        except Exception as e:
            error_msg = f"Error getting series 2 components for heading: {str(e)}"
            messages.error(request, error_msg)
        
        return JsonResponse(response_data)

    @staticmethod
    def parse_variation_name(variation_name):
        """
        Parse variation_name to extract technology_signature and dimension.
        
        Format: technology_signature_dimension[numeric_digits]
        Returns: (technology_signature, full_dimension_name)
        """
        # Mapping from abbreviated dimension to full dimension name
        dimension_mapping = {
            'mul': 'multiplier',
            'cap': 'capex', 
            'fom': 'fom',
            'vom': 'vom',
            'lif': 'lifetime'
        }
        
        try:
            # Split by .
            parts = variation_name.split('.')
            if len(parts) < 2:
                return None, None
                
            technology_signature = parts[0]
            
            # The second part contains dimension followed by numeric digits representing stages and step
            dimension_part = parts[1]
            
            # Extract dimension (characters before numeric digits)
            match = re.match(r'^([a-zA-Z]+)', dimension_part)
            if match:
                dimension_abbrev = match.group(1).lower()
                full_dimension = dimension_mapping.get(dimension_abbrev)
                return technology_signature, full_dimension
            
            return technology_signature, None
            
        except Exception as e:
            print(f"Error parsing variation name '{variation_name}': {str(e)}")
            return None, None

    @staticmethod
    def apply_variation_constraints(response_data, scenario_id, technology_signature, 
                                dimension, series_1_heading, series_2_heading, analysis_filter, request):
        """
        Apply constraints based on parsed variation name.
        """
        try:
            # Get technology name from Technologies table
            try:
                technology = Technologies.objects.get(technology_signature=technology_signature)
                technology_name = technology.technology_name
            except Technologies.DoesNotExist:
                # Return error if technology not found
                messages.error(request, f"Technology with signature '{technology_signature}' not found in Technologies table")
                response_data['series_1_headings'] = []
                response_data['series_1_components'] = []
                response_data['series_2_headings'] = []
                response_data['series_2_components'] = []
                return response_data
            
            # Series 1 component constraint: only the specific technology
            response_data['series_1_components'] = [technology_name]
            if dimension == 'multiplier':
                # Get the valid series 2 combinations for the variant and any component
                available_components = list(
                    Analysis.objects.filter(**analysis_filter)
                    .values_list('component', flat=True)
                    .distinct()
                    .order_by('component')
                )
                response_data['series_2_components'] = available_components
                response_data['series_2_headings'] = response_data['series_1_headings']
                # Get the valid series 1 combinations for the variant and the specific technology
                analysis_filter['component'] = technology_name
                valid_analysis = Analysis.objects.filter(**analysis_filter)
                response_data['series_1_headings'] = list(
                    valid_analysis.values_list('heading', flat=True).distinct().order_by('heading')
                )

            else:
                # For other dimensions (capex, fom, vom, lifetime), apply specific constraints
                # Series 2 component constraints: System Total, System Economics, Load Analysis
                allowed_series_2_components = ['System Economics']
                response_data['series_2_components'] = allowed_series_2_components
                if dimension == 'capex':
                    allowed_series_1_headings = ['Capital Cost']
                elif dimension in ['fom', 'vom']:
                    allowed_series_1_headings = ['Annual Cost']
                elif dimension == 'lifetime':
                    allowed_series_1_headings = ['Lifetime', 'Lifetime Cost']
                allowed_series_1_headings.extend(['Lifetime Cost', 'LCOG Cost', 'LCOE Cost', 'LCOE with CO2 Cost'])
                response_data['series_1_headings'] = allowed_series_1_headings
                response_data['series_2_headings'] = allowed_series_1_headings
        
        except Exception as e:
            print(f"Error applying variation constraints: {str(e)}")
            # Return error instead of falling back
            messages.error(request, f"Error applying variation constraints: {str(e)}")
            response_data['series_1_headings'] = []
            response_data['series_1_components'] = []
            response_data['series_2_headings'] = []
            response_data['series_2_components'] = []
        
        return response_data
