import altair as alt
import base64
from django.http import HttpResponse, JsonResponse
from django.views.generic import TemplateView
from django.shortcuts import render
from django.contrib import messages
from ..forms import PlotForm
import io
import logging
import matplotlib
matplotlib.use('Agg')  # Use the 'Agg' backend for non-interactive plotting
import matplotlib.pyplot as plt
from siren_web.models import Analysis, Scenarios, variations
import openpyxl
import pandas as pd

class PowerPlotHomeView(TemplateView):
    template_name = 'powerplotui_home.html'

    # Function to fetch data from the database

    def get_analysis_data(self, analysis_queryset):
        analysis_data = []
        for obj in analysis_queryset:
            obj_data = {}
            for field in obj._meta.fields:
                if field.name == 'idanalysis':
                    continue
                value = getattr(obj, field.name)
                if field.name == 'idscenarios':
                    obj_data['Scenario'] = value.title
                else:
                    obj_data[field.name.capitalize()] = value
            analysis_data.append(obj_data)
        return analysis_data
    
    def create_chart(self, request, plot_type):
        series_1 = request.POST.get('series_1')
        series_2 = request.POST.get('series_2')
        series_1_component = request.POST.get('series_1_component')
        series_2_component = request.POST.get('series_2_component')
        scenario = request.POST.get('scenario')
        variant = request.POST.get('variant')
        chart_type = request.POST.get('chart_type', 'line')
        chart_specialization = request.POST.get('chart_specialization', '')
        variation = variations.objects.get(pk=variant)

        analysis_queryset_1 = Analysis.objects.filter(
            idscenarios=scenario,
            variation__in=[variation.variation_name, 'Baseline'],
            heading=series_1,
            component=series_1_component,
        ).order_by('stage')

        analysis_queryset_2 = Analysis.objects.filter(
            idscenarios=scenario,
            variation__in=[variation.variation_name, 'Baseline'],
            heading=series_2,
            component=series_2_component,
        ).order_by('stage')
        
        analysis_data = []
        
        # Create a dictionary from analysis_queryset_2 with stage as the key
        stage_dict = {obj.stage: obj for obj in analysis_queryset_2}

        for analysis_obj_1 in analysis_queryset_1:
            # Check if the stage exists in the dictionary
            if analysis_obj_1.stage in stage_dict:
                analysis_obj_2 = stage_dict[analysis_obj_1.stage]
                analysis_data.append({
                    'stage': analysis_obj_1.stage,
                    'series_1_name': series_1,
                    'series_1_value': analysis_obj_1.quantity,
                    'series_2_name': series_2,
                    'series_2_value': analysis_obj_2.quantity,
                    'chart_type': chart_type,
                    'chart_specialization': chart_specialization
                })
        if (plot_type == 'Altair'):
            # Create an Altair chart
            df = pd.DataFrame([{'stage': item['stage'], 'series_1_name': item['series_1_name'], 'series_1_value': item['series_1_value']} for item in analysis_data])
            series_1_name = analysis_data[0]['series_1_name']

            chart = alt.Chart(df).mark_line().encode(
                x='stage',
                y=alt.Y('series_1_value', title=df['series_1_name'].iloc[0]),
                color='series_1_name'
            )

            # Save the chart as an HTML string
            html = chart.to_html()

            context = {
                'chart_html': html,
            }
            return render(request, 'altair.html', context)
        else:
            # Set the logging level to WARNING to suppress DEBUG and INFO messages
            logging.getLogger('matplotlib').setLevel(logging.WARNING)
            # Create a Matplotlib figure and plot the data
            fig, ax = plt.subplots()
            ax.plot([item['stage']for item in analysis_data], [item['series_1_value'] for item in analysis_data])
            ax.plot([item['stage']for item in analysis_data], [item['series_2_value'] for item in analysis_data])
            ax.set_xlabel(analysis_data[0]['series_1_name'])
            ax.set_ylabel(analysis_data[0]['series_2_name'])
            ax.set_title('Analysis Plot')

            # Save the figure to a BytesIO object
            buf = io.BytesIO()
            fig.savefig(buf, format='png')
            buf.seek(0)

            # Encode the image data as a base64 string
            image_data = base64.b64encode(buf.getvalue()).decode('utf-8')

            context = {
                'image_data': image_data,
                # ... (other context variables)
            }

            return render(request, 'matplotlib.html', context)
        
    def export_to_excel(self, request):
        # Get the selected parameters from the request.POST
        idscenarios = request.POST.get('scenario')
        idvariant = request.POST.get('variant')

        # Filter the Analysis data based on the selected parameters
        analysis_queryset = Analysis.objects.filter(
            idscenarios_id=idscenarios,
            variation__in=[variations.objects.get(pk=idvariant).variation_name, 'Baseline'],
        ).order_by('idanalysis')
        stages = Analysis.objects.filter(
            idscenarios_id=idscenarios,
            variation__in=[variations.objects.get(pk=idvariant).variation_name, 'Baseline'],
        ).values_list('stage', flat=True).distinct().order_by('stage')
        # Create a new workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Write the column headers
        headers = [field.name for field in Analysis._meta.fields if field.name not in ['idanalysis', 'idscenarios']]
        # stages = analysis_queryset.values_list('stage', flat=True).distinct()
        headers = ['Statistic', 'Component']
        headers.extend(['Stage ' + str(stage) for stage in stages])
        worksheet.append(headers)

        # Write the data rows
        last_column = 0
        for analysis in analysis_queryset:
            column = analysis.stage + 3
            if column != last_column:
                row = 2
                last_column = column
            if column == 3:
                cell = worksheet.cell(row=row, column=1)
                cell.value = analysis.heading
                cell = worksheet.cell(row=row, column=2)
                cell.value = analysis.component
            cell = worksheet.cell(row=row, column=column)
            cell.value = analysis.quantity  # Set the value of the new column
            row = row + 1

        # Set the response headers
        file_name = Scenarios.objects.get(pk=idscenarios).title + '_' + variations.objects.get(pk=idvariant).variation_name
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename={file_name}.xlsx"

        # Save the workbook to the response
        workbook.save(response)
        return response
        
    def get(self, request):
        analysis_queryset = Analysis.objects.all()[:8]
        analysis_data = self.get_analysis_data(analysis_queryset)
            
        plotform = PlotForm()
        context = {
            'analysis_data' : analysis_data,
            'plotform': plotform,
        }
        return render(request, 'powerplotui_home.html', context)

    def post(self, request):
        plotform = PlotForm(request.POST)
        idscenarios = request.POST.get('scenario')
        idvariant = request.POST.get('variant')
        
        # Validate required fields first
        if not idvariant or idvariant == '':
            messages.error(request, 'Please select a variant first.')
            return self.render_form_with_error(request, plotform)
        
        if not idscenarios or idscenarios == '':
            messages.error(request, 'Please select a scenario first.')
            return self.render_form_with_error(request, plotform)
        
        # Validate that variant exists
        try:
            variant = variations.objects.get(pk=idvariant)
        except variations.DoesNotExist:
            messages.error(request, 'Selected variant does not exist.')
            return self.render_form_with_error(request, plotform)
        
        # Get form data
        series_1 = request.POST.get('series_1')
        series_2 = request.POST.get('series_2')
        series_1_component = request.POST.get('series_1_component')
        series_2_component = request.POST.get('series_2_component')
        plot_type = request.POST.get('plot_type', '')
        
        # Handle different form actions
        if plot_type:
            return self.handle_plot_action(
                request, plot_type, idscenarios, idvariant, 
                series_1, series_2, series_1_component, series_2_component
            )
        
        elif 'export' in request.POST:
            return self.handle_export_action(request, idscenarios, idvariant)
        
        # If no specific action, return to form
        return self.render_form_with_error(request, plotform)

    def render_form_with_error(self, request, plotform=None):
        """Render the form with any error messages"""
        if plotform is None:
            plotform = PlotForm()
        
        # Get some default analysis data to display
        analysis_queryset = Analysis.objects.all()[:8]
        analysis_data = self.get_analysis_data(analysis_queryset)
        
        context = {
            'plotform': plotform,
            'analysis_data': analysis_data,
        }
        return render(request, 'powerplotui_home.html', context)

    def handle_plot_action(self, request, plot_type, idscenarios, idvariant, series_1, series_2, series_1_component, series_2_component):
        """Handle plot generation actions"""
        if plot_type in ['Altair', 'Matplotlib']:
            return self.create_chart(request, plot_type)
        
        elif plot_type == 'Echart':
            chart_type = request.POST.get('chart_type')
            chart_specialization = request.POST.get('chart_specialization')
            
            context = {
                'series_1': series_1,
                'series_2': series_2,
                'series_1_component': series_1_component,
                'series_2_component': series_2_component,
                'scenario': idscenarios,
                'variant': idvariant,
                'chart_type': chart_type,
                'chart_specialization': chart_specialization,
            }
            return render(request, 'echarts.html', context)
        
        else:
            messages.error(request, f'Unknown plot type: {plot_type}')
            return self.render_form_with_error(request)

    def handle_export_action(self, request, idscenarios, idvariant):
        """Handle export to Excel action"""
        try:
            # Get scenario and variant names for validation
            scenario = Scenarios.objects.get(pk=idscenarios)
            variant = variations.objects.get(pk=idvariant)
            
            # Generate and return the file response directly
            return self.export_to_excel(request)
            
        except Scenarios.DoesNotExist:
            messages.error(request, 'Selected scenario does not exist.')
            return self.render_form_with_error(request)
        except variations.DoesNotExist:
            messages.error(request, 'Selected variant does not exist.')
            return self.render_form_with_error(request)
        except Exception as e:
            messages.error(request, f'Error exporting to Excel: {str(e)}')
            return self.render_form_with_error(request)
    
    @staticmethod
    def get_valid_choices(request):
        scenario_id = request.GET.get('scenario')
        variant_id = request.GET.get('variant')
        series_1_heading = request.GET.get('series_1_heading')
        series_2_heading = request.GET.get('series_2_heading')
        
        response_data = {
            'variants': [],
            'headings': [],
            'components': [],
            'series_1_components': [],
            'series_2_components': []
        }
        
        try:
            # Get variants for the selected scenario
            if scenario_id:
                variants_queryset = variations.objects.filter(idscenarios=scenario_id)
                variant_count = variants_queryset.count()
                
                if variant_count == 0:
                    # No variants available
                    response_data['variants'] = [
                        {'id': '', 'name': 'No variants available for this scenario'}
                    ]
                elif variant_count == 1:
                    # Only one variant - auto-select it
                    variant = variants_queryset.first()
                    response_data['variants'] = [
                        {'id': variant.idvariations, 'name': variant.variation_name, 'selected': True}
                    ]
                else:
                    # Multiple variants - show selection prompt
                    response_data['variants'] = [
                        {'id': '', 'name': 'Select a Variant'}
                    ] + [
                        {'id': variant.idvariations, 'name': variant.variation_name} 
                        for variant in variants_queryset
                    ]
                
                # Filter Analysis based on scenario
                analysis_filter = {'idscenarios': scenario_id}
                
                # If a specific variant is selected
                if variant_id:
                    try:
                        selected_variant = variations.objects.get(idvariations=variant_id)
                        analysis_filter['variation'] = selected_variant.variation_name
                    except variations.DoesNotExist:
                        pass
                
                # Get valid headings
                valid_analysis = Analysis.objects.filter(**analysis_filter)
                response_data['headings'] = list(valid_analysis.values_list('heading', flat=True).distinct().order_by('heading'))
                
                # For general components, we don't filter by heading since no heading is selected yet
                # This will be empty initially and populated when headings are selected
                response_data['components'] = []
                
                # Get components for specific headings
                if series_1_heading:
                    series_1_filter = analysis_filter.copy()
                    series_1_filter['heading'] = series_1_heading
                    response_data['series_1_components'] = list(
                        Analysis.objects.filter(**series_1_filter).values_list('component', flat=True).distinct().order_by('component')
                    )
                
                if series_2_heading:
                    series_2_filter = analysis_filter.copy()
                    series_2_filter['heading'] = series_2_heading
                    response_data['series_2_components'] = list(
                        Analysis.objects.filter(**series_2_filter).values_list('component', flat=True).distinct().order_by('component')
                    )
        
        except Exception as e:
            # Log the error in production
            pass
        
        return JsonResponse(response_data)
