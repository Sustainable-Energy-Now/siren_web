"""
Views for comparing SCADA (actual) vs SupplyFactors (simulated) generation data.

This module provides views to overlay and compare actual SCADA generation data
with simulated SupplyFactors data for facilities, facility groups, and technologies.
"""

from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from siren_web.models import facilities, FacilityScada, supplyfactors, Technologies
import openpyxl
from openpyxl.utils import get_column_letter

from ..services.generation_utils import (
    get_hour_range_from_months,
    aggregate_by_period,
    aggregate_multiple_fields,
    calculate_correlation_metrics,
    calculate_error_metrics,
    get_x_label,
    PEAK_HOUR_PRESETS
)
from ..services.time_series_aligner import TimeSeriesAligner

# Initialize the time series aligner
aligner = TimeSeriesAligner()


def generation_comparison_view(request):
    """Main view for SCADA vs SupplyFactors comparison.

    Shows only facilities and technologies that have both SCADA and SupplyFactors data.
    """
    # Get years with data in both sources
    common_years = aligner.get_comparable_years(FacilityScada, supplyfactors)

    # Get facilities that have both SCADA and SupplyFactors data
    comparable_facilities = aligner.get_comparable_facilities(
        facilities, FacilityScada, supplyfactors
    )

    # Get technologies that have comparable facilities
    comparable_technologies = aligner.get_comparable_technologies(
        Technologies, facilities, FacilityScada, supplyfactors
    )

    context = {
        'facilities': comparable_facilities,
        'technologies': comparable_technologies,
        'years': common_years,
    }
    return render(request, 'generation_comparison.html', context)


def get_facility_scada_vs_supply(request):
    """API endpoint to compare SCADA vs SupplyFactors for a single facility."""
    facility_id = request.GET.get('facility_id')
    year = request.GET.get('year')
    aggregation = request.GET.get('aggregation', 'hour')
    start_hour = request.GET.get('start_hour')
    end_hour = request.GET.get('end_hour')
    peak_filter = request.GET.get('peak_filter', 'all')  # Peak hours filter

    # Validation
    if not facility_id:
        return JsonResponse({'error': 'facility_id is required'}, status=400)
    if not year:
        return JsonResponse({'error': 'year is required'}, status=400)

    try:
        facility = facilities.objects.select_related('idtechnologies').get(idfacilities=facility_id)
        year = int(year)
    except facilities.DoesNotExist:
        return JsonResponse({'error': 'Facility not found'}, status=404)
    except ValueError:
        return JsonResponse({'error': 'Invalid year format'}, status=400)

    # Parse hour range
    start_hour_int = None
    end_hour_int = None
    if start_hour and end_hour:
        try:
            start_hour_int = int(start_hour)
            end_hour_int = int(end_hour)
        except ValueError:
            return JsonResponse({'error': 'Invalid hour range'}, status=400)

    # Get SCADA data and convert to hourly
    scada_qs = FacilityScada.objects.filter(
        facility=facility,
        dispatch_interval__year=year
    ).order_by('dispatch_interval')

    if not scada_qs.exists():
        return JsonResponse({
            'error': f'No SCADA data found for {facility.facility_name} in {year}'
        }, status=404)

    scada_hourly = aligner.convert_scada_to_hourly(scada_qs, year, start_hour_int, end_hour_int)

    # Get SupplyFactors data
    supply_qs = supplyfactors.objects.filter(
        idfacilities=facility,
        year=year
    )
    if start_hour_int and end_hour_int:
        supply_qs = supply_qs.filter(hour__gte=start_hour_int, hour__lte=end_hour_int)

    if not supply_qs.exists():
        return JsonResponse({
            'error': f'No SupplyFactors data found for {facility.facility_name} in {year}'
        }, status=404)

    supply_data = aligner.get_supply_data_as_dict(supply_qs, start_hour_int, end_hour_int)

    # Align the datasets
    aligned = aligner.align_scada_and_supply(scada_hourly, supply_data)

    if not aligned:
        return JsonResponse({
            'error': f'No overlapping data found for {facility.facility_name} in {year}'
        }, status=404)

    # Apply peak hours filter if specified
    if peak_filter and peak_filter != 'all':
        aligned = aligner.filter_aligned_to_peak_hours(aligned, peak_filter)
        if not aligned:
            return JsonResponse({
                'error': f'No data found for {facility.facility_name} in {year} after applying {peak_filter} filter'
            }, status=404)

    # Aggregate by period if needed
    if aggregation != 'hour':
        # Create hour_data format for aggregation
        hour_data = [
            {'hour': h, 'scada': s, 'supply': p}
            for h, s, p in zip(aligned['hours'], aligned['scada_values'], aligned['supply_values'])
        ]
        aggregated = aggregate_multiple_fields(hour_data, aggregation, ['scada', 'supply'])
        periods = aggregated['periods']
        scada_values = aggregated['scada']
        supply_values = aggregated['supply']
    else:
        periods = aligned['hours']
        scada_values = aligned['scada_values']
        supply_values = aligned['supply_values']

    # Calculate metrics
    correlation_metrics = calculate_correlation_metrics(scada_values, supply_values)
    error_metrics = calculate_error_metrics(scada_values, supply_values)

    # Build response
    response_data = {
        'facility_name': facility.facility_name,
        'facility_code': facility.facility_code,
        'technology': facility.idtechnologies.technology_name,
        'year': year,
        'aggregation': aggregation,
        'x_label': get_x_label(aggregation),
        'periods': periods,
        'scada_values': scada_values,
        'supply_values': supply_values,
        'correlation_metrics': correlation_metrics,
        'error_metrics': error_metrics,
        'overlap_percentage': aligned['overlap_percentage'],
        'data_points': len(periods),
        'scada_only_hours': aligned['scada_only_hours'],
        'supply_only_hours': aligned['supply_only_hours'],
    }

    # Add peak filter info if applied
    if peak_filter and peak_filter != 'all':
        response_data['peak_filter'] = peak_filter
        response_data['peak_filter_label'] = aligned.get('peak_filter_label', '')
        response_data['filter_retention_pct'] = aligned.get('filter_retention_pct', 100)

    return JsonResponse(response_data)


def get_facility_group_scada_vs_supply(request):
    """API endpoint to compare SCADA vs SupplyFactors for a group of facilities."""
    facility_ids = request.GET.getlist('facility_id[]')
    year = request.GET.get('year')
    aggregation = request.GET.get('aggregation', 'hour')
    start_hour = request.GET.get('start_hour')
    end_hour = request.GET.get('end_hour')
    peak_filter = request.GET.get('peak_filter', 'all')  # Peak hours filter

    # Validation
    if not facility_ids or len(facility_ids) == 0:
        return JsonResponse({'error': 'At least one facility must be selected'}, status=400)
    if not year:
        return JsonResponse({'error': 'year is required'}, status=400)

    try:
        selected_facilities = facilities.objects.filter(
            idfacilities__in=facility_ids
        ).select_related('idtechnologies')

        if not selected_facilities.exists():
            return JsonResponse({'error': 'No valid facilities found'}, status=404)

        year = int(year)
    except ValueError:
        return JsonResponse({'error': 'Invalid year format'}, status=400)

    # Parse hour range
    start_hour_int = None
    end_hour_int = None
    if start_hour and end_hour:
        try:
            start_hour_int = int(start_hour)
            end_hour_int = int(end_hour)
        except ValueError:
            return JsonResponse({'error': 'Invalid hour range'}, status=400)

    # Get SCADA data aggregated across facilities
    scada_qs = FacilityScada.objects.filter(
        facility__in=selected_facilities,
        dispatch_interval__year=year
    ).order_by('dispatch_interval')

    if not scada_qs.exists():
        return JsonResponse({
            'error': f'No SCADA data found for selected facilities in {year}'
        }, status=404)

    scada_hourly = aligner.convert_scada_to_hourly_aggregated(
        scada_qs, year, start_hour_int, end_hour_int
    )

    # Get SupplyFactors data aggregated across facilities
    supply_qs = supplyfactors.objects.filter(
        idfacilities__in=selected_facilities,
        year=year
    )
    if start_hour_int and end_hour_int:
        supply_qs = supply_qs.filter(hour__gte=start_hour_int, hour__lte=end_hour_int)

    if not supply_qs.exists():
        return JsonResponse({
            'error': f'No SupplyFactors data found for selected facilities in {year}'
        }, status=404)

    supply_data = aligner.get_supply_data_aggregated(supply_qs, start_hour_int, end_hour_int)

    # Align the datasets
    aligned = aligner.align_scada_and_supply(scada_hourly, supply_data)

    if not aligned:
        return JsonResponse({
            'error': f'No overlapping data found for selected facilities in {year}'
        }, status=404)

    # Apply peak hours filter if specified
    if peak_filter and peak_filter != 'all':
        aligned = aligner.filter_aligned_to_peak_hours(aligned, peak_filter)
        if not aligned:
            return JsonResponse({
                'error': f'No data found for selected facilities in {year} after applying {peak_filter} filter'
            }, status=404)

    # Aggregate by period if needed
    if aggregation != 'hour':
        hour_data = [
            {'hour': h, 'scada': s, 'supply': p}
            for h, s, p in zip(aligned['hours'], aligned['scada_values'], aligned['supply_values'])
        ]
        aggregated = aggregate_multiple_fields(hour_data, aggregation, ['scada', 'supply'])
        periods = aggregated['periods']
        scada_values = aggregated['scada']
        supply_values = aggregated['supply']
    else:
        periods = aligned['hours']
        scada_values = aligned['scada_values']
        supply_values = aligned['supply_values']

    # Calculate metrics
    correlation_metrics = calculate_correlation_metrics(scada_values, supply_values)
    error_metrics = calculate_error_metrics(scada_values, supply_values)

    # Get facility names and technology breakdown
    facility_names = list(selected_facilities.values_list('facility_name', flat=True))
    tech_breakdown = []
    technologies = selected_facilities.values_list('idtechnologies__technology_name', flat=True).distinct()
    for tech_name in technologies:
        count = selected_facilities.filter(idtechnologies__technology_name=tech_name).count()
        tech_breakdown.append({'name': tech_name, 'facility_count': count})

    response_data = {
        'facility_names': facility_names[:10],  # Limit display
        'facility_count': selected_facilities.count(),
        'technology_breakdown': tech_breakdown,
        'year': year,
        'aggregation': aggregation,
        'x_label': get_x_label(aggregation),
        'periods': periods,
        'scada_values': scada_values,
        'supply_values': supply_values,
        'correlation_metrics': correlation_metrics,
        'error_metrics': error_metrics,
        'overlap_percentage': aligned['overlap_percentage'],
        'data_points': len(periods),
    }

    # Add peak filter info if applied
    if peak_filter and peak_filter != 'all':
        response_data['peak_filter'] = peak_filter
        response_data['peak_filter_label'] = aligned.get('peak_filter_label', '')
        response_data['filter_retention_pct'] = aligned.get('filter_retention_pct', 100)

    return JsonResponse(response_data)


def get_technology_scada_vs_supply(request):
    """API endpoint to compare SCADA vs SupplyFactors aggregated by technology."""
    technology_ids = request.GET.getlist('technology_id[]')
    year = request.GET.get('year')
    aggregation = request.GET.get('aggregation', 'hour')
    start_hour = request.GET.get('start_hour')
    end_hour = request.GET.get('end_hour')
    peak_filter = request.GET.get('peak_filter', 'all')  # Peak hours filter

    # Validation
    if not technology_ids or len(technology_ids) == 0:
        return JsonResponse({'error': 'At least one technology must be selected'}, status=400)
    if not year:
        return JsonResponse({'error': 'year is required'}, status=400)

    try:
        technologies = Technologies.objects.filter(idtechnologies__in=technology_ids)
        if not technologies.exists():
            return JsonResponse({'error': 'No valid technologies found'}, status=404)

        year = int(year)
    except ValueError:
        return JsonResponse({'error': 'Invalid year format'}, status=400)

    # Parse hour range
    start_hour_int = None
    end_hour_int = None
    if start_hour and end_hour:
        try:
            start_hour_int = int(start_hour)
            end_hour_int = int(end_hour)
        except ValueError:
            return JsonResponse({'error': 'Invalid hour range'}, status=400)

    # Get facilities with these technologies
    tech_facilities = facilities.objects.filter(idtechnologies__in=technologies)
    facility_count = tech_facilities.count()

    if facility_count == 0:
        return JsonResponse({
            'error': 'No facilities found for selected technologies'
        }, status=404)

    # Get SCADA data aggregated across facilities
    scada_qs = FacilityScada.objects.filter(
        facility__in=tech_facilities,
        dispatch_interval__year=year
    ).order_by('dispatch_interval')

    if not scada_qs.exists():
        tech_names = ', '.join(technologies.values_list('technology_name', flat=True))
        return JsonResponse({
            'error': f'No SCADA data found for {tech_names} in {year}'
        }, status=404)

    scada_hourly = aligner.convert_scada_to_hourly_aggregated(
        scada_qs, year, start_hour_int, end_hour_int
    )

    # Get SupplyFactors data aggregated
    supply_qs = supplyfactors.objects.filter(
        idfacilities__idtechnologies__in=technologies,
        year=year
    )
    if start_hour_int and end_hour_int:
        supply_qs = supply_qs.filter(hour__gte=start_hour_int, hour__lte=end_hour_int)

    if not supply_qs.exists():
        tech_names = ', '.join(technologies.values_list('technology_name', flat=True))
        return JsonResponse({
            'error': f'No SupplyFactors data found for {tech_names} in {year}'
        }, status=404)

    supply_data = aligner.get_supply_data_aggregated(supply_qs, start_hour_int, end_hour_int)

    # Align the datasets
    aligned = aligner.align_scada_and_supply(scada_hourly, supply_data)

    if not aligned:
        return JsonResponse({
            'error': f'No overlapping data found for selected technologies in {year}'
        }, status=404)

    # Apply peak hours filter if specified
    if peak_filter and peak_filter != 'all':
        aligned = aligner.filter_aligned_to_peak_hours(aligned, peak_filter)
        if not aligned:
            return JsonResponse({
                'error': f'No data found for selected technologies in {year} after applying {peak_filter} filter'
            }, status=404)

    # Aggregate by period if needed
    if aggregation != 'hour':
        hour_data = [
            {'hour': h, 'scada': s, 'supply': p}
            for h, s, p in zip(aligned['hours'], aligned['scada_values'], aligned['supply_values'])
        ]
        aggregated = aggregate_multiple_fields(hour_data, aggregation, ['scada', 'supply'])
        periods = aggregated['periods']
        scada_values = aggregated['scada']
        supply_values = aggregated['supply']
    else:
        periods = aligned['hours']
        scada_values = aligned['scada_values']
        supply_values = aligned['supply_values']

    # Calculate metrics
    correlation_metrics = calculate_correlation_metrics(scada_values, supply_values)
    error_metrics = calculate_error_metrics(scada_values, supply_values)

    # Get technology breakdown
    technology_names = list(technologies.values_list('technology_name', flat=True))
    tech_breakdown = []
    for tech in technologies:
        count = tech_facilities.filter(idtechnologies=tech).count()
        tech_breakdown.append({'name': tech.technology_name, 'facility_count': count})

    response_data = {
        'technology_names': technology_names,
        'technology_breakdown': tech_breakdown,
        'facility_count': facility_count,
        'year': year,
        'aggregation': aggregation,
        'x_label': get_x_label(aggregation),
        'periods': periods,
        'scada_values': scada_values,
        'supply_values': supply_values,
        'correlation_metrics': correlation_metrics,
        'error_metrics': error_metrics,
        'overlap_percentage': aligned['overlap_percentage'],
        'data_points': len(periods),
    }

    # Add peak filter info if applied
    if peak_filter and peak_filter != 'all':
        response_data['peak_filter'] = peak_filter
        response_data['peak_filter_label'] = aligned.get('peak_filter_label', '')
        response_data['filter_retention_pct'] = aligned.get('filter_retention_pct', 100)

    return JsonResponse(response_data)


def get_technology_group_scada_vs_supply(request):
    """API endpoint to compare two technology groups, each with SCADA vs SupplyFactors."""
    technology1_ids = request.GET.getlist('technology1_id[]')
    technology2_ids = request.GET.getlist('technology2_id[]')
    year = request.GET.get('year')
    aggregation = request.GET.get('aggregation', 'hour')
    start_hour = request.GET.get('start_hour')
    end_hour = request.GET.get('end_hour')

    # Validation
    if not technology1_ids or len(technology1_ids) == 0:
        return JsonResponse({
            'error': 'At least one technology must be selected for Group 1'
        }, status=400)
    if not technology2_ids or len(technology2_ids) == 0:
        return JsonResponse({
            'error': 'At least one technology must be selected for Group 2'
        }, status=400)
    if not year:
        return JsonResponse({'error': 'year is required'}, status=400)

    try:
        technologies1 = Technologies.objects.filter(idtechnologies__in=technology1_ids)
        technologies2 = Technologies.objects.filter(idtechnologies__in=technology2_ids)

        if not technologies1.exists():
            return JsonResponse({'error': 'No valid technologies found for Group 1'}, status=404)
        if not technologies2.exists():
            return JsonResponse({'error': 'No valid technologies found for Group 2'}, status=404)

        year = int(year)
    except ValueError:
        return JsonResponse({'error': 'Invalid year format'}, status=400)

    # Parse hour range
    start_hour_int = None
    end_hour_int = None
    if start_hour and end_hour:
        try:
            start_hour_int = int(start_hour)
            end_hour_int = int(end_hour)
        except ValueError:
            return JsonResponse({'error': 'Invalid hour range'}, status=400)

    # Helper to get data for a technology group
    def get_tech_group_data(technologies):
        tech_facilities = facilities.objects.filter(idtechnologies__in=technologies)
        if not tech_facilities.exists():
            return None, None, "No facilities found"

        # SCADA
        scada_qs = FacilityScada.objects.filter(
            facility__in=tech_facilities,
            dispatch_interval__year=year
        ).order_by('dispatch_interval')

        if not scada_qs.exists():
            return None, None, "No SCADA data"

        scada_hourly = aligner.convert_scada_to_hourly_aggregated(
            scada_qs, year, start_hour_int, end_hour_int
        )

        # SupplyFactors
        supply_qs = supplyfactors.objects.filter(
            idfacilities__idtechnologies__in=technologies,
            year=year
        )
        if start_hour_int and end_hour_int:
            supply_qs = supply_qs.filter(hour__gte=start_hour_int, hour__lte=end_hour_int)

        if not supply_qs.exists():
            return None, None, "No SupplyFactors data"

        supply_data = aligner.get_supply_data_aggregated(supply_qs, start_hour_int, end_hour_int)

        # Align
        aligned = aligner.align_scada_and_supply(scada_hourly, supply_data)
        if not aligned:
            return None, None, "No overlapping data"

        return aligned, tech_facilities, None

    # Get data for both groups
    aligned1, facilities1, error1 = get_tech_group_data(technologies1)
    aligned2, facilities2, error2 = get_tech_group_data(technologies2)

    if error1:
        tech_names = ', '.join(technologies1.values_list('technology_name', flat=True))
        return JsonResponse({
            'error': f'{error1} for {tech_names} in {year}'
        }, status=404)

    if error2:
        tech_names = ', '.join(technologies2.values_list('technology_name', flat=True))
        return JsonResponse({
            'error': f'{error2} for {tech_names} in {year}'
        }, status=404)

    # Find common hours across all data
    common_hours = sorted(set(aligned1['hours']) & set(aligned2['hours']))
    if not common_hours:
        return JsonResponse({
            'error': 'No common time periods between the two technology groups'
        }, status=404)

    # Extract values for common hours
    def get_values_for_hours(aligned, hours):
        lookup_scada = dict(zip(aligned['hours'], aligned['scada_values']))
        lookup_supply = dict(zip(aligned['hours'], aligned['supply_values']))
        return (
            [lookup_scada[h] for h in hours],
            [lookup_supply[h] for h in hours]
        )

    scada1, supply1 = get_values_for_hours(aligned1, common_hours)
    scada2, supply2 = get_values_for_hours(aligned2, common_hours)

    # Aggregate by period if needed
    if aggregation != 'hour':
        hour_data = [
            {'hour': h, 'scada1': s1, 'supply1': p1, 'scada2': s2, 'supply2': p2}
            for h, s1, p1, s2, p2 in zip(common_hours, scada1, supply1, scada2, supply2)
        ]
        aggregated = aggregate_multiple_fields(
            hour_data, aggregation, ['scada1', 'supply1', 'scada2', 'supply2']
        )
        periods = aggregated['periods']
        scada1, supply1 = aggregated['scada1'], aggregated['supply1']
        scada2, supply2 = aggregated['scada2'], aggregated['supply2']
    else:
        periods = common_hours

    # Calculate metrics for each group
    correlation1 = calculate_correlation_metrics(scada1, supply1)
    error_metrics1 = calculate_error_metrics(scada1, supply1)
    correlation2 = calculate_correlation_metrics(scada2, supply2)
    error_metrics2 = calculate_error_metrics(scada2, supply2)

    # Build response
    tech1_names = list(technologies1.values_list('technology_name', flat=True))
    tech2_names = list(technologies2.values_list('technology_name', flat=True))

    return JsonResponse({
        'technology_group1': {
            'names': tech1_names,
            'facility_count': facilities1.count(),
            'scada_values': scada1,
            'supply_values': supply1,
            'correlation_metrics': correlation1,
            'error_metrics': error_metrics1,
        },
        'technology_group2': {
            'names': tech2_names,
            'facility_count': facilities2.count(),
            'scada_values': scada2,
            'supply_values': supply2,
            'correlation_metrics': correlation2,
            'error_metrics': error_metrics2,
        },
        'year': year,
        'aggregation': aggregation,
        'x_label': get_x_label(aggregation),
        'periods': periods,
        'data_points': len(periods),
    })


def export_generation_comparison_to_excel(request):
    """Export SCADA vs SupplyFactors comparison data to Excel."""
    export_type = request.GET.get('type', 'single')  # single, group, technology, techcompare
    year = request.GET.get('year')
    aggregation = request.GET.get('aggregation', 'hour')
    start_hour = request.GET.get('start_hour')
    end_hour = request.GET.get('end_hour')
    peak_filter = request.GET.get('peak_filter', 'all')

    if not year:
        return JsonResponse({'error': 'Year is required'}, status=400)

    try:
        year = int(year)
    except ValueError:
        return JsonResponse({'error': 'Invalid year format'}, status=400)

    # Parse hour range
    start_hour_int = None
    end_hour_int = None
    if start_hour and end_hour:
        try:
            start_hour_int = int(start_hour)
            end_hour_int = int(end_hour)
        except ValueError:
            pass

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    if export_type == 'single':
        facility_id = request.GET.get('facility_id')
        if not facility_id:
            return JsonResponse({'error': 'Facility ID is required'}, status=400)

        try:
            facility = facilities.objects.select_related('idtechnologies').get(idfacilities=facility_id)
        except facilities.DoesNotExist:
            return JsonResponse({'error': 'Facility not found'}, status=404)

        # Get SCADA data
        scada_qs = FacilityScada.objects.filter(
            facility=facility,
            dispatch_interval__year=year
        ).order_by('dispatch_interval')

        scada_hourly = aligner.convert_scada_to_hourly(scada_qs, year, start_hour_int, end_hour_int)

        # Get SupplyFactors data
        supply_qs = supplyfactors.objects.filter(
            idfacilities=facility,
            year=year
        )
        if start_hour_int and end_hour_int:
            supply_qs = supply_qs.filter(hour__gte=start_hour_int, hour__lte=end_hour_int)

        supply_data = aligner.get_supply_data_as_dict(supply_qs, start_hour_int, end_hour_int)

        # Align datasets
        aligned = aligner.align_scada_and_supply(scada_hourly, supply_data)

        if not aligned:
            return JsonResponse({'error': 'No overlapping data found'}, status=404)

        # Apply peak filter
        if peak_filter and peak_filter != 'all':
            aligned = aligner.filter_aligned_to_peak_hours(aligned, peak_filter)
            if not aligned:
                return JsonResponse({'error': 'No data after applying peak filter'}, status=404)

        # Aggregate if needed
        if aggregation != 'hour':
            hour_data = [
                {'hour': h, 'scada': s, 'supply': p}
                for h, s, p in zip(aligned['hours'], aligned['scada_values'], aligned['supply_values'])
            ]
            aggregated = aggregate_multiple_fields(hour_data, aggregation, ['scada', 'supply'])
            periods = aggregated['periods']
            scada_values = aggregated['scada']
            supply_values = aggregated['supply']
        else:
            periods = aligned['hours']
            scada_values = aligned['scada_values']
            supply_values = aligned['supply_values']

        worksheet.title = 'SCADA vs Simulated'

        headers = [get_x_label(aggregation), 'SCADA (Actual)', 'Simulated (SupplyFactors)']
        worksheet.append(headers)

        for i, period in enumerate(periods):
            worksheet.append([period, scada_values[i], supply_values[i]])

        # Add metrics sheet
        metrics_sheet = workbook.create_sheet('Metrics')
        correlation_metrics = calculate_correlation_metrics(scada_values, supply_values)
        error_metrics = calculate_error_metrics(scada_values, supply_values)

        metrics_sheet.append(['Metric', 'Value'])
        metrics_sheet.append(['Facility', facility.facility_name])
        metrics_sheet.append(['Technology', facility.idtechnologies.technology_name])
        metrics_sheet.append(['Year', year])
        metrics_sheet.append(['Aggregation', aggregation])
        metrics_sheet.append(['Peak Filter', peak_filter])
        metrics_sheet.append(['Data Points', len(periods)])
        metrics_sheet.append(['Overlap %', aligned.get('overlap_percentage', '-')])
        metrics_sheet.append([''])
        metrics_sheet.append(['Correlation', correlation_metrics.get('correlation', '-')])
        metrics_sheet.append(['MAE', error_metrics.get('mae', '-')])
        metrics_sheet.append(['RMSE', error_metrics.get('rmse', '-')])
        metrics_sheet.append(['MBE', error_metrics.get('mbe', '-')])
        metrics_sheet.append(['MAPE', error_metrics.get('mape', '-')])

        filename = f"generation_comparison_{facility.facility_code or facility.facility_name}_{year}"

    elif export_type == 'group':
        facility_ids = request.GET.getlist('facility_id[]')
        if not facility_ids:
            return JsonResponse({'error': 'At least one facility must be selected'}, status=400)

        selected_facilities = facilities.objects.filter(
            idfacilities__in=facility_ids
        ).select_related('idtechnologies')

        if not selected_facilities.exists():
            return JsonResponse({'error': 'No valid facilities found'}, status=404)

        # Get aggregated SCADA data
        scada_qs = FacilityScada.objects.filter(
            facility__in=selected_facilities,
            dispatch_interval__year=year
        ).order_by('dispatch_interval')

        scada_hourly = aligner.convert_scada_to_hourly_aggregated(
            scada_qs, year, start_hour_int, end_hour_int
        )

        # Get aggregated SupplyFactors data
        supply_qs = supplyfactors.objects.filter(
            idfacilities__in=selected_facilities,
            year=year
        )
        if start_hour_int and end_hour_int:
            supply_qs = supply_qs.filter(hour__gte=start_hour_int, hour__lte=end_hour_int)

        supply_data = aligner.get_supply_data_aggregated(supply_qs, start_hour_int, end_hour_int)

        # Align datasets
        aligned = aligner.align_scada_and_supply(scada_hourly, supply_data)

        if not aligned:
            return JsonResponse({'error': 'No overlapping data found'}, status=404)

        # Apply peak filter
        if peak_filter and peak_filter != 'all':
            aligned = aligner.filter_aligned_to_peak_hours(aligned, peak_filter)
            if not aligned:
                return JsonResponse({'error': 'No data after applying peak filter'}, status=404)

        # Aggregate if needed
        if aggregation != 'hour':
            hour_data = [
                {'hour': h, 'scada': s, 'supply': p}
                for h, s, p in zip(aligned['hours'], aligned['scada_values'], aligned['supply_values'])
            ]
            aggregated = aggregate_multiple_fields(hour_data, aggregation, ['scada', 'supply'])
            periods = aggregated['periods']
            scada_values = aggregated['scada']
            supply_values = aggregated['supply']
        else:
            periods = aligned['hours']
            scada_values = aligned['scada_values']
            supply_values = aligned['supply_values']

        worksheet.title = 'Facility Group Comparison'

        headers = [get_x_label(aggregation), 'SCADA (Actual)', 'Simulated (SupplyFactors)']
        worksheet.append(headers)

        for i, period in enumerate(periods):
            worksheet.append([period, scada_values[i], supply_values[i]])

        # Add metrics sheet
        metrics_sheet = workbook.create_sheet('Metrics')
        correlation_metrics = calculate_correlation_metrics(scada_values, supply_values)
        error_metrics = calculate_error_metrics(scada_values, supply_values)

        metrics_sheet.append(['Metric', 'Value'])
        metrics_sheet.append(['Facilities', ', '.join(list(selected_facilities.values_list('facility_name', flat=True))[:5])])
        metrics_sheet.append(['Facility Count', selected_facilities.count()])
        metrics_sheet.append(['Year', year])
        metrics_sheet.append(['Aggregation', aggregation])
        metrics_sheet.append(['Peak Filter', peak_filter])
        metrics_sheet.append(['Data Points', len(periods)])
        metrics_sheet.append([''])
        metrics_sheet.append(['Correlation', correlation_metrics.get('correlation', '-')])
        metrics_sheet.append(['MAE', error_metrics.get('mae', '-')])
        metrics_sheet.append(['RMSE', error_metrics.get('rmse', '-')])
        metrics_sheet.append(['MBE', error_metrics.get('mbe', '-')])
        metrics_sheet.append(['MAPE', error_metrics.get('mape', '-')])

        filename = f"generation_comparison_group_{year}"

    elif export_type == 'technology':
        technology_ids = request.GET.getlist('technology_id[]')
        if not technology_ids:
            return JsonResponse({'error': 'At least one technology must be selected'}, status=400)

        technologies = Technologies.objects.filter(idtechnologies__in=technology_ids)
        tech_facilities = facilities.objects.filter(idtechnologies__in=technologies)

        if not tech_facilities.exists():
            return JsonResponse({'error': 'No facilities found for selected technologies'}, status=404)

        # Get aggregated SCADA data
        scada_qs = FacilityScada.objects.filter(
            facility__in=tech_facilities,
            dispatch_interval__year=year
        ).order_by('dispatch_interval')

        scada_hourly = aligner.convert_scada_to_hourly_aggregated(
            scada_qs, year, start_hour_int, end_hour_int
        )

        # Get aggregated SupplyFactors data
        supply_qs = supplyfactors.objects.filter(
            idfacilities__idtechnologies__in=technologies,
            year=year
        )
        if start_hour_int and end_hour_int:
            supply_qs = supply_qs.filter(hour__gte=start_hour_int, hour__lte=end_hour_int)

        supply_data = aligner.get_supply_data_aggregated(supply_qs, start_hour_int, end_hour_int)

        # Align datasets
        aligned = aligner.align_scada_and_supply(scada_hourly, supply_data)

        if not aligned:
            return JsonResponse({'error': 'No overlapping data found'}, status=404)

        # Apply peak filter
        if peak_filter and peak_filter != 'all':
            aligned = aligner.filter_aligned_to_peak_hours(aligned, peak_filter)
            if not aligned:
                return JsonResponse({'error': 'No data after applying peak filter'}, status=404)

        # Aggregate if needed
        if aggregation != 'hour':
            hour_data = [
                {'hour': h, 'scada': s, 'supply': p}
                for h, s, p in zip(aligned['hours'], aligned['scada_values'], aligned['supply_values'])
            ]
            aggregated = aggregate_multiple_fields(hour_data, aggregation, ['scada', 'supply'])
            periods = aggregated['periods']
            scada_values = aggregated['scada']
            supply_values = aggregated['supply']
        else:
            periods = aligned['hours']
            scada_values = aligned['scada_values']
            supply_values = aligned['supply_values']

        worksheet.title = 'Technology Comparison'

        tech_names = ', '.join(list(technologies.values_list('technology_name', flat=True)))
        headers = [get_x_label(aggregation), 'SCADA (Actual)', 'Simulated (SupplyFactors)']
        worksheet.append(headers)

        for i, period in enumerate(periods):
            worksheet.append([period, scada_values[i], supply_values[i]])

        # Add metrics sheet
        metrics_sheet = workbook.create_sheet('Metrics')
        correlation_metrics = calculate_correlation_metrics(scada_values, supply_values)
        error_metrics = calculate_error_metrics(scada_values, supply_values)

        metrics_sheet.append(['Metric', 'Value'])
        metrics_sheet.append(['Technologies', tech_names])
        metrics_sheet.append(['Facility Count', tech_facilities.count()])
        metrics_sheet.append(['Year', year])
        metrics_sheet.append(['Aggregation', aggregation])
        metrics_sheet.append(['Peak Filter', peak_filter])
        metrics_sheet.append(['Data Points', len(periods)])
        metrics_sheet.append([''])
        metrics_sheet.append(['Correlation', correlation_metrics.get('correlation', '-')])
        metrics_sheet.append(['MAE', error_metrics.get('mae', '-')])
        metrics_sheet.append(['RMSE', error_metrics.get('rmse', '-')])
        metrics_sheet.append(['MBE', error_metrics.get('mbe', '-')])
        metrics_sheet.append(['MAPE', error_metrics.get('mape', '-')])

        filename = f"generation_comparison_technology_{year}"

    elif export_type == 'techcompare':
        technology1_ids = request.GET.getlist('technology1_id[]')
        technology2_ids = request.GET.getlist('technology2_id[]')

        if not technology1_ids or not technology2_ids:
            return JsonResponse({'error': 'Both technology groups are required'}, status=400)

        technologies1 = Technologies.objects.filter(idtechnologies__in=technology1_ids)
        technologies2 = Technologies.objects.filter(idtechnologies__in=technology2_ids)

        def get_tech_comparison_data(technologies):
            tech_facilities = facilities.objects.filter(idtechnologies__in=technologies)
            if not tech_facilities.exists():
                return None

            scada_qs = FacilityScada.objects.filter(
                facility__in=tech_facilities,
                dispatch_interval__year=year
            ).order_by('dispatch_interval')

            scada_hourly = aligner.convert_scada_to_hourly_aggregated(
                scada_qs, year, start_hour_int, end_hour_int
            )

            supply_qs = supplyfactors.objects.filter(
                idfacilities__idtechnologies__in=technologies,
                year=year
            )
            if start_hour_int and end_hour_int:
                supply_qs = supply_qs.filter(hour__gte=start_hour_int, hour__lte=end_hour_int)

            supply_data = aligner.get_supply_data_aggregated(supply_qs, start_hour_int, end_hour_int)

            aligned = aligner.align_scada_and_supply(scada_hourly, supply_data)
            return aligned, tech_facilities.count()

        result1 = get_tech_comparison_data(technologies1)
        result2 = get_tech_comparison_data(technologies2)

        if not result1 or not result1[0]:
            return JsonResponse({'error': 'No data for technology group 1'}, status=404)
        if not result2 or not result2[0]:
            return JsonResponse({'error': 'No data for technology group 2'}, status=404)

        aligned1, count1 = result1
        aligned2, count2 = result2

        # Find common hours
        common_hours = sorted(set(aligned1['hours']) & set(aligned2['hours']))
        if not common_hours:
            return JsonResponse({'error': 'No common time periods'}, status=404)

        # Extract values for common hours
        def get_values_for_hours(aligned, hours):
            lookup_scada = dict(zip(aligned['hours'], aligned['scada_values']))
            lookup_supply = dict(zip(aligned['hours'], aligned['supply_values']))
            return [lookup_scada[h] for h in hours], [lookup_supply[h] for h in hours]

        scada1, supply1 = get_values_for_hours(aligned1, common_hours)
        scada2, supply2 = get_values_for_hours(aligned2, common_hours)

        # Aggregate if needed
        if aggregation != 'hour':
            hour_data = [
                {'hour': h, 'scada1': s1, 'supply1': p1, 'scada2': s2, 'supply2': p2}
                for h, s1, p1, s2, p2 in zip(common_hours, scada1, supply1, scada2, supply2)
            ]
            aggregated = aggregate_multiple_fields(
                hour_data, aggregation, ['scada1', 'supply1', 'scada2', 'supply2']
            )
            periods = aggregated['periods']
            scada1, supply1 = aggregated['scada1'], aggregated['supply1']
            scada2, supply2 = aggregated['scada2'], aggregated['supply2']
        else:
            periods = common_hours

        worksheet.title = 'Tech Group Comparison'

        tech1_names = ', '.join(list(technologies1.values_list('technology_name', flat=True)))
        tech2_names = ', '.join(list(technologies2.values_list('technology_name', flat=True)))

        headers = [
            get_x_label(aggregation),
            f'Group 1 SCADA ({tech1_names})',
            f'Group 1 Simulated ({tech1_names})',
            f'Group 2 SCADA ({tech2_names})',
            f'Group 2 Simulated ({tech2_names})'
        ]
        worksheet.append(headers)

        for i, period in enumerate(periods):
            worksheet.append([period, scada1[i], supply1[i], scada2[i], supply2[i]])

        # Add metrics sheet
        metrics_sheet = workbook.create_sheet('Metrics')
        corr1 = calculate_correlation_metrics(scada1, supply1)
        error1 = calculate_error_metrics(scada1, supply1)
        corr2 = calculate_correlation_metrics(scada2, supply2)
        error2 = calculate_error_metrics(scada2, supply2)

        metrics_sheet.append(['Metric', 'Group 1', 'Group 2'])
        metrics_sheet.append(['Technologies', tech1_names, tech2_names])
        metrics_sheet.append(['Facility Count', count1, count2])
        metrics_sheet.append(['Year', year, year])
        metrics_sheet.append(['Data Points', len(periods), len(periods)])
        metrics_sheet.append([''])
        metrics_sheet.append(['Correlation', corr1.get('correlation', '-'), corr2.get('correlation', '-')])
        metrics_sheet.append(['MAE', error1.get('mae', '-'), error2.get('mae', '-')])
        metrics_sheet.append(['RMSE', error1.get('rmse', '-'), error2.get('rmse', '-')])
        metrics_sheet.append(['MBE', error1.get('mbe', '-'), error2.get('mbe', '-')])
        metrics_sheet.append(['MAPE', error1.get('mape', '-'), error2.get('mape', '-')])

        filename = f"generation_comparison_techgroups_{year}"

    else:
        return JsonResponse({'error': 'Invalid export type'}, status=400)

    # Auto-size columns for all sheets
    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column].width = adjusted_width

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    workbook.save(response)

    return response
